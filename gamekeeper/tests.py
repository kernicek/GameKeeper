"""Tests for the Mastersheet importers (DESIGN §12), the sleeves module (§5),
the purchases module (§6), the BGG sync engine (§8) and the §11 email
reminders."""

import datetime
import re
import smtplib
import tempfile
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path, PurePosixPath
from unittest import mock

import openpyxl
import requests
from PIL import Image

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core import mail
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import IntegrityError, transaction
from django.template.loader import render_to_string
from django.test import SimpleTestCase, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from gamekeeper.bgg import (
    BggAuthError, BggClient, BggError, extract_bgg_id, parse_collection,
    parse_collection_error, parse_collection_status_flags, parse_geekitem,
    parse_plays, parse_plays_error, parse_things,
)
from gamekeeper.bgg_sync import (
    IMPORT_ACTION_ARCHIVED, IMPORT_ACTION_COPY, IMPORT_ACTION_PREORDER,
    IMPORT_ACTION_WISHLIST, PUSH_CONFIRM_WINDOW, _status_from_flags,
    bgg_credentials_error, fetch_plays, group_candidates_by_action,
    import_action_from_flags, make_bgg_client, push_bgg_fortrade,
    push_bgg_status, push_is_pending, resolve_bgg_credentials, store_plays,
    sync_game,
)
from gamekeeper import crypto
from gamekeeper.context_processors import environment
from gamekeeper.cover_preview import PREVIEW_SIZE, render_square_preview

from gamekeeper.management.commands.import_sleeves import parse_dimensions
from gamekeeper.ntfy import send_ntfy
from gamekeeper.models import (
    Accessory,
    AccessoryCopy,
    AlternateName,
    BggLink,
    BggSyncDiff,
    CardSize,
    Copy,
    CopySleeveStatus,
    DigitalImplementation,
    Document,
    Edition,
    ExpansionSighting,
    ExternalLink,
    Family,
    Game,
    GameType,
    GameTag,
    Group,
    Invite,
    Loan,
    Location,
    Membership,
    PledgeManager,
    PledgePlan,
    PledgePlanBundle,
    PledgePlanItem,
    Play,
    PlayPlayer,
    Product,
    ProductSleeveRequirement,
    Purchase,
    ReminderLog,
    Series,
    ShareGrant,
    SleeveInventory,
    SleeveProduct,
    SleeveRequirement,
    Tag,
    ToolRun,
    Wave,
    WishlistEntry,
    sleeve_shortfall,
)
from gamekeeper.tasks import (
    push_bgg_fortrade_task, push_bgg_status_task, run_tool_command,
    send_reminder_emails,
)
from gamekeeper.views import INCOMING_STATUSES


def build_sheet(rows):
    """Write a minimal Overview workbook (data starts at row 4) to a temp file.

    Each row is a {column_number: value} dict; a (value, url) tuple sets the
    cell hyperlink the way the real sheet carries its BGG/CF/Drive links.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Overview"
    sheet.cell(row=1, column=2, value="Game")  # header rows 1-3 are skipped
    for offset, row in enumerate(rows):
        for col, value in row.items():
            cell = sheet.cell(row=4 + offset, column=col)
            if isinstance(value, tuple):
                cell.value, cell.hyperlink = value
            else:
                cell.value = value
    path = Path(tempfile.mkdtemp()) / "mastersheet.xlsx"
    workbook.save(path)
    return path


GAME_ROW = {
    2: "5 Minute Dungeon",
    3: ("BGG", "https://boardgamegeek.com/boardgame/207830/5-minute-dungeon"),
    4: ("CF", "https://www.kickstarter.com/projects/wiggles3d/5-minute-dungeon"),
    14: "Kickstarter Edition",
    16: "y",
    20: "M (Medium)",
    21: 1,
    216: "1 - Always Keep",
    218: 9,
    233: "Included",
    235: "To-do, nicer ones",
}

PNP_ROW = {
    2: "Lilypads (PnP)",
    3: ("BGG", "https://boardgamegeek.com/boardgame/329465/lilypads"),
    14: "PnP",
    15: "PnP",
}

# A self-built insert, as opposed to GAME_ROW's manufacturer-included one —
# issue #23: "Included" and "Done" must map to distinct UpgradeStatus values.
DONE_INSERT_ROW = {
    2: "Gloomhaven",
    3: ("BGG", "https://boardgamegeek.com/boardgame/174430/gloomhaven"),
    233: "Done",
}

JUNK_ROW = {2: "1"}  # the legend row: a title, no BGG link


class ExtractBggIdTests(TestCase):
    def test_extracts_id_from_thing_urls(self):
        self.assertEqual(
            extract_bgg_id("https://boardgamegeek.com/boardgame/207830/5-minute-dungeon"),
            207830,
        )
        self.assertEqual(
            extract_bgg_id("https://boardgamegeek.com/boardgameexpansion/12345/foo"),
            12345,
        )

    def test_rejects_non_bgg_urls(self):
        self.assertIsNone(extract_bgg_id(None))
        self.assertIsNone(extract_bgg_id("https://www.kickstarter.com/projects/x/y"))
        self.assertIsNone(extract_bgg_id("https://boardgamegeek.com/browse/boardgame"))


class ImportMastersheetTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")

    def run_import(self, path, **extra):
        out = StringIO()
        call_command("import_mastersheet", str(path), user="kernicek", stdout=out, **extra)
        return out.getvalue()

    def test_imports_game_edition_copy(self):
        output = self.run_import(build_sheet([GAME_ROW]))

        game = Game.objects.get()
        self.assertEqual(game.name, "5 Minute Dungeon")
        self.assertEqual(game.primary_bgg_link.bgg_id, 207830)
        link = ExternalLink.objects.get()
        self.assertEqual(link.link_type, ExternalLink.LinkType.KICKSTARTER)

        edition = Edition.objects.get()
        self.assertEqual(edition.name, "Kickstarter Edition")
        self.assertFalse(edition.is_default)
        self.assertFalse(edition.is_pnp)
        self.assertEqual(edition.size_category, Edition.SizeCategory.MEDIUM)
        self.assertEqual(edition.num_boxes, 1)

        copy = Copy.objects.get()
        self.assertEqual(copy.owner, self.user)
        self.assertEqual(copy.excitement, 9)
        self.assertEqual(copy.keep_status, Copy.KeepStatus.ALWAYS_KEEP)
        self.assertEqual(copy.insert_3d, Copy.UpgradeStatus.INCLUDED)
        self.assertEqual(copy.card_dividers, Copy.UpgradeStatus.TODO)
        self.assertIn("card dividers: To-do, nicer ones", copy.upgrades_note)
        self.assertEqual(copy.location.name, "Chynice")
        self.assertEqual(copy.location.group, Membership.objects.get(user=self.user).group)
        self.assertIn("games created: 1", output)

    def test_included_and_done_inserts_map_to_distinct_statuses(self):
        # Issue #23: "Included" (manufacturer-included) and "Done" (self-built)
        # used to collapse into the same UpgradeStatus.DONE value.
        self.run_import(build_sheet([DONE_INSERT_ROW]))

        copy = Copy.objects.get()
        self.assertEqual(copy.insert_3d, Copy.UpgradeStatus.DONE)

    def test_junk_row_without_bgg_link_is_skipped(self):
        output = self.run_import(build_sheet([JUNK_ROW, GAME_ROW]))

        self.assertEqual(Game.objects.count(), 1)
        self.assertIn("row 4 ('1'): no BGG link", output)

    def test_pnp_column_and_title_suffix(self):
        self.run_import(build_sheet([PNP_ROW]))

        game = Game.objects.get()
        self.assertEqual(game.name, "Lilypads")  # " (PnP)" stripped
        # PnP is an edition-level flag (#138): the literal "PnP" edition name is
        # a duplicate of it, not a name, so the edition stays the blank default.
        edition = Edition.objects.get()
        self.assertEqual(edition.name, "")
        self.assertTrue(edition.is_default)
        self.assertTrue(edition.is_pnp)
        self.assertTrue(game.has_pnp_edition)

    def test_pnp_inferred_from_title_when_column_empty(self):
        row = dict(PNP_ROW)
        del row[15]
        output = self.run_import(build_sheet([row]))

        self.assertTrue(Edition.objects.get().is_pnp)
        self.assertIn("inferred from title", output)

    def test_reimport_is_idempotent(self):
        path = build_sheet([GAME_ROW, PNP_ROW, JUNK_ROW])
        self.run_import(path)
        counts = {
            model: model.objects.count()
            for model in (Game, BggLink, Edition, Copy, ExternalLink)
        }
        self.assertEqual(counts[Game], 2)

        output = self.run_import(path)
        for model, count in counts.items():
            self.assertEqual(model.objects.count(), count)
        self.assertIn("games updated: 2", output)

    def test_dry_run_writes_nothing(self):
        output = self.run_import(build_sheet([GAME_ROW]), dry_run=True)

        self.assertEqual(Game.objects.count(), 0)
        self.assertEqual(Copy.objects.count(), 0)
        self.assertIn("DRY RUN", output)
        self.assertIn("games created: 1", output)


# ===========================================================================
# §5  Sleeves
# ===========================================================================

class SleeveModelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)
        cls.copy = Copy.objects.create(owner=cls.user, edition=cls.edition)
        cls.medium = CardSize.objects.create(
            width_mm=Decimal("57.5"), height_mm=Decimal("89.0"), name="Medium",
        )

    def test_card_size_dimensions_are_unique(self):
        with self.assertRaises(IntegrityError):
            CardSize.objects.create(
                width_mm=Decimal("57.5"), height_mm=Decimal("89.0"), name="Duplicate",
            )

    def test_one_requirement_per_edition_and_size(self):
        SleeveRequirement.objects.create(edition=self.edition, card_size=self.medium, count=365)
        with self.assertRaises(IntegrityError):
            SleeveRequirement.objects.create(edition=self.edition, card_size=self.medium, count=1)

    def test_one_status_per_copy_and_size(self):
        CopySleeveStatus.objects.create(copy=self.copy, card_size=self.medium)
        with self.assertRaises(IntegrityError):
            CopySleeveStatus.objects.create(copy=self.copy, card_size=self.medium)

    def test_shortfall_rounds_up_to_whole_packs(self):
        product = SleeveProduct.objects.create(
            brand="Tlama", name="Diamond Orange", card_size=self.medium,
        )
        SleeveRequirement.objects.create(edition=self.edition, card_size=self.medium, count=365)
        CopySleeveStatus.objects.create(
            copy=self.copy, card_size=self.medium,
            status=CopySleeveStatus.Status.TO_SLEEVE,
        )
        SleeveInventory.objects.create(owner=self.user, product=product, packs=1, loose=50)

        (entry,) = sleeve_shortfall(self.user)
        self.assertEqual(entry["card_size"], self.medium)
        self.assertEqual(entry["to_sleeve"], 365)
        self.assertEqual(entry["in_inventory"], 150)  # 1 pack of 100 + 50 loose
        self.assertEqual(entry["shortfall"], 215)
        self.assertEqual(entry["packs_to_buy"], 3)    # ceil(215 / 100)

    def test_shortfall_breaks_down_to_sleeve_count_by_game(self):
        # Issue #93: a shared card size should list which games contribute.
        other_game = Game.objects.create(name="Ark Nova")
        other_edition = Edition.objects.create(game=other_game, is_default=True)
        other_copy = Copy.objects.create(owner=self.user, edition=other_edition)

        SleeveRequirement.objects.create(edition=self.edition, card_size=self.medium, count=100)
        SleeveRequirement.objects.create(edition=other_edition, card_size=self.medium, count=200)
        CopySleeveStatus.objects.create(
            copy=self.copy, card_size=self.medium,
            status=CopySleeveStatus.Status.TO_SLEEVE,
        )
        CopySleeveStatus.objects.create(
            copy=other_copy, card_size=self.medium,
            status=CopySleeveStatus.Status.TO_SLEEVE,
        )

        (entry,) = sleeve_shortfall(self.user)
        self.assertEqual(entry["to_sleeve"], 300)
        self.assertEqual(entry["games"], [
            {"label": "Ark Nova", "game_pk": other_game.pk, "to_sleeve": 200},
            {"label": "5 Minute Dungeon", "game_pk": self.game.pk, "to_sleeve": 100},
        ])

    def test_shortfall_ignores_sleeved_and_archived_copies(self):
        SleeveRequirement.objects.create(edition=self.edition, card_size=self.medium, count=365)
        CopySleeveStatus.objects.create(
            copy=self.copy, card_size=self.medium,
            status=CopySleeveStatus.Status.SLEEVED,
        )
        self.assertEqual(sleeve_shortfall(self.user), [])

        status = self.copy.sleeve_statuses.get()
        status.status = CopySleeveStatus.Status.TO_SLEEVE
        status.save()
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.save()
        self.assertEqual(sleeve_shortfall(self.user), [])

    def test_shortfall_ignores_not_ready_copies(self):
        # Issue #19: an unprinted PnP copy has nothing to sleeve yet.
        SleeveRequirement.objects.create(edition=self.edition, card_size=self.medium, count=365)
        CopySleeveStatus.objects.create(
            copy=self.copy, card_size=self.medium,
            status=CopySleeveStatus.Status.TO_SLEEVE,
        )
        self.copy.ready_status = Copy.ReadyStatus.NOT_READY
        self.copy.save()
        self.assertEqual(sleeve_shortfall(self.user), [])


class AccessoryModelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)

    def test_standalone_accessory_has_no_game_or_edition(self):
        accessory = Accessory.objects.create(name="Generic Neoprene Playmat")
        self.assertIsNone(accessory.game)
        self.assertIsNone(accessory.edition)

    def test_game_linked_accessory(self):
        accessory = Accessory.objects.create(
            name="Burgle Bros 2 Playmat", game=self.game, bgg_id=343021,
        )
        self.assertEqual(accessory.game, self.game)
        self.assertIn(accessory, self.game.accessories.all())

    def test_edition_linked_accessory(self):
        accessory = Accessory.objects.create(name="Collector's Insert", edition=self.edition)
        self.assertEqual(accessory.edition, self.edition)
        self.assertIn(accessory, self.edition.accessories.all())

    def test_game_and_edition_are_mutually_exclusive(self):
        with self.assertRaises(IntegrityError):
            Accessory.objects.create(name="Bad Row", game=self.game, edition=self.edition)

    def test_bgg_id_is_unique(self):
        Accessory.objects.create(name="Burgle Bros 2 Playmat", bgg_id=343021)
        with self.assertRaises(IntegrityError):
            Accessory.objects.create(name="Duplicate", bgg_id=343021)

    def test_accessory_copy_ownership(self):
        accessory = Accessory.objects.create(name="Upgraded Tokens", game=self.game)
        copy = AccessoryCopy.objects.create(owner=self.user, accessory=accessory)
        self.assertEqual(copy.owner, self.user)
        self.assertEqual(copy.accessory, accessory)
        self.assertIn(copy, accessory.copies.all())
        self.assertIn(copy, self.user.accessory_copies.all())

    def test_product_accessory_copy_defaults_to_none(self):
        purchase = Purchase.objects.create(owner=self.user, name="Some Campaign")
        wave = Wave.objects.create(purchase=purchase, number=1)
        product = Product.objects.create(
            wave=wave, name="Playmat", kind=Product.Kind.ACCESSORY,
        )
        self.assertIsNone(product.accessory_copy)
        self.assertIsNone(product.game)

    def test_product_accessory_copy_can_be_linked_without_a_game(self):
        # Unlike Product.copy (which the product_convert view gates on
        # product.game_id), accessory_copy is a plain admin-set field with
        # no such requirement.
        accessory = Accessory.objects.create(name="Upgraded Tokens")
        accessory_copy = AccessoryCopy.objects.create(owner=self.user, accessory=accessory)
        purchase = Purchase.objects.create(owner=self.user, name="Some Campaign")
        wave = Wave.objects.create(purchase=purchase, number=1)
        product = Product.objects.create(
            wave=wave, name="Playmat", kind=Product.Kind.ACCESSORY,
            accessory_copy=accessory_copy,
        )
        self.assertIsNone(product.game)
        self.assertEqual(product.accessory_copy, accessory_copy)
        self.assertIn(product, accessory_copy.source_products.all())


class ParseDimensionsTests(TestCase):
    def test_parses_plain_decimal_and_suffixed_sizes(self):
        self.assertEqual(parse_dimensions("41x63"), (Decimal("41.0"), Decimal("63.0")))
        self.assertEqual(parse_dimensions("57.5x89"), (Decimal("57.5"), Decimal("89.0")))
        self.assertEqual(parse_dimensions("70x120-clear"), (Decimal("70.0"), Decimal("120.0")))
        self.assertEqual(parse_dimensions("76x101,5"), (Decimal("76.0"), Decimal("101.5")))

    def test_rejects_noise(self):
        self.assertIsNone(parse_dimensions("diam.85"))
        self.assertIsNone(parse_dimensions("different ones"))
        self.assertIsNone(parse_dimensions("num"))


def build_sleeves_workbook(overview_rows, catalog_columns=(), packs_row=20):
    """Write a minimal workbook with "Overview" + "Sleeves" sheets.

    overview_rows: as in build_sheet — {column: value | (value, url)} dicts.
    catalog_columns: per Sleeves-sheet size column, a dict with the pivot
    row values: {1: name, 2: "WxH", 20: packs, 40: product, 41: url}.
    """
    workbook = openpyxl.Workbook()
    overview = workbook.active
    overview.title = "Overview"
    for offset, row in enumerate(overview_rows):
        for col, value in row.items():
            cell = overview.cell(row=4 + offset, column=col)
            if isinstance(value, tuple):
                cell.value, cell.hyperlink = value
            else:
                cell.value = value

    sleeves = workbook.create_sheet("Sleeves")
    for offset, column in enumerate(catalog_columns):
        for sheet_row, value in column.items():
            sleeves.cell(row=sheet_row, column=3 + offset, value=value)
    path = Path(tempfile.mkdtemp()) / "mastersheet.xlsx"
    workbook.save(path)
    return path


# Catalog: Mini has a Tlama product + 2 packs owned; 45x68 is named by its
# Tlama colour ("Azur") and must alias-merge with the Overview name "Teal".
CATALOG = [
    {1: "Mini", 2: "41x63", 20: 2.0, 40: "Diamond Yellow",
     41: "https://www.tlamagames.com/doplnky/diamond-yellow/"},
    {1: "Azur", 2: "45x68", 20: 0.0},
]

# 100 Mini sleeved with Tlama, 27 Teal with no code, 50 Medium to-sleeve,
# 46 cards of a free-text "Other" size 63x63.
SLEEVE_ROW = {
    2: "5 Minute Dungeon",
    3: ("BGG", "https://boardgamegeek.com/boardgame/207830/5-minute-dungeon"),
    130: 100.0, 131: "tl",
    136: 27.0,
    142: 50.0, 143: "?",
    160: 46.0, 162: "63x63",
}


class ImportSleevesTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)
        cls.copy = Copy.objects.create(owner=cls.user, edition=cls.edition)

    def run_import(self, path, **extra):
        out = StringIO()
        call_command("import_sleeves", str(path), user="kernicek", stdout=out, **extra)
        return out.getvalue()

    def size(self, width, height):
        return CardSize.objects.get(width_mm=Decimal(width), height_mm=Decimal(height))

    def test_imports_requirements_and_statuses(self):
        self.run_import(build_sleeves_workbook([SLEEVE_ROW], CATALOG))

        requirements = {
            (r.card_size.width_mm, r.card_size.height_mm): r.count
            for r in self.edition.sleeve_requirements.all()
        }
        self.assertEqual(requirements, {
            (Decimal("41.0"), Decimal("63.0")): 100,
            (Decimal("45.0"), Decimal("68.0")): 27,
            (Decimal("57.5"), Decimal("89.0")): 50,
            (Decimal("63.0"), Decimal("63.0")): 46,  # from the Other pair
        })

        mini = self.copy.sleeve_statuses.get(card_size=self.size("41", "63"))
        self.assertEqual(mini.status, CopySleeveStatus.Status.SLEEVED)
        self.assertEqual(mini.product.brand, "Tlama")
        self.assertEqual(mini.product.name, "Diamond Yellow")  # catalog product reused

        teal = self.copy.sleeve_statuses.get(card_size=self.size("45", "68"))
        self.assertEqual(teal.status, CopySleeveStatus.Status.NOT_SLEEVED)
        self.assertIsNone(teal.product)

        medium = self.copy.sleeve_statuses.get(card_size=self.size("57.5", "89"))
        self.assertEqual(medium.status, CopySleeveStatus.Status.TO_SLEEVE)

    def test_imports_catalog_products_and_inventory(self):
        self.run_import(build_sleeves_workbook([], CATALOG))

        product = SleeveProduct.objects.get(brand="Tlama", name="Diamond Yellow")
        self.assertEqual(product.card_size, self.size("41", "63"))
        self.assertEqual(product.pack_size, 100)
        self.assertIn("tlamagames.com", product.url)

        inventory = SleeveInventory.objects.get()
        self.assertEqual(inventory.owner, self.user)
        self.assertEqual(inventory.product, product)
        self.assertEqual(inventory.packs, 2)
        self.assertEqual(inventory.loose, 0)

    def test_overview_name_wins_and_colour_name_becomes_alias(self):
        self.run_import(build_sleeves_workbook([SLEEVE_ROW], CATALOG))

        size = self.size("45", "68")
        self.assertEqual(size.name, "Teal")
        self.assertIn("Azur", size.alias_list)

    def test_same_size_in_fixed_and_other_column_sums_counts(self):
        # Arydia-style row: 60×63x88 sleeved with Tlama in the Standard
        # column plus 1280×"63x88" in the Other pair.
        row = {
            2: "5 Minute Dungeon",
            3: ("BGG", "https://boardgamegeek.com/boardgame/207830/5-minute-dungeon"),
            146: 60.0, 147: "tl",
            160: 1280.0, 162: "63x88",
        }
        output = self.run_import(build_sleeves_workbook([row], CATALOG))

        requirement = SleeveRequirement.objects.get()
        self.assertEqual(requirement.count, 1340)
        status = CopySleeveStatus.objects.get()
        # Status comes from the largest batch (1280 without a code).
        self.assertEqual(status.status, CopySleeveStatus.Status.NOT_SLEEVED)
        self.assertIn("appears in several batches", output)

    def test_rows_without_matching_game_or_copy_are_skipped(self):
        unknown = dict(SLEEVE_ROW)
        unknown[2] = "Unknown Game"
        unknown[3] = ("BGG", "https://boardgamegeek.com/boardgame/999999/unknown")
        no_copy_game = Game.objects.create(name="No Copy")
        BggLink.objects.create(game=no_copy_game, bgg_id=888888, is_primary=True)
        no_copy = dict(SLEEVE_ROW)
        no_copy[2] = "No Copy"
        no_copy[3] = ("BGG", "https://boardgamegeek.com/boardgame/888888/no-copy")

        output = self.run_import(build_sleeves_workbook([unknown, no_copy], CATALOG))

        self.assertEqual(SleeveRequirement.objects.count(), 0)
        self.assertIn("BGG id 999999 not in database", output)
        self.assertIn("owns no Copy", output)

    def test_reimport_is_idempotent(self):
        path = build_sleeves_workbook([SLEEVE_ROW], CATALOG)
        self.run_import(path)
        counts = {
            model: model.objects.count()
            for model in (CardSize, SleeveProduct, SleeveRequirement,
                          SleeveInventory, CopySleeveStatus)
        }
        self.assertEqual(counts[SleeveRequirement], 4)

        output = self.run_import(path)
        for model, count in counts.items():
            self.assertEqual(model.objects.count(), count)
        self.assertIn("requirements updated: 4", output)

    def test_dry_run_writes_nothing(self):
        output = self.run_import(build_sleeves_workbook([SLEEVE_ROW], CATALOG), dry_run=True)

        self.assertEqual(CardSize.objects.count(), 0)
        self.assertEqual(SleeveRequirement.objects.count(), 0)
        self.assertEqual(SleeveInventory.objects.count(), 0)
        self.assertIn("DRY RUN", output)
        self.assertIn("requirements created: 4", output)


# ===========================================================================
# §6  Purchases / crowdfunding
# ===========================================================================

class PurchaseModelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.purchase = Purchase.objects.create(
            owner=cls.user, name="Trickerion KS", status=Purchase.Status.COMMITTED,
        )
        cls.wave = Wave.objects.create(purchase=cls.purchase, number=1)
        cls.medium = CardSize.objects.create(
            width_mm=Decimal("57.5"), height_mm=Decimal("89.0"), name="Medium",
        )
        # Seeded by migration 0049 (issue #181) — fetching rather than
        # creating also verifies the seed data itself.
        cls.gamefound = PledgeManager.objects.get(name="Gamefound")
        cls.other_pm = PledgeManager.objects.get(name="Other")

    def test_one_purchase_per_owner_and_name(self):
        with self.assertRaises(IntegrityError):
            Purchase.objects.create(owner=self.user, name="Trickerion KS")

    def test_one_wave_per_purchase_and_number(self):
        with self.assertRaises(IntegrityError):
            Wave.objects.create(purchase=self.purchase, number=1)

    def test_one_product_per_wave_and_name(self):
        Product.objects.create(wave=self.wave, name="Trickerion")
        with self.assertRaises(IntegrityError):
            Product.objects.create(wave=self.wave, name="Trickerion")

    def test_one_sleeve_requirement_per_product_and_size(self):
        product = Product.objects.create(wave=self.wave, name="Trickerion")
        ProductSleeveRequirement.objects.create(
            product=product, card_size=self.medium, count=100,
        )
        with self.assertRaises(IntegrityError):
            ProductSleeveRequirement.objects.create(
                product=product, card_size=self.medium, count=1,
            )

    def test_wave_delay_and_purchase_fulfillment_are_derived(self):
        self.wave.original_eta = datetime.date(2024, 1, 1)
        self.wave.arrived_date = datetime.date(2024, 3, 1)
        self.wave.status = Wave.Status.ARRIVED
        self.wave.save()
        self.assertEqual(self.wave.delay_days, 60)
        self.assertTrue(self.purchase.is_fulfilled)

        Wave.objects.create(purchase=self.purchase, number=2)
        self.assertFalse(self.purchase.is_fulfilled)

    def test_pledge_manager_effective_url_prefers_purchase_specific_link(self):
        # Issue #159: a purchase-specific pledge_manager_url wins over the
        # PM's shared default.
        self.purchase.pledge_manager = self.gamefound
        self.purchase.pledge_manager_url = "https://gamefound.com/trickerion/pm"
        self.assertEqual(
            self.purchase.get_pledge_manager_effective_url(),
            "https://gamefound.com/trickerion/pm",
        )

    def test_pledge_manager_effective_url_falls_back_to_pm_default(self):
        self.purchase.pledge_manager = self.gamefound
        self.assertEqual(
            self.purchase.get_pledge_manager_effective_url(),
            "https://gamefound.com/en/users/dashboard",
        )

    def test_pledge_manager_effective_url_blank_for_pm_without_default(self):
        self.purchase.pledge_manager = self.other_pm
        self.assertEqual(self.purchase.get_pledge_manager_effective_url(), "")

    def test_shortfall_toggle_includes_pending_preorder_needs(self):
        product = Product.objects.create(wave=self.wave, name="Trickerion")
        ProductSleeveRequirement.objects.create(
            product=product, card_size=self.medium, count=250,
        )

        self.assertEqual(sleeve_shortfall(self.user), [])
        (entry,) = sleeve_shortfall(self.user, include_preorders=True)
        self.assertEqual(entry["card_size"], self.medium)
        self.assertEqual(entry["to_sleeve"], 250)
        self.assertEqual(entry["shortfall"], 250)
        self.assertEqual(entry["packs_to_buy"], 3)
        # Issue #93: a preorder Product not yet linked to a Game falls back
        # to its own name in the breakdown, with no game to link to.
        self.assertEqual(entry["games"], [
            {"label": "Trickerion", "game_pk": None, "to_sleeve": 250},
        ])

    def test_shortfall_toggle_skips_arrived_waves_and_watched_purchases(self):
        product = Product.objects.create(wave=self.wave, name="Trickerion")
        ProductSleeveRequirement.objects.create(
            product=product, card_size=self.medium, count=250,
        )

        self.wave.status = Wave.Status.ARRIVED
        self.wave.save()
        self.assertEqual(sleeve_shortfall(self.user, include_preorders=True), [])

        self.wave.status = Wave.Status.PENDING
        self.wave.save()
        self.purchase.status = Purchase.Status.WATCHING
        self.purchase.save()
        self.assertEqual(sleeve_shortfall(self.user, include_preorders=True), [])

        # Placeholder ($1) pledges DO count — the sheet counts them.
        self.purchase.status = Purchase.Status.PLACEHOLDER
        self.purchase.save()
        self.assertEqual(len(sleeve_shortfall(self.user, include_preorders=True)), 1)


class PledgePlanModelTests(TestCase):
    """Issue #186: pre-backing bundle comparison scoped to one Purchase."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.purchase = Purchase.objects.create(owner=cls.user, name="AC Brotherhood KS")
        cls.plan = PledgePlan.objects.create(purchase=cls.purchase, currency="EUR")

    def test_one_plan_per_purchase(self):
        with self.assertRaises(IntegrityError):
            PledgePlan.objects.create(purchase=self.purchase, currency="USD")

    def test_one_item_per_plan_and_name(self):
        PledgePlanItem.objects.create(plan=self.plan, name="Core Box", price=Decimal("50"))
        with self.assertRaises(IntegrityError):
            PledgePlanItem.objects.create(plan=self.plan, name="Core Box", price=Decimal("60"))

    def test_one_bundle_per_plan_and_name(self):
        PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("100"))
        with self.assertRaises(IntegrityError):
            PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("50"))

    def test_bundle_total_cost_uses_the_bundles_own_price_not_item_prices(self):
        # A bundle's advertised pledge price is usually a discount off the
        # sum of its items' individual prices — items don't drive the total.
        core = PledgePlanItem.objects.create(
            plan=self.plan, name="Core Box", price=Decimal("50"),
            want_priority=Game.WishlistPriority.MUST_HAVE,
        )
        expansion = PledgePlanItem.objects.create(
            plan=self.plan, name="Expansion", price=Decimal("25"),
            want_priority=Game.WishlistPriority.LOVE_TO_HAVE,
        )
        bundle = PledgePlanBundle.objects.create(
            plan=self.plan, name="All-in", price=Decimal("65"), shipping_cost=Decimal("10"),
        )
        bundle.items.set([core, expansion])

        self.assertEqual(bundle.total_cost, Decimal("75"))

    def test_total_cost_equals_price_with_no_shipping_or_vat(self):
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="Empty", price=Decimal("50"))
        self.assertEqual(bundle.total_cost, Decimal("50"))

    def test_total_cost_czk_is_none_without_a_rate(self):
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("50"))
        self.assertIsNone(bundle.total_cost_czk)

    def test_total_cost_czk_uses_the_plans_manual_rate(self):
        self.plan.czk_rate = Decimal("24.5")
        self.plan.save()
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("50"))

        self.assertEqual(bundle.total_cost_czk, Decimal("1225.0"))

    def test_vat_amount_is_zero_without_a_rate(self):
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("50"))

        self.assertEqual(bundle.vat_amount, Decimal("0"))
        self.assertEqual(bundle.total_cost, Decimal("50"))

    def test_vat_is_applied_to_price_plus_shipping(self):
        # Mirrors the real sheets' VAT = (price + shipping) * rate.
        self.plan.vat_rate = Decimal("21")
        self.plan.save()
        bundle = PledgePlanBundle.objects.create(
            plan=self.plan, name="All-in", price=Decimal("100"), shipping_cost=Decimal("27.29"),
        )

        self.assertEqual(bundle.vat_amount, Decimal("26.7309"))
        self.assertEqual(bundle.total_cost, Decimal("154.0209"))

    def test_total_cost_czk_reflects_the_vat_inclusive_total(self):
        self.plan.vat_rate = Decimal("21")
        self.plan.czk_rate = Decimal("10")
        self.plan.save()
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("100"))

        self.assertEqual(bundle.total_cost, Decimal("121"))
        self.assertEqual(bundle.total_cost_czk, Decimal("1210"))

    def test_value_sums_included_items_individual_prices(self):
        core = PledgePlanItem.objects.create(plan=self.plan, name="Core Box", price=Decimal("50"))
        expansion = PledgePlanItem.objects.create(
            plan=self.plan, name="Expansion", price=Decimal("25"),
        )
        PledgePlanItem.objects.create(plan=self.plan, name="Not included", price=Decimal("100"))
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("65"))
        bundle.items.set([core, expansion])

        self.assertEqual(bundle.value, Decimal("75"))
        self.assertEqual(bundle.savings, Decimal("10"))

    def test_value_is_none_when_no_included_item_has_a_price(self):
        exclusive = PledgePlanItem.objects.create(plan=self.plan, name="Exclusive Mini")
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("65"))
        bundle.items.add(exclusive)

        self.assertIsNone(bundle.value)
        self.assertIsNone(bundle.savings)

    def test_value_skips_items_with_no_individual_price(self):
        core = PledgePlanItem.objects.create(plan=self.plan, name="Core Box", price=Decimal("50"))
        exclusive = PledgePlanItem.objects.create(plan=self.plan, name="Exclusive Mini")
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="All-in", price=Decimal("40"))
        bundle.items.set([core, exclusive])

        self.assertEqual(bundle.value, Decimal("50"))

    def test_priority_coverage_counts_included_vs_total_at_each_level(self):
        core = PledgePlanItem.objects.create(
            plan=self.plan, name="Core Box", price=Decimal("50"),
            want_priority=Game.WishlistPriority.MUST_HAVE,
        )
        PledgePlanItem.objects.create(
            plan=self.plan, name="Also must-have", price=Decimal("20"),
            want_priority=Game.WishlistPriority.MUST_HAVE,
        )
        PledgePlanItem.objects.create(
            plan=self.plan, name="Unprioritised", price=Decimal("5"),
        )
        bundle = PledgePlanBundle.objects.create(plan=self.plan, name="Core only", price=Decimal("50"))
        bundle.items.add(core)

        self.assertEqual(
            bundle.priority_coverage, {Game.WishlistPriority.MUST_HAVE: (1, 2)},
        )


class PledgePlanViewTests(TestCase):
    """Owner-scoped CRUD for the pledge planner (issue #186)."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.purchase = Purchase.objects.create(owner=cls.user, name="AC Brotherhood KS")
        cls.plan = PledgePlan.objects.create(purchase=cls.purchase, currency="EUR")
        cls.item = PledgePlanItem.objects.create(
            plan=cls.plan, name="Core Box", price=Decimal("50"),
            want_priority=Game.WishlistPriority.MUST_HAVE,
        )
        cls.bundle = PledgePlanBundle.objects.create(
            plan=cls.plan, name="All-in", price=Decimal("45"),
        )

        cls.foreign_purchase = Purchase.objects.create(
            owner=cls.other, name="Pavel's Secret KS",
        )
        cls.foreign_plan = PledgePlan.objects.create(
            purchase=cls.foreign_purchase, currency="USD",
        )

    def test_pledge_plan_add_creates_plan_and_redirects_to_detail(self):
        purchase = Purchase.objects.create(owner=self.user, name="Second KS")
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            reverse("pledge_plan_add", args=[purchase.pk]),
            {"currency": "usd"},
        )
        purchase.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.url, reverse("pledge_plan_detail", args=[purchase.pk]),
        )
        self.assertEqual(purchase.pledge_plan.currency, "USD")

    def test_pledge_plan_add_404s_for_a_foreign_purchase(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(
            reverse("pledge_plan_add", args=[self.foreign_purchase.pk]),
        )
        self.assertEqual(response.status_code, 404)

    def test_pledge_plan_detail_404s_for_a_foreign_plan(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(reverse("pledge_plan_detail", args=[self.foreign_purchase.pk]))
        self.assertEqual(response.status_code, 404)

    def test_pledge_plan_edit_updates_currency_and_rates(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            reverse("pledge_plan_edit", args=[self.purchase.pk]),
            {"currency": "usd", "vat_rate": "21", "czk_rate": "24.5"},
        )
        self.assertEqual(response.status_code, 302)
        self.plan.refresh_from_db()
        self.assertEqual(self.plan.currency, "USD")
        self.assertEqual(self.plan.vat_rate, Decimal("21.00"))
        self.assertEqual(self.plan.czk_rate, Decimal("24.5000"))

    def test_pledge_plan_edit_404s_for_a_foreign_plan(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(reverse("pledge_plan_edit", args=[self.foreign_purchase.pk]))
        self.assertEqual(response.status_code, 404)

    def test_item_add_edit_and_delete_round_trip(self):
        self.client.login(username="kernicek", password="pass")

        response = self.client.post(
            reverse("pledge_plan_item_add", args=[self.purchase.pk]),
            {
                "name": "Expansion", "category": "expansion", "price": "25",
                "want_priority": "", "exclusive": "on",
            },
        )
        self.assertEqual(response.status_code, 302)
        item = PledgePlanItem.objects.get(plan=self.plan, name="Expansion")
        self.assertEqual(item.category, PledgePlanItem.Category.EXPANSION)
        self.assertTrue(item.exclusive)

        response = self.client.post(
            reverse("pledge_plan_item_edit", args=[item.pk]),
            {"name": "Expansion", "category": "expansion", "price": "30", "want_priority": ""},
        )
        self.assertEqual(response.status_code, 302)
        item.refresh_from_db()
        self.assertEqual(item.price, Decimal("30"))
        self.assertFalse(item.exclusive)

        response = self.client.post(reverse("pledge_plan_item_delete", args=[item.pk]))
        self.assertEqual(response.status_code, 302)
        self.assertFalse(PledgePlanItem.objects.filter(pk=item.pk).exists())

    def test_item_add_defaults_exclusive_to_false(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            reverse("pledge_plan_item_add", args=[self.purchase.pk]),
            {"name": "Retail Available", "category": "expansion", "price": "10", "want_priority": ""},
        )
        self.assertEqual(response.status_code, 302)
        item = PledgePlanItem.objects.get(plan=self.plan, name="Retail Available")
        self.assertFalse(item.exclusive)

    def test_item_add_allows_a_blank_individual_price(self):
        # Exclusive add-ons with no separate price (issue #186 grill-me).
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            reverse("pledge_plan_item_add", args=[self.purchase.pk]),
            {"name": "Exclusive Mini", "category": "accessory", "price": "", "want_priority": ""},
        )
        self.assertEqual(response.status_code, 302)
        item = PledgePlanItem.objects.get(plan=self.plan, name="Exclusive Mini")
        self.assertIsNone(item.price)

    def test_item_edit_404s_for_a_foreign_item(self):
        foreign_item = PledgePlanItem.objects.create(
            plan=self.foreign_plan, name="Foreign Item", price=Decimal("1"),
        )
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(reverse("pledge_plan_item_edit", args=[foreign_item.pk]))
        self.assertEqual(response.status_code, 404)

    def test_duplicate_item_name_in_same_plan_is_rejected(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            reverse("pledge_plan_item_add", args=[self.purchase.pk]),
            {"name": "Core Box", "category": "board_game", "price": "1", "want_priority": ""},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "already has an item with that name")

    def test_bundle_add_edit_and_delete_round_trip(self):
        self.client.login(username="kernicek", password="pass")

        response = self.client.post(
            reverse("pledge_plan_bundle_add", args=[self.purchase.pk]),
            {"name": "Core only", "price": "40", "shipping_cost": "10"},
        )
        self.assertEqual(response.status_code, 302)
        bundle = PledgePlanBundle.objects.get(plan=self.plan, name="Core only")
        self.assertEqual(bundle.price, Decimal("40"))

        response = self.client.post(
            reverse("pledge_plan_bundle_edit", args=[bundle.pk]),
            {"name": "Core only", "price": "45", "shipping_cost": "15"},
        )
        self.assertEqual(response.status_code, 302)
        bundle.refresh_from_db()
        self.assertEqual(bundle.price, Decimal("45"))
        self.assertEqual(bundle.shipping_cost, Decimal("15"))

        response = self.client.post(reverse("pledge_plan_bundle_delete", args=[bundle.pk]))
        self.assertEqual(response.status_code, 302)
        self.assertFalse(PledgePlanBundle.objects.filter(pk=bundle.pk).exists())

    def test_bundle_item_toggle_adds_then_removes(self):
        self.client.login(username="kernicek", password="pass")
        url = reverse(
            "pledge_plan_bundle_item_toggle", args=[self.bundle.pk, self.item.pk],
        )

        self.client.post(url)
        self.assertTrue(self.bundle.items.filter(pk=self.item.pk).exists())

        self.client.post(url)
        self.assertFalse(self.bundle.items.filter(pk=self.item.pk).exists())

    def test_bundle_item_toggle_404s_for_a_foreign_bundle(self):
        foreign_item = PledgePlanItem.objects.create(
            plan=self.foreign_plan, name="Foreign Item", price=Decimal("1"),
        )
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            reverse(
                "pledge_plan_bundle_item_toggle",
                args=[self.bundle.pk, foreign_item.pk],
            ),
        )
        self.assertEqual(response.status_code, 404)

    def test_pledge_plan_never_leaks_template_comment_markers(self):
        self.client.login(username="kernicek", password="pass")
        body = self.client.get(
            reverse("pledge_plan_detail", args=[self.purchase.pk]),
        ).content.decode()

        self.assertNotIn("{#", body)
        self.assertNotIn("{% comment", body)

    def test_plan_detail_shows_exclusive_icon_only_for_exclusive_items(self):
        # self.item (from setUpTestData) is non-exclusive.
        PledgePlanItem.objects.create(
            plan=self.plan, name="Exclusive Mini", price=Decimal("5"), exclusive=True,
        )
        self.client.login(username="kernicek", password="pass")
        body = self.client.get(
            reverse("pledge_plan_detail", args=[self.purchase.pk]),
        ).content.decode()

        # One icon in the column header, one more in the exclusive item's
        # row — none in the non-exclusive item's row.
        self.assertEqual(body.count("bi-award-fill"), 2)

    def test_shortlist_toggle_flips_flag(self):
        self.client.login(username="kernicek", password="pass")
        url = reverse("pledge_plan_bundle_shortlist_toggle", args=[self.bundle.pk])

        self.client.post(url)
        self.bundle.refresh_from_db()
        self.assertTrue(self.bundle.is_shortlisted)

        self.client.post(url)
        self.bundle.refresh_from_db()
        self.assertFalse(self.bundle.is_shortlisted)

    def test_shortlist_toggle_404s_for_a_foreign_bundle(self):
        foreign_bundle = PledgePlanBundle.objects.create(
            plan=self.foreign_plan, name="Foreign Bundle", price=Decimal("1"),
        )
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            reverse("pledge_plan_bundle_shortlist_toggle", args=[foreign_bundle.pk]),
        )
        self.assertEqual(response.status_code, 404)

    def test_plan_detail_narrows_to_shortlisted_bundles_once_any_are_starred(self):
        other_bundle = PledgePlanBundle.objects.create(
            plan=self.plan, name="Core only", price=Decimal("30"),
        )
        self.client.login(username="kernicek", password="pass")

        # Nothing starred yet — both bundles show.
        body = self.client.get(reverse("pledge_plan_detail", args=[self.purchase.pk])).content.decode()
        self.assertIn("All-in", body)
        self.assertIn("Core only", body)

        self.client.post(reverse("pledge_plan_bundle_shortlist_toggle", args=[self.bundle.pk]))

        body = self.client.get(reverse("pledge_plan_detail", args=[self.purchase.pk])).content.decode()
        self.assertIn("All-in", body)
        self.assertNotIn("Core only", body)
        self.assertIn("show all", body)

        body = self.client.get(
            reverse("pledge_plan_detail", args=[self.purchase.pk]), {"view": "all"},
        ).content.decode()
        self.assertIn("All-in", body)
        self.assertIn("Core only", body)
        # Viewing all with an existing shortlist offers a way back to it,
        # not the "nothing starred yet" message.
        self.assertIn("show shortlisted only", body)
        self.assertNotIn("Star a bundle below", body)


class PledgeManagerModelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.pm = PledgeManager.objects.create(
            name="TestPM", default_url="https://testpm.example.com/dashboard",
        )

    def test_str_returns_name(self):
        self.assertEqual(str(self.pm), "TestPM")

    def test_name_is_unique(self):
        with self.assertRaises(IntegrityError):
            PledgeManager.objects.create(name="TestPM")

    def test_purchase_fk_and_effective_url_fallback(self):
        purchase = Purchase.objects.create(
            owner=self.user, name="Trickerion KS", pledge_manager=self.pm,
        )
        self.assertIn(purchase, self.pm.purchases.all())
        self.assertEqual(
            purchase.get_pledge_manager_effective_url(),
            "https://testpm.example.com/dashboard",
        )


def build_preorders_workbook(rows):
    """Write a minimal "(Pre)orders" workbook (data starts at row 4) to a temp
    file. Same conventions as build_sheet: {column: value | (value, url)}."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "(Pre)orders"
    sheet.cell(row=1, column=1, value="Purchase")  # header rows 1-3 are skipped
    for offset, row in enumerate(rows):
        for col, value in row.items():
            cell = sheet.cell(row=4 + offset, column=col)
            if isinstance(value, tuple):
                cell.value, cell.hyperlink = value
            else:
                cell.value = value
    path = Path(tempfile.mkdtemp()) / "mastersheet.xlsx"
    workbook.save(path)
    return path


# An arrived campaign delivering a game the user already owns; carries sleeve
# counts in the Standard column (37 -> 63x88) and an Other pair (44/45).
ARRIVED_ROW = {
    1: "5 Minute Dungeon KS",
    2: "5 Minute Dungeon",
    3: ("BGG", "https://boardgamegeek.com/boardgame/207830/5-minute-dungeon"),
    4: ("KS", "https://www.kickstarter.com/projects/wiggles3d/5-minute-dungeon"),
    7: "KS",
    8: "Game+Expansion(s)",
    9: "Arrived",
    10: datetime.datetime(2016, 12, 2),
    11: datetime.datetime(2017, 4, 1),
    12: datetime.datetime(2017, 5, 3),
    20: ("BackerKit", "https://www.backerkit.com/backer_accounts"),
    21: "Filled out",
    24: "Chynice",
    25: "smooth campaign",
    27: 4.0,
    29: "Yes",
    30: "Yes",
    37: 100.0,
    44: 46.0,
    45: "63x63",
}

# A running campaign for a game not in the DB; product name carries a "(x2)"
# multiplier; excitement is numeric; counts in the Medium column (35 -> 57.5x89
# despite the sheet header claiming 56x89).
NEW_GAME_ROW = {
    1: "Cyberpunk 2077 GF",
    2: "Cyberpunk 2077 - Deluxe Core Box (x2)",
    3: ("BGG", "https://boardgamegeek.com/boardgame/417542/cyberpunk-2077"),
    7: "GF",
    8: "Board Game(s)",
    9: "Pre-production",
    10: datetime.datetime(2025, 5, 1),
    11: datetime.datetime(2026, 2, 1),
    12: datetime.datetime(2026, 3, 1),
    26: 5.0,
    35: 50.0,
}

WAVE1_ROW = {
    1: "HZD KS (Wave 1)",
    2: "Horizon Zero Dawn: The Board Game",
    3: ("BGG", "https://boardgamegeek.com/boardgame/260410/horizon-zero-dawn"),
    7: "KS",
    8: "Board Game(s)",
    9: "Arrived",
    12: datetime.datetime(2020, 10, 31),
}

WAVE2_ROW = {
    1: "HZD KS (Wave 2)",
    2: "HZD: Kickstarter Exclusives",
    3: ("BGG", "https://boardgamegeek.com/boardgameexpansion/315288/hzd-ks"),
    7: "KS",
    8: "Expansion(s)",
    9: "Production",
}

PLACEHOLDER_ROW = {
    1: "Voidfall KS",
    2: "Voidfall - $1 Pledge",
    3: ("BGG", "https://boardgamegeek.com/boardgame/337627/voidfall"),
    7: "KS",
    8: "$1 Pledge-just support",
    9: "Just support",
}

PNP_ROW_PREORDER = {
    1: "Waypoints KS",
    2: "Waypoints (PnP)",
    3: ("BGG", "https://boardgamegeek.com/boardgame/385292/waypoints"),
    7: "KS",
    8: "Game(s)-PnP",
    9: "Arrived",
    5: ("Drive", "https://drive.google.com/drive/folders/waypoints"),
}

ACCESSORY_ROW = {
    1: "5 Minute Dungeon KS",
    2: "5 Minute Dungeon - Playmat",
    3: ("BGG", "https://boardgamegeek.com/boardgameaccessory/999001/playmat"),
    7: "KS",
    8: "Accessories",
    9: "Arrived",
}


class ImportPreordersTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)
        cls.copy = Copy.objects.create(owner=cls.user, edition=cls.edition)

    def run_import(self, path, **extra):
        out = StringIO()
        call_command("import_preorders", str(path), user="kernicek", stdout=out, **extra)
        return out.getvalue()

    def size(self, width, height):
        return CardSize.objects.get(width_mm=Decimal(width), height_mm=Decimal(height))

    def test_imports_purchase_wave_product_hierarchy(self):
        self.run_import(build_preorders_workbook([ARRIVED_ROW]))

        purchase = Purchase.objects.get()
        self.assertEqual(purchase.name, "5 Minute Dungeon")  # "KS" suffix stripped (#32)
        self.assertEqual(purchase.owner, self.user)
        self.assertEqual(purchase.platform, Purchase.Platform.KICKSTARTER)
        self.assertEqual(purchase.status, Purchase.Status.COMMITTED)
        self.assertEqual(purchase.ordered_date, datetime.date(2016, 12, 2))
        self.assertEqual(purchase.pledge_manager, PledgeManager.objects.get(name="BackerKit"))
        self.assertIn("backerkit.com", purchase.pledge_manager_url)
        self.assertEqual(
            purchase.pledge_manager_status, Purchase.PledgeManagerStatus.FILLED_OUT,
        )
        self.assertIn("kickstarter.com", purchase.campaign_url)

        wave = Wave.objects.get()
        self.assertEqual(wave.number, 1)
        self.assertEqual(wave.status, Wave.Status.ARRIVED)
        self.assertEqual(wave.delivery_type, Wave.DeliveryType.PHYSICAL)
        self.assertEqual(wave.original_eta, datetime.date(2017, 4, 1))
        self.assertEqual(wave.arrived_date, datetime.date(2017, 5, 3))
        self.assertIsNone(wave.expected_arrival)
        self.assertEqual(wave.address, "Chynice")
        self.assertEqual(wave.delay_days, 32)

        product = Product.objects.get()
        self.assertEqual(product.kind, Product.Kind.GAME_AND_EXPANSIONS)
        self.assertEqual(product.game, self.game)      # matched via BGG id
        self.assertEqual(product.copy, self.copy)      # arrived -> converted
        self.assertEqual(product.contains_cards, Product.TriState.YES)
        self.assertEqual(product.needs_sleeves, Product.TriState.YES)
        self.assertEqual(product.miniatures_count, 4)
        self.assertEqual(product.notes, "smooth campaign")

        needs = {
            (r.card_size.width_mm, r.card_size.height_mm): r.count
            for r in product.sleeve_requirements.all()
        }
        self.assertEqual(needs, {
            (Decimal("63.0"), Decimal("88.0")): 100,
            (Decimal("63.0"), Decimal("63.0")): 46,  # from the Other pair
        })

    def test_wave_suffix_splits_waves(self):
        self.run_import(build_preorders_workbook([WAVE1_ROW, WAVE2_ROW]))

        purchase = Purchase.objects.get()
        self.assertEqual(purchase.name, "HZD")  # "KS" suffix stripped (#32)
        waves = {wave.number: wave for wave in purchase.waves.all()}
        self.assertEqual(set(waves), {1, 2})
        self.assertEqual(waves[1].status, Wave.Status.ARRIVED)
        self.assertEqual(waves[2].status, Wave.Status.PRODUCTION)
        self.assertFalse(purchase.is_fulfilled)

    def test_new_game_created_without_copy(self):
        self.run_import(build_preorders_workbook([NEW_GAME_ROW]))

        game = Game.objects.exclude(pk=self.game.pk).get()
        self.assertEqual(game.name, "Cyberpunk 2077 - Deluxe Core Box")  # (x2) stripped
        self.assertEqual(game.type, Game.Type.BASE)
        self.assertEqual(game.primary_bgg_link.bgg_id, 417542)
        self.assertEqual(Copy.objects.count(), 1)  # only the pre-existing one

        purchase = Purchase.objects.get()
        self.assertEqual(purchase.platform, Purchase.Platform.GAMEFOUND)
        self.assertEqual(purchase.excitement, Decimal("5.0"))
        product = Product.objects.get()
        self.assertIsNone(product.copy)  # not arrived
        wave = product.wave
        self.assertEqual(wave.status, Wave.Status.PRE_PRODUCTION)
        self.assertEqual(wave.expected_arrival, datetime.date(2026, 3, 1))
        self.assertIsNone(wave.arrived_date)

    def test_placeholder_pledge_and_pnp_digital_wave(self):
        self.run_import(build_preorders_workbook([PLACEHOLDER_ROW, PNP_ROW_PREORDER]))

        placeholder = Purchase.objects.get(name="Voidfall")  # "KS" stripped (#32)
        self.assertEqual(placeholder.status, Purchase.Status.PLACEHOLDER)
        self.assertEqual(
            placeholder.waves.get().products.get().kind,
            Product.Kind.PLACEHOLDER_PLEDGE,
        )

        pnp = Purchase.objects.get(name="Waypoints")  # "KS" stripped (#32)
        self.assertEqual(pnp.status, Purchase.Status.COMMITTED)
        wave = pnp.waves.get()
        self.assertEqual(wave.delivery_type, Wave.DeliveryType.DIGITAL)
        game = Game.objects.get(bgg_links__bgg_id=385292)
        self.assertEqual(game.name, "Waypoints")  # "(PnP)" stripped
        # PnP is now edition-level (#138): a PnP product materializes a default
        # PnP edition and links to it, rather than flagging the whole title.
        edition = Edition.objects.get(game=game)
        self.assertTrue(edition.is_default)
        self.assertTrue(edition.is_pnp)
        self.assertTrue(game.has_pnp_edition)
        product = wave.products.get()
        self.assertEqual(product.kind, Product.Kind.PNP_GAME)
        self.assertEqual(product.edition, edition)
        self.assertIn("drive.google.com", product.drive_url)

    def test_accessory_keeps_bgg_url_but_no_game(self):
        self.run_import(build_preorders_workbook([ACCESSORY_ROW]))

        product = Product.objects.get()
        self.assertEqual(product.kind, Product.Kind.ACCESSORY)
        self.assertIsNone(product.game)
        self.assertIn("boardgamegeek.com", product.bgg_url)
        self.assertEqual(Game.objects.count(), 1)  # nothing new created

    def test_shortfall_toggle_on_imported_data(self):
        self.run_import(build_preorders_workbook([ARRIVED_ROW, NEW_GAME_ROW]))

        # Arrived needs (100x 63x88, 46x 63x63) are excluded; only the running
        # Cyberpunk wave's 50x Medium count.
        self.assertEqual(sleeve_shortfall(self.user), [])
        (entry,) = sleeve_shortfall(self.user, include_preorders=True)
        self.assertEqual(entry["card_size"], self.size("57.5", "89"))
        self.assertEqual(entry["to_sleeve"], 50)

    def test_reimport_is_idempotent(self):
        path = build_preorders_workbook([
            ARRIVED_ROW, NEW_GAME_ROW, WAVE1_ROW, WAVE2_ROW,
            PLACEHOLDER_ROW, PNP_ROW_PREORDER, ACCESSORY_ROW,
        ])
        self.run_import(path)
        counts = {
            model: model.objects.count()
            for model in (Purchase, Wave, Product, ProductSleeveRequirement,
                          Game, BggLink, Copy, CardSize)
        }
        self.assertEqual(counts[Purchase], 5)
        self.assertEqual(counts[Wave], 6)
        self.assertEqual(counts[Product], 7)

        output = self.run_import(path)
        for model, count in counts.items():
            self.assertEqual(model.objects.count(), count)
        self.assertIn("purchases updated: 5", output)
        self.assertIn("products updated: 7", output)

    def test_dry_run_writes_nothing(self):
        output = self.run_import(
            build_preorders_workbook([ARRIVED_ROW, NEW_GAME_ROW]), dry_run=True,
        )

        self.assertEqual(Purchase.objects.count(), 0)
        self.assertEqual(Product.objects.count(), 0)
        self.assertEqual(Game.objects.count(), 1)  # only the pre-existing one
        self.assertEqual(ProductSleeveRequirement.objects.count(), 0)
        self.assertIn("DRY RUN", output)
        self.assertIn("purchases created: 2", output)


# ===========================================================================
# §11  Dashboard views
# ===========================================================================

class DashboardViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.medium = CardSize.objects.create(
            width_mm=Decimal("57.5"), height_mm=Decimal("89.0"), name="Medium",
        )
        cls.purchase = Purchase.objects.create(
            owner=cls.user, name="Cyberpunk 2077 GF",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.NOT_YET,
        )
        cls.wave = Wave.objects.create(
            purchase=cls.purchase, number=1,
            status=Wave.Status.PRE_PRODUCTION,
            original_eta=datetime.date(2026, 2, 1),
            expected_arrival=datetime.date(2026, 3, 1),
        )
        product = Product.objects.create(wave=cls.wave, name="Deluxe Core Box")
        ProductSleeveRequirement.objects.create(
            product=product, card_size=cls.medium, count=450,
        )

    def test_anonymous_users_are_redirected_to_login(self):
        for url in ("/", "/dashboard/", "/dashboard/shortfall/"):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 302)
            self.assertIn("/accounts/login/", response.url)

    def test_dashboard_shows_incoming_waves_and_pm_actions(self):
        # A sent-out PM is actionable and belongs in the PM section; the
        # NOT_YET setup purchase must stay out of it (#59).
        sent_out = Purchase.objects.create(
            owner=self.user, name="Sent Out KS",
            status=Purchase.Status.COMMITTED,
            pledge_manager=PledgeManager.objects.get(name="CrowdOx"),
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
        )
        Wave.objects.create(
            purchase=sent_out, number=1, status=Wave.Status.PRE_PRODUCTION,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, "Cyberpunk 2077 GF")
        self.assertContains(response, "Pre-production")
        self.assertContains(response, "+28d")  # 2026-02-01 -> 2026-03-01
        # Issue #33: the incoming-wave purchase name links to its detail page.
        self.assertContains(response, f'href="/purchases/{self.purchase.pk}/"')
        # Issue #159: the PM card's Purchase-name cell links to its detail
        # page too, and the new PM column links to the PM's default URL
        # (no purchase-specific pledge_manager_url set on this fixture).
        self.assertContains(response, f'href="/purchases/{sent_out.pk}/"')
        self.assertContains(response, 'href="http://portal.crowdox.com/"')
        self.assertContains(response, "CrowdOx")
        # Issue #59: only sent-out PMs are actionable — the "Not yet" badge
        # (rendered solely by the PM section) must not appear.
        self.assertContains(response, "Sent out")
        self.assertNotContains(response, "Not yet")

    def test_dashboard_shows_the_to_craft_card(self):
        # Issue #19: not-ready copies surface on a dedicated card; ready
        # copies never appear there.
        needs_prep = Game.objects.create(name="Unprinted Prototype")
        needs_prep_edition = Edition.objects.create(game=needs_prep, is_default=True)
        Copy.objects.create(
            owner=self.user, edition=needs_prep_edition,
            ready_status=Copy.ReadyStatus.NOT_READY,
        )
        ready = Game.objects.create(name="Printed Copy")
        ready_edition = Edition.objects.create(game=ready, is_default=True)
        Copy.objects.create(owner=self.user, edition=ready_edition)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, "To craft")
        self.assertContains(response, "Unprinted Prototype")
        self.assertNotContains(response, "Printed Copy")

    def test_incoming_waves_colour_each_status_distinctly(self):
        # Issue #11: the regression the screenshot showed — different wave
        # statuses all rendering the same grey. A pre-production wave (cyan)
        # and a production wave (blue) must now carry different badge classes.
        # setUpTestData's wave is PRE_PRODUCTION -> text-bg-info.
        production = Purchase.objects.create(
            owner=self.user, name="In Production KS",
            status=Purchase.Status.COMMITTED,
        )
        Wave.objects.create(
            purchase=production, number=1, status=Wave.Status.PRODUCTION,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        # Incoming-waves card sits between its own header and the To craft header.
        body = response.content.decode()
        waves_card = body[body.index("Incoming waves"): body.index("To craft")]
        self.assertIn("text-bg-info", waves_card)      # pre-production
        self.assertIn("text-bg-primary", waves_card)   # production

    def test_dashboard_never_leaks_template_comment_markers(self):
        # Multi-line {# #} comments are NOT stripped by Django (its comment
        # lexer doesn't span newlines) and render verbatim into the page. Guard
        # against any comment marker reaching the UI.
        sent_out = Purchase.objects.create(
            owner=self.user, name="Sent Out KS",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
        )
        Wave.objects.create(
            purchase=sent_out, number=1, status=Wave.Status.PRE_PRODUCTION,
        )

        self.client.login(username="kernicek", password="pass")
        body = self.client.get("/dashboard/").content.decode()

        self.assertNotIn("{#", body)
        self.assertNotIn("{% comment", body)
        self.assertNotIn("higher priority", body)

    def test_pm_card_colours_sent_out_and_filled_out_differently(self):
        today = timezone.localdate()
        sent_out = Purchase.objects.create(
            owner=self.user, name="Sent Out KS",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
        )
        Wave.objects.create(
            purchase=sent_out, number=1, status=Wave.Status.PRE_PRODUCTION,
        )
        filled_open = Purchase.objects.create(
            owner=self.user, name="Filled Open GF",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.FILLED_OUT,
            pledge_manager_close_date=today + datetime.timedelta(days=5),
        )
        Wave.objects.create(
            purchase=filled_open, number=1, status=Wave.Status.PRODUCTION,
        )

        self.client.login(username="kernicek", password="pass")
        pm_card = self._pm_card(self.client.get("/dashboard/"))

        self.assertIn("text-bg-warning", pm_card)  # sent out
        self.assertIn("text-bg-info", pm_card)     # filled out, still open

    def test_campaign_card_shows_a_coloured_status_badge(self):
        today = timezone.localdate()
        Purchase.objects.create(
            owner=self.user, name="Watched GF",
            status=Purchase.Status.WATCHING,
            campaign_end_date=today + datetime.timedelta(days=3),
        )

        self.client.login(username="kernicek", password="pass")
        body = self.client.get("/dashboard/").content.decode()
        campaign_card = body[body.index("Campaigns ending soon"):]
        # Watching purchases carry the info status badge (#11).
        self.assertIn("Watching", campaign_card)
        self.assertIn("text-bg-info", campaign_card)

    @staticmethod
    def _pm_card(response):
        # The PM purchases are also COMMITTED with incoming waves, so they show
        # in the "Incoming waves" card too; scope assertions to just the PM
        # card (from its header up to the next card's header).
        body = response.content.decode()
        start = body.index("Pledge managers needing action")
        end = body.index("Campaigns ending soon", start)
        return body[start:end]

    def test_dashboard_pm_card_lists_open_filled_out_at_bottom(self):
        # A filled-out PM whose close date is still ahead stays visible as a
        # "can still be revised" reminder, below the sent-out ones (#85).
        today = timezone.localdate()
        sent_out = Purchase.objects.create(
            owner=self.user, name="Sent Out KS",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
        )
        Wave.objects.create(
            purchase=sent_out, number=1, status=Wave.Status.PRE_PRODUCTION,
        )
        filled_open = Purchase.objects.create(
            owner=self.user, name="Filled Open GF",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.FILLED_OUT,
            pledge_manager_close_date=today + datetime.timedelta(days=5),
        )
        Wave.objects.create(
            purchase=filled_open, number=1, status=Wave.Status.PRODUCTION,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        pm_card = self._pm_card(response)
        self.assertIn("Sent Out KS", pm_card)
        self.assertIn("Filled Open GF", pm_card)
        self.assertIn("Filled out", pm_card)
        # Sent-out PMs (needing action) sort above the still-open filled-out
        # revisable reminders.
        self.assertLess(
            pm_card.index("Sent Out KS"), pm_card.index("Filled Open GF"),
        )

    def test_dashboard_pm_card_drops_closed_filled_out(self):
        # Once the close date passes, a filled-out PM is done and drops off.
        today = timezone.localdate()
        closed = Purchase.objects.create(
            owner=self.user, name="Filled Closed GF",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.FILLED_OUT,
            pledge_manager_close_date=today - datetime.timedelta(days=1),
        )
        Wave.objects.create(
            purchase=closed, number=1, status=Wave.Status.PRE_PRODUCTION,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertNotIn("Filled Closed GF", self._pm_card(response))

    def test_dashboard_pm_card_drops_undated_filled_out(self):
        # A filled-out PM with no known close date can't be "still open", so it
        # stays out of the card (#85).
        undated = Purchase.objects.create(
            owner=self.user, name="Filled Undated GF",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.FILLED_OUT,
        )
        Wave.objects.create(
            purchase=undated, number=1, status=Wave.Status.PRE_PRODUCTION,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertNotIn("Filled Undated GF", self._pm_card(response))

    def test_dashboard_lists_watched_campaigns_ending_soon(self):
        today = timezone.localdate()
        Purchase.objects.create(
            owner=self.user, name="Shiny New GF",
            status=Purchase.Status.WATCHING,
            campaign_url="https://gamefound.com/shiny",
            campaign_end_date=today + datetime.timedelta(days=3),
        )
        # Ended, backed and undated ones stay out of the table; the undated
        # watch shows up only as the fill-it-in nudge count.
        Purchase.objects.create(
            owner=self.user, name="Missed GF",
            status=Purchase.Status.WATCHING,
            campaign_end_date=today - datetime.timedelta(days=1),
        )
        Purchase.objects.create(
            owner=self.user, name="Backed GF",
            status=Purchase.Status.COMMITTED,
            campaign_end_date=today + datetime.timedelta(days=2),
        )
        Purchase.objects.create(
            owner=self.user, name="Someday GF",
            status=Purchase.Status.WATCHING,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, "Shiny New GF")
        self.assertContains(response, "https://gamefound.com/shiny")
        self.assertContains(response, "3d")  # (end date - today).days
        self.assertNotContains(response, "Missed GF")
        self.assertNotContains(response, "Backed GF")
        self.assertNotContains(response, "Someday GF")
        self.assertContains(response, "1 watched campaign")
        self.assertContains(response, "without an end date")

    def test_shortfall_partial_toggles_preorder_needs(self):
        self.client.login(username="kernicek", password="pass")

        without = self.client.get("/dashboard/shortfall/")
        self.assertContains(without, "No shortfall")

        with_preorders = self.client.get("/dashboard/shortfall/", {"preorders": "on"})
        self.assertContains(with_preorders, "450")
        self.assertContains(with_preorders, "5 packs")  # ceil(450/100)
        # Issue #93: the size row expands to a per-game/product breakdown.
        self.assertContains(with_preorders, 'data-bs-toggle="collapse"')
        self.assertContains(with_preorders, "Deluxe Core Box")

    def test_dashboard_lists_unreviewed_sync_diffs(self):
        """Issue #62: the §11 widget shows the owner's unreviewed diffs —
        dismissed rows and other owners' rows stay out."""
        now = timezone.now()
        game = Game.objects.create(name="Nemesis")
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.MISSING_FROM_BGG,
            game=game, bgg_id=167355,
            note="no linked purchase — review", last_seen_at=now,
        )
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.SUGGEST_ADD,
            bgg_id=91072, bgg_name="Mystery Cult", last_seen_at=now,
        )
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.SUGGEST_ADD,
            bgg_id=999999, bgg_name="Dismissed Cult",
            last_seen_at=now, dismissed_at=now,
        )
        rival = get_user_model().objects.create_user(username="rival")
        BggSyncDiff.objects.create(
            owner=rival, category=BggSyncDiff.Category.SUGGEST_ADD,
            bgg_id=555555, bgg_name="Rival Game", last_seen_at=now,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, "BGG sync diffs")
        self.assertContains(response, f'href="/games/{game.pk}/"')
        self.assertContains(response, "Nemesis")
        self.assertContains(response, "In the app, missing from BGG")
        self.assertContains(response, "no linked purchase — review")
        # Every row links its BGG item; the game-less suggest-add row names
        # it via bgg_name.
        self.assertContains(response, "Mystery Cult")
        self.assertContains(
            response, 'href="https://boardgamegeek.com/boardgame/91072"',
        )
        self.assertContains(
            response, 'href="https://boardgamegeek.com/boardgame/167355"',
        )
        self.assertNotContains(response, "Dismissed Cult")
        self.assertNotContains(response, "Rival Game")

    def test_sync_diff_widget_colours_each_category(self):
        # Each reconciliation category gets its own severity colour, not the
        # blanket amber it used to share.
        now = timezone.now()
        expected = {
            BggSyncDiff.Category.SUGGEST_ADD: "text-bg-info",
            BggSyncDiff.Category.MISSING_FROM_BGG: "text-bg-warning",
            BggSyncDiff.Category.PREV_OWNED_ACTIVE: "text-bg-danger",
            BggSyncDiff.Category.ARCHIVED_ON_BGG: "text-bg-secondary",
        }
        for i, (category, css) in enumerate(expected.items()):
            diff = BggSyncDiff.objects.create(
                owner=self.user, category=category,
                bgg_id=1000 + i, bgg_name="X", last_seen_at=now,
            )
            html = render_to_string(
                "partials/sync_diff_widget.html", {"sync_diffs": [diff]},
            )
            label = diff.get_category_display()
            # The badge class sits immediately before the category label.
            badge = html[html.index(label) - 60:html.index(label)]
            self.assertIn(css, badge, f"{category} should render {css}")

    def test_dashboard_sync_diff_empty_state(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, "No unreviewed sync diffs.")

    def test_dashboard_lists_unreviewed_new_expansions(self):
        """Issue #64: the §11 widget shows the owner's unreviewed
        new-expansion diffs — dismissed rows and other owners' rows stay out,
        and non-new-expansion diffs don't leak into this widget."""
        now = timezone.now()
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.NEW_EXPANSION,
            bgg_id=999001, bgg_name="Another New Threat",
            note="New expansion for 5 Minute Dungeon", last_seen_at=now,
        )
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.NEW_EXPANSION,
            bgg_id=999002, bgg_name="Dismissed Threat",
            last_seen_at=now, dismissed_at=now,
        )
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.SUGGEST_ADD,
            bgg_id=91072, bgg_name="Mystery Cult", last_seen_at=now,
        )
        rival = get_user_model().objects.create_user(username="rival")
        BggSyncDiff.objects.create(
            owner=rival, category=BggSyncDiff.Category.NEW_EXPANSION,
            bgg_id=555555, bgg_name="Rival Threat", last_seen_at=now,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, "New expansions")
        self.assertContains(response, "Another New Threat")
        self.assertContains(response, "New expansion for 5 Minute Dungeon")
        self.assertContains(
            response, 'href="https://boardgamegeek.com/boardgameexpansion/999001"',
        )
        self.assertNotContains(response, "Dismissed Threat")
        self.assertNotContains(response, "Rival Threat")
        # The SUGGEST_ADD diff legitimately renders too, in the separate BGG
        # sync diffs widget on the same page.
        self.assertContains(response, "Mystery Cult")

    def test_dashboard_new_expansion_empty_state(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, "No new expansions.")

    def test_dashboard_card_headers_link_to_full_pages(self):
        # Issue #83: each card header links to its full-list page (the sleeve
        # shortfall reuses the existing /sleeves/ workbench).
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, 'href="/dashboard/incoming-waves/"')
        self.assertContains(response, 'href="/dashboard/pledge-managers/"')
        self.assertContains(response, 'href="/dashboard/campaigns-ending/"')
        self.assertContains(response, 'href="/dashboard/sync-diffs/"')
        self.assertContains(response, 'href="/dashboard/new-expansions/"')
        self.assertContains(response, 'href="/sleeves/"')

    @staticmethod
    def _make_incoming(user, count, start_day=2):
        # Committed purchases with a dated incoming wave; ascending ETAs keep a
        # deterministic sort order for the cap assertions.
        for i in range(count):
            purchase = Purchase.objects.create(
                owner=user, name=f"Incoming {i:02d} GF",
                status=Purchase.Status.COMMITTED,
            )
            Wave.objects.create(
                purchase=purchase, number=1, status=Wave.Status.PRODUCTION,
                expected_arrival=datetime.date(2026, 4, 1)
                + datetime.timedelta(days=start_day + i),
            )

    def test_dashboard_incoming_card_caps_at_limit_with_hint(self):
        # Issue #83: the card shows at most DASHBOARD_CARD_LIMIT rows with a
        # "showing N of M" hint; overflow rows only appear on the full page.
        # 25 new + the setUpTestData wave = 26 incoming, capped at 20.
        self._make_incoming(self.user, 25)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertContains(response, "Showing 20 of 26")
        self.assertContains(response, "Incoming 00 GF")   # earliest ETA, kept
        self.assertNotContains(response, "Incoming 24 GF")  # latest ETA, cut

    def test_dashboard_incoming_card_no_hint_when_under_limit(self):
        # Only the single setUpTestData wave — no truncation, no hint.
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")

        self.assertNotContains(response, "Showing 20 of")


class SyncDiffDismissTests(TestCase):
    """POST /dashboard/sync-diffs/<pk>/dismiss/ (issue #62): §8 per-user
    mark-seen — owner-scoped 404s, already-dismissed rows 404 too, and the
    response re-renders the widget partial so the count stays honest."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.rival = get_user_model().objects.create_user(
            username="rival", password="pass",
        )
        cls.game = Game.objects.create(name="Nemesis")
        now = timezone.now()
        cls.diff = BggSyncDiff.objects.create(
            owner=cls.user, category=BggSyncDiff.Category.MISSING_FROM_BGG,
            game=cls.game, bgg_id=167355, last_seen_at=now,
        )
        cls.rival_diff = BggSyncDiff.objects.create(
            owner=cls.rival, category=BggSyncDiff.Category.SUGGEST_ADD,
            bgg_id=555555, bgg_name="Rival Game", last_seen_at=now,
        )

    def test_dismiss_marks_row_and_rerenders_widget(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(f"/dashboard/sync-diffs/{self.diff.pk}/dismiss/")

        self.assertEqual(response.status_code, 200)
        self.diff.refresh_from_db()
        self.assertIsNotNone(self.diff.dismissed_at)
        # The re-rendered widget no longer lists the row.
        self.assertContains(response, "No unreviewed sync diffs.")
        self.assertNotContains(response, "Nemesis")

    def test_dismiss_scoping_404s(self):
        self.client.login(username="kernicek", password="pass")

        # Someone else's diff: 404, untouched.
        response = self.client.post(
            f"/dashboard/sync-diffs/{self.rival_diff.pk}/dismiss/",
        )
        self.assertEqual(response.status_code, 404)
        self.rival_diff.refresh_from_db()
        self.assertIsNone(self.rival_diff.dismissed_at)

        # An already-dismissed row is out of the widget — a second POST
        # (stale button) 404s instead of moving the timestamp.
        self.diff.dismissed_at = timezone.now()
        self.diff.save(update_fields=["dismissed_at"])
        response = self.client.post(f"/dashboard/sync-diffs/{self.diff.pk}/dismiss/")
        self.assertEqual(response.status_code, 404)

    def test_dismiss_requires_login_and_post(self):
        response = self.client.post(f"/dashboard/sync-diffs/{self.diff.pk}/dismiss/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/dashboard/sync-diffs/{self.diff.pk}/dismiss/")
        self.assertEqual(response.status_code, 405)


class SyncDiffAcceptTests(TestCase):
    """POST /dashboard/sync-diffs/<pk>/accept/ (issue #168): the BGG->app
    pull direction — mutates the actual Copy instead of just hiding the
    diff, then deletes the row since its condition is resolved."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.rival = get_user_model().objects.create_user(
            username="rival", password="pass",
        )
        now = timezone.now()

        cls.active_game = Game.objects.create(name="Nemesis")
        BggLink.objects.create(game=cls.active_game, bgg_id=167355, is_primary=True)
        cls.active_edition = Edition.objects.create(game=cls.active_game, name="")
        cls.active_copy = Copy.objects.create(
            owner=cls.user, edition=cls.active_edition,
            archive_status=Copy.ArchiveStatus.ACTIVE,
        )
        cls.prev_owned_diff = BggSyncDiff.objects.create(
            owner=cls.user, category=BggSyncDiff.Category.PREV_OWNED_ACTIVE,
            game=cls.active_game, bgg_id=167355, last_seen_at=now,
        )

        cls.archived_game = Game.objects.create(name="Gloomhaven")
        cls.archived_edition = Edition.objects.create(game=cls.archived_game, name="")
        cls.older_copy = Copy.objects.create(
            owner=cls.user, edition=cls.archived_edition,
            archive_status=Copy.ArchiveStatus.ARCHIVED,
            archive_reason=Copy.ArchiveReason.SOLD,
            archive_date=datetime.date(2024, 1, 1),
        )
        cls.newer_copy = Copy.objects.create(
            owner=cls.user, edition=cls.archived_edition,
            archive_status=Copy.ArchiveStatus.ARCHIVED,
            archive_reason=Copy.ArchiveReason.CULLED,
            archive_date=datetime.date(2024, 6, 1),
        )
        cls.archived_diff = BggSyncDiff.objects.create(
            owner=cls.user, category=BggSyncDiff.Category.ARCHIVED_ON_BGG,
            game=cls.archived_game, bgg_id=174430, last_seen_at=now,
            note="BGG still says own — fix by hand on BGG",
        )

        cls.other_diff = BggSyncDiff.objects.create(
            owner=cls.user, category=BggSyncDiff.Category.MISSING_FROM_BGG,
            game=cls.active_game, bgg_id=999999, last_seen_at=now,
        )
        cls.rival_diff = BggSyncDiff.objects.create(
            owner=cls.rival, category=BggSyncDiff.Category.PREV_OWNED_ACTIVE,
            game=cls.active_game, bgg_id=167355, last_seen_at=now,
        )

    def test_accept_prev_owned_active_archives_copy_and_deletes_diff(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            f"/dashboard/sync-diffs/{self.prev_owned_diff.pk}/accept/",
        )

        self.assertEqual(response.status_code, 200)
        self.active_copy.refresh_from_db()
        self.assertEqual(self.active_copy.archive_status, Copy.ArchiveStatus.ARCHIVED)
        self.assertEqual(self.active_copy.archive_reason, "")
        self.assertIsNone(self.active_copy.archive_date)
        self.assertFalse(
            BggSyncDiff.objects.filter(pk=self.prev_owned_diff.pk).exists(),
        )

    def test_accept_archived_on_bgg_reactivates_most_recent_copy(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            f"/dashboard/sync-diffs/{self.archived_diff.pk}/accept/",
        )

        self.assertEqual(response.status_code, 200)
        self.newer_copy.refresh_from_db()
        self.older_copy.refresh_from_db()
        self.assertEqual(self.newer_copy.archive_status, Copy.ArchiveStatus.ACTIVE)
        self.assertEqual(self.newer_copy.archive_reason, "")
        self.assertIsNone(self.newer_copy.archive_date)
        # The older archived copy is untouched.
        self.assertEqual(self.older_copy.archive_status, Copy.ArchiveStatus.ARCHIVED)
        self.assertEqual(self.older_copy.archive_reason, Copy.ArchiveReason.SOLD)
        self.assertFalse(
            BggSyncDiff.objects.filter(pk=self.archived_diff.pk).exists(),
        )

    def test_accept_clears_push_failed_diff_for_the_same_game(self):
        push_failed = BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.PUSH_FAILED,
            game=self.active_game, bgg_id=167355, last_seen_at=timezone.now(),
        )
        self.client.login(username="kernicek", password="pass")
        self.client.post(f"/dashboard/sync-diffs/{self.prev_owned_diff.pk}/accept/")

        self.assertFalse(BggSyncDiff.objects.filter(pk=push_failed.pk).exists())

    def test_accept_scoping_and_category_404s(self):
        self.client.login(username="kernicek", password="pass")

        # Someone else's diff: 404, untouched.
        response = self.client.post(
            f"/dashboard/sync-diffs/{self.rival_diff.pk}/accept/",
        )
        self.assertEqual(response.status_code, 404)
        self.assertTrue(BggSyncDiff.objects.filter(pk=self.rival_diff.pk).exists())

        # A category with no Copy-side accept action: 404, untouched.
        response = self.client.post(
            f"/dashboard/sync-diffs/{self.other_diff.pk}/accept/",
        )
        self.assertEqual(response.status_code, 404)
        self.assertTrue(BggSyncDiff.objects.filter(pk=self.other_diff.pk).exists())

    def test_accept_requires_login_and_post(self):
        response = self.client.post(
            f"/dashboard/sync-diffs/{self.prev_owned_diff.pk}/accept/",
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get(
            f"/dashboard/sync-diffs/{self.prev_owned_diff.pk}/accept/",
        )
        self.assertEqual(response.status_code, 405)


class NewExpansionDismissTests(TestCase):
    """POST /dashboard/new-expansions/<pk>/dismiss/ (issue #64): same
    per-user mark-seen semantics as sync_diff_dismiss, scoped to the
    NEW_EXPANSION category (kept as its own view so that already-tested one
    doesn't need to branch)."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.rival = get_user_model().objects.create_user(
            username="rival", password="pass",
        )
        now = timezone.now()
        cls.diff = BggSyncDiff.objects.create(
            owner=cls.user, category=BggSyncDiff.Category.NEW_EXPANSION,
            bgg_id=999001, bgg_name="Another New Threat", last_seen_at=now,
        )
        cls.rival_diff = BggSyncDiff.objects.create(
            owner=cls.rival, category=BggSyncDiff.Category.NEW_EXPANSION,
            bgg_id=555555, bgg_name="Rival Threat", last_seen_at=now,
        )
        cls.other_category_diff = BggSyncDiff.objects.create(
            owner=cls.user, category=BggSyncDiff.Category.SUGGEST_ADD,
            bgg_id=91072, bgg_name="Mystery Cult", last_seen_at=now,
        )

    def test_dismiss_marks_row_and_rerenders_widget(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(f"/dashboard/new-expansions/{self.diff.pk}/dismiss/")

        self.assertEqual(response.status_code, 200)
        self.diff.refresh_from_db()
        self.assertIsNotNone(self.diff.dismissed_at)
        self.assertContains(response, "No new expansions.")
        self.assertNotContains(response, "Another New Threat")

    def test_dismiss_scoping_404s(self):
        self.client.login(username="kernicek", password="pass")

        response = self.client.post(f"/dashboard/new-expansions/{self.rival_diff.pk}/dismiss/")
        self.assertEqual(response.status_code, 404)
        self.rival_diff.refresh_from_db()
        self.assertIsNone(self.rival_diff.dismissed_at)

        # A different category's diff isn't a new-expansion row -> 404 too.
        response = self.client.post(
            f"/dashboard/new-expansions/{self.other_category_diff.pk}/dismiss/",
        )
        self.assertEqual(response.status_code, 404)

        self.diff.dismissed_at = timezone.now()
        self.diff.save(update_fields=["dismissed_at"])
        response = self.client.post(f"/dashboard/new-expansions/{self.diff.pk}/dismiss/")
        self.assertEqual(response.status_code, 404)

    def test_dismiss_requires_login_and_post(self):
        response = self.client.post(f"/dashboard/new-expansions/{self.diff.pk}/dismiss/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/dashboard/new-expansions/{self.diff.pk}/dismiss/")
        self.assertEqual(response.status_code, 405)


class WishlistAddTests(TestCase):
    """POST /dashboard/new-expansions/<pk>/wishlist/ (issue #64): creates a
    stub Game + WishlistEntry from the owner-scoped diff row, then dismisses
    it too — DESIGN §8 frames dismiss and add-to-wishlist as alternative ways
    to handle the same row."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.rival = get_user_model().objects.create_user(
            username="rival", password="pass",
        )
        cls.base = Game.objects.create(name="5 Minute Dungeon")
        cls.diff = BggSyncDiff.objects.create(
            owner=cls.user, category=BggSyncDiff.Category.NEW_EXPANSION,
            bgg_id=999001, bgg_name="Another New Threat",
            note="New expansion for 5 Minute Dungeon", last_seen_at=timezone.now(),
        )

    def setUp(self):
        # Issue #117: wishlisting now enqueues a BGG push — the freshly
        # created stub Game gets a primary BggLink, so an unmocked .delay()
        # would try to reach a real broker. Mocked for every test here.
        patcher = mock.patch("gamekeeper.views.push_bgg_status_task.delay")
        self.mock_delay = patcher.start()
        self.addCleanup(patcher.stop)

    def test_add_creates_stub_game_and_wishlist_entry_and_dismisses_diff(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            f"/dashboard/new-expansions/{self.diff.pk}/wishlist/",
            {"priority": Game.WishlistPriority.LOVE_TO_HAVE.value},
        )

        self.assertEqual(response.status_code, 200)
        game = Game.objects.get(bgg_links__bgg_id=999001)
        self.assertEqual(game.name, "Another New Threat")
        self.assertEqual(game.type, Game.Type.EXPANSION)
        self.assertTrue(game.bgg_links.get(bgg_id=999001).is_primary)
        entry = WishlistEntry.objects.get(owner=self.user, game=game)
        self.assertEqual(entry.priority, Game.WishlistPriority.LOVE_TO_HAVE)
        # Issue #117: wishlisting pushes the same status to BGG.
        self.mock_delay.assert_called_once_with(
            game.pk, Game.BggCollectionStatus.WISHLIST, self.user.pk,
            priority=Game.WishlistPriority.LOVE_TO_HAVE,
        )

        self.diff.refresh_from_db()
        self.assertIsNotNone(self.diff.dismissed_at)
        self.assertContains(response, "No new expansions.")

    def test_add_backfills_matching_expansion_sighting(self):
        sighting = ExpansionSighting.objects.create(
            base=self.base, bgg_id=999001, bgg_name="Another New Threat",
        )
        self.client.login(username="kernicek", password="pass")
        self.client.post(
            f"/dashboard/new-expansions/{self.diff.pk}/wishlist/",
            {"priority": Game.WishlistPriority.LOVE_TO_HAVE.value},
        )

        sighting.refresh_from_db()
        self.assertIsNotNone(sighting.expansion)
        self.assertEqual(sighting.expansion.bgg_links.get().bgg_id, 999001)

    def test_add_reuses_existing_game_for_a_known_bgg_id(self):
        existing = Game.objects.create(name="Already Tracked", type=Game.Type.EXPANSION)
        BggLink.objects.create(game=existing, bgg_id=999001, is_primary=True)

        self.client.login(username="kernicek", password="pass")
        self.client.post(
            f"/dashboard/new-expansions/{self.diff.pk}/wishlist/",
            {"priority": Game.WishlistPriority.MUST_HAVE.value},
        )

        self.assertEqual(Game.objects.filter(bgg_links__bgg_id=999001).count(), 1)
        entry = WishlistEntry.objects.get(owner=self.user)
        self.assertEqual(entry.game, existing)

    def test_add_rejects_unknown_priority(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            f"/dashboard/new-expansions/{self.diff.pk}/wishlist/", {"priority": "99"},
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(WishlistEntry.objects.exists())

    def test_add_scoping_404s(self):
        rival_diff = BggSyncDiff.objects.create(
            owner=self.rival, category=BggSyncDiff.Category.NEW_EXPANSION,
            bgg_id=555555, bgg_name="Rival Threat", last_seen_at=timezone.now(),
        )
        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            f"/dashboard/new-expansions/{rival_diff.pk}/wishlist/",
            {"priority": Game.WishlistPriority.MUST_HAVE.value},
        )
        self.assertEqual(response.status_code, 404)

    def test_add_requires_login_and_post(self):
        response = self.client.post(
            f"/dashboard/new-expansions/{self.diff.pk}/wishlist/",
            {"priority": Game.WishlistPriority.MUST_HAVE.value},
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/dashboard/new-expansions/{self.diff.pk}/wishlist/")
        self.assertEqual(response.status_code, 405)


class WishlistListViewTests(TestCase):
    """GET /wishlist/ (issue #64): read-only list of the owner's entries —
    so WishlistEntry isn't a write-only black hole."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.rival = get_user_model().objects.create_user(
            username="rival", password="pass",
        )
        cls.game = Game.objects.create(name="Another New Threat", type=Game.Type.EXPANSION)
        WishlistEntry.objects.create(
            owner=cls.user, game=cls.game, priority=Game.WishlistPriority.MUST_HAVE,
        )
        rival_game = Game.objects.create(name="Rival's Pick")
        WishlistEntry.objects.create(
            owner=cls.rival, game=rival_game, priority=Game.WishlistPriority.MUST_HAVE,
        )

    def test_lists_only_the_owners_entries(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/wishlist/")

        self.assertContains(response, "Another New Threat")
        self.assertContains(response, "Must have")
        self.assertNotContains(response, "Rival's Pick")

    def test_requires_login(self):
        response = self.client.get("/wishlist/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)


class WishlistRemoveViewTests(TestCase):
    """POST /wishlist/<pk>/remove/ (issue #117): the first general-purpose
    removal path for WishlistEntry. Only pushes the drop to BGG when
    wishlist is actually the game's currently tracked BGG status."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.rival = get_user_model().objects.create_user(
            username="rival", password="pass",
        )

    def setUp(self):
        patcher = mock.patch("gamekeeper.views.push_bgg_status_task.delay")
        self.mock_delay = patcher.start()
        self.addCleanup(patcher.stop)

    def test_remove_deletes_entry_and_rerenders_table(self):
        game = Game.objects.create(name="Fading Interest")
        entry = WishlistEntry.objects.create(
            owner=self.user, game=game, priority=Game.WishlistPriority.MUST_HAVE,
        )
        self.client.login(username="kernicek", password="pass")

        response = self.client.post(f"/wishlist/{entry.pk}/remove/")

        self.assertEqual(response.status_code, 200)
        self.assertFalse(WishlistEntry.objects.filter(pk=entry.pk).exists())
        self.assertNotContains(response, "Fading Interest")
        self.mock_delay.assert_not_called()  # no primary BGG link -> nothing to push

    def test_remove_pushes_the_drop_when_wishlist_is_the_tracked_status(self):
        game = Game.objects.create(
            name="Tracked Want", bgg_collection_status=Game.BggCollectionStatus.WISHLIST,
        )
        BggLink.objects.create(game=game, bgg_id=123456, is_primary=True)
        entry = WishlistEntry.objects.create(
            owner=self.user, game=game, priority=Game.WishlistPriority.MUST_HAVE,
        )
        self.client.login(username="kernicek", password="pass")

        self.client.post(f"/wishlist/{entry.pk}/remove/")

        self.mock_delay.assert_called_once_with(game.pk, "", self.user.pk, priority=None)

    def test_remove_does_not_push_when_bgg_status_is_something_else(self):
        # Own on BGG (e.g. bought it) but a stale local wishlist entry lingers
        # — dropping the entry must not clear the real "own" status.
        game = Game.objects.create(
            name="Now Owned", bgg_collection_status=Game.BggCollectionStatus.OWN,
        )
        BggLink.objects.create(game=game, bgg_id=654321, is_primary=True)
        entry = WishlistEntry.objects.create(
            owner=self.user, game=game, priority=Game.WishlistPriority.MUST_HAVE,
        )
        self.client.login(username="kernicek", password="pass")

        self.client.post(f"/wishlist/{entry.pk}/remove/")

        self.mock_delay.assert_not_called()

    def test_remove_scoping_404s(self):
        game = Game.objects.create(name="Rival's Want")
        rival_entry = WishlistEntry.objects.create(
            owner=self.rival, game=game, priority=Game.WishlistPriority.MUST_HAVE,
        )
        self.client.login(username="kernicek", password="pass")

        response = self.client.post(f"/wishlist/{rival_entry.pk}/remove/")

        self.assertEqual(response.status_code, 404)
        self.assertTrue(WishlistEntry.objects.filter(pk=rival_entry.pk).exists())

    def test_remove_requires_login_and_post(self):
        game = Game.objects.create(name="Solo Want")
        entry = WishlistEntry.objects.create(
            owner=self.user, game=game, priority=Game.WishlistPriority.MUST_HAVE,
        )
        response = self.client.post(f"/wishlist/{entry.pk}/remove/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/wishlist/{entry.pk}/remove/")
        self.assertEqual(response.status_code, 405)


class DashboardFullListViewTests(TestCase):
    """Issue #83: the per-card full-list pages share the card's query + table
    partial but render every row uncapped."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def test_full_list_pages_require_login(self):
        for url in (
            "/dashboard/incoming-waves/",
            "/dashboard/pledge-managers/",
            "/dashboard/campaigns-ending/",
            "/dashboard/sync-diffs/",
            "/dashboard/to-craft/",
            "/dashboard/new-expansions/",
            "/wishlist/",
        ):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 302)
            self.assertIn("/accounts/login/", response.url)

    def test_incoming_full_page_lists_all_rows_uncapped(self):
        # More rows than the dashboard cap — all must render here.
        DashboardViewTests._make_incoming(self.user, 25)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/incoming-waves/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["incoming"]), 25)
        self.assertContains(response, "Incoming 00 GF")
        self.assertContains(response, "Incoming 24 GF")  # would be cut on the card
        self.assertNotContains(response, "Showing 20 of")

    def test_pledge_managers_full_page_lists_actionable(self):
        Purchase.objects.create(
            owner=self.user, name="Sent Out KS",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
        )
        actionable = Purchase.objects.filter(name="Sent Out KS").first()
        Wave.objects.create(
            purchase=actionable, number=1, status=Wave.Status.PRE_PRODUCTION,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/pledge-managers/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Sent Out KS")

    def test_pledge_managers_dated_sort_above_undated(self):
        # Issue #135: within the "Sent out" group, PMs with a close date sort
        # ascending above the date-less ones, which trail at the bottom.
        today = timezone.localdate()
        specs = [
            ("Undated A KS", None),
            ("Later KS", today + datetime.timedelta(days=20)),
            ("Undated B KS", None),
            ("Sooner KS", today + datetime.timedelta(days=5)),
        ]
        for name, close in specs:
            purchase = Purchase.objects.create(
                owner=self.user, name=name,
                status=Purchase.Status.COMMITTED,
                pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
                pledge_manager_close_date=close,
            )
            Wave.objects.create(
                purchase=purchase, number=1, status=Wave.Status.PRE_PRODUCTION,
            )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/pledge-managers/")

        ordered = [p.name for p in response.context["pm_actions"]]
        self.assertEqual(
            ordered, ["Sooner KS", "Later KS", "Undated A KS", "Undated B KS"],
        )

    def test_campaigns_ending_full_page_lists_watched(self):
        today = timezone.localdate()
        Purchase.objects.create(
            owner=self.user, name="Shiny New GF",
            status=Purchase.Status.WATCHING,
            campaign_end_date=today + datetime.timedelta(days=3),
        )
        # The undated-campaigns footnote rides along on the full page too.
        Purchase.objects.create(
            owner=self.user, name="Someday GF",
            status=Purchase.Status.WATCHING,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/campaigns-ending/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Shiny New GF")
        self.assertContains(response, "1 watched campaign")

    def test_sync_diffs_full_page_lists_all_with_dismiss_buttons(self):
        game = Game.objects.create(name="Nemesis")
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.MISSING_FROM_BGG,
            game=game, bgg_id=167355, last_seen_at=timezone.now(),
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/sync-diffs/")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nemesis")
        # Dismiss buttons stay interactive here, posting in full scope so the
        # re-render stays uncapped.
        self.assertContains(response, "Dismiss")
        self.assertContains(response, "?full=1")

    def test_sync_diff_dismiss_full_scope_stays_uncapped(self):
        # A dismiss from the full page (?full=1) re-renders every remaining row,
        # not just the first DASHBOARD_CARD_LIMIT.
        now = timezone.now()
        diffs = [
            BggSyncDiff.objects.create(
                owner=self.user, category=BggSyncDiff.Category.SUGGEST_ADD,
                bgg_id=1000 + i, bgg_name=f"Diff {i:02d}", last_seen_at=now,
            )
            for i in range(25)
        ]

        self.client.login(username="kernicek", password="pass")
        response = self.client.post(
            f"/dashboard/sync-diffs/{diffs[0].pk}/dismiss/?full=1",
        )

        self.assertEqual(response.status_code, 200)
        # 24 remain; all render (a capped re-render would drop the tail).
        self.assertContains(response, "Diff 24")
        self.assertNotContains(response, "Showing 20 of")

    def test_to_craft_full_page_lists_all_not_ready_copies_uncapped(self):
        # Issue #19: more not-ready copies than the dashboard cap — all must
        # render here.
        game = Game.objects.create(name="PnP Backlog")
        edition = Edition.objects.create(game=game, is_default=True)
        for i in range(25):
            other = Game.objects.create(name=f"PnP {i:02d}")
            other_edition = Edition.objects.create(game=other, is_default=True)
            Copy.objects.create(
                owner=self.user, edition=other_edition,
                ready_status=Copy.ReadyStatus.NOT_READY,
            )
        Copy.objects.create(
            owner=self.user, edition=edition,
            ready_status=Copy.ReadyStatus.READY,
        )

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/to-craft/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["to_craft"]), 25)
        self.assertContains(response, "PnP 00")
        self.assertContains(response, "PnP 24")  # would be cut on the card
        self.assertNotContains(response, "PnP Backlog")  # ready, not in the list


class StatusBadgeFilterTests(SimpleTestCase):
    """Issue #11: one shared source of truth maps each status value to a
    Bootstrap ``text-bg-*`` class, so badges stay consistent across pages."""

    def test_wave_status_class_is_a_distinct_colour_per_status(self):
        from gamekeeper.templatetags.status_badges import wave_status_class

        expected = {
            "pending": "text-bg-secondary",
            "pre_production": "text-bg-info",
            "production": "text-bg-primary",
            "fulfilment": "text-bg-warning",
            "arrived": "text-bg-success",
            "never_arrived": "text-bg-danger",
            "cancelled": "text-bg-dark",
        }
        # Every Wave.Status value is mapped...
        self.assertEqual(set(expected), set(Wave.Status.values))
        # ...and each of the seven maps to its own colour (no collisions).
        self.assertEqual(len(set(expected.values())), len(expected))
        for status, css in expected.items():
            self.assertEqual(wave_status_class(status), css)

    def test_pm_status_class(self):
        from gamekeeper.templatetags.status_badges import pm_status_class

        self.assertEqual(pm_status_class("sent_out"), "text-bg-warning")
        self.assertEqual(pm_status_class("filled_out"), "text-bg-info")
        self.assertEqual(pm_status_class("not_yet"), "text-bg-secondary")

    def test_purchase_status_class_matches_the_purchase_table_palette(self):
        from gamekeeper.templatetags.status_badges import purchase_status_class

        self.assertEqual(purchase_status_class("committed"), "text-bg-success")
        self.assertEqual(purchase_status_class("watching"), "text-bg-info")
        self.assertEqual(purchase_status_class("refunded"), "text-bg-warning")
        self.assertEqual(
            purchase_status_class("never_delivered"), "text-bg-danger"
        )

    def test_sync_diff_category_class_is_distinct_per_category(self):
        from gamekeeper.templatetags.status_badges import (
            sync_diff_category_class,
        )

        expected = {
            "suggest_add": "text-bg-info",
            "missing_from_bgg": "text-bg-warning",
            "prev_owned_active": "text-bg-danger",
            "archived_on_bgg": "text-bg-secondary",
            "new_expansion": "text-bg-success",
            "push_failed": "text-bg-primary",
        }
        self.assertEqual(set(expected), set(BggSyncDiff.Category.values))
        self.assertEqual(len(set(expected.values())), len(expected))
        for category, css in expected.items():
            self.assertEqual(sync_diff_category_class(category), css)

    def test_unknown_value_falls_back_to_secondary(self):
        from gamekeeper.templatetags.status_badges import (
            pm_status_class,
            purchase_status_class,
            sync_diff_category_class,
            wave_status_class,
        )

        for fn in (
            wave_status_class,
            pm_status_class,
            purchase_status_class,
            sync_diff_category_class,
        ):
            self.assertEqual(fn("bogus"), "text-bg-secondary")
            self.assertEqual(fn(None), "text-bg-secondary")


class WaveFulfilmentStatusTests(TestCase):
    """The non-terminal 'Fulfilment' wave status (#122): it sits between
    Production and Arrived, counts as incoming, and reads as in-progress."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.purchase = Purchase.objects.create(
            owner=cls.user, name="In Transit KS",
            status=Purchase.Status.COMMITTED,
        )
        cls.wave = Wave.objects.create(
            purchase=cls.purchase, number=1,
            status=Wave.Status.FULFILMENT,
            expected_arrival=datetime.date(2026, 9, 1),
        )

    def test_fulfilment_is_a_valid_non_terminal_status(self):
        self.assertIn("fulfilment", Wave.Status.values)
        self.assertNotIn(Wave.Status.FULFILMENT, Wave.TERMINAL_STATUSES)
        self.assertIn(Wave.Status.FULFILMENT, INCOMING_STATUSES)

    def test_fulfilment_wave_appears_in_dashboard_incoming_list(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/dashboard/")
        self.assertContains(response, "In Transit KS")
        self.assertContains(response, "Fulfilment")

    def test_purchase_detail_renders_fulfilment_badge_in_amber(self):
        # #11: the shared wave palette gives fulfilment its own amber slot,
        # distinct from production (blue), rather than sharing one colour.
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/purchases/{self.purchase.pk}/")
        self.assertContains(response, "text-bg-warning")
        self.assertContains(response, "Fulfilment")


class PurchasesViewTests(TestCase):
    """The §6 purchase-pipeline browse view (/purchases/) and per-purchase
    detail: owner-scoped, in-flight purchases first, filterable by status
    and platform via the same htmx rig as the curation table."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )

        # Committed, half arrived: wave 1 landed (its game product became a
        # Copy), wave 2 is a month late and still in production.
        cls.committed = Purchase.objects.create(
            owner=cls.user, name="Committed KS",
            platform=Purchase.Platform.KICKSTARTER,
            status=Purchase.Status.COMMITTED,
            pledge_manager=PledgeManager.objects.get(name="Gamefound"),
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
            pledge_manager_close_date=datetime.date(2026, 8, 15),
        )
        wave1 = Wave.objects.create(
            purchase=cls.committed, number=1, status=Wave.Status.ARRIVED,
            arrived_date=datetime.date(2026, 5, 1),
        )
        game = Game.objects.create(name="Delivered Game")
        edition = Edition.objects.create(game=game, is_default=True)
        copy = Copy.objects.create(owner=cls.user, edition=edition)
        Product.objects.create(
            wave=wave1, name="Delivered Game", kind=Product.Kind.GAME,
            game=game, edition=edition, copy=copy,
        )
        cls.wave2 = Wave.objects.create(
            purchase=cls.committed, number=2, status=Wave.Status.PRODUCTION,
            original_eta=datetime.date(2026, 8, 1),
            expected_arrival=datetime.date(2026, 9, 1),
        )
        Product.objects.create(
            wave=cls.wave2, name="Metal Coins", kind=Product.Kind.ACCESSORY,
        )

        # Committed and completely arrived -> derived "fulfilled". Its PM close
        # date is moot now that everything landed (#36) and must stay hidden.
        cls.done = Purchase.objects.create(
            owner=cls.user, name="Done KS",
            platform=Purchase.Platform.KICKSTARTER,
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.FILLED_OUT,
            pledge_manager_close_date=datetime.date(2027, 1, 1),
        )
        Wave.objects.create(
            purchase=cls.done, number=1, status=Wave.Status.ARRIVED,
            arrived_date=datetime.date(2026, 1, 10),
        )

        cls.watched = Purchase.objects.create(
            owner=cls.user, name="Watched GF",
            platform=Purchase.Platform.GAMEFOUND,
            status=Purchase.Status.WATCHING,
        )
        Wave.objects.create(purchase=cls.watched, number=1)

        cls.placeholder = Purchase.objects.create(
            owner=cls.user, name="Placeholder BK",
            platform=Purchase.Platform.BACKERKIT,
            status=Purchase.Status.PLACEHOLDER,
        )
        Wave.objects.create(purchase=cls.placeholder, number=1)

        # Passed: decided against. Its wave and PM close date are both moot —
        # nothing arrives (#94) and the date shouldn't render (#36).
        cls.dead = Purchase.objects.create(
            owner=cls.user, name="Dead KS",
            platform=Purchase.Platform.KICKSTARTER,
            status=Purchase.Status.PASSED,
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
            pledge_manager_close_date=datetime.date(2027, 2, 2),
        )
        Wave.objects.create(purchase=cls.dead, number=1)

        # Someone else's purchase must never show up.
        cls.foreign = Purchase.objects.create(
            owner=cls.other, name="Pavel's Secret KS",
            status=Purchase.Status.COMMITTED,
        )

    def get(self, params=None, **extra):
        self.client.login(username="kernicek", password="pass")
        return self.client.get("/purchases/", params or {}, **extra)

    def names(self, response):
        return [row["purchase"].name for row in response.context["rows"]]

    def test_anonymous_is_sent_to_login(self):
        for url in ("/purchases/", f"/purchases/{self.committed.pk}/"):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 302)
            self.assertIn("/accounts/login/", response["Location"])

    def test_pipeline_order_and_owner_scoping(self):
        response = self.get()
        # Active purchases first by stage (committed-in-flight, placeholder,
        # watching), then the settled ones sink to the bottom: the fully
        # fulfilled buy and finally the passed/dead one (#36). A placeholder
        # still needing a decision now outranks the fulfilled "Done KS".
        # Pavel's never shows.
        self.assertEqual(
            self.names(response),
            ["Committed KS", "Placeholder BK", "Watched GF", "Done KS", "Dead KS"],
        )
        self.assertContains(response, "5 of 5 purchases")
        self.assertNotContains(response, "Pavel")

    def test_placeholder_and_passed_hide_wave_progress(self):
        # Nothing arrives for a placeholder ($1 hold) or a passed buy, so their
        # Waves cell is blanked even when waves exist (#94).
        response = self.get()
        rows = {row["purchase"].name: row for row in response.context["rows"]}
        self.assertTrue(rows["Placeholder BK"]["hide_waves"])
        self.assertTrue(rows["Dead KS"]["hide_waves"])
        # An in-flight watch still shows its progress.
        self.assertFalse(rows["Watched GF"]["hide_waves"])

    def test_close_date_hidden_once_fulfilled_or_passed(self):
        # The PM close date only matters while a decision/delivery pends (#36):
        # gone for the fulfilled and the passed buy, kept for the in-flight one.
        response = self.get()
        self.assertContains(response, "closes 15 Aug 2026")  # Committed KS
        self.assertNotContains(response, "closes 1 Jan 2027")  # Done KS (fulfilled)
        self.assertNotContains(response, "closes 2 Feb 2027")  # Dead KS (passed)

    def test_wave_progress_and_fulfilled_badge(self):
        response = self.get()
        self.assertContains(response, "1/2 arrived")
        self.assertContains(response, "next ETA Sep 2026")
        self.assertContains(response, "fulfilled")
        self.assertContains(response, "closes 15 Aug 2026")

    def test_status_filter(self):
        response = self.get({"status": "watching"})
        self.assertEqual(self.names(response), ["Watched GF"])
        self.assertContains(response, "1 of 5 purchases")

    def test_platform_filter(self):
        response = self.get({"platform": "kickstarter"})
        self.assertEqual(
            self.names(response), ["Committed KS", "Done KS", "Dead KS"],
        )

    def test_name_search(self):
        response = self.get({"q": "committed"})
        self.assertEqual(self.names(response), ["Committed KS"])
        self.assertContains(response, "1 of 5 purchases")

    def test_name_search_folds_case_and_accents(self):
        response = self.get({"q": "PLACEHOLDÉR"})
        self.assertEqual(self.names(response), ["Placeholder BK"])

    def test_htmx_request_returns_just_the_table(self):
        response = self.get({"status": "watching"}, HTTP_HX_REQUEST="true")
        self.assertContains(response, "Watched GF")
        self.assertNotContains(response, "<form")

    def test_detail_shows_waves_products_and_links(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/purchases/{self.committed.pk}/")

        self.assertContains(response, "Committed KS")
        self.assertContains(response, "Wave 1")
        self.assertContains(response, "Wave 2")
        self.assertContains(response, "+31d vs. original ETA")
        self.assertContains(response, "Metal Coins")
        # The arrived game product links into the collection.
        game = Game.objects.get(name="Delivered Game")
        self.assertContains(response, f"/games/{game.pk}/")
        self.assertContains(response, "in collection")
        self.assertContains(response, "Sent out")

    def test_detail_items_link_to_detail_and_edit(self):
        # Issue #38: the item name opens the read-only product page, the
        # compact Edit button its edit page.
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/purchases/{self.committed.pk}/")
        coins = Product.objects.get(name="Metal Coins")
        self.assertContains(response, f'href="/products/{coins.pk}/"')
        self.assertContains(response, f"/products/{coins.pk}/edit/")

    def test_detail_of_fulfilled_purchase_shows_the_badge(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/purchases/{self.done.pk}/")
        self.assertContains(response, "fulfilled")

    def test_someone_elses_detail_is_404(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/purchases/{self.foreign.pk}/")
        self.assertEqual(response.status_code, 404)


class PurchaseEditingTests(TestCase):
    """Issue #5: user-facing purchase editing — the add and edit pages for
    purchases, the per-wave forms and the product line-items. Settled
    pages like copy_edit: whole-form POSTs and redirects, owner-scoped
    404s, 400s for values the form's own inputs constrain, inline error
    only for the one mistake a form can't prevent (duplicate name)."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.purchase = Purchase.objects.create(
            owner=cls.user, name="Editable KS",
            platform=Purchase.Platform.KICKSTARTER,
            status=Purchase.Status.COMMITTED,
        )
        cls.wave = Wave.objects.create(
            purchase=cls.purchase, number=1, status=Wave.Status.PRODUCTION,
        )
        cls.product = Product.objects.create(
            wave=cls.wave, name="Core Box", kind=Product.Kind.GAME,
        )
        cls.foreign = Purchase.objects.create(
            owner=cls.other, name="Pavel's KS",
        )
        cls.foreign_wave = Wave.objects.create(purchase=cls.foreign, number=1)
        cls.foreign_product = Product.objects.create(
            wave=cls.foreign_wave, name="Pavel's Box",
        )

    def login(self):
        self.client.login(username="kernicek", password="pass")

    def purchase_form(self, **overrides):
        # platform/status always ride along — the selects have no blank
        # option, so an absent value is browser-impossible.
        data = {"name": "Editable KS", "platform": "kickstarter",
                "status": "committed"}
        data.update(overrides)
        return data

    def wave_row(self, index, wave, **overrides):
        # One WaveFormSet row's POST data, defaulted from the wave's current
        # DB values so an "edit" test that doesn't care about waves is a
        # no-op on them (issue #136: purchase + waves save together now).
        row = {
            f"waves-{index}-id": str(wave.pk),
            f"waves-{index}-status": wave.status,
            f"waves-{index}-delivery_type": wave.delivery_type,
            f"waves-{index}-original_eta":
                wave.original_eta.isoformat() if wave.original_eta else "",
            f"waves-{index}-expected_arrival":
                wave.expected_arrival.isoformat() if wave.expected_arrival else "",
            f"waves-{index}-arrived_date":
                wave.arrived_date.isoformat() if wave.arrived_date else "",
            f"waves-{index}-address": wave.address,
            f"waves-{index}-tracking_url": wave.tracking_url,
        }
        for field, value in overrides.items():
            row[f"waves-{index}-{field}"] = value
        return row

    def edit_form(self, waves=(), purchase_overrides=None, wave_overrides=None,
                  delete_products=()):
        waves = list(waves)
        data = self.purchase_form(**(purchase_overrides or {}))
        data.update({
            "waves-TOTAL_FORMS": str(len(waves)),
            "waves-INITIAL_FORMS": str(len(waves)),
            "waves-MIN_NUM_FORMS": "0",
            "waves-MAX_NUM_FORMS": "1000",
        })
        wave_overrides = wave_overrides or {}
        for i, wave in enumerate(waves):
            data.update(self.wave_row(i, wave, **wave_overrides.get(i, {})))
        if delete_products:
            data["delete_products"] = [str(pk) for pk in delete_products]
        return data

    def edit_post(self, pk=None, **kwargs):
        return self.client.post(
            f"/purchases/{pk or self.purchase.pk}/edit/", self.edit_form(**kwargs))

    # --- add ---

    def test_add_requires_login(self):
        response = self.client.get("/purchases/add/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_add_get_renders_the_campaign_form(self):
        self.login()
        response = self.client.get("/purchases/add/")
        self.assertContains(response, "New purchase")
        self.assertContains(response, 'name="pledge_manager_status"')

    def test_add_creates_purchase_with_wave_one(self):
        self.login()
        gamefound = PledgeManager.objects.get(name="Gamefound")
        response = self.client.post("/purchases/add/", self.purchase_form(
            name="Shiny New GF", platform="gamefound", status="watching",
            campaign_url="https://gamefound.com/shiny",
            campaign_end_date="2026-09-30", excitement="8.5",
            pledge_manager=str(gamefound.pk), pledge_manager_status="not_yet",
            comments="looks great",
        ))
        purchase = Purchase.objects.get(owner=self.user, name="Shiny New GF")
        self.assertRedirects(response, f"/purchases/{purchase.pk}/edit/")
        self.assertEqual(purchase.platform, Purchase.Platform.GAMEFOUND)
        self.assertEqual(purchase.status, Purchase.Status.WATCHING)
        self.assertEqual(purchase.campaign_end_date, datetime.date(2026, 9, 30))
        self.assertEqual(purchase.excitement, Decimal("8.5"))
        self.assertEqual(purchase.comments, "looks great")
        self.assertEqual(purchase.pledge_manager, gamefound)
        # DESIGN §6: a purchase auto-creates "Wave 1".
        self.assertEqual(
            list(purchase.waves.values_list("number", flat=True)), [1],
        )

    def test_add_duplicate_name_renders_inline_error(self):
        self.login()
        response = self.client.post("/purchases/add/", self.purchase_form(
            comments="typed a lot here",
        ))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "already have a purchase with that name")
        # The typed values survive the re-render.
        self.assertContains(response, "typed a lot here")
        self.assertEqual(
            Purchase.objects.filter(owner=self.user, name="Editable KS").count(), 1,
        )

    def test_add_bad_values_are_rejected_inline(self):
        # PurchaseForm (issue #136): invalid input re-renders at 200 with
        # inline errors, same as the duplicate-name case above.
        self.login()
        for overrides in (
            {"name": "   "},
            {"name": "X", "status": "solvent"},
            {"name": "X", "platform": "etsy"},
            {"name": "X", "campaign_end_date": "soon"},
            {"name": "X", "excitement": "great"},
            {"name": "X", "excitement": "11"},
            {"name": "X", "campaign_url": "javascript:alert(1)"},
        ):
            response = self.client.post(
                "/purchases/add/", self.purchase_form(**overrides))
            self.assertEqual(response.status_code, 200, overrides)
        self.assertFalse(Purchase.objects.filter(name="X").exists())

    # --- edit ---

    def test_edit_get_shows_the_waves_and_products(self):
        self.login()
        response = self.client.get(f"/purchases/{self.purchase.pk}/edit/")
        self.assertContains(response, "Editable KS")
        self.assertContains(response, "Wave 1")
        self.assertContains(response, "Core Box")
        self.assertContains(response, "Add wave")

    def test_edit_items_link_to_detail_and_edit(self):
        # Issue #38: the item name opens the read-only detail page, the
        # compact Edit button the edit page.
        self.login()
        response = self.client.get(f"/purchases/{self.purchase.pk}/edit/")
        self.assertContains(response, f'href="/products/{self.product.pk}/"')
        self.assertContains(response, f"/products/{self.product.pk}/edit/")

    def test_edit_saves_and_returns_to_detail(self):
        self.login()
        response = self.edit_post(
            waves=[self.wave],
            purchase_overrides={
                "name": "Renamed KS", "pledge_manager_status": "sent_out",
                "pledge_manager_close_date": "2026-08-15", "ordered_date": "2026-01-05",
            },
        )
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/")
        self.purchase.refresh_from_db()
        self.assertEqual(self.purchase.name, "Renamed KS")
        self.assertEqual(self.purchase.pledge_manager_status,
                         Purchase.PledgeManagerStatus.SENT_OUT)
        self.assertEqual(self.purchase.pledge_manager_close_date,
                         datetime.date(2026, 8, 15))
        self.assertEqual(self.purchase.ordered_date, datetime.date(2026, 1, 5))

    def test_edit_keeping_the_name_is_not_a_duplicate(self):
        self.login()
        response = self.edit_post(waves=[self.wave])
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/")

    def test_edit_renaming_onto_another_purchase_is_an_inline_error(self):
        Purchase.objects.create(owner=self.user, name="Taken KS")
        self.login()
        response = self.edit_post(
            waves=[self.wave], purchase_overrides={"name": "Taken KS"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "already have a purchase with that name")
        self.purchase.refresh_from_db()
        self.assertEqual(self.purchase.name, "Editable KS")

    def test_editing_campaign_and_wave_together_saves_both(self):
        # The actual point of #136: editing two sections and saving once
        # persists both instead of one silently dropping the other.
        self.login()
        response = self.edit_post(
            waves=[self.wave],
            purchase_overrides={"name": "Renamed KS"},
            wave_overrides={0: {"status": "arrived"}},
        )
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/")
        self.purchase.refresh_from_db()
        self.wave.refresh_from_db()
        self.assertEqual(self.purchase.name, "Renamed KS")
        self.assertEqual(self.wave.status, Wave.Status.ARRIVED)

    def test_invalid_wave_field_does_not_drop_the_campaign_edit(self):
        # All-or-nothing (issue #136): a mistake in one section doesn't
        # silently lose a valid, in-progress edit in another.
        self.login()
        response = self.edit_post(
            waves=[self.wave],
            purchase_overrides={"name": "Renamed KS"},
            wave_overrides={0: {"status": "teleported"}},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="Renamed KS"')
        self.purchase.refresh_from_db()
        self.assertEqual(self.purchase.name, "Editable KS")

    def test_someone_elses_purchase_is_404_everywhere(self):
        self.login()
        for method, url in (
            ("get", f"/purchases/{self.foreign.pk}/edit/"),
            ("post", f"/purchases/{self.foreign.pk}/edit/"),
            ("post", f"/purchases/{self.foreign.pk}/waves/add/"),
            ("post", f"/waves/{self.foreign_wave.pk}/products/add/"),
            ("get", f"/products/{self.foreign_product.pk}/edit/"),
            ("post", f"/products/{self.foreign_product.pk}/edit/"),
        ):
            response = getattr(self.client, method)(url)
            self.assertEqual(response.status_code, 404, url)

    # --- waves ---

    def test_wave_add_appends_the_next_number(self):
        self.login()
        response = self.client.post(f"/purchases/{self.purchase.pk}/waves/add/")
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/edit/")
        self.assertEqual(
            list(self.purchase.waves.values_list("number", flat=True)), [1, 2],
        )

    def test_wave_add_is_post_only(self):
        self.login()
        response = self.client.get(f"/purchases/{self.purchase.pk}/waves/add/")
        self.assertEqual(response.status_code, 405)

    def test_wave_edit_saves_the_card(self):
        self.login()
        response = self.edit_post(waves=[self.wave], wave_overrides={0: {
            "status": "arrived", "delivery_type": "physical",
            "original_eta": "2026-02-01", "expected_arrival": "2026-03-01",
            "arrived_date": "2026-03-05", "address": "Home, Prague",
            "tracking_url": "https://tracking.example/abc",
        }})
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/")
        self.wave.refresh_from_db()
        self.assertEqual(self.wave.status, Wave.Status.ARRIVED)
        self.assertEqual(self.wave.arrived_date, datetime.date(2026, 3, 5))
        self.assertEqual(self.wave.address, "Home, Prague")
        self.assertEqual(self.wave.tracking_url, "https://tracking.example/abc")

    def test_wave_edit_blank_dates_clear(self):
        self.wave.original_eta = datetime.date(2026, 2, 1)
        self.wave.save()
        self.login()
        self.edit_post(waves=[self.wave], wave_overrides={0: {
            "status": "production", "delivery_type": "physical",
            "original_eta": "",
        }})
        self.wave.refresh_from_db()
        self.assertIsNone(self.wave.original_eta)

    def test_wave_edit_bad_values_are_rejected_inline(self):
        self.login()
        for overrides in (
            {"status": "teleported"},
            {"delivery_type": "carrier_pigeon"},
            {"arrived_date": "yesterday"},
            {"tracking_url": "ftp://tracking"},
        ):
            response = self.edit_post(
                waves=[self.wave], wave_overrides={0: overrides})
            self.assertEqual(response.status_code, 200, overrides)
        self.wave.refresh_from_db()
        self.assertEqual(self.wave.status, Wave.Status.PRODUCTION)

    def test_wave_delete_only_deletes_empty_waves(self):
        self.login()
        response = self.edit_post(
            waves=[self.wave], wave_overrides={0: {"DELETE": "on"}})
        self.assertEqual(response.status_code, 200)  # has Core Box on it
        self.assertContains(response, "still has items on it")
        self.assertTrue(Wave.objects.filter(pk=self.wave.pk).exists())

        empty = Wave.objects.create(purchase=self.purchase, number=2)
        response = self.edit_post(
            waves=[self.wave, empty], wave_overrides={1: {"DELETE": "on"}})
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/")
        self.assertFalse(Wave.objects.filter(pk=empty.pk).exists())

    # --- products ---

    def test_product_add_creates_and_jumps_to_its_edit_page(self):
        self.login()
        response = self.client.post(
            f"/waves/{self.wave.pk}/products/add/",
            {"name": "Metal Coins", "kind": "accessory"},
        )
        product = Product.objects.get(wave=self.wave, name="Metal Coins")
        self.assertRedirects(response, f"/products/{product.pk}/edit/")
        self.assertEqual(product.kind, Product.Kind.ACCESSORY)

    def test_product_add_bad_values_are_400s(self):
        self.login()
        for data in (
            {"name": "   "},
            {"name": "Core Box"},               # duplicate on this wave
            {"name": "Thing", "kind": "spaceship"},
        ):
            response = self.client.post(
                f"/waves/{self.wave.pk}/products/add/", data)
            self.assertEqual(response.status_code, 400, data)

    def test_product_edit_saves_game_edition_and_contents(self):
        game = Game.objects.create(name="Delivered Game")
        edition = Edition.objects.create(game=game, is_default=True)
        self.login()
        response = self.client.post(f"/products/{self.product.pk}/edit/", {
            "name": "Core Box Deluxe", "kind": "game",
            "game": str(game.pk), "edition": str(edition.pk),
            "contains_cards": "yes", "needs_sleeves": "unknown",
            "miniatures_count": "12", "fits_sleeved_note": "probably",
            "insert_3d_note": "print the organizer",
            "bgg_url": "https://boardgamegeek.com/boardgame/1",
            "notes": "deluxe pledge",
        })
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/edit/")
        self.product.refresh_from_db()
        self.assertEqual(self.product.name, "Core Box Deluxe")
        self.assertEqual(self.product.game, game)
        self.assertEqual(self.product.edition, edition)
        self.assertEqual(self.product.contains_cards, Product.TriState.YES)
        self.assertEqual(self.product.miniatures_count, 12)
        self.assertEqual(self.product.insert_3d_note, "print the organizer")

    def test_product_edit_edition_of_another_game_is_cleared(self):
        # The edition select trails the saved game, so switching the game
        # posts a stale edition — it clears instead of erroring.
        game = Game.objects.create(name="Game A")
        other_game = Game.objects.create(name="Game B")
        stale = Edition.objects.create(game=other_game, is_default=True)
        self.product.game = other_game
        self.product.edition = stale
        self.product.save()
        self.login()
        self.client.post(f"/products/{self.product.pk}/edit/", {
            "name": "Core Box", "kind": "game",
            "game": str(game.pk), "edition": str(stale.pk),
        })
        self.product.refresh_from_db()
        self.assertEqual(self.product.game, game)
        self.assertIsNone(self.product.edition)

    def test_product_edit_clearing_the_game_clears_the_edition(self):
        game = Game.objects.create(name="Game A")
        edition = Edition.objects.create(game=game, is_default=True)
        self.product.game = game
        self.product.edition = edition
        self.product.save()
        self.login()
        self.client.post(f"/products/{self.product.pk}/edit/", {
            "name": "Core Box", "kind": "other",
            "game": "", "edition": str(edition.pk),
        })
        self.product.refresh_from_db()
        self.assertIsNone(self.product.game)
        self.assertIsNone(self.product.edition)

    def test_product_edit_bad_values_are_400s(self):
        Product.objects.create(wave=self.wave, name="Taken")
        self.login()
        base = {"name": "Core Box", "kind": "game"}
        for overrides in (
            {"name": ""},
            {"name": "Taken"},                  # duplicate on this wave
            {"kind": "spaceship"},
            {"game": "999999"},
            {"contains_cards": "maybe"},
            {"miniatures_count": "lots"},
            {"miniatures_count": "-3"},
            {"drive_url": "file:///etc/passwd"},
        ):
            response = self.client.post(
                f"/products/{self.product.pk}/edit/", {**base, **overrides})
            self.assertEqual(response.status_code, 400, overrides)

    def test_product_delete(self):
        self.login()
        response = self.edit_post(
            waves=[self.wave], delete_products=[self.product.pk])
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/")
        self.assertFalse(Product.objects.filter(pk=self.product.pk).exists())

    def test_converted_products_cannot_be_deleted(self):
        game = Game.objects.create(name="Delivered Game")
        edition = Edition.objects.create(game=game, is_default=True)
        copy = Copy.objects.create(owner=self.user, edition=edition)
        self.product.copy = copy
        self.product.save()
        self.login()
        response = self.edit_post(
            waves=[self.wave], delete_products=[self.product.pk])
        self.assertRedirects(response, f"/purchases/{self.purchase.pk}/")
        self.assertTrue(Product.objects.filter(pk=self.product.pk).exists())


class ProductDetailTests(TestCase):
    """Issue #38: the read-only product detail page, linked from the item
    name on the purchase edit table."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.game = Game.objects.create(name="Boxed Game")
        cls.purchase = Purchase.objects.create(
            owner=cls.user, name="Detail KS",
            status=Purchase.Status.COMMITTED,
        )
        cls.wave = Wave.objects.create(purchase=cls.purchase, number=1)
        cls.product = Product.objects.create(
            wave=cls.wave, name="Core Box", kind=Product.Kind.GAME,
            game=cls.game, contains_cards=Product.TriState.YES,
            miniatures_count=12, notes="deluxe pledge",
        )
        foreign = Purchase.objects.create(owner=cls.other, name="Pavel's KS")
        foreign_wave = Wave.objects.create(purchase=foreign, number=1)
        cls.foreign_product = Product.objects.create(
            wave=foreign_wave, name="Pavel's Box",
        )

    def test_requires_login(self):
        response = self.client.get(f"/products/{self.product.pk}/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_detail_shows_fields_and_links(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/products/{self.product.pk}/")
        self.assertContains(response, "Core Box")
        self.assertContains(response, "Board game")
        self.assertContains(response, f"/games/{self.game.pk}/")
        self.assertContains(response, f"/products/{self.product.pk}/edit/")
        # Back link to the owning purchase's detail page.
        self.assertContains(response, f"/purchases/{self.purchase.pk}/")
        self.assertContains(response, "Yes")
        self.assertContains(response, "12")
        self.assertContains(response, "deluxe pledge")

    def test_someone_elses_product_is_404(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/products/{self.foreign_product.pk}/")
        self.assertEqual(response.status_code, 404)


class ProductConvertTests(TestCase):
    """The §6 arrival seam (issue #5): an arrived game product converts
    into a Copy and lands on the copy edit page, carrying the purchase
    origin so save returns there (#45). Provisional product data (arrival
    date, 3D-insert plan, sleeve counts) moves along."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(name="Arrived Game")
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)
        cls.purchase = Purchase.objects.create(
            owner=cls.user, name="Arrived KS",
            status=Purchase.Status.COMMITTED,
        )
        cls.wave = Wave.objects.create(
            purchase=cls.purchase, number=1, status=Wave.Status.ARRIVED,
            arrived_date=datetime.date(2026, 6, 15),
        )
        cls.product = Product.objects.create(
            wave=cls.wave, name="Arrived Game", kind=Product.Kind.GAME,
            game=cls.game, insert_3d_note="print the tray",
        )

    def convert(self, product=None, data=None):
        self.client.login(username="kernicek", password="pass")
        target = product or self.product
        return self.client.post(f"/products/{target.pk}/convert/", data or {})

    def test_convert_creates_the_copy_and_links_the_product(self):
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            response = self.convert(data={"edition": str(self.edition.pk)})
        copy = Copy.objects.get(owner=self.user, edition=self.edition)
        # Edit the fresh copy, carrying the purchase origin so save returns
        # there (#45).
        self.assertRedirects(
            response, f"/copies/{copy.pk}/edit/?from_purchase={self.purchase.pk}")
        # Provisional data moved over: arrival date and the 3D-insert plan.
        self.assertEqual(copy.acquired_date, datetime.date(2026, 6, 15))
        self.assertEqual(copy.upgrades_note, "print the tray")
        self.product.refresh_from_db()
        self.assertEqual(self.product.copy, copy)
        self.assertEqual(self.product.edition, self.edition)
        # self.game has no BggLink — nothing to push.
        delay.assert_not_called()

    def test_convert_pushes_own_status_to_bgg_when_the_game_is_linked(self):
        """Issue #117: an arrived preorder becoming an active copy is the
        same "own" transition as copy_add."""
        BggLink.objects.create(game=self.game, bgg_id=207830, is_primary=True)
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.convert(data={"edition": str(self.edition.pk)})
        delay.assert_called_once_with(
            self.game.pk, Game.BggCollectionStatus.OWN, self.user.pk, priority=None,
        )

    def test_convert_creates_a_ready_copy_for_a_non_pnp_edition(self):
        self.convert(data={"edition": str(self.edition.pk)})
        copy = Copy.objects.get(owner=self.user, edition=self.edition)
        self.assertEqual(copy.ready_status, Copy.ReadyStatus.READY)

    def test_convert_creates_a_not_ready_copy_for_a_pnp_edition(self):
        # Issue #19: a PnP arrival is files, not a playable game yet.
        self.edition.is_pnp = True
        self.edition.save()
        self.convert(data={"edition": str(self.edition.pk)})
        copy = Copy.objects.get(owner=self.user, edition=self.edition)
        self.assertEqual(copy.ready_status, Copy.ReadyStatus.NOT_READY)

    def test_convert_onto_an_existing_copy_does_not_force_not_ready(self):
        # Reactivating/linking an existing copy (the common "already added by
        # hand" case) leaves its ready_status alone even for a PnP edition —
        # only newly-created copies get the not-ready default.
        self.edition.is_pnp = True
        self.edition.save()
        existing = Copy.objects.create(owner=self.user, edition=self.edition)
        self.convert(data={"edition": str(self.edition.pk)})
        existing.refresh_from_db()
        self.assertEqual(existing.ready_status, Copy.ReadyStatus.READY)

    def test_convert_uses_the_products_pinned_edition(self):
        self.product.edition = self.edition
        self.product.save()
        self.convert()  # no edition posted — the product already knows
        self.product.refresh_from_db()
        self.assertEqual(self.product.copy.edition, self.edition)

    def test_convert_links_an_existing_active_copy(self):
        existing = Copy.objects.create(owner=self.user, edition=self.edition)
        response = self.convert(data={"edition": str(self.edition.pk)})
        self.assertRedirects(
            response,
            f"/copies/{existing.pk}/edit/?from_purchase={self.purchase.pk}")
        self.product.refresh_from_db()
        self.assertEqual(self.product.copy, existing)
        self.assertEqual(Copy.objects.count(), 1)

    def test_convert_onto_an_archived_copy_is_the_rebuy_stance(self):
        Copy.objects.create(
            owner=self.user, edition=self.edition,
            archive_status=Copy.ArchiveStatus.ARCHIVED,
            archive_reason=Copy.ArchiveReason.SOLD,
        )
        response = self.convert(data={"edition": str(self.edition.pk)})
        self.assertEqual(response.status_code, 400)
        self.assertEqual(Copy.objects.count(), 1)

    def test_convert_editionless_game_creates_the_default_edition(self):
        bare = Game.objects.create(name="Bare Game")
        product = Product.objects.create(
            wave=self.wave, name="Bare Game", kind=Product.Kind.GAME, game=bare,
        )
        self.convert(product=product)
        product.refresh_from_db()
        self.assertTrue(product.edition.is_default)
        self.assertEqual(product.copy.edition.game, bare)

    def test_convert_with_editions_but_no_pick_is_a_400(self):
        Edition.objects.create(game=self.game, name="Collector's")
        response = self.convert()
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Copy.objects.exists())

    def test_convert_guards(self):
        # Not arrived yet.
        pending = Wave.objects.create(
            purchase=self.purchase, number=2, status=Wave.Status.PRODUCTION,
        )
        early = Product.objects.create(
            wave=pending, name="Early", kind=Product.Kind.GAME, game=self.game,
        )
        self.assertEqual(self.convert(product=early).status_code, 400)
        # Not a game.
        coins = Product.objects.create(
            wave=self.wave, name="Coins", kind=Product.Kind.ACCESSORY,
        )
        self.assertEqual(self.convert(product=coins).status_code, 400)
        # Already converted.
        copy = Copy.objects.create(owner=self.user, edition=self.edition)
        self.product.copy = copy
        self.product.save()
        self.assertEqual(self.convert().status_code, 400)

    def test_convert_transfers_sleeve_requirements_to_the_edition(self):
        standard = CardSize.objects.create(
            width_mm=Decimal("63.5"), height_mm=Decimal("88.0"), name="Standard",
        )
        mini = CardSize.objects.create(
            width_mm=Decimal("44.0"), height_mm=Decimal("68.0"), name="Mini",
        )
        ProductSleeveRequirement.objects.create(
            product=self.product, card_size=standard, count=120,
        )
        ProductSleeveRequirement.objects.create(
            product=self.product, card_size=mini, count=40,
        )
        # An already-curated edition count wins over the provisional one.
        SleeveRequirement.objects.create(
            edition=self.edition, card_size=standard, count=150,
        )
        self.convert(data={"edition": str(self.edition.pk)})
        counts = {
            r.card_size_id: r.count
            for r in SleeveRequirement.objects.filter(edition=self.edition)
        }
        self.assertEqual(counts, {standard.pk: 150, mini.pk: 40})

    def test_someone_elses_product_is_404(self):
        other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        theirs = Purchase.objects.create(owner=other, name="Pavel's KS")
        their_wave = Wave.objects.create(
            purchase=theirs, number=1, status=Wave.Status.ARRIVED,
        )
        their_product = Product.objects.create(
            wave=their_wave, name="Theirs", kind=Product.Kind.GAME,
            game=self.game,
        )
        self.assertEqual(self.convert(product=their_product).status_code, 404)

    def test_detail_page_offers_convert_only_when_it_applies(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/purchases/{self.purchase.pk}/")
        self.assertContains(response, f"/products/{self.product.pk}/convert/")

        self.product.copy = Copy.objects.create(
            owner=self.user, edition=self.edition,
        )
        self.product.save()
        response = self.client.get(f"/purchases/{self.purchase.pk}/")
        self.assertNotContains(response, f"/products/{self.product.pk}/convert/")
        self.assertContains(response, "in collection")


class PreorderPushTests(TestCase):
    """Issue #166: BGG write-back for 'preordered', derived from the same
    COMMITTED-purchase + non-terminal-wave rule the dashboard's "incoming"
    card already uses (_incoming_rows, views.py:118-135). Recomputed from
    scratch on every product_edit/purchase_edit save rather than tracked as
    a delta, so a second in-flight preorder of the same game is a harmless
    no-op and dropping one of two leaves the status intact. Mirrors the
    view-level push assertions in ProductConvertTests/CurationArchiveTests —
    mocks push_bgg_status_task.delay and asserts the call, or its absence."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(name="Preordered Game")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)
        cls.purchase = Purchase.objects.create(
            owner=cls.user, name="Incoming KS",
            status=Purchase.Status.COMMITTED,
        )
        cls.wave = Wave.objects.create(
            purchase=cls.purchase, number=1, status=Wave.Status.PRODUCTION,
        )
        cls.product = Product.objects.create(
            wave=cls.wave, name="Core Box", kind=Product.Kind.GAME,
        )

    def login(self):
        self.client.login(username="kernicek", password="pass")

    def edit_product(self, product=None, **overrides):
        target = product or self.product
        data = {"name": target.name, "kind": target.kind or Product.Kind.GAME}
        data.update(overrides)
        return self.client.post(f"/products/{target.pk}/edit/", data)

    def wave_row(self, index, wave, **overrides):
        row = {
            f"waves-{index}-id": str(wave.pk),
            f"waves-{index}-status": wave.status,
            f"waves-{index}-delivery_type": wave.delivery_type,
            f"waves-{index}-original_eta":
                wave.original_eta.isoformat() if wave.original_eta else "",
            f"waves-{index}-expected_arrival":
                wave.expected_arrival.isoformat() if wave.expected_arrival else "",
            f"waves-{index}-arrived_date":
                wave.arrived_date.isoformat() if wave.arrived_date else "",
            f"waves-{index}-address": wave.address,
            f"waves-{index}-tracking_url": wave.tracking_url,
        }
        for field, value in overrides.items():
            row[f"waves-{index}-{field}"] = value
        return row

    def edit_purchase(self, purchase=None, waves=(), purchase_overrides=None,
                       wave_overrides=None, delete_products=()):
        purchase = purchase or self.purchase
        waves = list(waves)
        data = {
            "name": purchase.name, "platform": purchase.platform or Purchase.Platform.OTHER,
            "status": purchase.status,
        }
        data.update(purchase_overrides or {})
        data.update({
            "waves-TOTAL_FORMS": str(len(waves)),
            "waves-INITIAL_FORMS": str(len(waves)),
            "waves-MIN_NUM_FORMS": "0",
            "waves-MAX_NUM_FORMS": "1000",
        })
        wave_overrides = wave_overrides or {}
        for i, wave in enumerate(waves):
            data.update(self.wave_row(i, wave, **wave_overrides.get(i, {})))
        if delete_products:
            data["delete_products"] = [str(pk) for pk in delete_products]
        self.login()
        return self.client.post(f"/purchases/{purchase.pk}/edit/", data)

    def test_product_edit_linking_a_game_pushes_preordered(self):
        self.login()
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_product(game=str(self.game.pk))
        delay.assert_called_once_with(
            self.game.pk, Game.BggCollectionStatus.PREORDERED, self.user.pk,
            priority=None,
        )

    def test_product_edit_linking_a_game_without_bgglink_does_not_push(self):
        unlinked = Game.objects.create(name="No BGG Link")
        self.login()
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_product(game=str(unlinked.pk))
        delay.assert_not_called()

    def test_product_edit_switching_the_game_clears_the_old_and_pushes_the_new(self):
        self.product.game = self.game
        self.product.save()
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREORDERED
        self.game.save(update_fields=["bgg_collection_status"])
        other = Game.objects.create(name="Other Preorder")
        BggLink.objects.create(game=other, bgg_id=300001, is_primary=True)
        self.login()
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_product(game=str(other.pk))
        self.assertEqual(delay.call_count, 2)
        delay.assert_any_call(
            self.game.pk, "", self.user.pk, priority=None,
        )
        delay.assert_any_call(
            other.pk, Game.BggCollectionStatus.PREORDERED, self.user.pk, priority=None,
        )

    def test_product_edit_unlinking_the_game_clears_preordered(self):
        self.product.game = self.game
        self.product.save()
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREORDERED
        self.game.save(update_fields=["bgg_collection_status"])
        self.login()
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_product(game="")
        delay.assert_called_once_with(
            self.game.pk, "", self.user.pk, priority=None,
        )

    def test_purchase_status_becoming_committed_pushes_preordered(self):
        self.purchase.status = Purchase.Status.WATCHING
        self.purchase.save(update_fields=["status"])
        self.product.game = self.game
        self.product.save()
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_purchase(
                waves=[self.wave], purchase_overrides={"status": "committed"},
            )
        delay.assert_called_once_with(
            self.game.pk, Game.BggCollectionStatus.PREORDERED, self.user.pk,
            priority=None,
        )

    def test_wave_cancelled_clears_preordered(self):
        self.product.game = self.game
        self.product.save()
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREORDERED
        self.game.save(update_fields=["bgg_collection_status"])
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_purchase(
                waves=[self.wave], wave_overrides={0: {"status": "cancelled"}},
            )
        delay.assert_called_once_with(
            self.game.pk, "", self.user.pk, priority=None,
        )

    def test_purchase_passed_clears_preordered(self):
        self.product.game = self.game
        self.product.save()
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREORDERED
        self.game.save(update_fields=["bgg_collection_status"])
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_purchase(
                waves=[self.wave], purchase_overrides={"status": "passed"},
            )
        delay.assert_called_once_with(
            self.game.pk, "", self.user.pk, priority=None,
        )

    def test_deleting_the_last_incoming_product_clears_preordered(self):
        self.product.game = self.game
        self.product.save()
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREORDERED
        self.game.save(update_fields=["bgg_collection_status"])
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_purchase(
                waves=[self.wave], delete_products=[self.product.pk],
            )
        delay.assert_called_once_with(
            self.game.pk, "", self.user.pk, priority=None,
        )

    def test_second_incoming_purchase_for_the_same_game_is_a_no_op_repush(self):
        # Design question #3 (issue #166): a second in-flight preorder of a
        # game already tracked as preordered doesn't need to push again.
        self.product.game = self.game
        self.product.save()
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREORDERED
        self.game.save(update_fields=["bgg_collection_status"])
        other_purchase = Purchase.objects.create(
            owner=self.user, name="Second KS", status=Purchase.Status.COMMITTED,
        )
        other_wave = Wave.objects.create(
            purchase=other_purchase, number=1, status=Wave.Status.PENDING,
        )
        self.login()
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.client.post(
                f"/waves/{other_wave.pk}/products/add/",
                {"name": "Second Copy", "kind": "game"},
            )
            second_product = Product.objects.get(wave=other_wave, name="Second Copy")
            self.edit_product(product=second_product, game=str(self.game.pk))
        delay.assert_not_called()

    def test_removing_one_of_two_incoming_purchases_leaves_preordered_intact(self):
        self.product.game = self.game
        self.product.save()
        other_purchase = Purchase.objects.create(
            owner=self.user, name="Second KS", status=Purchase.Status.COMMITTED,
        )
        other_wave = Wave.objects.create(
            purchase=other_purchase, number=1, status=Wave.Status.PENDING,
        )
        other_product = Product.objects.create(
            wave=other_wave, name="Second Copy", kind=Product.Kind.GAME, game=self.game,
        )
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREORDERED
        self.game.save(update_fields=["bgg_collection_status"])
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.edit_purchase(
                purchase=other_purchase, waves=[other_wave],
                delete_products=[other_product.pk],
            )
        delay.assert_not_called()


# ===========================================================================
# §8  BGG sync engine
# ===========================================================================

COLLECTION_XML = """<?xml version="1.0" encoding="utf-8"?>
<items totalitems="2" termsofuse="https://boardgamegeek.com/xmlapi/termsofuse">
  <item objecttype="thing" objectid="207830" subtype="boardgame" collid="1">
    <name sortindex="1">5-Minute Dungeon</name>
    <yearpublished>2017</yearpublished>
    <image>https://cf.geekdo-images.com/large/5md.jpg</image>
    <thumbnail>https://cf.geekdo-images.com/thumb/5md.jpg</thumbnail>
    <stats minplayers="2" maxplayers="5" minplaytime="5" maxplaytime="30"
           playingtime="30" numowned="12345">
      <rating value="N/A">
        <average value="7.1234"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="1500" bayesaverage="6.9"/>
        </ranks>
      </rating>
    </stats>
    <status own="1" lastmodified="2024-01-01 00:00:00"/>
    <numplays>54</numplays>
  </item>
  <item objecttype="thing" objectid="999999" subtype="boardgame" collid="2">
    <name sortindex="1">Mystery Cult</name>
    <yearpublished>0</yearpublished>
    <stats minplayers="0" maxplayers="0" minplaytime="0" maxplaytime="0">
      <rating value="N/A">
        <average value="0"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="Not Ranked"/>
        </ranks>
      </rating>
    </stats>
    <status own="1" lastmodified="2024-01-01 00:00:00"/>
  </item>
</items>"""

THING_XML = """<?xml version="1.0" encoding="utf-8"?>
<items termsofuse="https://boardgamegeek.com/xmlapi/termsofuse">
  <item type="boardgame" id="207830">
    <thumbnail>https://cf.geekdo-images.com/thumb/5md.jpg</thumbnail>
    <image>https://cf.geekdo-images.com/large/5md.jpg</image>
    <name type="primary" sortindex="1" value="5-Minute Dungeon"/>
    <name type="alternate" sortindex="1" value="5-Minuten Dungeon"/>
    <yearpublished value="2017"/>
    <minplayers value="2"/>
    <maxplayers value="5"/>
    <minplaytime value="5"/>
    <maxplaytime value="30"/>
    <link type="boardgamemechanic" id="2023" value="Cooperative Game"/>
    <link type="boardgameexpansion" id="228552" value="5-Minute Dungeon: Curses! Foiled Again!"/>
    <statistics page="1">
      <ratings>
        <average value="7.1234"/>
        <averageweight value="1.5321"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="1500"/>
        </ranks>
      </ratings>
    </statistics>
  </item>
  <item type="boardgame" id="337627">
    <thumbnail>https://cf.geekdo-images.com/thumb/voidfall.jpg</thumbnail>
    <image>https://cf.geekdo-images.com/large/voidfall.jpg</image>
    <name type="primary" sortindex="1" value="Voidfall"/>
    <yearpublished value="2023"/>
    <minplayers value="1"/>
    <maxplayers value="4"/>
    <minplaytime value="60"/>
    <maxplaytime value="240"/>
    <statistics page="1">
      <ratings>
        <average value="8.041"/>
        <averageweight value="4.4"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="220"/>
        </ranks>
      </ratings>
    </statistics>
  </item>
  <item type="boardgameexpansion" id="228552">
    <thumbnail>https://cf.geekdo-images.com/thumb/curses.jpg</thumbnail>
    <image>https://cf.geekdo-images.com/large/curses.jpg</image>
    <name type="primary" sortindex="1" value="5-Minute Dungeon: Curses! Foiled Again!"/>
    <yearpublished value="2018"/>
    <minplayers value="2"/>
    <maxplayers value="6"/>
    <minplaytime value="5"/>
    <maxplaytime value="30"/>
    <link type="boardgameexpansion" id="207830" value="5-Minute Dungeon" inbound="true"/>
    <link type="boardgameexpansion" id="888888" value="Base Not In App" inbound="true"/>
    <statistics page="1">
      <ratings>
        <average value="7.9"/>
        <averageweight value="1.61"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="Not Ranked"/>
        </ranks>
      </ratings>
    </statistics>
  </item>
</items>"""


EMPTY_COLLECTION_XML = """<?xml version="1.0" encoding="utf-8"?>
<items totalitems="0" termsofuse="https://boardgamegeek.com/xmlapi/termsofuse"/>"""

# Plays history (issue #65): two plays of 5-Minute Dungeon (207830, in the app)
# plus one of an unknown thing (555555, skipped on join). Play 100 carries a
# winner + scores; play 101 is incomplete with no player detail.
PLAYS_XML = """<?xml version="1.0" encoding="utf-8"?>
<plays username="kernicek" userid="1" total="3" page="1">
  <play id="100" date="2024-06-01" quantity="1" length="25" incomplete="0" nowinstats="0" location="Home">
    <item name="5-Minute Dungeon" objecttype="thing" objectid="207830">
      <subtypes><subtype value="boardgame"/></subtypes>
    </item>
    <players>
      <player username="kernicek" userid="1" name="Vojta" startposition="1" color="Red" score="42" new="0" rating="0" win="1"/>
      <player username="" userid="0" name="Bob" startposition="2" color="Blue" score="30" new="1" rating="0" win="0"/>
    </players>
    <comments>Great fun</comments>
  </play>
  <play id="101" date="2024-05-15" quantity="2" length="0" incomplete="1" nowinstats="0" location="">
    <item name="5-Minute Dungeon" objecttype="thing" objectid="207830"/>
  </play>
  <play id="102" date="2024-04-01" quantity="1" length="60" incomplete="0" location="Cafe">
    <item name="Unknown Game" objecttype="thing" objectid="555555"/>
    <players>
      <player name="Solo" score="99" win="1"/>
    </players>
  </play>
</plays>"""

EMPTY_PLAYS_XML = """<?xml version="1.0" encoding="utf-8"?>
<plays username="kernicek" userid="1" total="0" page="1"/>"""

# geekitems JSON for the expansion 228552 (same ids as THING_XML): the
# expandsboardgame links name one base in the app (207830) and one not
# (888888). Used by the sync_expansion_links stopgap command.
GEEKITEM_JSON = """{"item": {"itemid": "228552", "objecttype": "thing", "links": {
  "expandsboardgame": [
    {"objectid": "207830", "name": "5-Minute Dungeon", "objecttype": "thing"},
    {"objectid": "888888", "name": "Base Not In App", "objecttype": "thing"}
  ],
  "boardgameexpansion": []
}}}"""

# Voidfall as a preordered collection item (same stats the thing pass would
# deliver) and Bardsung as previously owned.
PREORDERED_COLLECTION_XML = """<?xml version="1.0" encoding="utf-8"?>
<items totalitems="1" termsofuse="https://boardgamegeek.com/xmlapi/termsofuse">
  <item objecttype="thing" objectid="337627" subtype="boardgame" collid="3">
    <name sortindex="1">Voidfall</name>
    <yearpublished>2023</yearpublished>
    <image>https://cf.geekdo-images.com/large/voidfall.jpg</image>
    <thumbnail>https://cf.geekdo-images.com/thumb/voidfall.jpg</thumbnail>
    <stats minplayers="1" maxplayers="4" minplaytime="60" maxplaytime="240">
      <rating value="N/A">
        <average value="8.041"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="220"/>
        </ranks>
      </rating>
    </stats>
    <status own="0" preordered="1" lastmodified="2024-01-01 00:00:00"/>
  </item>
</items>"""

PREVOWNED_COLLECTION_XML = """<?xml version="1.0" encoding="utf-8"?>
<items totalitems="2" termsofuse="https://boardgamegeek.com/xmlapi/termsofuse">
  <item objecttype="thing" objectid="245638" subtype="boardgame" collid="4">
    <name sortindex="1">Bardsung</name>
    <yearpublished>2021</yearpublished>
    <image>https://cf.geekdo-images.com/large/bardsung.jpg</image>
    <thumbnail>https://cf.geekdo-images.com/thumb/bardsung.jpg</thumbnail>
    <stats minplayers="1" maxplayers="5" minplaytime="60" maxplaytime="90">
      <rating value="N/A">
        <average value="7.5"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="2500"/>
        </ranks>
      </rating>
    </stats>
    <status own="0" prevowned="1" lastmodified="2024-01-01 00:00:00"/>
  </item>
  <item objecttype="thing" objectid="207830" subtype="boardgame" collid="5">
    <name sortindex="1">5-Minute Dungeon</name>
    <yearpublished>2017</yearpublished>
    <image>https://cf.geekdo-images.com/large/5md.jpg</image>
    <thumbnail>https://cf.geekdo-images.com/thumb/5md.jpg</thumbnail>
    <stats minplayers="2" maxplayers="5" minplaytime="5" maxplaytime="30">
      <rating value="N/A">
        <average value="7.1234"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="1500"/>
        </ranks>
      </rating>
    </stats>
    <status own="1" prevowned="1" lastmodified="2024-01-01 00:00:00"/>
  </item>
</items>"""


# Wishlist payload: one game already in the app (137637) and one that is
# not (222222). The sync mirrors the status onto the former and must NEVER
# suggest adding the latter.
WISHLIST_COLLECTION_XML = """<?xml version="1.0" encoding="utf-8"?>
<items totalitems="2" termsofuse="https://boardgamegeek.com/xmlapi/termsofuse">
  <item objecttype="thing" objectid="137637" subtype="boardgame" collid="6">
    <name sortindex="1">Wishlisted In App</name>
    <yearpublished>2020</yearpublished>
    <image>https://cf.geekdo-images.com/large/wish.jpg</image>
    <thumbnail>https://cf.geekdo-images.com/thumb/wish.jpg</thumbnail>
    <stats minplayers="2" maxplayers="4" minplaytime="30" maxplaytime="60">
      <rating value="N/A">
        <average value="7.7"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="800"/>
        </ranks>
      </rating>
    </stats>
    <status own="0" wishlist="1" wishlistpriority="2" lastmodified="2024-01-01 00:00:00"/>
    <numplays>3</numplays>
  </item>
  <item objecttype="thing" objectid="222222" subtype="boardgame" collid="7">
    <name sortindex="1">Wishlisted Not In App</name>
    <yearpublished>2022</yearpublished>
    <stats minplayers="1" maxplayers="6" minplaytime="20" maxplaytime="40">
      <rating value="N/A"><average value="6.5"/></rating>
    </stats>
    <status own="0" wishlist="1" lastmodified="2024-01-01 00:00:00"/>
  </item>
</items>"""


def _import_item_xml(bgg_id, name, status_attrs, year=2020):
    """One minimal collection <item> for the issue #81 import fixtures."""
    return (
        f'  <item objecttype="thing" objectid="{bgg_id}" subtype="boardgame"'
        f' collid="{bgg_id}">\n'
        f'    <name sortindex="1">{name}</name>\n'
        f'    <yearpublished>{year}</yearpublished>\n'
        f'    <status {status_attrs} lastmodified="2024-01-01 00:00:00"/>\n'
        f'  </item>'
    )


def _import_collection_xml(*items):
    return (
        '<?xml version="1.0" encoding="utf-8"?>\n'
        f'<items totalitems="{len(items)}" '
        'termsofuse="https://boardgamegeek.com/xmlapi/termsofuse">\n'
        + "\n".join(items) + "\n</items>"
    )


# Issue #81 preview payload: multi-flag rows exercising the import-action
# precedence (a BGG row can carry several flags at once).
IMPORT_STATUS_XML = _import_collection_xml(
    _import_item_xml(101, "Alpha Owned Wished", 'own="1" wishlist="1"'),
    _import_item_xml(102, "Bravo Rebought", 'prevowned="1" preordered="1"'),
    _import_item_xml(103, "Charlie Preordered", 'preordered="1"'),
    _import_item_xml(104, "Delta Wished", 'wishlist="1" wishlistpriority="2"'),
    _import_item_xml(105, "Echo For Trade", 'fortrade="1"'),
    _import_item_xml(106, "Foxtrot Wanted", 'want="1"'),
)

# BGG's 200-with-errors payload for an unknown username.
BGG_ERRORS_XML = """<?xml version="1.0" encoding="utf-8"?>
<errors>
  <error>
    <message>Invalid username specified</message>
  </error>
</errors>"""


class BggParserTests(TestCase):
    def test_parse_collection_maps_fields_and_unknown_markers(self):
        items = parse_collection(COLLECTION_XML)

        dungeon = items[207830]
        self.assertEqual(dungeon["bgg_name"], "5-Minute Dungeon")
        self.assertEqual(dungeon["year_published"], 2017)
        self.assertEqual(dungeon["image_url"], "https://cf.geekdo-images.com/large/5md.jpg")
        self.assertEqual(dungeon["thumbnail_url"], "https://cf.geekdo-images.com/thumb/5md.jpg")
        self.assertEqual(dungeon["min_players"], 2)
        self.assertEqual(dungeon["max_players"], 5)
        self.assertEqual(dungeon["min_playtime"], 5)
        self.assertEqual(dungeon["max_playtime"], 30)
        self.assertEqual(dungeon["bgg_rating"], Decimal("7.123"))
        self.assertEqual(dungeon["bgg_rank"], 1500)
        self.assertEqual(dungeon["bgg_numplays"], 54)  # token-free <numplays>
        self.assertNotIn("weight", dungeon)  # the collection payload has none

        # BGG's unknown markers (0 / "Not Ranked") come through as None,
        # and a missing <numplays> (0 plays) is None too.
        mystery = items[999999]
        self.assertIsNone(mystery["year_published"])
        self.assertIsNone(mystery["min_players"])
        self.assertIsNone(mystery["bgg_rating"])
        self.assertIsNone(mystery["bgg_rank"])
        self.assertIsNone(mystery["bgg_numplays"])
        self.assertEqual(mystery["image_url"], "")

    def test_parse_things_includes_weight_and_primary_name(self):
        items = parse_things(THING_XML)

        dungeon = items[207830]
        self.assertEqual(dungeon["bgg_name"], "5-Minute Dungeon")  # not the alternate
        self.assertEqual(dungeon["weight"], Decimal("1.53"))
        self.assertEqual(dungeon["bgg_rating"], Decimal("7.123"))
        self.assertEqual(dungeon["bgg_rank"], 1500)
        self.assertEqual(items[337627]["weight"], Decimal("4.40"))

    def test_parse_things_extracts_mechanic_tags(self):
        """DESIGN §10: mechanics are Tag(kind=mechanic)-bound by name, since
        BGG's mechanic link has no field on the Tag model to key by id."""
        items = parse_things(THING_XML)

        self.assertEqual(items[207830]["mechanics"], ["Cooperative Game"])
        # No boardgamemechanic link on this item.
        self.assertEqual(items[337627]["mechanics"], [])

    def test_parse_things_extracts_inbound_expansion_links(self):
        """Issue #40: an expansion's boardgameexpansion links with
        inbound="true" name its base game(s); the base game's OUTBOUND
        link to the expansion must not count."""
        items = parse_things(THING_XML)

        self.assertEqual(items[228552]["expands_bgg_ids"], [207830, 888888])
        # The base game links its expansion outbound (no inbound attr).
        self.assertEqual(items[207830]["expands_bgg_ids"], [])
        self.assertEqual(items[337627]["expands_bgg_ids"], [])

    def test_parse_collection_status_flags_reads_membership(self):
        """Issue #44: the per-game refresh reads all status flags off one
        unfiltered item instead of a request per status."""
        flags = parse_collection_status_flags(COLLECTION_XML)
        self.assertEqual(flags[207830], {
            "own": True, "preordered": False,
            "prevowned": False, "wishlist": False,
            "fortrade": False, "want": False,
            "wanttoplay": False, "wanttobuy": False,
            "wishlist_priority": None,
        })

    def test_parse_collection_status_flags_reads_all_flags_and_priority(self):
        """Issue #81: the four extra BGG flags and the wishlist priority."""
        flags = parse_collection_status_flags(IMPORT_STATUS_XML)
        self.assertTrue(flags[101]["own"])
        self.assertTrue(flags[101]["wishlist"])
        self.assertTrue(flags[102]["prevowned"])
        self.assertTrue(flags[102]["preordered"])
        self.assertTrue(flags[105]["fortrade"])
        self.assertTrue(flags[106]["want"])
        self.assertEqual(flags[104]["wishlist_priority"], 2)
        self.assertIsNone(flags[101]["wishlist_priority"])

    def test_parse_collection_error_message(self):
        """BGG answers 200 with an <errors> document for a bad username —
        it must not masquerade as an empty collection (issue #81)."""
        self.assertEqual(
            parse_collection_error(BGG_ERRORS_XML), "Invalid username specified",
        )
        self.assertEqual(parse_collection_error(COLLECTION_XML), "")
        self.assertEqual(parse_collection_error(EMPTY_COLLECTION_XML), "")

    def test_parse_geekitem_extracts_base_game_ids(self):
        """The undocumented geekitems JSON yields the same expands_bgg_ids
        key as parse_things, so both sources feed identical linking code."""
        self.assertEqual(
            parse_geekitem(GEEKITEM_JSON)["expands_bgg_ids"], [207830, 888888],
        )
        # A base game's payload (or one without links) parses to no bases.
        self.assertEqual(parse_geekitem('{"item": {}}')["expands_bgg_ids"], [])

    def test_parse_geekitem_extracts_outbound_expansion_links(self):
        """Issue #64: a base game's OWN links.boardgameexpansion list (the
        direction parse_things deliberately ignores) names its expansions,
        each as {bgg_id, name}."""
        self.assertEqual(
            parse_geekitem(BASE_GEEKITEM_WITH_EXPANSIONS_JSON)["expansions"],
            [
                {"bgg_id": 228552, "name": "5-Minute Dungeon: Curses! Foiled Again!"},
                {"bgg_id": 654321, "name": "5-Minute Dungeon: New Threat"},
            ],
        )
        # An expansion's own payload has no outbound expansions of its own.
        self.assertEqual(parse_geekitem(GEEKITEM_JSON)["expansions"], [])
        self.assertEqual(parse_geekitem('{"item": {}}')["expansions"], [])

    def test_parse_plays_maps_fields_players_winner_and_total(self):
        """Issue #65: /plays maps play + player fields, and reports the <plays
        total> so the caller can paginate."""
        plays, total = parse_plays(PLAYS_XML)
        self.assertEqual(total, 3)
        self.assertEqual(len(plays), 3)

        first = plays[0]
        self.assertEqual(first["external_id"], "100")
        self.assertEqual(first["objectid"], 207830)
        self.assertEqual(first["play_date"], datetime.date(2024, 6, 1))
        self.assertEqual(first["quantity"], 1)
        self.assertEqual(first["length_minutes"], 25)
        self.assertEqual(first["location"], "Home")
        self.assertFalse(first["incomplete"])
        self.assertEqual(first["comments"], "Great fun")

        winner, loser = first["players"]
        self.assertEqual(winner["name"], "Vojta")
        self.assertEqual(winner["username"], "kernicek")
        self.assertEqual(winner["score"], "42")
        self.assertTrue(winner["won"])
        self.assertFalse(winner["is_new"])
        self.assertEqual(winner["color"], "Red")
        self.assertEqual(winner["start_position"], "1")
        self.assertFalse(loser["won"])
        self.assertTrue(loser["is_new"])

        # Incomplete play: quantity>1, no <length> -> None, no players.
        second = plays[1]
        self.assertEqual(second["quantity"], 2)
        self.assertIsNone(second["length_minutes"])
        self.assertTrue(second["incomplete"])
        self.assertEqual(second["location"], "")
        self.assertEqual(second["players"], [])

    def test_parse_plays_tolerates_blank_and_bad_dates(self):
        xml = (
            '<plays total="1" page="1">'
            '<play id="7" date="0000-00-00" quantity="1">'
            '<item objectid="42"/></play></plays>'
        )
        plays, _ = parse_plays(xml)
        self.assertIsNone(plays[0]["play_date"])

    def test_parse_plays_error_message(self):
        """Like the collection parser: a bad username answers 200 with an
        <errors> document, not a <plays> feed."""
        self.assertEqual(
            parse_plays_error(BGG_ERRORS_XML), "Invalid username specified",
        )
        self.assertEqual(parse_plays_error(PLAYS_XML), "")
        self.assertEqual(parse_plays_error(EMPTY_PLAYS_XML), "")


class BggImportServiceTests(TestCase):
    """Pure mapping logic of the bulk import (issue #81) — no network."""

    def test_import_action_precedence(self):
        # One action per row, however many flags BGG set.
        self.assertEqual(
            import_action_from_flags({"own": True, "wishlist": True}),
            IMPORT_ACTION_COPY,
        )
        self.assertEqual(
            import_action_from_flags({"fortrade": True}), IMPORT_ACTION_COPY,
        )
        # Import precedence puts prevowned OVER preordered (the archived copy
        # carries the history) — the reverse of the stored-status precedence.
        self.assertEqual(
            import_action_from_flags({"prevowned": True, "preordered": True}),
            IMPORT_ACTION_ARCHIVED,
        )
        self.assertEqual(
            import_action_from_flags({"preordered": True}), IMPORT_ACTION_PREORDER,
        )
        for flag in ("wishlist", "want", "wanttoplay", "wanttobuy"):
            self.assertEqual(
                import_action_from_flags({flag: True}), IMPORT_ACTION_WISHLIST,
            )
        self.assertEqual(import_action_from_flags({}), "")

    def test_status_from_flags_maps_new_flags(self):
        # Regression guard: without these mappings a re-sync would blank the
        # stored status of games imported from want-type / for-trade rows.
        self.assertEqual(
            _status_from_flags({"fortrade": True}), Game.BggCollectionStatus.OWN,
        )
        for flag in ("want", "wanttoplay", "wanttobuy"):
            self.assertEqual(
                _status_from_flags({flag: True}),
                Game.BggCollectionStatus.WISHLIST,
            )
        # Stored-status precedence keeps preordered over prevowned.
        self.assertEqual(
            _status_from_flags({"prevowned": True, "preordered": True}),
            Game.BggCollectionStatus.PREORDERED,
        )

    def _candidate(self, bgg_id, action, wishlist_priority=None):
        return {
            "bgg_id": bgg_id, "action": action,
            "wishlist_priority": wishlist_priority,
        }

    def test_group_candidates_by_action_buckets_in_display_order(self):
        # IMPORT_ACTIONS order is copy, archived_copy, preorder, wishlist —
        # the input order here is scrambled to prove the buckets, not the
        # input, decide the group order.
        candidates = [
            self._candidate(1, IMPORT_ACTION_WISHLIST),
            self._candidate(2, IMPORT_ACTION_COPY),
            self._candidate(3, IMPORT_ACTION_WISHLIST, wishlist_priority=2),
            self._candidate(4, IMPORT_ACTION_COPY),
        ]
        groups = group_candidates_by_action(candidates)

        self.assertEqual([g["action"] for g in groups], [IMPORT_ACTION_COPY, IMPORT_ACTION_WISHLIST])
        copy_group, wishlist_group = groups
        self.assertEqual([c["bgg_id"] for c in copy_group["candidates"]], [2, 4])
        self.assertFalse(copy_group["has_priority"])
        self.assertEqual([c["bgg_id"] for c in wishlist_group["candidates"]], [1, 3])
        self.assertTrue(wishlist_group["has_priority"])

    def test_group_candidates_by_action_skips_empty_actions(self):
        groups = group_candidates_by_action([self._candidate(1, IMPORT_ACTION_PREORDER)])
        self.assertEqual([g["action"] for g in groups], [IMPORT_ACTION_PREORDER])

    def test_group_candidates_by_action_empty_input(self):
        self.assertEqual(group_candidates_by_action([]), [])


class BggClientAuthTests(TestCase):
    """login() is the §8 seam: Bearer header with a token, credential POST
    without one. Fallback stays usable so token-less self-hosters still sync."""

    def test_token_sets_bearer_header_and_skips_login_post(self):
        client = BggClient("kernicek", "secret", token="tok-123")
        with mock.patch.object(client.session, "post") as post:
            client.login()
        post.assert_not_called()
        self.assertEqual(client.session.headers["Authorization"], "Bearer tok-123")

    def test_no_token_posts_credentials_and_sets_no_bearer_header(self):
        client = BggClient("kernicek", "secret")
        response = mock.Mock(status_code=200)
        with mock.patch.object(client.session, "post", return_value=response) as post:
            client.login()
        post.assert_called_once()
        url, kwargs = post.call_args[0][0], post.call_args[1]
        self.assertEqual(url, BggClient.LOGIN_URL)
        self.assertEqual(
            kwargs["json"],
            {"credentials": {"username": "kernicek", "password": "secret"}},
        )
        self.assertNotIn("Authorization", client.session.headers)


class BggPlaysClientTests(TestCase):
    """get_plays (issue #65): the /plays request builder and its 401 degrade."""

    def test_get_plays_builds_params_and_returns_body(self):
        client = BggClient("kernicek", "secret")
        response = mock.Mock(status_code=200, text=PLAYS_XML)
        with mock.patch.object(client.session, "get", return_value=response) as get:
            body = client.get_plays("kernicek", page=2, bgg_id=207830)
        self.assertEqual(body, PLAYS_XML)
        url, kwargs = get.call_args[0][0], get.call_args[1]
        self.assertEqual(url, BggClient.PLAYS_URL)
        self.assertEqual(
            kwargs["params"], {"username": "kernicek", "page": 2, "id": 207830},
        )

    def test_get_plays_401_raises_auth_error(self):
        """Private plays without the Bearer token 401 — surface it so the sync
        degrades instead of failing (like get_things)."""
        client = BggClient("kernicek", "secret")
        response = mock.Mock(status_code=401)
        with mock.patch.object(client.session, "get", return_value=response):
            with self.assertRaises(BggAuthError):
                client.get_plays("kernicek")


class BggClientPushTests(TestCase):
    """The write-back REST API (issue #157, replacing #117's unverified
    geekcollection.php guess): get_user_id / get_collection_item /
    put_collection_item, confirmed live against a real BGG session."""

    def test_get_user_id_returns_the_numeric_id(self):
        client = BggClient("kernicek", "secret")
        response = mock.Mock(status_code=200, json=lambda: {"loggedIn": True, "userid": 588107})
        with mock.patch.object(client.session, "get", return_value=response) as get:
            userid = client.get_user_id()
        self.assertEqual(userid, 588107)
        self.assertEqual(get.call_args[0][0], BggClient.USER_CURRENT_URL)

    def test_get_user_id_401_raises_auth_error(self):
        client = BggClient("kernicek", "secret")
        response = mock.Mock(status_code=401)
        with mock.patch.object(client.session, "get", return_value=response):
            with self.assertRaises(BggAuthError):
                client.get_user_id()

    def test_get_collection_item_builds_params_and_returns_the_item(self):
        client = BggClient("kernicek", "secret")
        item = {"collid": 31355791, "status": {"own": True}}
        response = mock.Mock(status_code=200, json=lambda: {"items": [item]})
        with mock.patch.object(client.session, "get", return_value=response) as get:
            result = client.get_collection_item(207830, 588107)
        self.assertEqual(result, item)
        url, kwargs = get.call_args[0][0], get.call_args[1]
        self.assertEqual(url, BggClient.COLLECTIONS_API_URL)
        self.assertEqual(
            kwargs["params"],
            {"objectid": 207830, "objecttype": "thing", "userid": 588107},
        )

    def test_get_collection_item_returns_none_when_not_in_collection(self):
        client = BggClient("kernicek", "secret")
        response = mock.Mock(status_code=200, json=lambda: {"items": []})
        with mock.patch.object(client.session, "get", return_value=response):
            self.assertIsNone(client.get_collection_item(207830, 588107))

    def test_get_collection_item_401_raises_auth_error(self):
        client = BggClient("kernicek", "secret")
        response = mock.Mock(status_code=401)
        with mock.patch.object(client.session, "get", return_value=response):
            with self.assertRaises(BggAuthError):
                client.get_collection_item(207830, 588107)

    def test_put_collection_item_puts_the_full_item_to_its_collid_url(self):
        client = BggClient("kernicek", "secret")
        item = {"collid": 31355791, "status": {"prevowned": True}}
        response = mock.Mock(status_code=200, json=lambda: {"message": "Item updated"})
        with mock.patch.object(client.session, "put", return_value=response) as put:
            body = client.put_collection_item(item)
        self.assertEqual(body, {"message": "Item updated"})
        url, kwargs = put.call_args[0][0], put.call_args[1]
        self.assertEqual(url, BggClient.COLLECTION_ITEM_URL.format(31355791))
        self.assertEqual(kwargs["json"], {"item": item})

    def test_put_collection_item_401_raises_auth_error(self):
        client = BggClient("kernicek", "secret")
        response = mock.Mock(status_code=401)
        with mock.patch.object(client.session, "put", return_value=response):
            with self.assertRaises(BggAuthError):
                client.put_collection_item({"collid": 31355791})


class BggCredentialResolutionTests(TestCase):
    """bgg_credentials_error / make_bgg_client resolve settings into an auth
    path — token preferred, password fallback, username always required."""

    @override_settings(BGG_USERNAME="kernicek", BGG_API_TOKEN="tok", BGG_PASSWORD="")
    def test_token_alone_satisfies_the_guard(self):
        self.assertEqual(bgg_credentials_error(), "")

    @override_settings(BGG_USERNAME="kernicek", BGG_API_TOKEN="", BGG_PASSWORD="pw")
    def test_password_alone_satisfies_the_guard(self):
        self.assertEqual(bgg_credentials_error(), "")

    @override_settings(BGG_USERNAME="kernicek", BGG_API_TOKEN="", BGG_PASSWORD="")
    def test_username_without_auth_fails(self):
        self.assertIn("credentials are not configured", bgg_credentials_error())

    @override_settings(BGG_USERNAME="", BGG_API_TOKEN="tok", BGG_PASSWORD="")
    def test_token_without_username_fails(self):
        # BGG_USERNAME also names whose collection to pull — required in both modes.
        self.assertIn("credentials are not configured", bgg_credentials_error())

    @override_settings(BGG_USERNAME="kernicek", BGG_API_TOKEN="tok", BGG_PASSWORD="pw")
    def test_make_client_threads_the_token_through(self):
        client = make_bgg_client()
        self.assertEqual(client.token, "tok")
        self.assertEqual(client.username, "kernicek")

    @override_settings(BGG_USERNAME="", BGG_API_TOKEN="", BGG_PASSWORD="")
    def test_user_stored_creds_satisfy_the_guard_with_blank_env(self):
        # A user's own stored account satisfies auth even when no env creds
        # are configured — the per-user path (issue #118).
        user = get_user_model().objects.create_user(username="alice")
        membership = user.membership
        membership.bgg_username = "alice_bgg"
        membership.set_bgg_password("s3cret")
        membership.save()
        self.assertEqual(bgg_credentials_error(user), "")
        client = make_bgg_client(user)
        self.assertEqual(client.username, "alice_bgg")
        self.assertEqual(client.password, "s3cret")

    @override_settings(BGG_USERNAME="", BGG_API_TOKEN="", BGG_PASSWORD="")
    def test_user_without_stored_creds_still_fails_on_blank_env(self):
        user = get_user_model().objects.create_user(username="bob")
        self.assertIn("credentials are not configured", bgg_credentials_error(user))


class CryptoTests(SimpleTestCase):
    """Fernet encrypt/decrypt for BGG passwords at rest (issue #118)."""

    def test_round_trip(self):
        self.assertEqual(crypto.decrypt(crypto.encrypt("hunter2")), "hunter2")

    def test_empty_in_empty_out(self):
        self.assertEqual(crypto.encrypt(""), "")
        self.assertEqual(crypto.encrypt(None), "")
        self.assertEqual(crypto.decrypt(""), "")

    def test_ciphertext_is_not_the_plaintext(self):
        token = crypto.encrypt("hunter2")
        self.assertNotEqual(token, "hunter2")
        self.assertNotIn("hunter2", token)

    def test_key_derived_from_secret_key_is_deterministic(self):
        # Two encrypts of the same value differ (Fernet nonces), but both
        # decrypt under the SECRET_KEY-derived key with no BGG_ENCRYPTION_KEY set.
        with override_settings(BGG_ENCRYPTION_KEY=""):
            self.assertEqual(crypto.decrypt(crypto.encrypt("x")), "x")

    def test_explicit_bgg_encryption_key_is_used(self):
        from cryptography.fernet import Fernet
        with override_settings(BGG_ENCRYPTION_KEY=Fernet.generate_key().decode()):
            self.assertEqual(crypto.decrypt(crypto.encrypt("y")), "y")

    def test_undecryptable_ciphertext_returns_empty(self):
        # Garbage, or a token from a different key, decrypts to "" rather than
        # raising — callers treat that as "no usable secret" and fall back.
        self.assertEqual(crypto.decrypt("not-a-real-token"), "")
        from cryptography.fernet import Fernet
        other = Fernet(Fernet.generate_key()).encrypt(b"z").decode()
        self.assertEqual(crypto.decrypt(other), "")


class ResolveBggCredentialsTests(TestCase):
    """resolve_bgg_credentials precedence: per-user creds over env (issue #118)."""

    @override_settings(BGG_USERNAME="envuser", BGG_PASSWORD="envpw", BGG_API_TOKEN="envtok")
    def test_no_user_returns_env(self):
        self.assertEqual(
            resolve_bgg_credentials(None), ("envuser", "envpw", "envtok"),
        )

    @override_settings(BGG_USERNAME="envuser", BGG_PASSWORD="envpw", BGG_API_TOKEN="envtok")
    def test_stored_username_and_password_win_and_drop_token(self):
        user = get_user_model().objects.create_user(username="alice")
        membership = user.membership
        membership.bgg_username = "alice_bgg"
        membership.set_bgg_password("alice_pw")
        membership.save()
        self.assertEqual(
            resolve_bgg_credentials(user), ("alice_bgg", "alice_pw", ""),
        )

    @override_settings(BGG_USERNAME="envuser", BGG_PASSWORD="envpw", BGG_API_TOKEN="envtok")
    def test_stored_username_only_keeps_env_password_and_token(self):
        user = get_user_model().objects.create_user(username="bob")
        membership = user.membership
        membership.bgg_username = "bob_bgg"
        membership.save(update_fields=["bgg_username"])
        self.assertEqual(
            resolve_bgg_credentials(user), ("bob_bgg", "envpw", "envtok"),
        )

    @override_settings(BGG_USERNAME="envuser", BGG_PASSWORD="envpw", BGG_API_TOKEN="envtok")
    def test_user_without_stored_creds_falls_back_to_env(self):
        user = get_user_model().objects.create_user(username="carol")
        self.assertEqual(
            resolve_bgg_credentials(user), ("envuser", "envpw", "envtok"),
        )


class SettingsPageTests(TestCase):
    """The Settings page (issue #137); its BGG account section carries the
    write-only password / set / clear behaviour from issue #118."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/settings/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_get_shows_no_password_stored_initially(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/")
        self.assertContains(response, "No password stored")

    def test_post_sets_password_encrypted_at_rest(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post("/settings/", {
            "bgg_username": "kernicek_bgg", "bgg_password": "topsecret",
        })
        self.assertEqual(response.status_code, 302)  # PRG
        membership = Membership.objects.get(user=self.user)
        self.assertEqual(membership.bgg_username, "kernicek_bgg")
        self.assertTrue(membership.bgg_password_encrypted)
        self.assertNotEqual(membership.bgg_password_encrypted, "topsecret")
        self.assertNotIn("topsecret", membership.bgg_password_encrypted)
        self.assertEqual(membership.get_bgg_password(), "topsecret")

    def test_password_is_write_only_never_rendered_back(self):
        self.client.login(username="kernicek", password="pass")
        self.client.post("/settings/", {
            "bgg_username": "kernicek_bgg", "bgg_password": "topsecret",
        })
        response = self.client.get("/settings/")
        self.assertContains(response, "Password is set")
        self.assertNotContains(response, "topsecret")

    def test_clear_password_removes_it(self):
        membership = self.user.membership
        membership.set_bgg_password("topsecret")
        membership.save()
        self.client.login(username="kernicek", password="pass")
        self.client.post("/settings/", {
            "bgg_username": "kernicek_bgg", "clear_password": "on",
        })
        self.assertEqual(
            Membership.objects.get(user=self.user).bgg_password_encrypted, "",
        )

    def test_blank_password_without_clear_keeps_stored_value(self):
        membership = self.user.membership
        membership.set_bgg_password("topsecret")
        membership.save()
        self.client.login(username="kernicek", password="pass")
        self.client.post("/settings/", {
            "bgg_username": "renamed", "bgg_password": "",
        })
        membership = Membership.objects.get(user=self.user)
        self.assertEqual(membership.bgg_username, "renamed")
        self.assertEqual(membership.get_bgg_password(), "topsecret")

    def test_post_ntfy_topic_sets_it(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.post("/settings/", {"ntfy_topic": "kernicek-reminders"})
        self.assertEqual(response.status_code, 302)  # PRG
        self.assertEqual(
            Membership.objects.get(user=self.user).ntfy_topic, "kernicek-reminders",
        )

    def test_get_shows_current_ntfy_topic(self):
        membership = self.user.membership
        membership.ntfy_topic = "kernicek-reminders"
        membership.save()
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/")
        self.assertContains(response, "kernicek-reminders")

    def test_saving_ntfy_topic_does_not_touch_bgg_username(self):
        membership = self.user.membership
        membership.bgg_username = "kernicek_bgg"
        membership.save()
        self.client.login(username="kernicek", password="pass")
        self.client.post("/settings/", {"ntfy_topic": "kernicek-reminders"})
        membership = Membership.objects.get(user=self.user)
        self.assertEqual(membership.bgg_username, "kernicek_bgg")
        self.assertEqual(membership.ntfy_topic, "kernicek-reminders")

    def test_saving_bgg_username_does_not_touch_ntfy_topic(self):
        membership = self.user.membership
        membership.ntfy_topic = "kernicek-reminders"
        membership.save()
        self.client.login(username="kernicek", password="pass")
        self.client.post("/settings/", {"bgg_username": "kernicek_bgg"})
        membership = Membership.objects.get(user=self.user)
        self.assertEqual(membership.bgg_username, "kernicek_bgg")
        self.assertEqual(membership.ntfy_topic, "kernicek-reminders")

    def test_settings_never_leaks_template_comment_markers(self):
        # Multi-line {# #} comments are NOT stripped by Django (its comment lexer
        # doesn't span newlines) and render verbatim; settings.html uses a
        # {% comment %} block instead. Guard that neither the markers nor the
        # comment's interior text reach the UI (CLAUDE.md).
        self.client.login(username="kernicek", password="pass")
        body = self.client.get("/settings/").content.decode()
        self.assertNotIn("{#", body)
        self.assertNotIn("{% comment", body)
        self.assertNotIn("stack as cards", body)


class SettingsNtfyTestButtonTests(TestCase):
    """The "Send test notification" button on the Settings page (issue #162):
    always tests the persisted Membership.ntfy_topic, never an unsaved edit
    from the POST body, and fails soft back to the settings page either way."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.post("/settings/ntfy/test/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_get_is_not_allowed(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/ntfy/test/")
        self.assertEqual(response.status_code, 405)

    def test_no_saved_topic_redirects_without_sending(self):
        self.client.login(username="kernicek", password="pass")
        with mock.patch("gamekeeper.views.ntfy.send_ntfy") as send_ntfy_mock:
            response = self.client.post("/settings/ntfy/test/")
        send_ntfy_mock.assert_not_called()
        self.assertRedirects(response, "/settings/")

    def test_ignores_unsaved_topic_in_post_body(self):
        membership = self.user.membership
        membership.ntfy_topic = "saved-topic"
        membership.save()
        self.client.login(username="kernicek", password="pass")
        with mock.patch(
            "gamekeeper.views.ntfy.send_ntfy", return_value=True,
        ) as send_ntfy_mock:
            self.client.post("/settings/ntfy/test/", {"ntfy_topic": "unsaved-edit"})
        send_ntfy_mock.assert_called_once_with(
            "saved-topic", "Test notification", mock.ANY,
        )

    def test_success_redirects_with_ok(self):
        membership = self.user.membership
        membership.ntfy_topic = "saved-topic"
        membership.save()
        self.client.login(username="kernicek", password="pass")
        with mock.patch("gamekeeper.views.ntfy.send_ntfy", return_value=True):
            response = self.client.post("/settings/ntfy/test/")
        self.assertRedirects(response, "/settings/?ntfy_test=ok")

    def test_failure_redirects_with_fail(self):
        membership = self.user.membership
        membership.ntfy_topic = "saved-topic"
        membership.save()
        self.client.login(username="kernicek", password="pass")
        with mock.patch("gamekeeper.views.ntfy.send_ntfy", return_value=False):
            with self.assertLogs("gamekeeper.views", level="WARNING"):
                response = self.client.post("/settings/ntfy/test/")
        self.assertRedirects(response, "/settings/?ntfy_test=fail")

    def test_ok_result_shows_success_alert(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/?ntfy_test=ok")
        self.assertContains(response, "Test notification sent.")

    def test_fail_result_shows_danger_alert(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/?ntfy_test=fail")
        self.assertContains(response, "Couldn't reach the ntfy server")

    def test_bogus_query_param_shows_no_alert(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/?ntfy_test=bogus")
        self.assertNotContains(response, "Test notification sent.")
        self.assertNotContains(response, "Couldn't reach the ntfy server")


class SettingsEmailTestButtonTests(TestCase):
    """The "Send test email" button on the Settings page (issue #171): sends
    a one-off test email to the user's account address, and fails soft back
    to the settings page either way."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.post("/settings/email/test/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_get_is_not_allowed(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/email/test/")
        self.assertEqual(response.status_code, 405)

    def test_no_email_redirects_without_sending(self):
        self.client.login(username="kernicek", password="pass")
        with mock.patch("gamekeeper.views.send_mail") as send_mail_mock:
            response = self.client.post("/settings/email/test/")
        send_mail_mock.assert_not_called()
        self.assertRedirects(response, "/settings/")

    def test_success_redirects_with_ok(self):
        self.user.email = "kernicek@example.com"
        self.user.save(update_fields=["email"])
        self.client.login(username="kernicek", password="pass")
        with mock.patch("gamekeeper.views.send_mail", return_value=1):
            response = self.client.post("/settings/email/test/")
        self.assertRedirects(response, "/settings/?email_test=ok")

    def test_failure_redirects_with_fail(self):
        self.user.email = "kernicek@example.com"
        self.user.save(update_fields=["email"])
        self.client.login(username="kernicek", password="pass")
        with mock.patch(
            "gamekeeper.views.send_mail", side_effect=smtplib.SMTPException,
        ):
            with self.assertLogs("gamekeeper.views", level="ERROR"):
                response = self.client.post("/settings/email/test/")
        self.assertRedirects(response, "/settings/?email_test=fail")

    def test_non_smtp_failure_also_redirects_with_fail(self):
        # Guards against a misconfigured EMAIL_BACKEND raising something
        # outside the old narrow (SMTPException, OSError) catch (issue #180).
        self.user.email = "kernicek@example.com"
        self.user.save(update_fields=["email"])
        self.client.login(username="kernicek", password="pass")
        with mock.patch(
            "gamekeeper.views.send_mail", side_effect=ValueError("bad backend"),
        ):
            with self.assertLogs("gamekeeper.views", level="ERROR"):
                response = self.client.post("/settings/email/test/")
        self.assertRedirects(response, "/settings/?email_test=fail")

    def test_ok_result_shows_success_alert(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/?email_test=ok")
        self.assertContains(response, "Test email sent.")

    def test_fail_result_shows_danger_alert(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/?email_test=fail")
        self.assertContains(response, "Couldn't send the test email")

    def test_bogus_query_param_shows_no_alert(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/settings/?email_test=bogus")
        self.assertNotContains(response, "Test email sent.")
        self.assertNotContains(response, "Couldn't send the test email")


class SettingsNavbarDropdownTests(TestCase):
    """Issue #137: account actions live under one far-right username dropdown,
    with Sharing (owner-only) and Tools (superuser-only) folded in."""

    @classmethod
    def setUpTestData(cls):
        cls.owner = get_user_model().objects.create_user(
            username="owner1", password="pass",
        )
        cls.superuser = get_user_model().objects.create_superuser(
            username="admin", password="pass",
        )
        cls.member = get_user_model().objects.create_user(
            username="member1", password="pass",
        )
        # Every user is auto-owner of a personal group (signal); downgrade this
        # one to a plain member so the Sharing gate has something to hide from.
        m = cls.member.membership
        m.role = Membership.Role.MEMBER
        m.save(update_fields=["role"])

    def test_username_dropdown_replaces_the_loose_right_nav_links(self):
        self.client.login(username="owner1", password="pass")
        body = self.client.get("/settings/").content.decode()
        # A Bootstrap dropdown toggle labelled with the username.
        self.assertIn('data-bs-toggle="dropdown"', body)
        self.assertIn("owner1", body)
        # Settings and Log out are always in the menu; logout is a POST button.
        self.assertIn('class="dropdown-item" href="/settings/">Settings', body)
        self.assertIn('class="dropdown-item">Log out', body)
        # The old flat right-nav links are gone.
        self.assertNotIn("Sign out (", body)
        self.assertNotIn("BGG account</a>", body)

    def test_sharing_moves_into_the_dropdown_off_the_left_nav(self):
        self.client.login(username="owner1", password="pass")
        body = self.client.get("/settings/").content.decode()
        self.assertIn(
            'class="dropdown-item" href="/g/owner1/settings/">Sharing', body,
        )
        self.assertNotIn('class="nav-link" href="/g/owner1/settings/"', body)

    def test_sharing_hidden_for_non_owner_members(self):
        self.client.login(username="member1", password="pass")
        body = self.client.get("/settings/").content.decode()
        self.assertNotIn("Sharing", body)

    def test_tools_only_shown_to_superusers(self):
        self.client.login(username="owner1", password="pass")
        owner_body = self.client.get("/settings/").content.decode()
        self.assertNotIn('href="/tools/"', owner_body)
        self.client.login(username="admin", password="pass")
        admin_body = self.client.get("/settings/").content.decode()
        self.assertIn('class="dropdown-item" href="/tools/">Tools', admin_body)


class PushBggStatusTests(TestCase):
    """bgg_sync.push_bgg_status (issue #117, rebuilt on the live-verified
    REST API for issue #157), with BggClient mocked out — the read-modify-
    write mechanics themselves are pinned down in BggClientPushTests; this
    only covers push_bgg_status's own contract: never raises, never stamps
    last_synced_at, records failures as diffs."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)

    def _client(self, item=None):
        """A mock BggClient whose get_collection_item returns `item` (a real
        dict, since push_bgg_status mutates it — the default Mock() auto-
        attribute isn't subscriptable)."""
        client = mock.Mock()
        client.get_user_id.return_value = 588107
        client.get_collection_item.return_value = item
        return client

    def test_success_writes_status_and_marker_without_touching_last_synced_at(self):
        client = self._client({"collid": 999, "status": {}})
        result = push_bgg_status(
            self.game, Game.BggCollectionStatus.OWN, client=client, user=self.user,
        )

        self.assertTrue(result.ok)
        self.assertEqual(result.new_status, Game.BggCollectionStatus.OWN)
        self.assertIsNotNone(result.pushed_at)
        client.get_collection_item.assert_called_once_with(207830, 588107)
        client.put_collection_item.assert_called_once_with(
            {"collid": 999, "status": {"own": True}},
        )
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, Game.BggCollectionStatus.OWN)
        self.assertEqual(self.game.bgg_status_pushed, Game.BggCollectionStatus.OWN)
        self.assertIsNotNone(self.game.bgg_status_pushed_at)
        self.assertIsNone(self.game.last_synced_at)  # push is not a read sync

    def test_wishlist_success_sets_and_non_wishlist_clears_priority(self):
        client = self._client({"collid": 999, "status": {}})
        push_bgg_status(
            self.game, Game.BggCollectionStatus.WISHLIST, priority=2,
            client=client, user=self.user,
        )
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_wishlist_priority, 2)
        client.put_collection_item.assert_called_with(
            {"collid": 999, "status": {"wishlist": True}, "wishlistpriority": 2},
        )

        client.get_collection_item.return_value = {
            "collid": 999, "status": {"wishlist": True}, "wishlistpriority": 2,
        }
        push_bgg_status(
            self.game, Game.BggCollectionStatus.OWN, client=client, user=self.user,
        )
        self.game.refresh_from_db()
        self.assertIsNone(self.game.bgg_wishlist_priority)

    def test_success_clears_an_existing_push_failed_diff(self):
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.PUSH_FAILED,
            bgg_id=207830, game=self.game, last_seen_at=timezone.now(),
        )
        push_bgg_status(
            self.game, Game.BggCollectionStatus.OWN,
            client=self._client({"collid": 999, "status": {}}), user=self.user,
        )
        self.assertFalse(
            BggSyncDiff.objects.filter(category=BggSyncDiff.Category.PUSH_FAILED).exists(),
        )

    def test_client_failure_leaves_status_untouched_and_records_a_diff(self):
        client = self._client({"collid": 999, "status": {}})
        client.put_collection_item.side_effect = BggError("BGG kept answering 503.")

        result = push_bgg_status(
            self.game, Game.BggCollectionStatus.PREV_OWNED,
            client=client, user=self.user,
        )

        self.assertFalse(result.ok)
        self.assertIn("503", result.error)
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, "")
        self.assertIsNone(self.game.bgg_status_pushed_at)
        diff = BggSyncDiff.objects.get(category=BggSyncDiff.Category.PUSH_FAILED)
        self.assertEqual(diff.owner, self.user)
        self.assertEqual(diff.game, self.game)
        self.assertEqual(diff.bgg_id, 207830)

    def test_auth_failure_is_reported_like_any_other_failure(self):
        client = self._client()
        client.get_collection_item.side_effect = BggAuthError("refused (401)")
        result = push_bgg_status(
            self.game, Game.BggCollectionStatus.OWN, client=client, user=self.user,
        )
        self.assertFalse(result.ok)
        self.assertIn("401", result.error)

    def test_no_primary_link_short_circuits_before_any_network_call(self):
        linkless = Game.objects.create(name="No BGG Link")
        client = self._client()
        result = push_bgg_status(
            linkless, Game.BggCollectionStatus.OWN, client=client, user=self.user,
        )
        self.assertTrue(result.no_primary_link)
        client.get_collection_item.assert_not_called()

    def test_invalid_status_is_rejected(self):
        result = push_bgg_status(
            self.game, "bogus_status", client=self._client(), user=self.user,
        )
        self.assertTrue(result.invalid_status)

    def test_empty_string_removes_the_status(self):
        self.game.bgg_collection_status = Game.BggCollectionStatus.WISHLIST
        self.game.bgg_wishlist_priority = 1
        self.game.save(update_fields=["bgg_collection_status", "bgg_wishlist_priority"])

        result = push_bgg_status(
            self.game, "", client=self._client({"collid": 999, "status": {"wishlist": True}}),
            user=self.user,
        )

        self.assertTrue(result.ok)
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, "")
        self.assertIsNone(self.game.bgg_wishlist_priority)

    def test_empty_string_is_a_noop_when_already_absent_from_bgg(self):
        client = self._client(None)
        result = push_bgg_status(self.game, "", client=client, user=self.user)
        self.assertTrue(result.ok)
        client.put_collection_item.assert_not_called()

    def test_not_in_collection_yet_refuses_to_guess_a_new_item_write(self):
        """Adding a brand-new collection item is untested (issue #157's still-
        open question) — push_bgg_status must refuse rather than guess."""
        client = self._client(None)
        result = push_bgg_status(
            self.game, Game.BggCollectionStatus.OWN, client=client, user=self.user,
        )
        self.assertFalse(result.ok)
        self.assertIn("brand-new", result.error)
        client.put_collection_item.assert_not_called()
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, "")
        diff = BggSyncDiff.objects.get(category=BggSyncDiff.Category.PUSH_FAILED)
        self.assertEqual(diff.game, self.game)


class PushBggFortradeTests(TestCase):
    """bgg_sync.push_bgg_fortrade (issue #82) — the merge-based sibling of
    push_bgg_status: fortrade is orthogonal to the membership status, so
    this must preserve whatever else is already on the item's status object
    rather than replacing it."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)

    def _client(self, item=None):
        client = mock.Mock()
        client.get_user_id.return_value = 588107
        client.get_collection_item.return_value = item
        return client

    def test_marking_fortrade_preserves_the_existing_own_flag(self):
        client = self._client({"collid": 999, "status": {"own": True}})
        result = push_bgg_fortrade(self.game, True, client=client, user=self.user)

        self.assertTrue(result.ok)
        client.put_collection_item.assert_called_once_with(
            {"collid": 999, "status": {"own": True, "fortrade": True}},
        )
        self.game.refresh_from_db()
        self.assertTrue(self.game.bgg_fortrade_pushed)
        self.assertIsNotNone(self.game.bgg_fortrade_pushed_at)
        self.assertIsNone(self.game.last_synced_at)  # push is not a read sync

    def test_clearing_fortrade_removes_only_that_flag(self):
        client = self._client({"collid": 999, "status": {"own": True, "fortrade": True}})
        result = push_bgg_fortrade(self.game, False, client=client, user=self.user)

        self.assertTrue(result.ok)
        client.put_collection_item.assert_called_once_with(
            {"collid": 999, "status": {"own": True}},
        )
        self.game.refresh_from_db()
        self.assertFalse(self.game.bgg_fortrade_pushed)

    def test_no_primary_link_short_circuits_before_any_network_call(self):
        linkless = Game.objects.create(name="No BGG Link")
        client = self._client()
        result = push_bgg_fortrade(linkless, True, client=client, user=self.user)
        self.assertTrue(result.no_primary_link)
        client.get_collection_item.assert_not_called()

    def test_client_failure_leaves_marker_untouched_and_records_a_diff(self):
        client = self._client({"collid": 999, "status": {"own": True}})
        client.put_collection_item.side_effect = BggError("BGG kept answering 503.")

        result = push_bgg_fortrade(self.game, True, client=client, user=self.user)

        self.assertFalse(result.ok)
        self.assertIn("503", result.error)
        self.game.refresh_from_db()
        self.assertFalse(self.game.bgg_fortrade_pushed)
        diff = BggSyncDiff.objects.get(category=BggSyncDiff.Category.PUSH_FAILED)
        self.assertEqual(diff.game, self.game)

    def test_clearing_is_a_noop_when_already_absent_from_bgg(self):
        client = self._client(None)
        result = push_bgg_fortrade(self.game, False, client=client, user=self.user)
        self.assertTrue(result.ok)
        client.put_collection_item.assert_not_called()


@override_settings(BGG_USERNAME="kernicek", BGG_PASSWORD="test-password")
class SyncBggTests(TestCase):
    """sync_bgg with the HTTP layer mocked out (no live BGG in tests)."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        # Owned + on BGG: gets the full backfill.
        cls.game = Game.objects.create(name="5 Minute Dungeon (localized)")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)
        cls.copy = Copy.objects.create(
            owner=cls.user, edition=cls.edition,
            excitement=Decimal("9.0"), keep_status=Copy.KeepStatus.ALWAYS_KEEP,
            notes="app-only note",
        )
        # Preorder game: not in the BGG collection, covered by the thing pass;
        # its pending purchase explains the absence in the reconciliation.
        cls.voidfall = Game.objects.create(name="Voidfall")
        BggLink.objects.create(game=cls.voidfall, bgg_id=337627, is_primary=True)
        voidfall_edition = Edition.objects.create(game=cls.voidfall, is_default=True)
        Copy.objects.create(owner=cls.user, edition=voidfall_edition)
        purchase = Purchase.objects.create(
            owner=cls.user, name="Voidfall KS", status=Purchase.Status.COMMITTED,
        )
        wave = Wave.objects.create(
            purchase=purchase, number=1, status=Wave.Status.PRODUCTION,
        )
        Product.objects.create(wave=wave, name="Voidfall Core", game=cls.voidfall)

    def run_sync(self, things=THING_XML, collections=None, plays=EMPTY_PLAYS_XML,
                 **extra):
        """call_command with BggClient mocked; returns (output, client mock).

        collections maps status param -> XML; statuses not given answer an
        empty collection (the sync asks own / preordered / prevowned). plays is
        the /plays body (a string, or an Exception for get_plays to raise, or a
        {page: xml} dict for the paginated case).
        """
        collections = {"own": COLLECTION_XML, **(collections or {})}
        out = StringIO()
        # The command builds its client via bgg_sync.make_bgg_client, which
        # constructs bgg_sync.BggClient — patch there to intercept it.
        with mock.patch(
            "gamekeeper.bgg_sync.BggClient",
        ) as client_class:
            client = client_class.return_value
            client.get_collection.side_effect = (
                lambda username, status="own": collections.get(status, EMPTY_COLLECTION_XML)
            )
            if isinstance(things, Exception):
                client.get_things.side_effect = things
            else:
                client.get_things.return_value = things
            if isinstance(plays, Exception):
                client.get_plays.side_effect = plays
            elif isinstance(plays, dict):
                client.get_plays.side_effect = (
                    lambda username, page=1, bgg_id=None: plays[page]
                )
            else:
                client.get_plays.return_value = plays
            call_command("sync_bgg", user="kernicek", stdout=out, **extra)
        return out.getvalue(), client

    def test_backfills_bgg_fields_from_collection_and_thing(self):
        output, client = self.run_sync()

        client.login.assert_called_once_with()
        # One request per membership status (BGG ANDs combined status
        # filters, so the union takes three) plus one for the wishlist.
        self.assertEqual(client.get_collection.call_count, 4)
        for status in ("own", "preordered", "prevowned", "wishlist"):
            client.get_collection.assert_any_call("kernicek", status=status)

        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_name, "5-Minute Dungeon")
        self.assertEqual(self.game.year_published, 2017)
        self.assertEqual(self.game.image_url, "https://cf.geekdo-images.com/large/5md.jpg")
        self.assertEqual(self.game.thumbnail_url, "https://cf.geekdo-images.com/thumb/5md.jpg")
        self.assertEqual(self.game.min_players, 2)
        self.assertEqual(self.game.max_players, 5)
        self.assertEqual(self.game.min_playtime, 5)
        self.assertEqual(self.game.max_playtime, 30)
        self.assertEqual(self.game.weight, Decimal("1.53"))  # thing pass
        self.assertEqual(self.game.bgg_rank, 1500)
        self.assertEqual(self.game.bgg_rating, Decimal("7.123"))
        self.assertEqual(self.game.bgg_numplays, 54)  # token-free play count
        self.assertIsNotNone(self.game.last_synced_at)
        # DESIGN §10: mechanic tags land via the thing pass too.
        self.assertEqual(
            sorted(gt.tag.name for gt in self.game.game_tags.all()
                   if gt.tag.kind == Tag.Kind.MECHANIC),
            ["Cooperative Game"],
        )
        self.assertIn("mechanic tags added: 1", output)

        # Not in the BGG collection, still backfilled via the thing pass.
        self.voidfall.refresh_from_db()
        self.assertEqual(self.voidfall.weight, Decimal("4.40"))
        self.assertEqual(self.voidfall.year_published, 2023)
        self.assertIn("voidfall.jpg", self.voidfall.image_url)
        self.assertIn("games synced from thing only (not in BGG collection): 1", output)
        # Voidfall's thing item carries no boardgamemechanic link.
        self.assertFalse(
            self.voidfall.game_tags.filter(tag__kind=Tag.Kind.MECHANIC).exists())

    def test_mechanic_tags_are_reconciled_not_just_added(self):
        """DESIGN §10: mechanics are fully BGG-driven (unlike expands), so a
        mechanic BGG no longer reports for a game is removed, not kept."""
        stale_tag = Tag.objects.create(kind=Tag.Kind.MECHANIC, name="Stale Mechanic")
        GameTag.objects.create(game=self.game, tag=stale_tag)

        self.run_sync()

        self.game.refresh_from_db()
        mechanic_names = sorted(
            gt.tag.name for gt in self.game.game_tags.all()
            if gt.tag.kind == Tag.Kind.MECHANIC
        )
        self.assertEqual(mechanic_names, ["Cooperative Game"])

    def test_app_only_data_is_untouched(self):
        self.run_sync()

        self.game.refresh_from_db()
        self.copy.refresh_from_db()
        self.assertEqual(self.game.name, "5 Minute Dungeon (localized)")  # not bgg_name
        self.assertEqual(self.copy.excitement, Decimal("9.0"))
        self.assertEqual(self.copy.keep_status, Copy.KeepStatus.ALWAYS_KEEP)
        self.assertEqual(self.copy.notes, "app-only note")

    def test_dry_run_writes_nothing(self):
        output, _ = self.run_sync(dry_run=True)

        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_name, "")
        self.assertEqual(self.game.image_url, "")
        self.assertIsNone(self.game.last_synced_at)
        self.assertIn("DRY RUN", output)
        self.assertIn("games updated: 2", output)

    def test_resync_is_idempotent(self):
        first_output, _ = self.run_sync()
        self.assertIn("games updated: 2", first_output)
        self.game.refresh_from_db()
        first_values = (self.game.bgg_name, self.game.weight, self.game.bgg_rating)

        second_output, _ = self.run_sync()
        self.assertIn("games updated: 0", second_output)
        self.assertIn("games unchanged: 2", second_output)
        self.game.refresh_from_db()
        self.assertEqual(
            (self.game.bgg_name, self.game.weight, self.game.bgg_rating),
            first_values,
        )

    def _create_expansion(self):
        """An expansion in the DB whose thing item (228552) points at the
        owned base (207830) and at a base NOT in the app (888888)."""
        expansion = Game.objects.create(
            name="5-Minute Dungeon: Curses! Foiled Again!",
            type=Game.Type.EXPANSION,
        )
        BggLink.objects.create(game=expansion, bgg_id=228552, is_primary=True)
        return expansion

    def test_expansion_gains_expands_link_from_thing_pass(self):
        """Issue #40: the thing pass resolves inbound boardgameexpansion
        links to bases already in the DB and fills Game.expands."""
        expansion = self._create_expansion()

        output, _ = self.run_sync()

        self.assertEqual(list(expansion.expands.all()), [self.game])
        # The base's OUTBOUND link to its expansion must not reverse-link,
        # and the base id 888888 (not in the app) is skipped silently.
        self.assertEqual(self.game.expands.count(), 0)
        self.assertIn("expansion links added (Game.expands): 1", output)

        # Idempotent: a re-run adds no links and reports no updates.
        second_output, _ = self.run_sync()
        self.assertEqual(list(expansion.expands.all()), [self.game])
        self.assertNotIn("expansion links added", second_output)
        self.assertIn("games updated: 0", second_output)

    def test_expansion_links_are_add_only(self):
        """Hand-set expands links survive the sync — expands is structural,
        so the sync only ever adds (issue #40 decision)."""
        expansion = self._create_expansion()
        expansion.expands.add(self.voidfall)  # admin-curated, unknown to BGG

        self.run_sync()

        self.assertCountEqual(expansion.expands.all(), [self.game, self.voidfall])

    def test_dry_run_writes_no_expansion_links(self):
        expansion = self._create_expansion()

        self.run_sync(dry_run=True)

        self.assertEqual(expansion.expands.count(), 0)

    def test_reconciliation_reports_both_directions(self):
        output, _ = self.run_sync()

        # (a) On BGG, not in the app — suggested, never auto-added.
        self.assertIn("BGG 999999 ('Mystery Cult')", output)
        self.assertEqual(Game.objects.count(), 2)
        # (b) Owned Copy missing from BGG — flagged with the purchase
        # cross-reference, never auto-removed.
        self.assertIn("'Voidfall' (BGG 337627)", output)
        self.assertIn("Voidfall KS (wave 1: Production)", output)
        self.assertEqual(Copy.objects.count(), 2)

    def test_sync_persists_diff_records(self):
        """Issue #62: the reconciliation lists land as per-owner BggSyncDiff
        rows (unreviewed), feeding the §11 dashboard widget."""
        output, _ = self.run_sync()

        self.assertEqual(BggSyncDiff.objects.count(), 2)
        suggest = BggSyncDiff.objects.get(
            category=BggSyncDiff.Category.SUGGEST_ADD,
        )
        self.assertEqual(suggest.owner, self.user)
        self.assertEqual(suggest.bgg_id, 999999)
        self.assertEqual(suggest.bgg_name, "Mystery Cult")
        self.assertIsNone(suggest.game)
        self.assertIsNone(suggest.dismissed_at)
        self.assertIsNotNone(suggest.last_seen_at)
        missing = BggSyncDiff.objects.get(
            category=BggSyncDiff.Category.MISSING_FROM_BGG,
        )
        self.assertEqual(missing.game, self.voidfall)
        self.assertEqual(missing.bgg_id, 337627)
        self.assertIn("Voidfall KS (wave 1: Production)", missing.note)
        self.assertIn("sync diffs recorded (new, unreviewed): 2", output)
        self.assertIn("sync diffs open after run: 2", output)

    def test_sync_persists_prev_owned_and_archived_diffs(self):
        """Issue #62: the two §4-mapping diff directions persist too."""
        # Active copy, but prevowned on BGG (same setup as the report test).
        bardsung = Game.objects.create(name="Bardsung")
        BggLink.objects.create(game=bardsung, bgg_id=245638, is_primary=True)
        bardsung_edition = Edition.objects.create(game=bardsung, is_default=True)
        Copy.objects.create(owner=self.user, edition=bardsung_edition)
        # Archived copy of a game BGG still lists as own=1.
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.save(update_fields=["archive_status"])

        self.run_sync(collections={"prevowned": PREVOWNED_COLLECTION_XML})

        prev_owned = BggSyncDiff.objects.get(
            category=BggSyncDiff.Category.PREV_OWNED_ACTIVE,
        )
        self.assertEqual(prev_owned.game, bardsung)
        self.assertEqual(prev_owned.bgg_id, 245638)
        archived = BggSyncDiff.objects.get(
            category=BggSyncDiff.Category.ARCHIVED_ON_BGG,
        )
        self.assertEqual(archived.game, self.game)
        self.assertEqual(archived.bgg_id, 207830)
        self.assertIn("BGG still says owned", archived.note)

    def test_rerun_preserves_dismissed_diffs(self):
        """Issue #62 / §8 "never nag again": a dismissed diff that is still
        observed keeps its dismissal across syncs; the upsert only bumps
        last_seen_at."""
        self.run_sync()
        diff = BggSyncDiff.objects.get(
            category=BggSyncDiff.Category.MISSING_FROM_BGG,
        )
        diff.dismissed_at = timezone.now()
        diff.save(update_fields=["dismissed_at"])
        before = (diff.pk, diff.dismissed_at, diff.last_seen_at)

        second_output, _ = self.run_sync()

        diff.refresh_from_db()
        self.assertEqual(diff.pk, before[0])
        self.assertEqual(diff.dismissed_at, before[1])
        self.assertGreater(diff.last_seen_at, before[2])
        self.assertEqual(BggSyncDiff.objects.count(), 2)
        self.assertNotIn("sync diffs recorded (new", second_output)

    def test_resolved_diffs_are_deleted_and_reappear_unreviewed(self):
        """Issue #62: a diff no longer observed is deleted — even a
        dismissed one — so a later reappearance is a new, unreviewed
        occurrence that nags again."""
        self.run_sync()
        diff = BggSyncDiff.objects.get(
            category=BggSyncDiff.Category.MISSING_FROM_BGG,
        )
        diff.dismissed_at = timezone.now()
        diff.save(update_fields=["dismissed_at"])
        old_pk = diff.pk

        # Archiving the copy resolves the diff: no active copy, so Voidfall
        # is no longer expected on BGG.
        voidfall_copy = Copy.objects.get(edition__game=self.voidfall)
        voidfall_copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        voidfall_copy.save(update_fields=["archive_status"])
        output, _ = self.run_sync()

        self.assertFalse(BggSyncDiff.objects.filter(pk=old_pk).exists())
        self.assertIn("sync diffs resolved (removed): 1", output)

        # The copy comes back: the same divergence is a fresh occurrence.
        voidfall_copy.archive_status = Copy.ArchiveStatus.ACTIVE
        voidfall_copy.save(update_fields=["archive_status"])
        self.run_sync()

        reborn = BggSyncDiff.objects.get(
            category=BggSyncDiff.Category.MISSING_FROM_BGG,
        )
        self.assertNotEqual(reborn.pk, old_pk)
        self.assertIsNone(reborn.dismissed_at)

    def test_dry_run_persists_no_diffs(self):
        self.run_sync(dry_run=True)

        self.assertEqual(BggSyncDiff.objects.count(), 0)

    def test_preordered_and_prevowned_statuses_are_synced_and_marked(self):
        bardsung = Game.objects.create(name="Bardsung")
        BggLink.objects.create(game=bardsung, bgg_id=245638, is_primary=True)
        bardsung_edition = Edition.objects.create(game=bardsung, is_default=True)
        Copy.objects.create(owner=self.user, edition=bardsung_edition)

        output, _ = self.run_sync(collections={
            "preordered": PREORDERED_COLLECTION_XML,
            "prevowned": PREVOWNED_COLLECTION_XML,
        })

        # 207830 carries own=1 AND prevowned=1 — own wins the precedence.
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, Game.BggCollectionStatus.OWN)

        # The preorder now syncs from the collection payload itself, no
        # longer only via the (token-blocked) thing pass.
        self.voidfall.refresh_from_db()
        self.assertEqual(
            self.voidfall.bgg_collection_status, Game.BggCollectionStatus.PREORDERED,
        )
        self.assertIn("voidfall.jpg", self.voidfall.image_url)

        # Previously owned: stats synced, status stored, active copy kept
        # and reported — NOT flagged missing-from-BGG.
        bardsung.refresh_from_db()
        self.assertEqual(
            bardsung.bgg_collection_status, Game.BggCollectionStatus.PREV_OWNED,
        )
        self.assertIn("bardsung.jpg", bardsung.image_url)
        self.assertIn("active copies marked previously-owned on BGG: 1", output)
        self.assertNotIn("'Bardsung' (BGG 245638)", output)
        self.assertEqual(Copy.objects.count(), 3)

    def test_wishlist_status_applied_to_matched_games_only(self):
        # A wishlisted game that IS in the app gets the status + stats +
        # play count; a wishlisted game that is NOT is never suggested.
        wished = Game.objects.create(name="Wishlisted In App")
        BggLink.objects.create(game=wished, bgg_id=137637, is_primary=True)

        output, _ = self.run_sync(collections={"wishlist": WISHLIST_COLLECTION_XML})

        wished.refresh_from_db()
        self.assertEqual(
            wished.bgg_collection_status, Game.BggCollectionStatus.WISHLIST,
        )
        self.assertEqual(wished.year_published, 2020)
        self.assertIn("wish.jpg", wished.image_url)
        self.assertEqual(wished.bgg_numplays, 3)
        self.assertIn("games synced from wishlist (already in the app): 1", output)

        # The wishlist-only BGG item (222222) is NOT suggested for adding
        # and no Game was created for it.
        self.assertNotIn("222222", output)
        self.assertNotIn("Wishlisted Not In App", output)
        self.assertFalse(BggLink.objects.filter(bgg_id=222222).exists())

    def test_wishlist_priority_synced_from_wishlist_payload(self):
        # Issue #81: the priority rides the same payload — no extra request.
        wished = Game.objects.create(name="Wishlisted In App")
        BggLink.objects.create(game=wished, bgg_id=137637, is_primary=True)

        self.run_sync(collections={"wishlist": WISHLIST_COLLECTION_XML})

        wished.refresh_from_db()
        self.assertEqual(wished.bgg_wishlist_priority, 2)

        # Off the wishlist next run -> status and priority both clear.
        self.run_sync()
        wished.refresh_from_db()
        self.assertEqual(wished.bgg_collection_status, "")
        self.assertIsNone(wished.bgg_wishlist_priority)

    def test_membership_status_wins_over_wishlist(self):
        # 207830 is owned AND on the wishlist payload — own must win.
        owned_and_wished = WISHLIST_COLLECTION_XML.replace(
            'objectid="137637"', 'objectid="207830"',
        )
        self.run_sync(collections={"wishlist": owned_and_wished})

        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, Game.BggCollectionStatus.OWN)

    def test_stale_collection_status_is_cleared(self):
        # Was prevowned last sync; the mark has since been removed on BGG
        # (not in any status payload this run).
        self.voidfall.bgg_collection_status = Game.BggCollectionStatus.PREV_OWNED
        self.voidfall.save(update_fields=["bgg_collection_status"])

        output, _ = self.run_sync()

        self.voidfall.refresh_from_db()
        self.assertEqual(self.voidfall.bgg_collection_status, "")
        self.assertIn("games whose BGG collection status was cleared: 1", output)

    def test_archived_copy_still_owned_on_bgg_is_flagged(self):
        # §4 archive mapping, reverse direction: the copy left the shelf in
        # the app but BGG still says own=1. The API cannot write collections,
        # so the report nudges a manual BGG fix — nothing is changed.
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.archive_reason = Copy.ArchiveReason.SOLD
        self.copy.save(update_fields=["archive_status", "archive_reason"])

        output, _ = self.run_sync()

        self.assertIn("archived copies still owned on BGG: 1", output)
        self.assertIn("'5 Minute Dungeon (localized)': BGG says owned", output)
        # The BGG status still syncs (evidence, not authority) and the
        # archive is untouched — the diff is report-only.
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, Game.BggCollectionStatus.OWN)
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.archive_status, Copy.ArchiveStatus.ARCHIVED)
        # And the archived copy no longer counts as an active-copy diff.
        self.assertIn("your active copies (games): 1", output)

    def test_archived_copy_with_an_active_rebuy_is_not_flagged(self):
        # An active Copy of the same game (upgraded to another edition)
        # makes own=1 consistent — the archived sibling raises no reverse
        # diff. (owner, edition) is unique, so the rebuy gets its own row.
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.save(update_fields=["archive_status"])
        collectors = Edition.objects.create(game=self.game, name="Collector's")
        Copy.objects.create(owner=self.user, edition=collectors)

        output, _ = self.run_sync()

        self.assertIn("archived copies still owned on BGG: 0", output)

    def test_archived_copy_marked_prevowned_on_bgg_is_consistent(self):
        # Archived here, previously-owned there: the two signals agree, so
        # neither direction of the report flags the game.
        bardsung = Game.objects.create(name="Bardsung")
        BggLink.objects.create(game=bardsung, bgg_id=245638, is_primary=True)
        bardsung_edition = Edition.objects.create(game=bardsung, is_default=True)
        Copy.objects.create(
            owner=self.user, edition=bardsung_edition,
            archive_status=Copy.ArchiveStatus.ARCHIVED,
            archive_reason=Copy.ArchiveReason.SOLD,
        )

        output, _ = self.run_sync(
            collections={"prevowned": PREVOWNED_COLLECTION_XML},
        )

        self.assertIn("archived copies still owned on BGG: 0", output)
        self.assertIn("active copies marked previously-owned on BGG: 0", output)
        self.assertNotIn("'Bardsung' (BGG 245638)", output)

    def test_thing_pass_degrades_when_session_is_refused(self):
        output, _ = self.run_sync(
            things=BggAuthError("BGG refused the session on /thing (401)."),
        )

        # Collection data still lands; weight stays blocked.
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_name, "5-Minute Dungeon")
        self.assertIsNone(self.game.weight)
        self.voidfall.refresh_from_db()
        self.assertEqual(self.voidfall.image_url, "")
        self.assertIn("Thing pass blocked", output)
        self.assertIn("DESIGN §15", output)

    @override_settings(BGG_USERNAME="", BGG_PASSWORD="", BGG_API_TOKEN="")
    def test_missing_credentials_fail_loudly_without_touching_bgg(self):
        with self.assertRaisesMessage(CommandError, "credentials are not configured"):
            call_command("sync_bgg", user="kernicek")

    # --- Plays history (issue #65) ---------------------------------------

    def test_syncs_plays_joined_to_game_with_players(self):
        output, client = self.run_sync(plays=PLAYS_XML)

        client.get_plays.assert_called_once_with("kernicek", page=1, bgg_id=None)
        # Two plays join to 5-Minute Dungeon (207830); the third (555555) has
        # no game in the app and is skipped.
        self.assertEqual(self.game.plays.count(), 2)
        self.assertIn("plays synced: 2", output)
        self.assertIn("plays skipped (no matching game): 1", output)
        self.assertFalse(Play.objects.filter(external_id="102").exists())

        play = self.game.plays.get(external_id="100")
        self.assertEqual(play.source, Play.Source.BGG)
        self.assertEqual(play.play_date, datetime.date(2024, 6, 1))
        self.assertEqual(play.location, "Home")
        self.assertEqual(play.length_minutes, 25)
        winner = play.players.get(won=True)
        self.assertEqual(winner.name, "Vojta")
        self.assertEqual(winner.score, "42")
        self.assertEqual(play.players.count(), 2)

        incomplete = self.game.plays.get(external_id="101")
        self.assertTrue(incomplete.incomplete)
        self.assertEqual(incomplete.quantity, 2)
        self.assertEqual(incomplete.players.count(), 0)

    def test_re_sync_is_idempotent_and_replaces_players(self):
        self.run_sync(plays=PLAYS_XML)
        self.run_sync(plays=PLAYS_XML)

        # No duplicate Play/PlayPlayer rows — update_or_create keys on
        # (source, external_id) and players are replaced, not appended.
        self.assertEqual(self.game.plays.count(), 2)
        self.assertEqual(
            self.game.plays.get(external_id="100").players.count(), 2,
        )
        self.assertEqual(Play.objects.count(), 2)
        self.assertEqual(PlayPlayer.objects.count(), 2)

    def test_plays_pass_degrades_when_session_is_refused(self):
        output, _ = self.run_sync(
            plays=BggAuthError("BGG refused the session on /plays (401)."),
        )

        # Collection data still lands; only the plays pass is blocked.
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_name, "5-Minute Dungeon")
        self.assertEqual(Play.objects.count(), 0)
        self.assertIn("Plays pass blocked", output)

    def test_plays_pagination_follows_the_total(self):
        page1 = (
            '<plays username="kernicek" total="2" page="1">'
            '<play id="200" date="2024-01-01" quantity="1">'
            '<item objectid="207830"/></play></plays>'
        )
        page2 = (
            '<plays username="kernicek" total="2" page="2">'
            '<play id="201" date="2024-01-02" quantity="1">'
            '<item objectid="207830"/></play></plays>'
        )
        with mock.patch("gamekeeper.bgg_sync.time.sleep"):
            output, client = self.run_sync(plays={1: page1, 2: page2})

        self.assertEqual(client.get_plays.call_count, 2)
        client.get_plays.assert_any_call("kernicek", page=2, bgg_id=None)
        self.assertEqual(self.game.plays.count(), 2)

    # --- write-back reconciliation (issue #117) ---------------------------

    def test_pending_push_survives_a_disagreeing_read(self):
        """A push_bgg_status write still within its confirmation window must
        not be stale-cleared or diff-flagged just because BGG's export
        hasn't caught up (self.game carries own=1 in the default fixture)."""
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.save(update_fields=["archive_status"])
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREV_OWNED
        self.game.bgg_status_pushed = Game.BggCollectionStatus.PREV_OWNED
        self.game.bgg_status_pushed_at = timezone.now()
        self.game.save(update_fields=[
            "bgg_collection_status", "bgg_status_pushed", "bgg_status_pushed_at",
        ])

        self.run_sync()

        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, Game.BggCollectionStatus.PREV_OWNED)
        self.assertEqual(self.game.bgg_status_pushed, Game.BggCollectionStatus.PREV_OWNED)
        self.assertIsNotNone(self.game.bgg_status_pushed_at)
        self.assertFalse(
            BggSyncDiff.objects.filter(category=BggSyncDiff.Category.ARCHIVED_ON_BGG).exists(),
        )

    def test_confirmed_push_clears_the_marker_and_any_push_failed_diff(self):
        """Once a read agrees with what we pushed, the marker (and a stale
        PUSH_FAILED row from an earlier failed attempt) both clear."""
        self.game.bgg_status_pushed = Game.BggCollectionStatus.OWN
        self.game.bgg_status_pushed_at = timezone.now()
        self.game.save(update_fields=["bgg_status_pushed", "bgg_status_pushed_at"])
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.PUSH_FAILED,
            bgg_id=207830, game=self.game, last_seen_at=timezone.now(),
        )

        self.run_sync()  # default collection carries own=1 for 207830

        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_status_pushed, "")
        self.assertIsNone(self.game.bgg_status_pushed_at)
        self.assertFalse(
            BggSyncDiff.objects.filter(category=BggSyncDiff.Category.PUSH_FAILED).exists(),
        )

    def test_expired_pending_push_lets_the_read_win(self):
        """Past PUSH_CONFIRM_WINDOW, a still-disagreeing read is trusted
        again — the marker clears and normal reconciliation resumes."""
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.save(update_fields=["archive_status"])
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREV_OWNED
        self.game.bgg_status_pushed = Game.BggCollectionStatus.PREV_OWNED
        self.game.bgg_status_pushed_at = timezone.now() - PUSH_CONFIRM_WINDOW - datetime.timedelta(hours=1)
        self.game.save(update_fields=[
            "bgg_collection_status", "bgg_status_pushed", "bgg_status_pushed_at",
        ])

        self.run_sync()  # default collection still carries own=1 for 207830

        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, Game.BggCollectionStatus.OWN)
        self.assertEqual(self.game.bgg_status_pushed, "")
        self.assertIsNone(self.game.bgg_status_pushed_at)
        self.assertTrue(
            BggSyncDiff.objects.filter(category=BggSyncDiff.Category.ARCHIVED_ON_BGG).exists(),
        )

    def test_persist_diffs_preserves_rows_it_does_not_manage(self):
        """PUSH_FAILED (issue #117) and NEW_EXPANSION rows are written by
        other code paths (push_bgg_status / the new-expansion widget) — a
        sync run must not wipe them just because it never observes them."""
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.PUSH_FAILED,
            bgg_id=555555, last_seen_at=timezone.now(),
        )
        BggSyncDiff.objects.create(
            owner=self.user, category=BggSyncDiff.Category.NEW_EXPANSION,
            bgg_id=666666, last_seen_at=timezone.now(),
        )

        self.run_sync()

        self.assertTrue(
            BggSyncDiff.objects.filter(
                category=BggSyncDiff.Category.PUSH_FAILED, bgg_id=555555,
            ).exists(),
        )
        self.assertTrue(
            BggSyncDiff.objects.filter(
                category=BggSyncDiff.Category.NEW_EXPANSION, bgg_id=666666,
            ).exists(),
        )


class SyncExpansionLinksTests(TestCase):
    """sync_expansion_links — the manual geekitems stopgap for issue #40,
    with the HTTP layer mocked out (no live BGG in tests)."""

    @classmethod
    def setUpTestData(cls):
        cls.base = Game.objects.create(name="5 Minute Dungeon (localized)")
        BggLink.objects.create(game=cls.base, bgg_id=207830, is_primary=True)
        cls.expansion = Game.objects.create(
            name="5-Minute Dungeon: Curses! Foiled Again!",
            type=Game.Type.EXPANSION,
        )
        BggLink.objects.create(game=cls.expansion, bgg_id=228552, is_primary=True)

    def run_command(self, payloads=None, **extra):
        """call_command with BggClient mocked; returns (output, client mock).

        payloads maps bgg_id -> geekitems JSON (or an exception to raise);
        default answers 228552 with GEEKITEM_JSON. Requesting an id not in
        the map is a test bug and KeyErrors."""
        if payloads is None:
            payloads = {228552: GEEKITEM_JSON}

        def get_geekitem(bgg_id):
            result = payloads[bgg_id]
            if isinstance(result, Exception):
                raise result
            return result

        out = StringIO()
        with mock.patch(
            "gamekeeper.management.commands.sync_expansion_links.BggClient",
        ) as client_class, mock.patch(
            "gamekeeper.management.commands.sync_expansion_links.time.sleep",
        ):
            client = client_class.return_value
            client.get_geekitem.side_effect = get_geekitem
            call_command("sync_expansion_links", stdout=out, **extra)
        return out.getvalue(), client

    def test_links_expansion_to_base_without_logging_in(self):
        output, client = self.run_command()

        client.login.assert_not_called()  # geekitems answers anonymously
        self.assertEqual(list(self.expansion.expands.all()), [self.base])
        self.assertIn("links added: 1", output)
        # The other base (888888) is not in the app — reported, not linked.
        self.assertIn("expands BGG 888888", output)
        self.assertIn("base games not in the app: 1", output)

    def test_default_skips_already_linked_expansions(self):
        self.expansion.expands.add(self.base)

        output, client = self.run_command(payloads={})  # any fetch KeyErrors

        client.get_geekitem.assert_not_called()
        self.assertIn("Fetching geekitems for 0 expansion(s)", output)

    def test_all_refreshes_without_removing_hand_set_links(self):
        hand_set = Game.objects.create(name="Hand-linked Base")
        self.expansion.expands.add(hand_set)  # admin-curated, unknown to BGG

        output, _ = self.run_command(all=True)

        self.assertCountEqual(self.expansion.expands.all(), [self.base, hand_set])
        self.assertIn("links added: 1", output)

    def test_dry_run_writes_nothing(self):
        output, _ = self.run_command(dry_run=True)

        self.assertEqual(self.expansion.expands.count(), 0)
        self.assertIn("DRY RUN", output)
        self.assertIn("links added: 1", output)  # reported, then rolled back

    def test_expansion_without_primary_link_is_reported_not_fetched(self):
        Game.objects.create(name="Orphan Expansion", type=Game.Type.EXPANSION)

        output, _ = self.run_command()

        self.assertIn("'Orphan Expansion': no primary BGG link", output)
        self.assertIn("expansions without a primary BGG link: 1", output)

    def test_bgg_error_aborts_fetching_but_keeps_earlier_links(self):
        """Politeness: an error stops the fetch loop instead of hammering
        on; whatever was already fetched still lands."""
        refused = Game.objects.create(
            name="Refused Expansion", type=Game.Type.EXPANSION,
        )
        BggLink.objects.create(game=refused, bgg_id=333333, is_primary=True)

        # name-ordered: 228552 fetches fine first, then 333333 errors.
        output, _ = self.run_command(payloads={
            228552: GEEKITEM_JSON,
            333333: BggError("BGG kept answering 503."),
        })

        self.assertEqual(list(self.expansion.expands.all()), [self.base])
        self.assertEqual(refused.expands.count(), 0)
        self.assertIn("ABORTED", output)
        self.assertIn("BGG kept answering 503.", output)


class SyncNewExpansionsTests(TestCase):
    """sync_new_expansions — the manual geekitems stopgap for issue #64,
    with the HTTP layer mocked out (no live BGG in tests)."""

    @classmethod
    def setUpTestData(cls):
        cls.owner = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.base = Game.objects.create(name="5 Minute Dungeon")
        BggLink.objects.create(game=cls.base, bgg_id=207830, is_primary=True)
        edition = Edition.objects.create(game=cls.base, is_default=True)
        Copy.objects.create(owner=cls.owner, edition=edition)

    def run_command(self, payloads=None, **extra):
        """call_command with BggClient mocked; returns (output, client mock).

        payloads maps bgg_id -> geekitems JSON (or an exception to raise);
        default answers 207830 with the two-expansion baseline fixture."""
        if payloads is None:
            payloads = {207830: BASE_GEEKITEM_WITH_EXPANSIONS_JSON}

        def get_geekitem(bgg_id):
            result = payloads[bgg_id]
            if isinstance(result, Exception):
                raise result
            return result

        out = StringIO()
        with mock.patch(
            "gamekeeper.management.commands.sync_new_expansions.BggClient",
        ) as client_class, mock.patch(
            "gamekeeper.management.commands.sync_new_expansions.time.sleep",
        ):
            client = client_class.return_value
            client.get_geekitem.side_effect = get_geekitem
            call_command("sync_new_expansions", stdout=out, **extra)
        return out.getvalue(), client

    def test_first_sync_seeds_baseline_without_notifying(self):
        output, client = self.run_command()

        client.login.assert_not_called()  # geekitems answers anonymously
        self.assertEqual(
            set(ExpansionSighting.objects.values_list("bgg_id", flat=True)),
            {228552, 654321},
        )
        self.assertEqual(BggSyncDiff.objects.count(), 0)
        self.assertIn("seeded 2 baseline expansion(s)", output)
        self.assertIn("new expansions discovered: 0", output)

    def test_later_sync_notifies_every_active_owner_of_genuinely_new_expansion(self):
        self.run_command()  # establish the baseline first

        second_owner = get_user_model().objects.create_user(
            username="second", password="pass",
        )
        Copy.objects.create(owner=second_owner, edition=self.base.editions.get())

        output, _ = self.run_command(
            payloads={207830: BASE_GEEKITEM_WITH_NEW_EXPANSION_JSON},
        )

        self.assertIn("new expansions discovered: 1", output)
        sighting = ExpansionSighting.objects.get(bgg_id=999001)
        self.assertEqual(sighting.base, self.base)
        diffs = BggSyncDiff.objects.filter(
            category=BggSyncDiff.Category.NEW_EXPANSION, bgg_id=999001,
        )
        self.assertEqual(
            set(diffs.values_list("owner", flat=True)),
            {self.owner.pk, second_owner.pk},
        )
        for diff in diffs:
            self.assertEqual(diff.bgg_name, "5-Minute Dungeon: Another New Threat")
            self.assertEqual(diff.note, "New expansion for 5 Minute Dungeon")

    def test_owner_added_after_baseline_is_not_spammed(self):
        """The global first-seen table (not per-owner) means a game already
        known to the app doesn't dump its whole expansion list on a new
        owner — only a genuinely new sighting notifies."""
        self.run_command()  # baseline established with only self.owner

        late_owner = get_user_model().objects.create_user(
            username="late", password="pass",
        )
        Copy.objects.create(owner=late_owner, edition=self.base.editions.get())

        # Same expansions as the baseline — nothing new to discover.
        self.run_command()

        self.assertEqual(BggSyncDiff.objects.filter(owner=late_owner).count(), 0)

    def test_base_without_primary_link_is_reported_not_fetched(self):
        Game.objects.create(name="Orphan Base")  # no Copy at all -> not even queried
        edition = Edition.objects.create(
            game=Game.objects.create(name="Linkless Owned Base"), is_default=True,
        )
        Copy.objects.create(owner=self.owner, edition=edition)

        output, _ = self.run_command()

        self.assertIn("'Linkless Owned Base': no primary BGG link", output)
        self.assertIn("base games without a primary BGG link: 1", output)

    def test_bgg_error_aborts_fetching_but_keeps_earlier_sightings(self):
        other_base = Game.objects.create(name="Refused Base")
        BggLink.objects.create(game=other_base, bgg_id=333333, is_primary=True)
        other_edition = Edition.objects.create(game=other_base, is_default=True)
        Copy.objects.create(owner=self.owner, edition=other_edition)

        # name-ordered ("5 Minute Dungeon" < "Refused Base"): 207830 fetches
        # fine first, then 333333 errors.
        output, _ = self.run_command(payloads={
            207830: BASE_GEEKITEM_WITH_EXPANSIONS_JSON,
            333333: BggError("BGG kept answering 503."),
        })

        self.assertEqual(ExpansionSighting.objects.filter(base=self.base).count(), 2)
        self.assertEqual(ExpansionSighting.objects.filter(base=other_base).count(), 0)
        self.assertIn("ABORTED", output)
        self.assertIn("BGG kept answering 503.", output)

    def test_dry_run_writes_nothing(self):
        output, _ = self.run_command(dry_run=True)

        self.assertEqual(ExpansionSighting.objects.count(), 0)
        self.assertIn("DRY RUN", output)
        self.assertIn("seeded 2 baseline expansion(s)", output)  # reported, then rolled back


@override_settings(BGG_USERNAME="kernicek", BGG_PASSWORD="test-password")
class GameBggSyncViewTests(TestCase):
    """Per-game on-demand 'Sync from BGG' button (issue #44): one id-filtered
    collection request (no full-collection download) settles the game, and
    geekitems backfills expansion links, with the HTTP layer mocked out."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(name="5 Minute Dungeon (localized)")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)

    def post_sync(self, pk=None, *, collection=COLLECTION_XML, geekitem=GEEKITEM_JSON,
                  login_error=None, collection_error=None, geekitem_error=None,
                  plays=EMPTY_PLAYS_XML, plays_error=None):
        """POST the sync endpoint with bgg_sync.BggClient mocked; returns
        (response, client mock). One unfiltered id-scoped collection call
        carries the game's stats + status flags."""
        with mock.patch("gamekeeper.bgg_sync.BggClient") as client_class:
            client = client_class.return_value
            if login_error is not None:
                client.login.side_effect = login_error
            if collection_error is not None:
                client.get_collection.side_effect = collection_error
            else:
                client.get_collection.return_value = collection
            if geekitem_error is not None:
                client.get_geekitem.side_effect = geekitem_error
            else:
                client.get_geekitem.return_value = geekitem
            if plays_error is not None:
                client.get_plays.side_effect = plays_error
            else:
                client.get_plays.return_value = plays
            response = self.client.post(f"/games/{pk or self.game.pk}/sync/")
        return response, client

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.post(f"/games/{self.game.pk}/sync/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_get_is_not_allowed(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/games/{self.game.pk}/sync/")
        self.assertEqual(response.status_code, 405)

    def test_syncs_bgg_fields_from_one_id_filtered_request(self):
        self.client.login(username="kernicek", password="pass")
        response, client = self.post_sync()

        self.assertEqual(response.status_code, 200)
        client.login.assert_called_once_with()
        # ONE collection request (id-filtered, no status filter) — not the
        # whole collection, and not one request per status.
        client.get_collection.assert_called_once_with(
            "kernicek", status=None, bgg_id=207830,
        )
        # Base game: geekitems (expansion links) is skipped entirely.
        client.get_geekitem.assert_not_called()

        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_name, "5-Minute Dungeon")
        self.assertEqual(self.game.min_players, 2)
        self.assertEqual(self.game.bgg_rating, Decimal("7.123"))
        self.assertEqual(self.game.bgg_numplays, 54)
        self.assertIsNone(self.game.weight)  # /thing-only; sync_game never calls /thing
        self.assertEqual(self.game.bgg_collection_status, Game.BggCollectionStatus.OWN)
        self.assertIsNotNone(self.game.last_synced_at)
        # Curated field survives the sync.
        self.assertEqual(self.game.name, "5 Minute Dungeon (localized)")
        self.assertContains(response, "Synced")

    def test_resync_reports_already_up_to_date(self):
        self.client.login(username="kernicek", password="pass")
        self.post_sync()
        response, _ = self.post_sync()
        self.assertContains(response, "Already up to date")

    def test_missing_from_collection_clears_stale_status(self):
        self.game.bgg_collection_status = Game.BggCollectionStatus.OWN
        self.game.save(update_fields=["bgg_collection_status"])
        self.client.login(username="kernicek", password="pass")

        # BGG lists it in no status — the stale 'own' is cleared (a change).
        response, _ = self.post_sync(collection=EMPTY_COLLECTION_XML)
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, "")
        self.assertContains(response, "Synced")

    def test_sync_writes_and_clears_wishlist_priority(self):
        # Issue #81: the priority is a synced field — set from the payload's
        # wishlistpriority attribute, cleared when the game leaves BGG.
        wished_xml = COLLECTION_XML.replace(
            '<status own="1" lastmodified="2024-01-01 00:00:00"/>',
            '<status own="0" wishlist="1" wishlistpriority="3" '
            'lastmodified="2024-01-01 00:00:00"/>',
        )
        self.client.login(username="kernicek", password="pass")

        self.post_sync(collection=wished_xml)
        self.game.refresh_from_db()
        self.assertEqual(
            self.game.bgg_collection_status, Game.BggCollectionStatus.WISHLIST,
        )
        self.assertEqual(self.game.bgg_wishlist_priority, 3)

        self.post_sync(collection=EMPTY_COLLECTION_XML)
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_collection_status, "")
        self.assertIsNone(self.game.bgg_wishlist_priority)

    def test_not_in_collection_reports_nothing_to_update(self):
        self.client.login(username="kernicek", password="pass")
        response, _ = self.post_sync(collection=EMPTY_COLLECTION_XML)
        self.game.refresh_from_db()
        self.assertContains(response, "Not found in your BGG collection")
        # Nothing to write, so no phantom last_synced_at.
        self.assertIsNone(self.game.last_synced_at)

    def test_game_without_primary_link_reports_no_link(self):
        linkless = Game.objects.create(name="Linkless")
        self.client.login(username="kernicek", password="pass")
        with mock.patch("gamekeeper.bgg_sync.BggClient") as client_class:
            response = self.client.post(f"/games/{linkless.pk}/sync/")
        self.assertContains(response, "No primary BGG link")
        client_class.assert_not_called()  # short-circuits before any network

    def test_login_failure_reports_error_and_writes_nothing(self):
        self.client.login(username="kernicek", password="pass")
        response, _ = self.post_sync(
            login_error=BggAuthError("BGG login failed: bad credentials"),
        )
        self.game.refresh_from_db()
        self.assertContains(response, "BGG login failed")
        self.assertEqual(self.game.bgg_name, "")
        self.assertIsNone(self.game.last_synced_at)

    def test_throttle_reports_error_and_writes_nothing(self):
        self.client.login(username="kernicek", password="pass")
        response, _ = self.post_sync(
            collection_error=BggError("BGG kept answering 429 after 7 attempts."),
        )
        self.game.refresh_from_db()
        self.assertContains(response, "BGG sync failed")
        self.assertEqual(self.game.bgg_name, "")
        self.assertIsNone(self.game.last_synced_at)

    def test_expansion_links_base_from_geekitems(self):
        """Issue #40 on demand: an expansion's geekitems payload resolves its
        base links to Games already in the app and fills Game.expands."""
        expansion = Game.objects.create(
            name="Curses! Foiled Again!", type=Game.Type.EXPANSION,
        )
        BggLink.objects.create(game=expansion, bgg_id=228552, is_primary=True)
        self.client.login(username="kernicek", password="pass")

        # Not in the collection; geekitems still supplies the base links.
        response, client = self.post_sync(
            pk=expansion.pk, collection=EMPTY_COLLECTION_XML,
        )
        client.get_geekitem.assert_called_once_with(228552)
        # GEEKITEM_JSON bases: 207830 (self.game, in app) + 888888 (not).
        self.assertEqual(list(expansion.expands.all()), [self.game])
        expansion.refresh_from_db()
        self.assertIsNotNone(expansion.last_synced_at)  # link counts as a touch
        self.assertContains(response, "Synced")

    def test_geekitems_failure_does_not_fail_the_sync(self):
        expansion = Game.objects.create(name="Exp", type=Game.Type.EXPANSION)
        BggLink.objects.create(game=expansion, bgg_id=228552, is_primary=True)
        self.client.login(username="kernicek", password="pass")

        response, _ = self.post_sync(
            pk=expansion.pk, collection=EMPTY_COLLECTION_XML,
            geekitem_error=BggError("geekitems kept answering 503."),
        )
        self.assertEqual(expansion.expands.count(), 0)
        self.assertContains(response, "Expansion links skipped")  # links_note
        self.assertNotContains(response, "text-danger")  # not a fatal error

    @override_settings(BGG_USERNAME="", BGG_PASSWORD="")
    def test_missing_credentials_reports_config_error(self):
        self.client.login(username="kernicek", password="pass")
        with mock.patch("gamekeeper.bgg_sync.BggClient") as client_class:
            response = self.client.post(f"/games/{self.game.pk}/sync/")
        self.assertContains(response, "credentials are not configured")
        client_class.assert_not_called()

    def test_per_game_refresh_pulls_this_games_plays(self):
        """Issue #65: the detail-page refresh also pulls the game's plays,
        id-filtered — best-effort, so a plays hiccup never fails the sync."""
        self.client.login(username="kernicek", password="pass")
        response, client = self.post_sync(plays=PLAYS_XML)

        self.assertEqual(response.status_code, 200)
        client.get_plays.assert_called_once_with("kernicek", page=1, bgg_id=207830)
        self.assertEqual(self.game.plays.count(), 2)  # 555555 not this game

    def test_per_game_refresh_survives_a_blocked_plays_pass(self):
        self.client.login(username="kernicek", password="pass")
        response, _ = self.post_sync(
            plays_error=BggAuthError("BGG refused the session on /plays (401)."),
        )
        # Collection sync still succeeds; plays just don't land.
        self.game.refresh_from_db()
        self.assertEqual(self.game.bgg_name, "5-Minute Dungeon")
        self.assertEqual(self.game.plays.count(), 0)
        self.assertContains(response, "Synced")


# geekitems JSON for a BASE game: no expandsboardgame links, so the add-game
# type probe keeps Type.BASE.
BASE_GEEKITEM_JSON = """{"item": {"itemid": "207830", "objecttype": "thing",
  "links": {"expandsboardgame": [], "boardgameexpansion": []}}}"""

# geekitems JSON for base game 207830 listing its own expansions outbound
# (issue #64) — one already in the app (228552) and one brand new (654321).
BASE_GEEKITEM_WITH_EXPANSIONS_JSON = """{"item": {"itemid": "207830", "objecttype": "thing",
  "links": {"expandsboardgame": [], "boardgameexpansion": [
    {"objectid": "228552", "name": "5-Minute Dungeon: Curses! Foiled Again!", "objecttype": "thing"},
    {"objectid": "654321", "name": "5-Minute Dungeon: New Threat", "objecttype": "thing"}
  ]}}}"""

# Same base, one further expansion (999001) added — simulates a later sync
# discovering something genuinely new after the baseline above.
BASE_GEEKITEM_WITH_NEW_EXPANSION_JSON = """{"item": {"itemid": "207830", "objecttype": "thing",
  "links": {"expandsboardgame": [], "boardgameexpansion": [
    {"objectid": "228552", "name": "5-Minute Dungeon: Curses! Foiled Again!", "objecttype": "thing"},
    {"objectid": "654321", "name": "5-Minute Dungeon: New Threat", "objecttype": "thing"},
    {"objectid": "999001", "name": "5-Minute Dungeon: Another New Threat", "objecttype": "thing"}
  ]}}}"""

# The expansion 228552 as an owned collection item (GEEKITEM_JSON's subject),
# so the add-game flow can create it through the sync path.
EXPANSION_COLLECTION_XML = """<?xml version="1.0" encoding="utf-8"?>
<items totalitems="1" termsofuse="https://boardgamegeek.com/xmlapi/termsofuse">
  <item objecttype="thing" objectid="228552" subtype="boardgame" collid="9">
    <name sortindex="1">5-Minute Dungeon: Curses! Foiled Again!</name>
    <yearpublished>2018</yearpublished>
    <image>https://cf.geekdo-images.com/large/curses.jpg</image>
    <thumbnail>https://cf.geekdo-images.com/thumb/curses.jpg</thumbnail>
    <stats minplayers="2" maxplayers="6" minplaytime="5" maxplaytime="30">
      <rating value="N/A">
        <average value="7.9"/>
        <ranks>
          <rank type="subtype" id="1" name="boardgame"
                friendlyname="Board Game Rank" value="Not Ranked"/>
        </ranks>
      </rating>
    </stats>
    <status own="1" lastmodified="2024-01-01 00:00:00"/>
  </item>
</items>"""


@override_settings(BGG_USERNAME="kernicek", BGG_PASSWORD="test-password")
class GameAddViewTests(TestCase):
    """Add-game page (issue #55): a BGG id or pasted URL in, a game created
    through the existing sync path out — HTTP layer mocked like the per-game
    sync tests. Failures never leave a nameless orphan behind."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def post_add(self, value, *, collection=COLLECTION_XML,
                 geekitem=BASE_GEEKITEM_JSON, login_error=None):
        """POST the add form with bgg_sync.BggClient mocked; returns
        (response, client mock). The default geekitem payload carries no
        expandsboardgame links, so the type probe keeps BASE."""
        with mock.patch("gamekeeper.bgg_sync.BggClient") as client_class:
            client = client_class.return_value
            if login_error is not None:
                client.login.side_effect = login_error
            client.get_collection.return_value = collection
            client.get_geekitem.return_value = geekitem
            # Issue #65: create-from-BGG runs sync_game, which now also pulls
            # plays; no plays in these fixtures.
            client.get_plays.return_value = EMPTY_PLAYS_XML
            response = self.client.post("/games/add/", {"bgg": value})
        return response, client, client_class

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/games/add/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_get_renders_the_form(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/games/add/")
        self.assertContains(response, "BGG id or URL")

    def test_creates_game_from_bare_id(self):
        self.client.login(username="kernicek", password="pass")
        response, client, _ = self.post_add("207830")

        game = Game.objects.get()
        self.assertRedirects(response, f"/games/{game.pk}/")
        # Curated name seeded from the fetched canonical name; BGG fields
        # populated through the ordinary sync path.
        self.assertEqual(game.name, "5-Minute Dungeon")
        self.assertEqual(game.bgg_name, "5-Minute Dungeon")
        self.assertEqual(game.type, Game.Type.BASE)
        self.assertEqual(game.year_published, 2017)
        self.assertEqual(game.min_players, 2)
        self.assertEqual(game.bgg_collection_status, Game.BggCollectionStatus.OWN)
        self.assertIsNotNone(game.last_synced_at)
        link = game.primary_bgg_link
        self.assertEqual(link.bgg_id, 207830)
        # One type probe + the sync's id-filtered collection request.
        client.get_geekitem.assert_called_once_with(207830)
        client.get_collection.assert_called_once_with(
            "kernicek", status=None, bgg_id=207830,
        )

    def test_creates_game_from_pasted_bgg_url(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_add(
            "https://boardgamegeek.com/boardgame/207830/5-minute-dungeon",
        )
        game = Game.objects.get()
        self.assertRedirects(response, f"/games/{game.pk}/")
        self.assertEqual(game.primary_bgg_link.bgg_id, 207830)

    def test_unparseable_input_shows_error_and_creates_nothing(self):
        self.client.login(username="kernicek", password="pass")
        response, _, client_class = self.post_add("not a bgg reference")
        self.assertContains(response, "Enter a numeric BGG id")
        self.assertEqual(Game.objects.count(), 0)
        client_class.assert_not_called()  # rejected before any network

    def test_already_linked_id_redirects_to_the_existing_game(self):
        existing = Game.objects.create(name="Already Here")
        BggLink.objects.create(game=existing, bgg_id=207830, is_primary=True)
        self.client.login(username="kernicek", password="pass")

        response, _, client_class = self.post_add("207830")
        self.assertRedirects(response, f"/games/{existing.pk}/")
        self.assertEqual(Game.objects.count(), 1)
        client_class.assert_not_called()  # dedup short-circuits the network

    def test_alternate_link_dedups_too(self):
        # bgg_id is only unique per game — an ALTERNATE link on another game
        # must also block creation, or two games would claim the same id.
        existing = Game.objects.create(name="Has Alternate")
        BggLink.objects.create(game=existing, bgg_id=207830, is_primary=False)
        self.client.login(username="kernicek", password="pass")

        response, _, _ = self.post_add("207830")
        self.assertRedirects(response, f"/games/{existing.pk}/")
        self.assertEqual(Game.objects.count(), 1)

    def test_id_not_in_collection_creates_nothing(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_add("207830", collection=EMPTY_COLLECTION_XML)
        self.assertContains(response, "not in the kernicek BGG collection")
        self.assertEqual(Game.objects.count(), 0)  # orphan deleted
        self.assertEqual(BggLink.objects.count(), 0)

    def test_login_failure_creates_nothing(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_add(
            "207830", login_error=BggAuthError("BGG login failed: bad credentials"),
        )
        self.assertContains(response, "BGG login failed")
        self.assertEqual(Game.objects.count(), 0)

    @override_settings(BGG_USERNAME="", BGG_PASSWORD="")
    def test_missing_credentials_reports_config_error(self):
        self.client.login(username="kernicek", password="pass")
        response, _, client_class = self.post_add("207830")
        self.assertContains(response, "credentials are not configured")
        self.assertEqual(Game.objects.count(), 0)
        client_class.assert_not_called()

    def test_expansion_id_creates_typed_expansion_with_base_links(self):
        # The geekitems probe sees expandsboardgame links -> Type.EXPANSION,
        # and the sync then resolves the bases already in the app (#40).
        base = Game.objects.create(name="5 Minute Dungeon (localized)")
        BggLink.objects.create(game=base, bgg_id=207830, is_primary=True)
        self.client.login(username="kernicek", password="pass")

        response, client, _ = self.post_add(
            "228552", collection=EXPANSION_COLLECTION_XML, geekitem=GEEKITEM_JSON,
        )
        expansion = Game.objects.get(bgg_links__bgg_id=228552)
        self.assertRedirects(response, f"/games/{expansion.pk}/")
        self.assertEqual(expansion.type, Game.Type.EXPANSION)
        self.assertEqual(expansion.name, "5-Minute Dungeon: Curses! Foiled Again!")
        self.assertEqual(list(expansion.expands.all()), [base])
        # Probe + the sync's own expansion-links call.
        self.assertEqual(client.get_geekitem.call_count, 2)


@override_settings(BGG_USERNAME="kernicek", BGG_PASSWORD="test-password")
class BggImportViewTests(TestCase):
    """Bulk collection import (issue #81): status-filtered preview, then the
    confirmed rows run through the single-game create path with ONE login —
    HTTP layer mocked like the other BGG view tests."""

    # Per-id payloads answering the confirm loop's id-filtered sync fetches.
    ITEM_XML = {
        101: _import_collection_xml(
            _import_item_xml(101, "Alpha Owned Wished", 'own="1" wishlist="1"')),
        102: _import_collection_xml(
            _import_item_xml(102, "Bravo Rebought", 'prevowned="1" preordered="1"')),
        103: _import_collection_xml(
            _import_item_xml(103, "Charlie Preordered", 'preordered="1"')),
        104: _import_collection_xml(
            _import_item_xml(104, "Delta Wished", 'wishlist="1" wishlistpriority="2"')),
        105: _import_collection_xml(
            _import_item_xml(105, "Echo For Trade", 'fortrade="1"')),
    }

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def post_import(self, data, *, collections=None, login_error=None):
        """POST /games/import/ with bgg_sync.BggClient mocked; returns
        (response, client mock, client class mock). collections maps a
        status param to preview XML (unlisted statuses answer an empty
        collection); the confirm loop's id-filtered fetches answer from
        ITEM_XML. The default geekitem payload keeps every game BASE."""
        collections = collections or {}

        def get_collection(username, status="own", bgg_id=None):
            if bgg_id is not None:
                return self.ITEM_XML.get(bgg_id, EMPTY_COLLECTION_XML)
            return collections.get(status, EMPTY_COLLECTION_XML)

        with mock.patch("gamekeeper.bgg_sync.BggClient") as client_class:
            client = client_class.return_value
            if login_error is not None:
                client.login.side_effect = login_error
            client.get_collection.side_effect = get_collection
            client.get_geekitem.return_value = BASE_GEEKITEM_JSON
            # Issue #65: import runs sync_game per item, which now also pulls
            # plays; no plays in these fixtures.
            client.get_plays.return_value = EMPTY_PLAYS_XML
            response = self.client.post("/games/import/", data)
        return response, client, client_class

    def preview(self, statuses=("own",), bgg_username="kernicek", **kwargs):
        return self.post_import({
            "step": "preview", "bgg_username": bgg_username,
            "status": list(statuses),
        }, **kwargs)

    # --- form ---------------------------------------------------------------

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/games/import/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_get_renders_form_with_prefill_and_default_checks(self):
        membership = self.user.membership
        membership.bgg_username = "profileuser"
        membership.save(update_fields=["bgg_username"])
        self.client.login(username="kernicek", password="pass")

        response = self.client.get("/games/import/")
        self.assertContains(response, 'value="profileuser"')
        for param in ("own", "preordered", "prevowned"):
            self.assertContains(response, f'value="{param}" checked')
        for param in ("fortrade", "wishlist", "want", "wanttoplay", "wanttobuy"):
            self.assertNotContains(response, f'value="{param}" checked')

    def test_get_falls_back_to_settings_username_when_membership_blank(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/games/import/")
        self.assertContains(response, 'value="kernicek"')

    # --- preview ------------------------------------------------------------

    def test_preview_saves_bgg_username_on_membership(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.preview(bgg_username="newname")
        # The empty-collection notice proves the flow got past the save.
        self.assertContains(response, "Nothing new to import")
        self.assertEqual(
            Membership.objects.get(user=self.user).bgg_username, "newname",
        )

    def test_preview_requests_one_fetch_per_status_and_logs_in_once(self):
        self.client.login(username="kernicek", password="pass")
        _, client, _ = self.preview(
            statuses=("own", "prevowned"),
            collections={"own": IMPORT_STATUS_XML},
        )
        client.login.assert_called_once_with()
        self.assertEqual(client.get_collection.call_count, 2)
        client.get_collection.assert_any_call("kernicek", status="own")
        client.get_collection.assert_any_call("kernicek", status="prevowned")

    def test_preview_resolves_actions_and_emits_hidden_inputs(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.preview(collections={"own": IMPORT_STATUS_XML})

        # Multi-flag precedence: own+wishlist and fortrade-only -> copy;
        # prevowned+preordered -> archived copy; want-only -> wishlist.
        self.assertContains(response, 'name="action_101" value="copy"')
        self.assertContains(response, 'name="action_102" value="archived_copy"')
        self.assertContains(response, 'name="action_103" value="preorder"')
        self.assertContains(response, 'name="action_104" value="wishlist"')
        self.assertContains(response, 'name="action_105" value="copy"')
        self.assertContains(response, 'name="action_106" value="wishlist"')
        self.assertContains(response, "priority 2")

    def test_preview_renders_select_all_and_per_action_group_toggles(self):
        # Issue #88: a master checkbox plus one group-toggle per action
        # actually present in this preview (copy/archived/preorder/wishlist).
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.preview(collections={"own": IMPORT_STATUS_XML})

        self.assertContains(response, 'id="select-all-master"')
        self.assertContains(response, 'data-group="copy"')
        self.assertContains(response, 'data-group="archived_copy"')
        self.assertContains(response, 'data-group="preorder"')
        self.assertContains(response, 'data-group="wishlist"')
        self.assertContains(response, "Game + copy")
        self.assertContains(response, "Game only (wishlist)")

    def test_preview_renders_wishlist_priority_filter_and_data_attribute(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.preview(collections={"own": IMPORT_STATUS_XML})

        self.assertContains(response, 'id="priority-filter"')
        self.assertContains(response, 'data-priority="2"')
        # 106 ("want" only) has no wishlistpriority — no data-priority for it.
        self.assertNotContains(response, 'value="106" data-group="wishlist" data-priority')

    def test_preview_hides_priority_filter_when_no_wishlist_candidates(self):
        self.client.login(username="kernicek", password="pass")
        no_wishlist_xml = _import_collection_xml(
            _import_item_xml(201, "Golf Owned", 'own="1"'),
        )
        response, _, _ = self.preview(collections={"own": no_wishlist_xml})

        self.assertNotContains(response, 'id="priority-filter"')
        self.assertNotContains(response, 'data-group="wishlist"')

    def test_preview_skips_games_already_linked_by_any_bgg_link(self):
        # An ALTERNATE link blocks re-import just like a primary one.
        existing = Game.objects.create(name="Already Here")
        BggLink.objects.create(game=existing, bgg_id=101, is_primary=False)
        self.client.login(username="kernicek", password="pass")

        response, _, _ = self.preview(collections={"own": IMPORT_STATUS_XML})
        self.assertNotContains(response, 'name="action_101"')
        self.assertContains(response, "Already in the app")
        self.assertContains(response, "Already Here")
        self.assertContains(response, "5 games to import")

    def test_preview_unknown_status_returns_400(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.preview(statuses=("own", "bogus"))
        self.assertEqual(response.status_code, 400)

    def test_preview_blank_username_shows_error(self):
        self.client.login(username="kernicek", password="pass")
        response, _, client_class = self.preview(bgg_username="")
        self.assertContains(response, "Enter your BGG username")
        client_class.assert_not_called()

    def test_preview_no_status_shows_error(self):
        self.client.login(username="kernicek", password="pass")
        response, _, client_class = self.preview(statuses=())
        self.assertContains(response, "Pick at least one collection status")
        client_class.assert_not_called()

    def test_preview_invalid_bgg_username_shows_bgg_message(self):
        # BGG answers 200 with an <errors> payload for a bad username — the
        # form must show the message, not an empty preview.
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.preview(collections={"own": BGG_ERRORS_XML})
        self.assertContains(response, "Invalid username specified")

    def test_preview_empty_collection_shows_notice(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.preview()
        self.assertContains(response, "Nothing new to import")

    def test_preview_bgg_failure_shows_error_and_writes_nothing(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.preview(
            login_error=BggAuthError("BGG login failed: bad credentials"),
        )
        self.assertContains(response, "BGG login failed")
        self.assertEqual(Game.objects.count(), 0)

    # --- confirm ------------------------------------------------------------

    def test_confirm_creates_game_edition_and_active_copy(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_import({
            "step": "import", "include": ["101"], "action_101": "copy",
        })

        game = Game.objects.get()
        self.assertEqual(game.name, "Alpha Owned Wished")
        self.assertEqual(game.bgg_collection_status, Game.BggCollectionStatus.OWN)
        self.assertEqual(game.primary_bgg_link.bgg_id, 101)
        edition = Edition.objects.get(game=game)
        self.assertTrue(edition.is_default)
        copy = Copy.objects.get(edition=edition)
        self.assertEqual(copy.owner, self.user)
        self.assertEqual(copy.archive_status, Copy.ArchiveStatus.ACTIVE)
        self.assertContains(response, "Created (1)")

    def test_confirm_fortrade_row_creates_copy_marked_will_leave(self):
        # Issue #82: the fortrade_<id> hidden input carries the flag through
        # to the confirm step, seeding keep_status on the created copy.
        self.client.login(username="kernicek", password="pass")
        self.post_import({
            "step": "import", "include": ["105"],
            "action_105": "copy", "fortrade_105": "1",
        })

        copy = Copy.objects.get()
        self.assertEqual(copy.keep_status, Copy.KeepStatus.WILL_LEAVE)

    def test_confirm_non_fortrade_row_leaves_keep_status_blank(self):
        self.client.login(username="kernicek", password="pass")
        self.post_import({
            "step": "import", "include": ["101"], "action_101": "copy",
        })

        copy = Copy.objects.get()
        self.assertEqual(copy.keep_status, "")

    def test_confirm_prevowned_creates_archived_copy(self):
        self.client.login(username="kernicek", password="pass")
        self.post_import({
            "step": "import", "include": ["102"], "action_102": "archived_copy",
        })

        game = Game.objects.get()
        copy = Copy.objects.get()
        self.assertEqual(copy.archive_status, Copy.ArchiveStatus.ARCHIVED)
        # Stored status keeps its own precedence (preordered > prevowned) —
        # the archived copy carries the previously-owned history instead.
        self.assertEqual(
            game.bgg_collection_status, Game.BggCollectionStatus.PREORDERED,
        )

    def test_confirm_preordered_creates_game_only(self):
        self.client.login(username="kernicek", password="pass")
        self.post_import({
            "step": "import", "include": ["103"], "action_103": "preorder",
        })

        game = Game.objects.get()
        self.assertEqual(
            game.bgg_collection_status, Game.BggCollectionStatus.PREORDERED,
        )
        self.assertEqual(Copy.objects.count(), 0)

    def test_confirm_wishlist_creates_game_only_with_priority(self):
        self.client.login(username="kernicek", password="pass")
        self.post_import({
            "step": "import", "include": ["104"], "action_104": "wishlist",
        })

        game = Game.objects.get()
        self.assertEqual(
            game.bgg_collection_status, Game.BggCollectionStatus.WISHLIST,
        )
        self.assertEqual(game.bgg_wishlist_priority, 2)
        self.assertEqual(Copy.objects.count(), 0)

    def test_confirm_logs_in_once_for_many_items(self):
        self.client.login(username="kernicek", password="pass")
        _, client, _ = self.post_import({
            "step": "import", "include": ["101", "103"],
            "action_101": "copy", "action_103": "preorder",
        })
        client.login.assert_called_once_with()
        self.assertEqual(Game.objects.count(), 2)

    def test_confirm_unchecked_rows_are_not_imported(self):
        self.client.login(username="kernicek", password="pass")
        self.post_import({
            "step": "import", "include": ["101"],
            "action_101": "copy", "action_103": "preorder",
        })
        self.assertEqual(Game.objects.count(), 1)
        self.assertFalse(BggLink.objects.filter(bgg_id=103).exists())

    def test_confirm_invalid_action_returns_400(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_import({
            "step": "import", "include": ["101"], "action_101": "bogus",
        })
        self.assertEqual(response.status_code, 400)
        # A missing action input is just as invalid.
        response, _, _ = self.post_import({"step": "import", "include": ["101"]})
        self.assertEqual(response.status_code, 400)
        self.assertEqual(Game.objects.count(), 0)

    def test_confirm_bad_id_returns_400(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_import({
            "step": "import", "include": ["abc"], "action_abc": "copy",
        })
        self.assertEqual(response.status_code, 400)

    def test_confirm_empty_selection_shows_error(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_import({"step": "import"})
        self.assertContains(response, "Nothing was selected to import")

    def test_confirm_partial_failure_reports_and_continues(self):
        # 999 is in no payload — its create fails (not in collection) and is
        # cleaned up; 101 before it still lands. Re-running would retry 999.
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_import({
            "step": "import", "include": ["101", "999"],
            "action_101": "copy", "action_999": "copy",
        })
        self.assertContains(response, "Created (1)")
        self.assertContains(response, "Failed (1)")
        self.assertEqual(Game.objects.count(), 1)  # no orphan for 999
        self.assertFalse(BggLink.objects.filter(bgg_id=999).exists())

    def test_confirm_already_linked_id_reports_skipped(self):
        # Double-submit safety: the preview excluded linked ids, but the
        # create path's own guard still catches a re-posted one.
        existing = Game.objects.create(name="Already Here")
        BggLink.objects.create(game=existing, bgg_id=101, is_primary=True)
        self.client.login(username="kernicek", password="pass")

        response, _, _ = self.post_import({
            "step": "import", "include": ["101"], "action_101": "copy",
        })
        self.assertContains(response, "Skipped")
        self.assertContains(response, "Already Here")
        self.assertEqual(Game.objects.count(), 1)

    def test_unknown_step_returns_400(self):
        self.client.login(username="kernicek", password="pass")
        response, _, _ = self.post_import({"step": "bogus"})
        self.assertEqual(response.status_code, 400)


# ===========================================================================
# §10  Taxonomy import
# ===========================================================================

# Theme header names live in row 1 of the theme/adapts columns; the importer
# reads them from there, so the builder plants a small subset.
TAXONOMY_THEME_HEADERS = {169: "Abstract", 180: "Fantasy", 206: "Adapts: \nBook"}


def build_taxonomy_workbook(rows, apps_rows=()):
    """An Overview sheet (data from row 4, theme headers in row 1) plus an
    APPs sheet (data from row 2), the shape import_taxonomy expects."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Overview"
    sheet.cell(row=1, column=2, value="Game")
    for col, name in TAXONOMY_THEME_HEADERS.items():
        sheet.cell(row=1, column=col, value=name)
    for offset, row in enumerate(rows):
        for col, value in row.items():
            cell = sheet.cell(row=4 + offset, column=col)
            if isinstance(value, tuple):
                cell.value, cell.hyperlink = value
            else:
                cell.value = value
    apps = workbook.create_sheet("APPs")
    apps.cell(row=1, column=2, value="Game")
    for offset, row in enumerate(apps_rows):
        for col, value in row.items():
            cell = apps.cell(row=2 + offset, column=col)
            if isinstance(value, tuple):
                cell.value, cell.hyperlink = value
            else:
                cell.value = value
    path = Path(tempfile.mkdtemp()) / "mastersheet.xlsx"
    workbook.save(path)
    return path


DUNGEON_URL = "https://boardgamegeek.com/boardgame/207830/5-minute-dungeon"

TAXONOMY_ROW = {
    2: "5 Minute Dungeon",
    3: ("BGG", DUNGEON_URL),
    17: "EN",
    18: "Easy",
    19: "doable - coop",
    33: "y",      # Competitive
    36: "opt",    # Solo (optional mode)
    43: 2,        # player conflict
    45: "opt",    # companion app optional
    46: "y",      # app version exists
    48: "y",      # soundtrack (timer)
    78: "y",      # Scenarios/missions
    79: "y",      # One-off
    169: "f",     # Abstract — favourite theme
    180: "y",     # Fantasy
    206: "y",     # Adapts: Book
}

APPS_ROW = {
    2: "5 Minute Dungeon",
    3: ("BGG", DUNGEON_URL),
    6: "y",  # Android
    7: "y",  # Steam
}


class ImportTaxonomyTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)
        Copy.objects.create(owner=cls.user, edition=cls.edition)

    def run_import(self, path, **extra):
        out = StringIO()
        call_command("import_taxonomy", str(path), user="kernicek", stdout=out, **extra)
        return out.getvalue()

    def test_imports_curated_fields_tags_and_apps(self):
        self.run_import(build_taxonomy_workbook([TAXONOMY_ROW], [APPS_ROW]))

        game = Game.objects.get()
        self.assertEqual(game.language_dependency, Game.LanguageDependency.EASY)
        self.assertEqual(game.language_dependency_note, "doable - coop")
        self.assertEqual(game.player_conflict, 2)
        self.assertEqual(game.player_conflict_note, "")
        self.assertEqual(game.companion_app, Game.AppUse.OPTIONAL)
        self.assertTrue(game.has_app_version)
        self.assertFalse(game.soundtrack_ambience)
        self.assertTrue(game.soundtrack_timer)
        self.assertFalse(game.is_campaign)
        self.assertFalse(game.is_legacy)
        self.assertTrue(game.has_scenarios)
        self.assertTrue(game.is_one_off)

        types = {pt.game_type: pt.qualifier for pt in game.game_types.all()}
        self.assertEqual(types, {
            GameType.Type.COMPETITIVE: "",
            GameType.Type.SOLO: GameType.Qualifier.OPTIONAL,
        })

        tags = {gt.tag.name: gt.is_favourite for gt in game.game_tags.select_related("tag")}
        self.assertEqual(tags, {"Abstract": True, "Fantasy": False, "Adapts: Book": False})
        # Mechanics stay a BGG-sync seam — nothing lands there from the sheet.
        self.assertFalse(Tag.objects.filter(kind=Tag.Kind.MECHANIC).exists())

        self.edition.refresh_from_db()
        self.assertEqual(
            self.edition.components_language, Edition.ComponentsLanguage.ENGLISH,
        )

        platforms = {di.platform for di in game.digital_implementations.all()}
        self.assertEqual(platforms, {
            DigitalImplementation.Platform.ANDROID,
            DigitalImplementation.Platform.STEAM,
        })

    def test_messy_values_kept_in_notes(self):
        row = dict(TAXONOMY_ROW)
        row.update({17: "?", 18: "?", 19: "story only", 43: "0-1?"})
        output = self.run_import(build_taxonomy_workbook([row]))

        game = Game.objects.get()
        self.assertEqual(game.language_dependency, "")
        self.assertEqual(game.language_dependency_note, "story only")
        self.assertIsNone(game.player_conflict)
        self.assertEqual(game.player_conflict_note, "0-1?")
        self.edition.refresh_from_db()
        self.assertEqual(self.edition.components_language, "")
        self.assertIn("story only", output)  # reported as ambiguous

    def test_rows_without_matching_game_are_skipped(self):
        row = {
            2: "Not Owned",
            3: ("BGG", "https://boardgamegeek.com/boardgame/999999/not-owned"),
            180: "y",
        }
        apps_row = {
            2: "Not Owned",
            3: ("BGG", "https://boardgamegeek.com/boardgame/999999/not-owned"),
            7: "y",
        }
        output = self.run_import(build_taxonomy_workbook([row], [apps_row]))

        self.assertIn("not in database", output)
        self.assertFalse(GameTag.objects.exists())
        self.assertFalse(DigitalImplementation.objects.exists())

    def test_reimport_is_idempotent(self):
        path = build_taxonomy_workbook([TAXONOMY_ROW], [APPS_ROW])
        self.run_import(path)
        state = (
            Tag.objects.count(), GameTag.objects.count(),
            GameType.objects.count(), DigitalImplementation.objects.count(),
        )

        self.run_import(path)

        self.assertEqual(state, (
            Tag.objects.count(), GameTag.objects.count(),
            GameType.objects.count(), DigitalImplementation.objects.count(),
        ))

    def test_dry_run_writes_nothing(self):
        output = self.run_import(
            build_taxonomy_workbook([TAXONOMY_ROW], [APPS_ROW]), dry_run=True,
        )

        self.assertIn("DRY RUN", output)
        self.assertFalse(Tag.objects.exists())
        self.assertFalse(GameType.objects.exists())
        game = Game.objects.get()
        self.assertEqual(game.language_dependency, "")


class GameTypeSoloDecisionTests(TestCase):
    """Guards the issue #124 decision: 'Solo' is kept as a game-type value even
    though it reads like a player-*count* fact (min_players == 1). It stays
    because it records what the player-count range cannot — a designed solo mode
    for a 2+ player game, and the optional/app qualifier. If someone ever wants
    to drop SOLO, these tests should fail loudly so it's a deliberate act, not a
    silent regression."""

    def test_solo_is_still_a_valid_game_type_choice(self):
        # Removing SOLO from the enum (the tempting "derive it from player count"
        # refactor) breaks this — see issue #124 for why that loses data.
        self.assertIn(GameType.Type.SOLO, GameType.Type.values)
        self.assertEqual(GameType.Type.SOLO, "solo")

    def test_solo_with_qualifier_on_a_multiplayer_game_round_trips(self):
        # The case min_players can't express: a 2+ player game with a curated,
        # optional solo mode. min_players == 1 would report this game as NOT
        # solo-capable, so the mark carries non-derivable information.
        game = Game.objects.create(name="Pandemic-like", min_players=2, max_players=4)
        mark = GameType.objects.create(
            game=game,
            game_type=GameType.Type.SOLO,
            qualifier=GameType.Qualifier.OPTIONAL,
        )

        self.assertGreater(game.min_players, 1)  # not derivable from player count
        self.assertEqual(mark.get_game_type_display(), "Solo")
        self.assertEqual(mark.get_qualifier_display(), "Optional")
        self.assertEqual(str(mark), "Pandemic-like: Solo (opt)")


class EffectivePlayerRangeTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def _own(self, game):
        """Give the game an active Copy so it counts as owned."""
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(owner=self.user, edition=edition)

    def test_owned_expansion_overrides_widen_the_base_range(self):
        base = Game.objects.create(name="Base", min_players=2, max_players=4)
        expansion = Game.objects.create(
            name="5-6 Player Expansion", type=Game.Type.EXPANSION,
            players_max_override=6,
        )
        expansion.expands.add(base)
        self._own(expansion)

        self.assertEqual(base.effective_player_range(), (2, 6))
        # The expansion itself falls back to the base-stat fields it lacks.
        self.assertEqual(expansion.effective_player_range(), (None, 6))

    def test_unowned_expansion_does_not_widen_the_base_range(self):
        # An expansion in the DB but owned by nobody (preorder / BGG-synced)
        # must not touch the base's effective range (DESIGN §4).
        base = Game.objects.create(name="Base", min_players=2, max_players=4)
        expansion = Game.objects.create(
            name="5-6 Player Expansion", type=Game.Type.EXPANSION,
            players_max_override=6,
        )
        expansion.expands.add(base)

        self.assertEqual(base.effective_player_range(), (2, 4))

    def test_archived_expansion_copy_does_not_widen_the_range(self):
        # A sold/culled expansion is no longer owned, so it stops widening.
        base = Game.objects.create(name="Base", min_players=2, max_players=4)
        expansion = Game.objects.create(
            name="5-6 Player Expansion", type=Game.Type.EXPANSION,
            players_max_override=6,
        )
        expansion.expands.add(base)
        copy = self._own(expansion)
        copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        copy.save(update_fields=["archive_status"])

        self.assertEqual(base.effective_player_range(), (2, 4))

    def test_unsynced_game_has_no_range(self):
        game = Game.objects.create(name="Preorder")
        self.assertEqual(game.effective_player_range(), (None, None))


class EffectiveGameTypesTests(TestCase):
    """Issue #134: owned-expansion game-types roll up onto the base game, the
    taxonomy parallel to EffectivePlayerRangeTests."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def _own(self, game):
        """Give the game an active Copy so it counts as owned."""
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(owner=self.user, edition=edition)

    def _expansion_of(self, base, name="Solo Expansion"):
        expansion = Game.objects.create(name=name, type=Game.Type.EXPANSION)
        expansion.expands.add(base)
        return expansion

    def _types(self, game):
        """{game_type: from_expansion} for concise assertions."""
        return {egt["game_type"]: egt["from_expansion"]
                for egt in game.effective_game_types()}

    def test_owned_expansion_adds_a_game_type_absent_on_the_base(self):
        base = Game.objects.create(name="Base")
        expansion = self._expansion_of(base)
        GameType.objects.create(game=expansion, game_type=GameType.Type.SOLO)
        self._own(expansion)

        effective = base.effective_game_types()
        self.assertEqual(len(effective), 1)
        chip = effective[0]
        self.assertEqual(chip["game_type"], GameType.Type.SOLO)
        self.assertEqual(chip["qualifier"], "")
        self.assertTrue(chip["from_expansion"])

    def test_unowned_expansion_game_types_do_not_roll_up(self):
        base = Game.objects.create(name="Base")
        expansion = self._expansion_of(base)
        GameType.objects.create(game=expansion, game_type=GameType.Type.SOLO)
        # No Copy → not owned; its marks must not surface (DESIGN §4 owned-only).

        self.assertEqual(base.effective_game_types(), [])

    def test_archived_expansion_copy_stops_the_roll_up(self):
        base = Game.objects.create(name="Base")
        expansion = self._expansion_of(base)
        GameType.objects.create(game=expansion, game_type=GameType.Type.SOLO)
        copy = self._own(expansion)
        copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        copy.save(update_fields=["archive_status"])

        self.assertEqual(base.effective_game_types(), [])

    def test_least_restrictive_qualifier_wins(self):
        # Base offers Solo only with the app; an owned expansion offers a
        # native solo mode — the merged chip shows the least-restrictive mark.
        base = Game.objects.create(name="Base")
        GameType.objects.create(
            game=base, game_type=GameType.Type.SOLO,
            qualifier=GameType.Qualifier.APP,
        )
        expansion = self._expansion_of(base)
        GameType.objects.create(game=expansion, game_type=GameType.Type.SOLO)
        self._own(expansion)

        effective = base.effective_game_types()
        self.assertEqual(len(effective), 1)
        chip = effective[0]
        self.assertEqual(chip["qualifier"], "")
        # Base carries the type natively, so it is not a rolled-up-only chip.
        self.assertFalse(chip["from_expansion"])

    def test_base_native_mark_is_not_flagged_from_expansion(self):
        base = Game.objects.create(name="Base")
        GameType.objects.create(
            game=base, game_type=GameType.Type.COOPERATIVE,
        )
        self.assertEqual(self._types(base), {GameType.Type.COOPERATIVE: False})

    def test_filter_parity_owned_expansion_surfaces_base_under_its_type(self):
        # GameChooser 'Solo' filter must match a base whose only solo comes
        # from an owned expansion — parity with the effective player range.
        base = Game.objects.create(name="Base Game", min_players=2, max_players=4)
        self._own(base)
        expansion = self._expansion_of(base)
        GameType.objects.create(game=expansion, game_type=GameType.Type.SOLO)
        self._own(expansion)
        # A base with no solo anywhere must stay filtered out.
        other = Game.objects.create(name="No Solo Game")
        self._own(other)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/", {"game_type": "solo", "show_unavailable": "1"})

        self.assertContains(response, "Base Game")
        self.assertNotContains(response, "No Solo Game")

    def test_detail_page_renders_rolled_up_chip_with_provenance_marker(self):
        base = Game.objects.create(name="Base Game")
        self._own(base)
        expansion = self._expansion_of(base)
        GameType.objects.create(game=expansion, game_type=GameType.Type.SOLO)
        self._own(expansion)

        self.client.login(username="kernicek", password="pass")
        response = self.client.get(reverse("game_detail", args=[base.pk]))

        self.assertContains(response, "Solo")
        self.assertContains(response, "From an owned expansion")
        self.assertContains(response, "text-bg-secondary")


# ===========================================================================
# §13  Collection cover grid
# ===========================================================================

class CollectionViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.synced = Game.objects.create(
            name="5 Minute Dungeon",
            image_url="https://cf.geekdo-images.com/large/5md.jpg",
            thumbnail_url="https://cf.geekdo-images.com/thumb/5md.jpg",
        )
        cls.unsynced = Game.objects.create(name="Imageless Prototype")

    def test_grid_shows_covers_and_text_fallback(self):
        self.client.login(username="kernicek", password="pass")
        # These fixtures carry no copies; show_unavailable opts past the
        # available-only default (issue #107) so cover rendering is under test.
        response = self.client.get("/", {"show_unavailable": "1"})

        # Synced game renders the full image (sharp), not the ~200px thumb.
        self.assertContains(response, "https://cf.geekdo-images.com/large/5md.jpg")
        self.assertContains(response, "5 Minute Dungeon")
        # Imageless game keeps the text tile.
        self.assertContains(response, "Imageless Prototype")
        self.assertContains(response, "2 games")

    def test_local_cover_beats_remote_urls(self):
        self.synced.cover_image.name = "covers/207830.jpg"
        self.synced.save(update_fields=["cover_image"])

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/", {"show_unavailable": "1"})

        self.assertContains(response, "/media/covers/207830.jpg")
        self.assertNotContains(response, "https://cf.geekdo-images.com/large/5md.jpg")

    def test_zoomed_out_cover_gets_the_fit_treatment(self):
        # §13 zoom-out: below 100 % the tile shows the whole art over the
        # game's letterbox colour; regular tiles get neither.
        self.synced.cover_zoom = 60
        self.synced.cover_fit_color = "#1a2b3c"
        self.synced.save(update_fields=["cover_zoom", "cover_fit_color"])

        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/", {"show_unavailable": "1"})

        self.assertContains(response, "cover-art-fit")
        self.assertContains(response, "transform: scale(0.6)")
        self.assertContains(
            response, 'cover-art-backdrop" style="background-color: #1a2b3c"')


# ===========================================================================
# §10  GameChooser filters
# ===========================================================================

class GameSortNameTests(TestCase):
    """Issue #6: the derived article-blind, lowercased sort key."""

    def test_save_maintains_sort_name(self):
        game = Game.objects.create(name="The Crew")
        self.assertEqual(game.sort_name, "crew")
        game.name = "An Age Contrived"
        game.save()
        self.assertEqual(game.sort_name, "age contrived")

    def test_update_fields_rename_carries_sort_name_along(self):
        # The sync_bgg save shape: save(update_fields=[...]).
        game = Game.objects.create(name="The Crew")
        game.name = "Arnak"
        game.save(update_fields=["name"])
        game.refresh_from_db()
        self.assertEqual(game.sort_name, "arnak")

    def test_only_whole_leading_articles_are_stripped(self):
        self.assertEqual(Game.compute_sort_name("Antique Bazaar"),
                         "antique bazaar")
        self.assertEqual(Game.compute_sort_name("Theseus"), "theseus")
        # A name that IS an article keeps itself rather than sorting as "".
        self.assertEqual(Game.compute_sort_name("The"), "the")


class GameShortNameTests(TestCase):
    """Issue #98: the derived expansion tail, base-prefix stripped."""

    def _expansion(self, name, base):
        exp = Game.objects.create(name=name, type=Game.Type.EXPANSION)
        exp.expands.add(base)
        return exp

    def test_base_game_returns_full_name(self):
        base = Game.objects.create(name="Spirit Island")
        self.assertEqual(base.short_name, "Spirit Island")

    def test_colon_prefix_is_stripped(self):
        base = Game.objects.create(name="Spirit Island")
        exp = self._expansion("Spirit Island: Jagged Earth", base)
        self.assertEqual(exp.short_name, "Jagged Earth")

    def test_dash_prefix_is_stripped_when_base_has_a_colon(self):
        # BGG switches ": " for " - " when the base title itself has a colon.
        base = Game.objects.create(name="Sleeping Gods")
        exp = self._expansion("Sleeping Gods - Distant Skies", base)
        self.assertEqual(exp.short_name, "Distant Skies")

    def test_en_dash_prefix_is_stripped(self):
        base = Game.objects.create(name="Gloomhaven")
        exp = self._expansion("Gloomhaven – Forgotten Circles", base)
        self.assertEqual(exp.short_name, "Forgotten Circles")

    def test_name_not_starting_with_base_falls_back_to_full(self):
        base = Game.objects.create(name="Spirit Island")
        exp = self._expansion("Jagged Earth Standalone", base)
        self.assertEqual(exp.short_name, "Jagged Earth Standalone")

    def test_expansion_without_base_link_falls_back_to_full(self):
        exp = Game.objects.create(
            name="Orphan Expansion", type=Game.Type.EXPANSION,
        )
        self.assertEqual(exp.short_name, "Orphan Expansion")

    def test_multiple_bases_strips_against_the_matching_one(self):
        wrong = Game.objects.create(name="Wingspan")
        right = Game.objects.create(name="Spirit Island")
        exp = Game.objects.create(
            name="Spirit Island: Jagged Earth", type=Game.Type.EXPANSION,
        )
        exp.expands.add(wrong, right)
        self.assertEqual(exp.short_name, "Jagged Earth")


class GameChooserViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        group = cls.user.membership.group  # auto-created on signup (§3)
        cls.shelf = Location.objects.create(group=group, name="Chynice")
        cls.lent = Location.objects.create(group=group, name="At Pavel's")
        cls.fantasy = Tag.objects.create(kind=Tag.Kind.THEME, name="Fantasy")
        cls.horror = Tag.objects.create(kind=Tag.Kind.THEME, name="Horror")

        cls.duel = cls._game(
            "Duel", min_players=2, max_players=2, min_playtime=30,
            game_type=GameType.Type.COMPETITIVE,
            tag=cls.horror, location=cls.shelf,
        )
        cls.epic = cls._game(
            "Epic Quest", min_players=1, max_players=4, min_playtime=90,
            game_type=GameType.Type.COOPERATIVE,
            tag=cls.fantasy, location=cls.lent,
        )
        # Issue #43: lending is tracked via Loan, not Location — "lent out"
        # is what makes Epic Quest's copy unavailable below.
        Loan.objects.create(
            copy=cls.epic.editions.get().copies.get(),
            direction=Loan.Direction.LENT_OUT, counterparty_name="Pavel",
        )
        # A 5-6 player expansion, owned (active Copy), widens Epic Quest's
        # effective range.
        expansion = Game.objects.create(
            name="Epic Quest: More Heroes", type=Game.Type.EXPANSION,
            players_min_override=5, players_max_override=6,
        )
        expansion.expands.add(cls.epic)
        exp_edition = Edition.objects.create(game=expansion, is_default=True)
        Copy.objects.create(owner=cls.user, edition=exp_edition)
        # No stats at all — the 55-unsynced-preorders case.
        cls.unsynced = Game.objects.create(name="Preorder Mystery")

    @classmethod
    def _game(cls, name, game_type, tag, location, **stats):
        game = Game.objects.create(name=name, **stats)
        GameType.objects.create(game=game, game_type=game_type)
        GameTag.objects.create(game=game, tag=tag)
        edition = Edition.objects.create(game=game, is_default=True)
        Copy.objects.create(owner=cls.user, edition=edition, location=location)
        return game

    def get(self, params=None, **extra):
        self.client.login(username="kernicek", password="pass")
        # The collection is the home page since issue #7.
        return self.client.get("/", params or {}, **extra)

    def names(self, response):
        return [game.name for game in response.context["games"]]

    def test_no_filters_shows_base_games_only(self):
        # show_unavailable opts out of the available-only default (issue #107)
        # so lent-out Epic Quest and copy-less Preorder Mystery appear too.
        response = self.get({"show_unavailable": "1"})
        self.assertEqual(
            self.names(response), ["Duel", "Epic Quest", "Preorder Mystery"],
        )
        self.assertContains(response, "3 of 3 games")

    def test_expansions_toggle_adds_expansions(self):
        response = self.get({"expansions": "1", "show_unavailable": "1"})
        self.assertEqual(len(self.names(response)), 4)
        self.assertIn("Epic Quest: More Heroes", self.names(response))

    def test_expansion_tiles_show_a_puzzle_badge(self):
        # Issue #46: expansion tiles carry a puzzle badge; base-only grids
        # (the default) never render one.
        self.assertNotContains(self.get(), "cover-flag-expansion")
        self.assertContains(self.get({"expansions": "1"}), "cover-flag-expansion")

    def test_players_filter_uses_effective_range(self):
        # 5 players: Epic Quest qualifies only through its expansion override;
        # stat-less games cannot promise a seat count and drop out.
        response = self.get({"players": "5", "show_unavailable": "1"})
        self.assertEqual(self.names(response), ["Epic Quest"])

        response = self.get({"players": "5", "expansions": "1", "show_unavailable": "1"})
        self.assertEqual(self.names(response), ["Epic Quest", "Epic Quest: More Heroes"])

        response = self.get({"players": "2", "show_unavailable": "1"})
        self.assertEqual(self.names(response), ["Duel", "Epic Quest"])

    def test_unowned_expansion_does_not_widen_the_filter(self):
        # A 6-player expansion of Duel (2-2) that nobody owns must not let
        # Duel answer a 6-player search (DESIGN §4 owned-only effective range).
        expansion = Game.objects.create(
            name="Duel: Pantheon", type=Game.Type.EXPANSION,
            players_max_override=6,
        )
        expansion.expands.add(self.duel)

        response = self.get({"players": "6"})
        self.assertNotIn("Duel", self.names(response))

    def test_playtime_filter_keeps_games_that_could_fit(self):
        response = self.get({"playtime": "45"})
        self.assertEqual(self.names(response), ["Duel"])

    def test_game_type_filter_matches_any_selected(self):
        response = self.get({"game_type": ["cooperative", "solo"],
                             "show_unavailable": "1"})
        self.assertEqual(self.names(response), ["Epic Quest"])

    def test_theme_filter(self):
        response = self.get({"theme": str(self.fantasy.pk),
                             "show_unavailable": "1"})
        self.assertEqual(self.names(response), ["Epic Quest"])

    def test_location_filter(self):
        response = self.get({"location": str(self.shelf.pk)})
        self.assertEqual(self.names(response), ["Duel"])

    def test_location_none_matches_unplaced_copies_only(self):
        # Issue #30: "none" finds copies with no location, so they can be
        # assigned one. Copy-less games (preorders) stay out — there is
        # nothing to place.
        self._game(
            "Homeless", game_type=GameType.Type.COMPETITIVE,
            tag=self.horror, location=None,
        )
        response = self.get({"location": "none"})
        self.assertEqual(self.names(response), ["Homeless"])
        self.assertContains(response, "(no location set)")

    def test_leaving_filter(self):
        # Issue #82: opt-in axis, only games with an active WILL_LEAVE copy.
        copy = Copy.objects.get(edition__game=self.duel)
        copy.keep_status = Copy.KeepStatus.WILL_LEAVE
        copy.save(update_fields=["keep_status"])
        response = self.get({"leaving": "1", "show_unavailable": "1"})
        self.assertEqual(self.names(response), ["Duel"])

    def test_leaving_copy_shows_a_grid_badge(self):
        self.assertNotContains(
            self.get({"show_unavailable": "1"}), "cover-flag-leaving",
        )
        copy = Copy.objects.get(edition__game=self.duel)
        copy.keep_status = Copy.KeepStatus.WILL_LEAVE
        copy.save(update_fields=["keep_status"])
        self.assertContains(
            self.get({"show_unavailable": "1"}), "cover-flag-leaving",
        )

    def test_search_box_filters_by_name_or_bgg_name(self):
        # Issue #20: case-blind substring on the user-facing name AND the
        # BGG canonical one (localized titles differ).
        Game.objects.create(name="6 bere!", bgg_name="6 nimmt!")
        # show_unavailable: Epic Quest is lent, "6 bere!" has no copy.
        self.assertEqual(
            self.names(self.get({"q": "epic", "show_unavailable": "1"})),
            ["Epic Quest"])
        self.assertEqual(
            self.names(self.get({"q": "NIMMT", "show_unavailable": "1"})),
            ["6 bere!"])
        self.assertEqual(self.names(self.get({"q": "zzz"})), [])

    def test_search_box_matches_alternate_names(self):
        # Issue #51: a curated localized title is searchable even when neither
        # the user-facing name nor bgg_name contains the query.
        beasty = Game.objects.create(name="Beasty Bar")
        AlternateName.objects.create(game=beasty, name="Safari Bar")
        self.assertEqual(
            self.names(self.get({"q": "Safari", "show_unavailable": "1"})),
            ["Beasty Bar"])

    def test_search_folds_diacritics(self):
        # Issue #126: name search folds Latin diacritics on both sides, so the
        # ASCII query 'selmy' finds the Czech title 'Šelmy' — and the accented
        # query still matches too (the change is purely additive).
        Game.objects.create(name="Šelmy")
        self.assertEqual(
            self.names(self.get({"q": "selmy", "show_unavailable": "1"})),
            ["Šelmy"])
        self.assertEqual(
            self.names(self.get({"q": "šelmy", "show_unavailable": "1"})),
            ["Šelmy"])

    def test_search_folds_diacritics_in_alternate_and_series_names(self):
        # Issue #126: the same folding covers the alternate-name (#51) and
        # series-name (#86) haystacks, not just the primary name.
        beasty = Game.objects.create(name="Beasty Bar")
        AlternateName.objects.create(game=beasty, name="Krčma")
        self.assertEqual(
            self.names(self.get({"q": "krcma", "show_unavailable": "1"})),
            ["Beasty Bar"])
        member = Game.objects.create(name="Lonely Member")
        series = Series.objects.create(name="Zvířata", primary_game=member)
        member.series = series
        member.save()
        self.assertEqual(
            self.names(self.get({"q": "zvirata", "show_unavailable": "1"})),
            ["Lonely Member"])

    def test_sort_ignores_leading_articles_and_case(self):
        # Issue #6: "The Crew" files under C; the key is lowercased too.
        self._game(
            "The Crew", game_type=GameType.Type.COOPERATIVE,
            tag=self.fantasy, location=self.shelf,
        )
        Game.objects.create(name="a Feast for Odin")
        response = self.get({"show_unavailable": "1"})
        self.assertEqual(self.names(response), [
            "The Crew", "Duel", "Epic Quest", "a Feast for Odin",
            "Preorder Mystery",
        ])

    def test_old_collection_url_redirects_home_with_filters(self):
        # Issue #7: /collection/ bookmarks (incl. hx-push-url'd filters)
        # land on the new home permanently.
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/collection/", {"players": "2"})
        self.assertEqual(response.status_code, 301)
        self.assertEqual(response["Location"], "/?players=2")

    def test_available_only_is_the_default_and_show_unavailable_opts_out(self):
        # Issue #107: a clean URL shows available games only — Epic Quest is at
        # Pavel's (lent) and Preorder Mystery has no copy, so both drop out.
        # Locationless copies still count as available. The opt-out param brings
        # the lent-out and copy-less games back.
        self.assertEqual(self.names(self.get()), ["Duel"])
        self.assertEqual(
            self.names(self.get({"show_unavailable": "1"})),
            ["Duel", "Epic Quest", "Preorder Mystery"],
        )

    def test_not_ready_copy_is_excluded_from_available_now(self):
        # Issue #19: an active copy that isn't ready to play (e.g. an
        # unprinted PnP copy) isn't "available now" even though it's not
        # lent out — but it's still visible with show_unavailable.
        game = Game.objects.create(name="Papercraft Prototype")
        edition = Edition.objects.create(game=game, is_default=True)
        Copy.objects.create(
            owner=self.user, edition=edition, location=self.shelf,
            ready_status=Copy.ReadyStatus.NOT_READY,
        )
        self.assertNotIn("Papercraft Prototype", self.names(self.get()))
        self.assertIn(
            "Papercraft Prototype",
            self.names(self.get({"show_unavailable": "1"})),
        )

    def test_series_collapsed_by_default_and_show_all_editions_opts_out(self):
        # Issue #107: series collapse is the default (a clean URL groups a
        # series into one tile); show_all_editions opts out into per-game tiles.
        members = [
            self._game(
                name, game_type=GameType.Type.COMPETITIVE,
                tag=self.fantasy, location=self.shelf,
            )
            for name in ("Twin Saga: One", "Twin Saga: Two")
        ]
        series = Series.objects.create(name="Twin Saga", primary_game=members[0])
        Game.objects.filter(pk__in=[m.pk for m in members]).update(series=series)

        default_tiles = self.get().context["tiles"]
        series_tiles = [t for t in default_tiles if t.get("series")]
        self.assertEqual([t["series"].name for t in series_tiles], ["Twin Saga"])

        opened = self.get({"show_all_editions": "1"}).context["tiles"]
        self.assertFalse(any(t.get("series") for t in opened))
        names = [t["game"].name for t in opened if t.get("game")]
        self.assertIn("Twin Saga: One", names)
        self.assertIn("Twin Saga: Two", names)

    def test_htmx_request_returns_just_the_grid(self):
        response = self.get({"players": "2", "show_unavailable": "1"},
                            HTTP_HX_REQUEST="true")
        self.assertContains(response, "2 of 3 games")
        self.assertNotContains(response, "<form")

    def test_default_view_is_grid_and_view_list_switches_to_a_table(self):
        # Issue #92: ?view=list is the only thing that swaps the partial —
        # a clean URL keeps the existing cover-grid behaviour.
        grid = self.get()
        self.assertContains(grid, 'class="cover-grid"')
        self.assertNotContains(grid, "<table")

        listing = self.get({"view": "list"})
        self.assertContains(listing, "<table")
        self.assertNotContains(listing, 'class="cover-grid"')

    def test_toggle_buttons_render_with_correct_state(self):
        def pressed(html, title):
            match = re.search(rf'aria-pressed="(true|false)"[^>]*title="{title}"', html)
            self.assertIsNotNone(match, f"{title} button not found")
            return match.group(1)

        grid_html = self.get().content.decode()
        self.assertEqual(pressed(grid_html, "Grid view"), "true")
        self.assertEqual(pressed(grid_html, "List view"), "false")

        list_html = self.get({"view": "list"}).content.decode()
        self.assertEqual(pressed(list_html, "Grid view"), "false")
        self.assertEqual(pressed(list_html, "List view"), "true")

    def test_list_view_columns_for_a_single_game(self):
        self._game(
            "Statful Game", min_players=2, max_players=4, min_playtime=60,
            game_type=GameType.Type.COMPETITIVE, tag=self.horror,
            location=self.shelf, weight=Decimal("3.5"),
            year_published=2019, bgg_rank=42,
        )
        response = self.get({"view": "list", "q": "Statful"})
        self.assertContains(response, "Statful Game")
        self.assertContains(response, "Base game")
        self.assertContains(response, "2019")
        self.assertContains(response, "2–4")
        self.assertContains(response, "60 min")
        self.assertContains(response, "3.50 / 5")
        self.assertContains(response, "42")
        self.assertContains(response, "Chynice")
        # No copies carry a keep_status in this fixture — falls back to "—".
        tile = response.context["tiles"][0]
        self.assertEqual(tile["copy_count"], 1)
        self.assertEqual(tile["keep_label"], "—")

    def test_copy_summary_uniform_vs_mixed(self):
        uniform = Game.objects.create(name="Uniform Game")
        Copy.objects.create(
            owner=self.user,
            edition=Edition.objects.create(game=uniform, is_default=True),
            location=self.shelf, keep_status=Copy.KeepStatus.KEEP,
        )
        Copy.objects.create(
            owner=self.user,
            edition=Edition.objects.create(game=uniform, name="Reprint"),
            location=self.shelf, keep_status=Copy.KeepStatus.KEEP,
        )

        mixed = Game.objects.create(name="Mixed Game")
        Copy.objects.create(
            owner=self.user,
            edition=Edition.objects.create(game=mixed, is_default=True),
            location=self.shelf, keep_status=Copy.KeepStatus.KEEP,
        )
        Copy.objects.create(
            owner=self.user,
            edition=Edition.objects.create(game=mixed, name="Reprint"),
            location=self.lent, keep_status=Copy.KeepStatus.WILL_LEAVE,
        )

        response = self.get({"view": "list", "show_unavailable": "1"})
        tiles_by_name = {t["game"].name: t for t in response.context["tiles"] if t.get("game")}

        self.assertEqual(tiles_by_name["Uniform Game"]["copy_count"], 2)
        self.assertEqual(tiles_by_name["Uniform Game"]["location_label"], "Chynice")
        self.assertEqual(tiles_by_name["Uniform Game"]["keep_label"], "Keep")

        self.assertEqual(tiles_by_name["Mixed Game"]["copy_count"], 2)
        self.assertEqual(tiles_by_name["Mixed Game"]["location_label"], "multiple")
        self.assertEqual(tiles_by_name["Mixed Game"]["keep_label"], "mixed")

    def test_htmx_list_request_returns_just_the_results(self):
        response = self.get({"view": "list", "show_unavailable": "1"},
                            HTTP_HX_REQUEST="true")
        self.assertContains(response, "of 3 games")
        self.assertContains(response, "<table")
        self.assertNotContains(response, "<form")

    def test_sort_by_weight_ascending_and_descending(self):
        # Issue #92 follow-up: clickable column sort. show_unavailable so
        # Epic Quest (lent) stays in the comparison too.
        Game.objects.filter(pk=self.duel.pk).update(weight=Decimal("2.5"))
        Game.objects.filter(pk=self.epic.pk).update(weight=Decimal("4.0"))

        ascending = self.get({"view": "list", "sort": "weight", "show_unavailable": "1"})
        weighted = [t["stat_game"].name for t in ascending.context["tiles"]
                    if t["stat_game"].weight is not None]
        self.assertEqual(weighted, ["Duel", "Epic Quest"])

        descending = self.get({"view": "list", "sort": "-weight", "show_unavailable": "1"})
        weighted = [t["stat_game"].name for t in descending.context["tiles"]
                    if t["stat_game"].weight is not None]
        self.assertEqual(weighted, ["Epic Quest", "Duel"])

    def test_sort_missing_values_sort_last_regardless_of_direction(self):
        # Preorder Mystery and Epic Quest carry no weight; Duel does. A
        # missing value should never jump to the top on a descending sort.
        Game.objects.filter(pk=self.duel.pk).update(weight=Decimal("3.0"))
        response = self.get({"view": "list", "sort": "-weight", "show_unavailable": "1"})
        names = [t["stat_game"].name for t in response.context["tiles"]]
        self.assertEqual(names[0], "Duel")

    def test_sort_column_headers_report_next_sort_and_direction(self):
        response = self.get({"view": "list", "sort": "-weight"})
        columns = {c["key"]: c for c in response.context["sort_columns"]}
        self.assertEqual(columns["weight"]["direction"], "desc")
        self.assertEqual(columns["weight"]["next_sort"], "weight")
        self.assertIsNone(columns["year"]["direction"])
        self.assertEqual(columns["year"]["next_sort"], "year")

    def test_view_and_sort_controls_include_the_chooser_form(self):
        # Regression: the Django test client sends query params directly, so
        # it can't catch a dropped filter the way a real toggle click would
        # (that's htmx/browser behavior — covered manually, not here). What
        # IS testable here: every control that re-requests the results must
        # hx-include="#chooser" so the browser actually attaches the rest of
        # the active filters, not just its own hx-vals override.
        grid_response = self.get()
        self.assertContains(grid_response, 'hx-include="#chooser"')

        list_response = self.get({"view": "list"})
        self.assertContains(list_response, 'hx-include="#chooser"',
                            count=2 + len(list_response.context["sort_columns"]))

    def test_weight_mechanic_and_played_axes_render_enabled(self):
        # Issues #63/#147: the BGG token landed, so these controls are no
        # longer disabled seams.
        response = self.get()
        self.assertNotContains(response, "awaiting BGG sync")
        self.assertNotContains(response, 'id="f-weight" disabled')
        self.assertNotContains(response, 'id="f-mechanic" disabled')
        self.assertContains(response, 'id="f-weight"')
        self.assertContains(response, 'id="f-mechanic"')
        self.assertContains(response, 'id="f-played"')

    def test_mechanic_filter(self):
        coop = Tag.objects.create(kind=Tag.Kind.MECHANIC, name="Cooperative Game")
        GameTag.objects.create(game=self.epic, tag=coop)
        response = self.get({"mechanic": str(coop.pk), "show_unavailable": "1"})
        self.assertEqual(self.names(response), ["Epic Quest"])

    def test_weight_filter_at_most_and_at_least(self):
        self.duel.weight = Decimal("1.50")
        self.duel.save(update_fields=["weight"])
        self.epic.weight = Decimal("3.80")
        self.epic.save(update_fields=["weight"])

        at_most = self.get({"weight": "2", "weight_mode": "max", "show_unavailable": "1"})
        self.assertEqual(self.names(at_most), ["Duel"])

        at_least = self.get({"weight": "2", "weight_mode": "min", "show_unavailable": "1"})
        self.assertEqual(self.names(at_least), ["Epic Quest"])

        # Preorder Mystery has no weight at all — excluded from either
        # threshold, never assumed to satisfy it.
        self.assertNotIn("Preorder Mystery", self.names(at_most))
        self.assertNotIn("Preorder Mystery", self.names(at_least))

    def test_weight_slider_value_is_ignored_without_a_mode(self):
        self.duel.weight = Decimal("1.50")
        self.duel.save(update_fields=["weight"])
        response = self.get({"weight": "2", "show_unavailable": "1"})
        self.assertEqual(
            self.names(response), ["Duel", "Epic Quest", "Preorder Mystery"],
        )

    def test_played_status_filter(self):
        self.duel.bgg_numplays = 3
        self.duel.save(update_fields=["bgg_numplays"])

        played = self.get({"played": "played", "show_unavailable": "1"})
        self.assertEqual(self.names(played), ["Duel"])

        # bgg_numplays NULL means "unsynced or zero" (models.py) — both
        # Epic Quest (unsynced) and Preorder Mystery count as never played.
        never = self.get({"played": "never", "show_unavailable": "1"})
        self.assertEqual(self.names(never), ["Epic Quest", "Preorder Mystery"])

    def test_tab_title_puts_the_page_first(self):
        # Issue #25: "<something> | GameKeeper", not the other way round.
        response = self.get()
        self.assertContains(response,
                            "<title>Collection | GameKeeper</title>")


# ===========================================================================
# §13  Local cover downloads
# ===========================================================================

class DownloadCoversTests(TestCase):
    def setUp(self):
        # A fresh MEDIA_ROOT per test: files are not transactional, so a
        # shared dir would leak covers between tests and trigger storage
        # name dedupe.
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)

    @classmethod
    def setUpTestData(cls):
        cls.game = Game.objects.create(
            name="5 Minute Dungeon",
            image_url="https://cf.geekdo-images.com/original/img/5md.png",
        )
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)
        cls.unsynced = Game.objects.create(name="Preorder Mystery")

    def run_command(self, content=None, **extra):
        out = StringIO()
        with mock.patch(
            "gamekeeper.management.commands.download_covers.requests.get",
        ) as get:
            get.return_value = mock.Mock(content=content or image_bytes())
            call_command("download_covers", stdout=out, **extra)
        return out.getvalue(), get

    def test_downloads_full_image_named_by_bgg_id(self):
        output, get = self.run_command()

        get.assert_called_once_with(
            "https://cf.geekdo-images.com/original/img/5md.png", timeout=30,
        )
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_image.name, "covers/207830.png")
        with self.game.cover_image.open() as stored:
            self.assertEqual(stored.read(), image_bytes())
        self.assertIn("downloaded: 1", output)
        self.assertIn("no image URL (not synced yet): 1", output)

    def test_download_records_art_dimensions(self):
        self.run_command()
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_width, 4)
        self.assertEqual(self.game.cover_height, 4)

    def test_download_bakes_the_grid_preview(self):
        # Issue #104: a freshly downloaded cover gets its square thumbnail.
        self.run_command()
        self.game.refresh_from_db()
        self.assertTrue(self.game.cover_preview)
        self.assertTrue(self.game.cover_preview.name.startswith("covers/previews/"))

    def test_unreadable_bytes_leave_dimensions_empty(self):
        self.run_command(content=b"not-an-image")
        self.game.refresh_from_db()
        self.assertTrue(self.game.cover_image)
        self.assertIsNone(self.game.cover_width)

    def test_skip_backfills_missing_dimensions(self):
        # Covers downloaded before the dimension fields existed (issue #1).
        self.run_command()
        Game.objects.filter(pk=self.game.pk).update(
            cover_width=None, cover_height=None,
        )

        output, get = self.run_command(dry_run=True)
        get.assert_not_called()
        self.assertIn("would record dimensions for existing file: 1", output)
        self.game.refresh_from_db()
        self.assertIsNone(self.game.cover_width)

        output, get = self.run_command()
        get.assert_not_called()
        self.assertIn("dimensions recorded for existing file: 1", output)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_width, 4)
        self.assertEqual(self.game.cover_height, 4)

    def test_existing_files_survive_without_force(self):
        self.run_command()
        output, get = self.run_command()

        get.assert_not_called()
        self.assertIn("already downloaded — skipped: 1", output)

    def test_force_replaces_the_file_under_the_same_name(self):
        self.run_command()
        _, get = self.run_command(force=True)

        get.assert_called_once()
        self.game.refresh_from_db()
        # Deleted before saving, so no storage dedupe suffix creeps in.
        self.assertEqual(self.game.cover_image.name, "covers/207830.png")

    def test_download_failure_is_reported_and_others_continue(self):
        out = StringIO()
        with mock.patch(
            "gamekeeper.management.commands.download_covers.requests.get",
        ) as get:
            get.side_effect = requests.RequestException("boom")
            call_command("download_covers", stdout=out)

        self.game.refresh_from_db()
        self.assertFalse(self.game.cover_image)
        self.assertIn("download failed", out.getvalue())

    def test_dry_run_fetches_nothing(self):
        output, get = self.run_command(dry_run=True)

        get.assert_not_called()
        self.assertIn("would download: 1", output)


# ===========================================================================
# §11  Curation / cull candidates
# ===========================================================================

class CurationViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )

        cls.bored = cls._copy("Boring Filler", excitement="3.0",
                              why_might_leave="Never hits the table.")
        cls.fine = cls._copy("Fine Game", excitement="6.5")
        cls.loved = cls._copy(
            "All-time Favourite", excitement="9.5", immune=True,
            keep_status=Copy.KeepStatus.ALWAYS_KEEP,
        )
        cls.unrated = cls._copy("Unrated Newcomer")
        cls.gone = cls._copy(
            "Already Culled", excitement="1.0",
            archive_status=Copy.ArchiveStatus.ARCHIVED,
        )
        # Someone else's low-excitement copy must never show up.
        cls._copy("Pavel's Problem", owner=cls.other, excitement="0.5")

    @classmethod
    def _copy(cls, name, owner=None, game_type=Game.Type.BASE, **fields):
        game = Game.objects.create(name=name, type=game_type)
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(
            owner=owner or cls.user, edition=edition, **fields,
        )

    def get(self, params=None, **extra):
        self.client.login(username="kernicek", password="pass")
        return self.client.get("/curation/", params or {}, **extra)

    def names(self, response):
        return [copy.edition.game.name for copy in response.context["copies"]]

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/curation/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_cull_priority_order_low_excitement_first_unrated_last(self):
        # Immune and archived copies are out; Pavel's copy is not ours.
        response = self.get()
        self.assertEqual(
            self.names(response),
            ["Boring Filler", "Fine Game", "Unrated Newcomer"],
        )
        self.assertContains(response, "Never hits the table.")
        self.assertContains(response, "3 of 4 active copies")

    def test_frozen_order_hidden_field_reflects_display_order(self):
        # Issue #24: the pinned-order hidden field always mirrors whatever
        # order the table actually rendered in, so the next inline edit can
        # echo it back and keep the row from jumping.
        response = self.get()
        pks = [copy.pk for copy in response.context["copies"]]
        self.assertContains(
            response,
            f'name="frozen_order" value="{",".join(str(pk) for pk in pks)}"',
        )

    def test_why_it_might_leave_breaks_excitement_ties(self):
        self._copy("Tied But Fine", excitement="3.0")
        response = self.get()
        # Same excitement: the copy with a filled-in reason sorts first.
        self.assertEqual(
            self.names(response)[:2], ["Boring Filler", "Tied But Fine"],
        )

    def test_immune_toggle_includes_protected_copies(self):
        response = self.get({"show_immune": "1"})
        self.assertEqual(
            self.names(response),
            ["Boring Filler", "Fine Game", "All-time Favourite", "Unrated Newcomer"],
        )
        self.assertContains(response, "immune")

    def test_keep_status_filter(self):
        response = self.get({"show_immune": "1", "keep": "always_keep"})
        self.assertEqual(self.names(response), ["All-time Favourite"])

    def test_htmx_request_returns_just_the_table(self):
        response = self.get({"show_immune": "1"}, HTTP_HX_REQUEST="true")
        self.assertContains(response, "4 of 4 active copies")
        self.assertNotContains(response, "<form")

    def test_prev_owned_rows_cull_as_sold(self):
        # §4 archive mapping: BGG already says the game left the collection,
        # so that row's Cull button pre-selects reason=sold; everything else
        # keeps the culled default.
        game = self.bored.edition.game
        game.bgg_collection_status = Game.BggCollectionStatus.PREV_OWNED
        game.save(update_fields=["bgg_collection_status"])

        response = self.get()
        self.assertContains(response, "hx-vals='{\"reason\": \"sold\"}'", count=1)
        self.assertContains(
            response,
            "Archive Boring Filler as sold? BGG already lists it as previously owned.",
        )
        self.assertContains(response, "Archive Fine Game as culled?")

    def test_borrowed_copy_is_excluded_from_cull_candidates(self):
        # Issue #43: you can't cull what you don't own.
        borrowed = self._copy("Borrowed Beauty", excitement="0.1", is_borrowed_in=True)
        Loan.objects.create(
            copy=borrowed, direction=Loan.Direction.BORROWED_IN,
            counterparty_name="Mira",
        )
        response = self.get()
        self.assertNotIn("Borrowed Beauty", self.names(response))
        self.assertContains(response, "3 of 4 active copies")

    def test_expansions_hidden_by_default(self):
        # Issue #39: expansions clutter the cull list — hidden by default,
        # same as immune copies, and still counted toward the total.
        self._copy(
            "Expansion Pack", game_type=Game.Type.EXPANSION, excitement="0.1",
        )
        response = self.get()
        self.assertNotIn("Expansion Pack", self.names(response))
        self.assertContains(response, "3 of 5 active copies")

    def test_show_expansions_toggle_includes_expansions(self):
        self._copy(
            "Expansion Pack", game_type=Game.Type.EXPANSION, excitement="0.1",
        )
        response = self.get({"show_expansions": "1"})
        self.assertIn("Expansion Pack", self.names(response))


class CurationEditTests(TestCase):
    """§11 in-place editing on the curation table: owner-scoped htmx POSTs
    update excitement / keep-status and return the re-rendered table (with
    the hx-include'd filters applied). Issue #24: the row order stays
    pinned to whatever frozen_order the client echoes back (as the real
    table's hidden field would); only a fresh page load or filter change
    re-applies the true cull-priority sort."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.copy = cls._copy("Boring Filler", excitement="3.0")
        cls.fine = cls._copy("Fine Game", excitement="6.5")
        cls.foreign = cls._copy("Pavel's Problem", owner=cls.other)

    @classmethod
    def _copy(cls, name, owner=None, **fields):
        game = Game.objects.create(name=name)
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(
            owner=owner or cls.user, edition=edition, **fields,
        )

    def post(self, data, copy=None):
        self.client.login(username="kernicek", password="pass")
        return self.client.post(
            f"/curation/copies/{(copy or self.copy).pk}/", data,
        )

    def names(self, response):
        return [copy.edition.game.name for copy in response.context["copies"]]

    def test_anonymous_is_sent_to_login(self):
        response = self.client.post(f"/curation/copies/{self.copy.pk}/", {})
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_not_allowed(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/curation/copies/{self.copy.pk}/")
        self.assertEqual(response.status_code, 405)

    def test_someone_elses_copy_is_404(self):
        response = self.post({"excitement": "1"}, copy=self.foreign)
        self.assertEqual(response.status_code, 404)
        self.foreign.refresh_from_db()
        self.assertIsNone(self.foreign.excitement)

    def test_archived_copy_is_404(self):
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.save()
        response = self.post({"excitement": "1"})
        self.assertEqual(response.status_code, 404)

    def _frozen_order(self):
        # Boring Filler (3.0) currently beats Fine Game (6.5) — the order
        # the real table's hidden field would echo back on any edit.
        return f"{self.copy.pk},{self.fine.pk}"

    def test_excitement_update_stays_pinned_when_order_is_frozen(self):
        response = self.post({
            "excitement": "8", "frozen_order": self._frozen_order(),
        })
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.excitement, Decimal("8.0"))
        # 8.0 would normally beat Fine Game's 6.5, but the posted
        # frozen_order pins the row in place until the next full load.
        self.assertEqual(self.names(response), ["Boring Filler", "Fine Game"])
        self.assertContains(response, 'value="8.0"')

    def test_excitement_update_resorts_when_no_order_is_pinned(self):
        # Fallback path: without a frozen_order (e.g. a stale client), the
        # table still falls back to a fresh cull-priority sort.
        response = self.post({"excitement": "8"})
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.excitement, Decimal("8.0"))
        self.assertEqual(self.names(response), ["Fine Game", "Boring Filler"])

    def test_blank_excitement_stays_pinned_when_order_is_frozen(self):
        response = self.post({
            "excitement": "", "frozen_order": self._frozen_order(),
        })
        self.copy.refresh_from_db()
        self.assertIsNone(self.copy.excitement)
        # Unrated copies would normally sink to the bottom, but the posted
        # frozen_order pins the row in place until the next full load.
        self.assertEqual(self.names(response), ["Boring Filler", "Fine Game"])

    def test_invalid_excitement_is_rejected(self):
        for bad in ("eleven", "15", "-1"):
            response = self.post({"excitement": bad})
            self.assertEqual(response.status_code, 400)
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.excitement, Decimal("3.0"))

    def test_keep_status_update_and_clear(self):
        response = self.post({"keep_status": "will_leave"})
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.keep_status, Copy.KeepStatus.WILL_LEAVE)
        self.assertContains(response, "Will leave")

        self.post({"keep_status": ""})
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.keep_status, "")

    def test_invalid_keep_status_is_rejected(self):
        response = self.post({"keep_status": "sell_it_all"})
        self.assertEqual(response.status_code, 400)
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.keep_status, "")

    def test_marking_will_leave_pushes_fortrade(self):
        # Issue #82: keep_status=WILL_LEAVE is the "for trade" marker —
        # setting it on a BGG-linked game's copy pushes fortrade=True.
        game = self.copy.edition.game
        BggLink.objects.create(game=game, bgg_id=300100, is_primary=True)
        with mock.patch("gamekeeper.views.push_bgg_fortrade_task.delay") as delay:
            self.post({"keep_status": "will_leave"})
        delay.assert_called_once_with(game.pk, True, self.user.pk)

    def test_clearing_will_leave_clears_the_fortrade_push(self):
        game = self.copy.edition.game
        BggLink.objects.create(game=game, bgg_id=300100, is_primary=True)
        game.bgg_fortrade_pushed = True
        game.save(update_fields=["bgg_fortrade_pushed"])
        with mock.patch("gamekeeper.views.push_bgg_fortrade_task.delay") as delay:
            self.post({"keep_status": ""})
        delay.assert_called_once_with(game.pk, False, self.user.pk)

    def test_unrelated_field_edit_does_not_push_fortrade(self):
        game = self.copy.edition.game
        BggLink.objects.create(game=game, bgg_id=300100, is_primary=True)
        with mock.patch("gamekeeper.views.push_bgg_fortrade_task.delay") as delay:
            self.post({"excitement": "4"})
        delay.assert_not_called()

    def test_included_filters_shape_the_returned_table(self):
        immune = self._copy("All-time Favourite", excitement="9.5", immune=True)
        # The filter form rides along via hx-include, so the swapped-in
        # table honours the caller's current view.
        response = self.post({"excitement": "2", "show_immune": "1"}, copy=immune)
        self.assertIn("All-time Favourite", self.names(response))

        response = self.post({"excitement": "2"}, copy=immune)
        self.assertNotIn("All-time Favourite", self.names(response))

    def test_edit_lands_in_the_copy_history(self):
        before = self.copy.history.count()
        self.post({"excitement": "1.5"})
        self.assertEqual(self.copy.history.count(), before + 1)
        self.assertEqual(
            self.copy.history.latest().excitement, Decimal("1.5"),
        )

    def test_immune_toggle_updates_the_flag(self):
        # Making a copy immune hides it unless the show-immune filter rides
        # along — it just left the cull race.
        response = self.post({"immune": "1"})
        self.copy.refresh_from_db()
        self.assertTrue(self.copy.immune)
        self.assertNotIn("Boring Filler", self.names(response))

        response = self.post({"immune": "0", "show_immune": "1"})
        self.copy.refresh_from_db()
        self.assertFalse(self.copy.immune)
        self.assertIn("Boring Filler", self.names(response))

    def test_invalid_immune_is_rejected(self):
        response = self.post({"immune": "yes"})
        self.assertEqual(response.status_code, 400)
        self.copy.refresh_from_db()
        self.assertFalse(self.copy.immune)

    def test_why_might_leave_update_stays_pinned_when_order_is_frozen(self):
        # On an excitement tie a filled-in reason would normally win the
        # cull priority — but the posted frozen_order pins the row in
        # place until the next full load.
        self.copy.excitement = Decimal("6.5")
        self.copy.save()
        frozen_order = f"{self.copy.pk},{self.fine.pk}"
        response = self.post(
            {"why_might_leave": "  Shelf queen.  ", "frozen_order": frozen_order},
            copy=self.fine,
        )
        self.fine.refresh_from_db()
        self.assertEqual(self.fine.why_might_leave, "Shelf queen.")
        self.assertEqual(self.names(response), ["Boring Filler", "Fine Game"])

        self.post(
            {"why_might_leave": "", "frozen_order": frozen_order}, copy=self.fine,
        )
        self.fine.refresh_from_db()
        self.assertEqual(self.fine.why_might_leave, "")

    def test_why_might_leave_update_resorts_when_no_order_is_pinned(self):
        # Fallback path: without a frozen_order, the table still falls
        # back to a fresh cull-priority sort.
        self.copy.excitement = Decimal("6.5")
        self.copy.save()
        response = self.post(
            {"why_might_leave": "  Shelf queen.  "}, copy=self.fine,
        )
        self.fine.refresh_from_db()
        self.assertEqual(self.fine.why_might_leave, "Shelf queen.")
        self.assertEqual(self.names(response), ["Fine Game", "Boring Filler"])


class CurationArchiveTests(TestCase):
    """§11 "cull this" from the curation table: the DESIGN §4 archive
    lifecycle endpoint — owner-scoped POSTs that archive the copy and
    return the re-rendered table without it."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.copy = cls._copy("Boring Filler", excitement="3.0")
        cls.fine = cls._copy("Fine Game", excitement="6.5")
        cls.foreign = cls._copy("Pavel's Problem", owner=cls.other)

    @classmethod
    def _copy(cls, name, owner=None, **fields):
        game = Game.objects.create(name=name)
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(
            owner=owner or cls.user, edition=edition, **fields,
        )

    def post(self, data=None, copy=None):
        self.client.login(username="kernicek", password="pass")
        return self.client.post(
            f"/curation/copies/{(copy or self.copy).pk}/archive/", data or {},
        )

    def names(self, response):
        return [copy.edition.game.name for copy in response.context["copies"]]

    def test_anonymous_is_sent_to_login(self):
        response = self.client.post(f"/curation/copies/{self.copy.pk}/archive/", {})
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_not_allowed(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/curation/copies/{self.copy.pk}/archive/")
        self.assertEqual(response.status_code, 405)

    def test_someone_elses_copy_is_404(self):
        response = self.post(copy=self.foreign)
        self.assertEqual(response.status_code, 404)
        self.foreign.refresh_from_db()
        self.assertEqual(self.foreign.archive_status, Copy.ArchiveStatus.ACTIVE)

    def test_already_archived_copy_is_404(self):
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.save()
        response = self.post()
        self.assertEqual(response.status_code, 404)

    def test_archive_defaults_to_culled_and_drops_the_row(self):
        response = self.post()
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.archive_status, Copy.ArchiveStatus.ARCHIVED)
        self.assertEqual(self.copy.archive_reason, Copy.ArchiveReason.CULLED)
        self.assertEqual(self.copy.archive_date, timezone.localdate())
        self.assertEqual(self.names(response), ["Fine Game"])
        self.assertContains(response, "1 of 1 active copy")

    def test_explicit_reason_is_stored(self):
        self.post({"reason": "sold"})
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.archive_reason, Copy.ArchiveReason.SOLD)

    def test_unknown_reason_is_rejected(self):
        response = self.post({"reason": "ate_it"})
        self.assertEqual(response.status_code, 400)
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.archive_status, Copy.ArchiveStatus.ACTIVE)

    def test_included_filters_shape_the_returned_table(self):
        self._copy("All-time Favourite", excitement="9.5", immune=True)
        response = self.post({"show_immune": "1"})
        self.assertEqual(
            self.names(response), ["Fine Game", "All-time Favourite"],
        )

    def test_archive_lands_in_the_copy_history(self):
        before = self.copy.history.count()
        self.post()
        self.assertEqual(self.copy.history.count(), before + 1)
        self.assertEqual(
            self.copy.history.latest().archive_status,
            Copy.ArchiveStatus.ARCHIVED,
        )

    def test_archive_pushes_prev_owned_when_the_game_is_linked(self):
        """Issue #117: archiving the last active copy of a game is the
        "previously owned" signal."""
        BggLink.objects.create(
            game=self.copy.edition.game, bgg_id=207830, is_primary=True,
        )
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.post()
        delay.assert_called_once_with(
            self.copy.edition.game.pk, Game.BggCollectionStatus.PREV_OWNED,
            self.user.pk, priority=None,
        )

    def test_archive_does_not_push_when_another_active_copy_of_the_game_remains(self):
        """A different edition of the same game staying active means the
        game is still owned overall — no "previously owned" push."""
        game = self.copy.edition.game
        BggLink.objects.create(game=game, bgg_id=207830, is_primary=True)
        other_edition = Edition.objects.create(game=game, name="Deluxe")
        Copy.objects.create(owner=self.user, edition=other_edition)

        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.post()

        delay.assert_not_called()

    def test_archiving_the_will_leave_copy_clears_the_fortrade_push(self):
        # Issue #82: the archived copy was the one carrying WILL_LEAVE, so
        # the previously-pushed fortrade flag needs to clear.
        game = self.copy.edition.game
        self.copy.keep_status = Copy.KeepStatus.WILL_LEAVE
        self.copy.save(update_fields=["keep_status"])
        BggLink.objects.create(game=game, bgg_id=207830, is_primary=True)
        game.bgg_fortrade_pushed = True
        game.save(update_fields=["bgg_fortrade_pushed"])

        with mock.patch("gamekeeper.views.push_bgg_fortrade_task.delay") as delay:
            self.post()

        delay.assert_called_once_with(game.pk, False, self.user.pk)

    def test_archiving_a_non_leaving_copy_does_not_touch_fortrade(self):
        game = self.copy.edition.game
        BggLink.objects.create(game=game, bgg_id=207830, is_primary=True)

        with mock.patch("gamekeeper.views.push_bgg_fortrade_task.delay") as delay:
            self.post()

        delay.assert_not_called()

    def test_archive_enqueue_failure_never_breaks_the_archive(self):
        BggLink.objects.create(
            game=self.copy.edition.game, bgg_id=207830, is_primary=True,
        )
        with mock.patch(
            "gamekeeper.views.push_bgg_status_task.delay",
            side_effect=RuntimeError("broker down"),
        ):
            response = self.post()
        self.assertEqual(response.status_code, 200)
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.archive_status, Copy.ArchiveStatus.ARCHIVED)
        diff = BggSyncDiff.objects.get(category=BggSyncDiff.Category.PUSH_FAILED)
        self.assertEqual(diff.game, self.copy.edition.game)


class CopyMarkReadyTests(TestCase):
    """Issue #19: the "flip back to ready" endpoint behind the to-craft
    backlog's "Mark as printed" button — owner/not-ready-scoped POSTs that
    re-render the to-craft table without the copy."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.copy = cls._copy("Papercraft Prototype")
        cls.foreign = cls._copy("Pavel's Papercraft", owner=cls.other)

    @classmethod
    def _copy(cls, name, owner=None, **fields):
        game = Game.objects.create(name=name)
        edition = Edition.objects.create(game=game, is_default=True)
        fields.setdefault("ready_status", Copy.ReadyStatus.NOT_READY)
        return Copy.objects.create(owner=owner or cls.user, edition=edition, **fields)

    def post(self, copy=None):
        self.client.login(username="kernicek", password="pass")
        return self.client.post(f"/copies/{(copy or self.copy).pk}/mark-ready/")

    def names(self, response):
        return [copy.edition.game.name for copy in response.context["to_craft"]]

    def test_anonymous_is_sent_to_login(self):
        response = self.client.post(f"/copies/{self.copy.pk}/mark-ready/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_not_allowed(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/copies/{self.copy.pk}/mark-ready/")
        self.assertEqual(response.status_code, 405)

    def test_someone_elses_copy_is_404(self):
        response = self.post(copy=self.foreign)
        self.assertEqual(response.status_code, 404)
        self.foreign.refresh_from_db()
        self.assertEqual(self.foreign.ready_status, Copy.ReadyStatus.NOT_READY)

    def test_already_ready_copy_is_404(self):
        self.copy.ready_status = Copy.ReadyStatus.READY
        self.copy.save()
        response = self.post()
        self.assertEqual(response.status_code, 404)

    def test_mark_ready_flips_status_and_drops_the_row(self):
        response = self.post()
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.ready_status, Copy.ReadyStatus.READY)
        self.assertNotIn("Papercraft Prototype", self.names(response))


class ArchivedCopiesViewTests(TestCase):
    """§4 archive shelf: archived copies stay findable in a browse-only,
    owner-scoped view — newest departures first, filterable by reason."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.culled = cls._archived(
            "Boring Filler", Copy.ArchiveReason.CULLED,
            datetime.date(2026, 6, 1), excitement="3.0",
            why_might_leave="Never hits the table.",
        )
        cls.sold = cls._archived(
            "Sold Star", Copy.ArchiveReason.SOLD, datetime.date(2026, 7, 1),
        )
        cls.undated = cls._archived("Undated Departure", "", None)
        # Still-active copies and other people's archives never show up.
        cls._copy("Active Keeper", excitement="8.0")
        cls._archived(
            "Pavel's Past", Copy.ArchiveReason.SOLD,
            datetime.date(2026, 7, 2), owner=cls.other,
        )

    @classmethod
    def _copy(cls, name, owner=None, **fields):
        game = Game.objects.create(name=name)
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(
            owner=owner or cls.user, edition=edition, **fields,
        )

    @classmethod
    def _archived(cls, name, reason, date, **fields):
        return cls._copy(
            name, archive_status=Copy.ArchiveStatus.ARCHIVED,
            archive_reason=reason, archive_date=date, **fields,
        )

    def get(self, params=None, **extra):
        self.client.login(username="kernicek", password="pass")
        return self.client.get("/curation/archived/", params or {}, **extra)

    def names(self, response):
        return [copy.edition.game.name for copy in response.context["copies"]]

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/curation/archived/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_own_archived_copies_newest_first_undated_last(self):
        response = self.get()
        self.assertEqual(
            self.names(response),
            ["Sold Star", "Boring Filler", "Undated Departure"],
        )
        self.assertContains(response, "3 of 3 archived copies")
        self.assertContains(response, "Never hits the table.")

    def test_reason_filter(self):
        response = self.get({"reason": "culled"})
        self.assertEqual(self.names(response), ["Boring Filler"])
        self.assertContains(response, "1 of 3 archived copies")

    def test_unknown_reason_shows_everything(self):
        response = self.get({"reason": "ate_it"})
        self.assertEqual(len(self.names(response)), 3)

    def test_htmx_request_returns_just_the_table(self):
        response = self.get({"reason": "sold"}, HTTP_HX_REQUEST="true")
        self.assertContains(response, "1 of 3 archived copies")
        self.assertNotContains(response, "<form")

    def test_culled_row_from_curation_lands_here(self):
        # The curation Cull button feeds this shelf end to end.
        active = self._copy("Fresh Cull", excitement="2.0")
        self.client.login(username="kernicek", password="pass")
        self.client.post(f"/curation/copies/{active.pk}/archive/", {})
        response = self.get()
        self.assertEqual(self.names(response)[0], "Fresh Cull")


# ===========================================================================
# §5  Sleeves workbench views
# ===========================================================================

class SleevesPageMixin:
    """Shared fixture: two sizes, a product per size, and copies whose
    editions carry sleeve requirements — the §5 worklist building blocks."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.standard = CardSize.objects.create(
            width_mm=Decimal("63.0"), height_mm=Decimal("88.0"), name="Standard",
        )
        cls.mini = CardSize.objects.create(
            width_mm=Decimal("41.0"), height_mm=Decimal("63.0"), name="Mini",
        )
        cls.green = SleeveProduct.objects.create(
            brand="Tlama", name="Diamond Green", card_size=cls.standard,
        )
        cls.yellow = SleeveProduct.objects.create(
            brand="Tlama", name="Diamond Yellow", card_size=cls.mini,
        )

        # Arnak: standard cards marked to-sleeve, minis with no status row yet.
        cls.arnak = cls._copy("Lost Ruins of Arnak")
        cls._require(cls.arnak, cls.standard, 195)
        cls._require(cls.arnak, cls.mini, 30)
        CopySleeveStatus.objects.create(
            copy=cls.arnak, card_size=cls.standard,
            status=CopySleeveStatus.Status.TO_SLEEVE,
        )
        # Roam: already sleeved, with the product recorded.
        cls.roam = cls._copy("Roam")
        cls._require(cls.roam, cls.standard, 60)
        CopySleeveStatus.objects.create(
            copy=cls.roam, card_size=cls.standard,
            status=CopySleeveStatus.Status.SLEEVED, product=cls.green,
        )
        # Noise that must never show: someone else's copy, an archived copy.
        foreign = cls._copy("Pavel's Problem", owner=cls.other)
        cls._require(foreign, cls.standard, 10)
        archived = cls._copy(
            "Already Culled", archive_status=Copy.ArchiveStatus.ARCHIVED,
        )
        cls._require(archived, cls.standard, 10)

    @classmethod
    def _copy(cls, name, owner=None, **fields):
        game = Game.objects.create(name=name)
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(
            owner=owner or cls.user, edition=edition, **fields,
        )

    @classmethod
    def _require(cls, copy, card_size, count):
        return SleeveRequirement.objects.get_or_create(
            edition=copy.edition, card_size=card_size,
            defaults={"count": count},
        )[0]

    def login(self):
        self.client.login(username="kernicek", password="pass")

    def rows(self, response):
        return [
            (row["copy"].edition.game.name, row["card_size"].name, row["status"])
            for row in response.context["sleeve_rows"]
        ]


class SleevesViewTests(SleevesPageMixin, TestCase):
    """§5 sleeves workbench: shortfall with covered sizes, inventory table,
    and the per-copy worklist built from requirements joined with statuses."""

    def get(self, params=None, **extra):
        self.login()
        return self.client.get("/sleeves/", params or {}, **extra)

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/sleeves/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_worklist_shows_own_active_slots_work_first(self):
        response = self.get()
        # Requirement slots without a status row render as not-sleeved;
        # foreign and archived copies never appear.
        self.assertEqual(self.rows(response), [
            ("Lost Ruins of Arnak", "Standard", "to_sleeve"),
            ("Lost Ruins of Arnak", "Mini", "not_sleeved"),
            ("Roam", "Standard", "sleeved"),
        ])
        self.assertContains(response, "3 of 3 card slots")
        self.assertContains(response, "195")

    def test_orphan_status_without_requirement_shows_unknown_count(self):
        SleeveRequirement.objects.filter(edition=self.roam.edition).delete()
        response = self.get()
        self.assertIn(("Roam", "Standard", "sleeved"), self.rows(response))
        self.assertContains(response, ">?</td>", html=False)

    def test_status_filter(self):
        response = self.get({"show": "to_sleeve"})
        self.assertEqual(
            self.rows(response), [("Lost Ruins of Arnak", "Standard", "to_sleeve")],
        )
        self.assertContains(response, "1 of 3 card slots")

    def test_size_filter(self):
        response = self.get({"size": str(self.mini.pk)})
        self.assertEqual(
            self.rows(response), [("Lost Ruins of Arnak", "Mini", "not_sleeved")],
        )

    def test_shortfall_keeps_covered_sizes_visible(self):
        # 195 standard to sleeve vs 300 in stock: covered, but still listed.
        SleeveInventory.objects.create(owner=self.user, product=self.green, packs=3)
        response = self.get()
        (entry,) = response.context["shortfall"]
        self.assertEqual(entry["card_size"], self.standard)
        self.assertEqual(entry["shortfall"], 0)

    def test_inventory_lists_every_product_with_own_stock_only(self):
        SleeveInventory.objects.create(owner=self.user, product=self.green, packs=3)
        SleeveInventory.objects.create(owner=self.other, product=self.yellow, packs=9)
        response = self.get()
        rows = {
            row["product"].pk: row["inventory"]
            for row in response.context["inventory_rows"]
        }
        self.assertEqual(set(rows), {self.green.pk, self.yellow.pk})
        self.assertEqual(rows[self.green.pk].packs, 3)
        self.assertIsNone(rows[self.yellow.pk])  # pavel's stock is not ours

    def test_htmx_request_returns_just_the_worklist_table(self):
        response = self.get({"show": "sleeved"}, HTTP_HX_REQUEST="true")
        self.assertContains(response, "1 of 3 card slots")
        self.assertNotContains(response, "<form")
        self.assertNotContains(response, "Inventory")


class SleeveInventoryEditTests(SleevesPageMixin, TestCase):
    """§5 in-place stock editing: (owner, product) rows are created on first
    edit, validated, and the whole inventory table is re-rendered."""

    def post(self, data, product=None):
        self.login()
        return self.client.post(
            f"/sleeves/inventory/{(product or self.green).pk}/", data,
        )

    def test_anonymous_is_sent_to_login(self):
        response = self.client.post(f"/sleeves/inventory/{self.green.pk}/", {})
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_not_allowed(self):
        self.login()
        response = self.client.get(f"/sleeves/inventory/{self.green.pk}/")
        self.assertEqual(response.status_code, 405)

    def test_unknown_product_is_404(self):
        response = self.post({"packs": "1"}, product=SleeveProduct(pk=99999))
        self.assertEqual(response.status_code, 404)

    def test_first_edit_creates_the_inventory_row(self):
        response = self.post({"packs": "4"})
        inventory = SleeveInventory.objects.get(owner=self.user, product=self.green)
        self.assertEqual(inventory.packs, 4)
        self.assertEqual(inventory.loose, 0)
        self.assertContains(response, "400")  # total = 4 packs of 100

    def test_loose_edit_keeps_packs(self):
        SleeveInventory.objects.create(owner=self.user, product=self.green, packs=2)
        self.post({"loose": "37"})
        inventory = SleeveInventory.objects.get(owner=self.user, product=self.green)
        self.assertEqual((inventory.packs, inventory.loose), (2, 37))

    def test_blank_means_zero(self):
        SleeveInventory.objects.create(owner=self.user, product=self.green, packs=2)
        self.post({"packs": ""})
        inventory = SleeveInventory.objects.get(owner=self.user, product=self.green)
        self.assertEqual(inventory.packs, 0)

    def test_garbage_and_negative_counts_are_rejected(self):
        for bad in ("many", "-1", "1.5"):
            response = self.post({"packs": bad})
            self.assertEqual(response.status_code, 400)
        self.assertFalse(SleeveInventory.objects.filter(owner=self.user).exists())

    def test_edits_touch_only_the_editors_stock(self):
        SleeveInventory.objects.create(owner=self.other, product=self.green, packs=9)
        self.post({"packs": "1"})
        self.assertEqual(
            SleeveInventory.objects.get(owner=self.other, product=self.green).packs, 9,
        )
        self.assertEqual(
            SleeveInventory.objects.get(owner=self.user, product=self.green).packs, 1,
        )


class EditionRequirementEditorTests(SleevesPageMixin, TestCase):
    """§5 in-app sleeve-requirement editor (issue #129): add / count-edit /
    delete CardSize→count rows on an edition, defining a new size inline."""

    def setUp(self):
        self.edition = self.arnak.edition

    def add(self, data):
        self.login()
        return self.client.post(
            f"/editions/{self.edition.pk}/requirements/add/", data)

    def edit(self, requirement, data):
        self.login()
        return self.client.post(f"/requirements/{requirement.pk}/edit/", data)

    def delete(self, requirement):
        self.login()
        return self.client.post(f"/requirements/{requirement.pk}/delete/")

    def std_req(self):
        return SleeveRequirement.objects.get(
            edition=self.edition, card_size=self.standard)

    # --- auth / method guards ------------------------------------------
    def test_anonymous_add_is_sent_to_login(self):
        response = self.client.post(
            f"/editions/{self.edition.pk}/requirements/add/", {})
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_anonymous_edit_and_delete_are_sent_to_login(self):
        req = self.std_req()
        for url in (f"/requirements/{req.pk}/edit/",
                    f"/requirements/{req.pk}/delete/"):
            response = self.client.post(url, {})
            self.assertEqual(response.status_code, 302)
            self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_not_allowed(self):
        self.login()
        response = self.client.get(
            f"/editions/{self.edition.pk}/requirements/add/")
        self.assertEqual(response.status_code, 405)

    # --- add -----------------------------------------------------------
    def test_add_existing_card_size_creates_requirement(self):
        blank = Edition.objects.create(game=Game.objects.create(name="Fresh"))
        self.login()
        response = self.client.post(
            f"/editions/{blank.pk}/requirements/add/",
            {"card_size": str(self.standard.pk), "count": "84"})
        req = SleeveRequirement.objects.get(edition=blank, card_size=self.standard)
        self.assertEqual(req.count, 84)
        self.assertContains(response, 'value="84"')

    def test_add_new_size_get_or_creates_the_card_size(self):
        response = self.add({
            "card_size": "new", "width_mm": "45", "height_mm": "68",
            "size_name": "Tarot-ish", "count": "18",
        })
        self.assertEqual(response.status_code, 200)
        size = CardSize.objects.get(
            width_mm=Decimal("45.0"), height_mm=Decimal("68.0"))
        self.assertEqual(size.name, "Tarot-ish")
        self.assertTrue(SleeveRequirement.objects.filter(
            edition=self.edition, card_size=size, count=18).exists())

    def test_add_new_size_reuses_an_existing_dimension_pair(self):
        # 63×88 already exists as Standard; "new" with those dims must reuse it
        # (unique_card_size_dimensions), not create a duplicate CardSize.
        fresh = Edition.objects.create(game=Game.objects.create(name="Fresh"))
        before = CardSize.objects.count()
        self.login()
        self.client.post(
            f"/editions/{fresh.pk}/requirements/add/",
            {"card_size": "new", "width_mm": "63", "height_mm": "88", "count": "10"})
        self.assertEqual(CardSize.objects.count(), before)
        self.assertTrue(SleeveRequirement.objects.filter(
            edition=fresh, card_size=self.standard, count=10).exists())

    def test_re_adding_a_size_updates_the_count(self):
        self.add({"card_size": str(self.standard.pk), "count": "500"})
        self.assertEqual(self.std_req().count, 500)
        self.assertEqual(SleeveRequirement.objects.filter(
            edition=self.edition, card_size=self.standard).count(), 1)

    def test_add_bad_count_is_rejected(self):
        for bad in ("0", "-3", "two", ""):
            response = self.add({"card_size": str(self.standard.pk), "count": bad})
            self.assertEqual(response.status_code, 400)

    def test_add_new_size_needs_positive_dimensions(self):
        for width, height in (("", "68"), ("45", ""), ("0", "68"), ("x", "68")):
            response = self.add({
                "card_size": "new", "width_mm": width, "height_mm": height,
                "count": "5"})
            self.assertEqual(response.status_code, 400)

    # --- edit / delete -------------------------------------------------
    def test_count_edit_updates_the_row(self):
        req = self.std_req()
        response = self.edit(req, {"count": "222"})
        req.refresh_from_db()
        self.assertEqual(req.count, 222)
        self.assertContains(response, 'value="222"')

    def test_edit_bad_count_is_rejected(self):
        req = self.std_req()
        response = self.edit(req, {"count": "0"})
        self.assertEqual(response.status_code, 400)
        req.refresh_from_db()
        self.assertEqual(req.count, 195)

    def test_delete_removes_the_row(self):
        req = self.std_req()
        response = self.delete(req)
        self.assertFalse(SleeveRequirement.objects.filter(pk=req.pk).exists())
        self.assertNotContains(response, 'value="195"')

    def test_unknown_requirement_is_404(self):
        self.login()
        self.assertEqual(self.client.post(
            "/requirements/99999/edit/", {"count": "5"}).status_code, 404)
        self.assertEqual(
            self.client.post("/requirements/99999/delete/").status_code, 404)

    # --- page render ---------------------------------------------------
    def test_edition_edit_page_shows_the_editor(self):
        self.login()
        response = self.client.get(f"/editions/{self.edition.pk}/edit/")
        self.assertContains(response, "Sleeve requirements")
        self.assertContains(response, 'value="195"')     # existing count
        self.assertContains(response, "Add requirement")
        self.assertContains(response, "＋ New size…")      # define-new option

    def test_editor_never_leaks_template_comment_markers(self):
        # The {% comment %} block in the partial and the {# #} lines in
        # edition_edit.html must be stripped, not rendered (CLAUDE.md).
        self.login()
        response = self.client.get(f"/editions/{self.edition.pk}/edit/")
        self.assertNotContains(response, "hx-swap=outerHTML")
        self.assertNotContains(response, "keyed off the saved edition pk")


class SleeveStatusEditTests(SleevesPageMixin, TestCase):
    """§5 in-place worklist editing: per-copy per-size status (and the
    product used), owner+active scoped, row created on first edit."""

    def post(self, data, copy=None, size=None):
        self.login()
        return self.client.post(
            f"/sleeves/copies/{(copy or self.arnak).pk}"
            f"/sizes/{(size or self.standard).pk}/",
            data,
        )

    def test_anonymous_is_sent_to_login(self):
        response = self.client.post(
            f"/sleeves/copies/{self.arnak.pk}/sizes/{self.standard.pk}/", {},
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_not_allowed(self):
        self.login()
        response = self.client.get(
            f"/sleeves/copies/{self.arnak.pk}/sizes/{self.standard.pk}/",
        )
        self.assertEqual(response.status_code, 405)

    def test_someone_elses_copy_is_404(self):
        foreign = Copy.objects.get(owner=self.other)
        response = self.post({"status": "sleeved"}, copy=foreign)
        self.assertEqual(response.status_code, 404)

    def test_archived_copy_is_404(self):
        archived = Copy.objects.get(archive_status=Copy.ArchiveStatus.ARCHIVED)
        response = self.post({"status": "sleeved"}, copy=archived)
        self.assertEqual(response.status_code, 404)

    def test_unknown_size_is_404(self):
        response = self.post({"status": "sleeved"}, size=CardSize(pk=99999))
        self.assertEqual(response.status_code, 404)

    def test_status_change_resorts_the_table(self):
        response = self.post({"status": "sleeved"})
        status = CopySleeveStatus.objects.get(
            copy=self.arnak, card_size=self.standard,
        )
        self.assertEqual(status.status, CopySleeveStatus.Status.SLEEVED)
        # Sleeved rows sink below the remaining work.
        self.assertEqual(self.rows(response)[0],
                         ("Lost Ruins of Arnak", "Mini", "not_sleeved"))

    def test_first_edit_creates_the_status_row(self):
        # Arnak's minis have a requirement but no status row yet.
        self.post({"status": "to_sleeve"}, size=self.mini)
        status = CopySleeveStatus.objects.get(copy=self.arnak, card_size=self.mini)
        self.assertEqual(status.status, CopySleeveStatus.Status.TO_SLEEVE)

    def test_unknown_status_is_rejected(self):
        response = self.post({"status": "shrinkwrapped"})
        self.assertEqual(response.status_code, 400)

    def test_product_used_is_recorded_and_cleared(self):
        self.post({"product": str(self.green.pk)})
        status = CopySleeveStatus.objects.get(
            copy=self.arnak, card_size=self.standard,
        )
        self.assertEqual(status.product, self.green)

        self.post({"product": ""})
        status.refresh_from_db()
        self.assertIsNone(status.product)

    def test_product_of_another_size_is_rejected(self):
        # Yellow fits minis; it cannot have been used on standard cards.
        response = self.post({"product": str(self.yellow.pk)})
        self.assertEqual(response.status_code, 400)

    def test_hx_included_filters_ride_along(self):
        # The edit param is status; the filter param is show — no collision.
        response = self.post({"status": "sleeved", "show": "sleeved"})
        self.assertEqual(self.rows(response), [
            ("Lost Ruins of Arnak", "Standard", "sleeved"),
            ("Roam", "Standard", "sleeved"),
        ])


class GameDetailSleeveCardTests(SleevesPageMixin, TestCase):
    """Issue #17: the read-only sleeve card on the game detail page — the
    viewer's own active copy's sizes + statuses, the same shared table as the
    §5 worklist, hidden when there is nothing to show."""

    def get(self, copy):
        self.login()
        return self.client.get(f"/games/{copy.edition.game.pk}/")

    def test_card_lists_the_viewers_copy_sizes_and_statuses(self):
        response = self.get(self.arnak)
        (entry,) = response.context["sleeve_copies"]
        rows = [(row["card_size"].name, row["status"]) for row in entry["rows"]]
        # Size-ordered: mini (41×63) before standard (63×88); the mini slot has
        # no status row yet so it reads not-sleeved.
        self.assertEqual(rows, [("Mini", "not_sleeved"), ("Standard", "to_sleeve")])
        self.assertContains(response, "Sleeves")
        self.assertContains(response, "To sleeve")
        self.assertContains(response, "195")  # standard card count

    def test_recorded_product_shows_read_only(self):
        response = self.get(self.roam)
        self.assertContains(response, "Tlama Diamond Green")
        # Read-only here: editing lives on the copy edit page.
        self.assertNotContains(response, "sleeve-status-select")

    def test_orphan_status_shows_unknown_count(self):
        SleeveRequirement.objects.filter(edition=self.roam.edition).delete()
        response = self.get(self.roam)
        self.assertContains(response, ">?</td>", html=False)

    def test_card_absent_without_owned_rows(self):
        # Pavel's copy of his own game — the viewer owns nothing here.
        foreign = Copy.objects.get(owner=self.other)
        response = self.get(foreign)
        self.assertEqual(response.context["sleeve_copies"], [])
        self.assertNotContains(response, 'data-bs-target="#sleeve-card"')


class CopyEditSleeveCardTests(SleevesPageMixin, TestCase):
    """Issue #17: the editable sleeve card on the copy edit page, and the
    scope=copy re-render that swaps just this copy's rows back in (rather than
    the whole-collection worklist)."""

    def get(self, copy=None):
        self.login()
        return self.client.get(f"/copies/{(copy or self.arnak).pk}/edit/")

    def post(self, data=None, copy=None, size=None):
        self.login()
        return self.client.post(
            f"/sleeves/copies/{(copy or self.arnak).pk}"
            f"/sizes/{(size or self.standard).pk}/",
            {"scope": "copy", **(data or {})},
        )

    def test_edit_page_renders_the_editable_table(self):
        response = self.get()
        self.assertContains(response, 'id="copy-sleeves"')
        self.assertContains(response, "sleeve-status-select")
        self.assertContains(response, '{"scope": "copy"}')
        self.assertContains(response, "195")

    def test_empty_state_when_edition_has_no_requirements(self):
        bare = self._copy("Bare Game")  # no requirements, no status rows
        response = self.get(bare)
        self.assertContains(response, "No sleeve requirements for this edition")
        self.assertNotContains(response, 'id="copy-sleeves"')

    def test_scope_copy_returns_just_this_copys_editable_rows(self):
        response = self.post({"status": "sleeved"}, size=self.mini)
        status = CopySleeveStatus.objects.get(copy=self.arnak, card_size=self.mini)
        self.assertEqual(status.status, CopySleeveStatus.Status.SLEEVED)
        # Copy-scoped re-render: no Game column, no other copy leaks in, and it
        # stays editable.
        self.assertNotContains(response, "<th>Game</th>")
        self.assertNotContains(response, "Roam")
        self.assertContains(response, "sleeve-status-select")


# ===========================================================================
# §13  Game detail page
# ===========================================================================

class GameDetailViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        group = cls.user.membership.group  # auto-created on signup (§3)
        lent = Location.objects.create(group=group, name="At Pavel's")

        cls.game = Game.objects.create(
            name="Epic Quest", year_published=2020,
            image_url="https://cf.geekdo-images.com/large/eq.jpg",
            min_players=1, max_players=4, min_playtime=60, max_playtime=120,
            is_campaign=True, has_scenarios=True,
            language_dependency=Game.LanguageDependency.EASY,
            language_dependency_note="easy (goals only, coop)",
            companion_app=Game.AppUse.OPTIONAL,
            player_conflict=0, soundtrack_ambience=True,
        )
        BggLink.objects.create(game=cls.game, bgg_id=12345, is_primary=True)
        ExternalLink.objects.create(
            game=cls.game, link_type=ExternalLink.LinkType.ZATROLENE,
            url="https://www.zatrolene-hry.cz/epic-quest/", label="Zatrolené",
        )
        DigitalImplementation.objects.create(
            game=cls.game, platform=DigitalImplementation.Platform.BGA,
            url="https://boardgamearena.com/epicquest",
        )
        theme = Tag.objects.create(kind=Tag.Kind.THEME, name="Fantasy")
        GameTag.objects.create(game=cls.game, tag=theme, is_favourite=True)
        GameType.objects.create(
            game=cls.game,
            game_type=GameType.Type.SOLO,
            qualifier=GameType.Qualifier.OPTIONAL,
        )

        edition = Edition.objects.create(game=cls.game, is_default=True)
        Copy.objects.create(
            owner=cls.user, edition=edition, location=lent,
            excitement="8.0", keep_status=Copy.KeepStatus.KEEP,
        )

        cls.expansion = Game.objects.create(
            name="Epic Quest: More Heroes", type=Game.Type.EXPANSION,
            players_min_override=5, players_max_override=6,
            playtime_delta_override=30,
        )
        cls.expansion.expands.add(cls.game)
        # Issue #16: the expansion is owned (has an active Copy), so it shows
        # in the base game's owned-expansions list.
        exp_edition = Edition.objects.create(game=cls.expansion, is_default=True)
        Copy.objects.create(owner=cls.user, edition=exp_edition)

    def get(self, pk=None):
        self.client.login(username="kernicek", password="pass")
        return self.client.get(f"/games/{pk or self.game.pk}/")

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_unknown_game_is_404(self):
        response = self.get(pk=99999)
        self.assertEqual(response.status_code, 404)

    def test_alternate_names_render_as_also_known_as(self):
        # Issue #51: curated alternate titles surface under the main title.
        self.assertNotContains(self.get(), "Also known as")
        AlternateName.objects.create(game=self.game, name="Heroická výprava")
        AlternateName.objects.create(game=self.game, name="Epische Suche")
        response = self.get()
        self.assertContains(response, "Also known as")
        self.assertContains(response, "Heroická výprava")
        self.assertContains(response, "Epische Suche")

    def test_hero_stats_and_taxonomy_chips(self):
        response = self.get()
        self.assertContains(response, "https://cf.geekdo-images.com/large/eq.jpg")
        # Effective players: the 5-6 expansion override widens 1-4 to 1-6.
        self.assertContains(response, "1–6")
        self.assertContains(response, "60–120 min")
        self.assertContains(response, "not yet synced")  # weight seam
        self.assertContains(response, "★ Fantasy")
        self.assertContains(response, "Solo (optional)")
        self.assertContains(response, "Campaign")
        self.assertContains(response, "Scenarios")
        self.assertContains(response, "easy (goals only, coop)")
        self.assertContains(response, "0 / 3")

    def test_will_leave_copy_shows_a_distinct_leaving_badge(self):
        # Issue #82: WILL_LEAVE gets its own badge, not the generic keep-
        # status one other values still use.
        copy = Copy.objects.get(edition__game=self.game)
        self.assertNotContains(self.get(), "Leaving")
        copy.keep_status = Copy.KeepStatus.WILL_LEAVE
        copy.save(update_fields=["keep_status"])
        response = self.get()
        self.assertContains(
            response, '<span class="badge text-bg-warning">Leaving</span>',
        )

    def test_bgg_rank_links_to_the_browse_page(self):
        # Issue #39: the rank number links to BGG's rank-browse page anchored
        # at that rank.
        self.game.bgg_rating = "7.500"
        self.game.bgg_rank = 1132
        self.game.save()
        response = self.get()
        self.assertContains(
            response,
            "https://boardgamegeek.com/browse/boardgame?sort=rank"
            "&amp;rankobjecttype=subtype&amp;rankobjectid=1"
            "&amp;rank=1132#1132",
        )

    def test_missing_bgg_rank_renders_no_link(self):
        # The setUpTestData game has no bgg_rank, so no browse link appears.
        self.assertNotContains(self.get(), "browse/boardgame?sort=rank")

    def test_copies_links_and_expansions(self):
        response = self.get()
        self.assertContains(response, "At Pavel&#x27;s")
        self.assertContains(response, "8.0")
        self.assertContains(response, "https://boardgamegeek.com/boardgame/12345")
        self.assertContains(response, "https://www.zatrolene-hry.cz/epic-quest/")
        self.assertContains(response, "https://boardgamearena.com/epicquest")
        # Issue #16: owned expansions listed with a link to their page, an
        # override hint and the owner. Issue #98: the base page implies the
        # base game, so the expansion shows its short (prefix-stripped) name.
        self.assertContains(response, "Expansions")
        self.assertContains(response, "More Heroes")
        self.assertNotContains(response, "Epic Quest: More Heroes")
        self.assertContains(response, f"/games/{self.expansion.pk}/")
        self.assertContains(response, "5–6 players, +30 min")

    def test_unowned_expansion_listed_and_distinct(self):
        # Issue #47: an expansion with no active Copy still appears — in muted
        # italic with a "not owned" marker — as a what-could-be-added overview,
        # while the owned expansion keeps its owner badge.
        orphan = Game.objects.create(
            name="Epic Quest: Phantom Pack", type=Game.Type.EXPANSION,
        )
        orphan.expands.add(self.game)
        response = self.get()
        # Issue #98: short (prefix-stripped) names in the base's expansion list.
        self.assertContains(response, "Phantom Pack")
        self.assertNotContains(response, "Epic Quest: Phantom Pack")
        self.assertContains(response, "fst-italic")
        self.assertContains(response, "not owned")
        # The owned expansion still lists its owner.
        self.assertContains(response, "More Heroes")
        self.assertContains(response, "kernicek")

    def test_expansion_links_back_to_its_base(self):
        response = self.get(pk=self.expansion.pk)
        self.assertContains(response, "Expansion")
        self.assertContains(response, f"/games/{self.game.pk}/")

    def test_grid_tiles_link_to_detail(self):
        self.client.login(username="kernicek", password="pass")
        # Epic Quest's only copy is lent out; show_unavailable opts past the
        # available-only default (issue #107) so the tile is on the grid.
        response = self.client.get("/", {"show_unavailable": "1"})
        self.assertContains(response, f'href="/games/{self.game.pk}/"')

    def test_previously_owned_mark_shows_on_detail_grid_and_curation(self):
        self.game.bgg_collection_status = Game.BggCollectionStatus.PREV_OWNED
        self.game.save(update_fields=["bgg_collection_status"])

        response = self.get()
        self.assertContains(response, "Previously owned on BGG")
        # Grid marker is now an icon flag (issue #8) with a title tooltip.
        response = self.client.get("/", {"show_unavailable": "1"})
        self.assertContains(response, 'title="Previously owned"')
        response = self.client.get("/curation/")
        self.assertContains(response, "prev. owned on BGG")

    def test_incoming_preorder_mark_shows_on_the_grid(self):
        """Issue #8: an unconverted purchase Product flags the game incoming
        on the collection grid; a converted one (copy set) does not."""
        self.client.login(username="kernicek", password="pass")
        purchase = Purchase.objects.create(
            owner=self.user, name="Epic Quest KS",
            status=Purchase.Status.COMMITTED,
        )
        wave = Wave.objects.create(
            purchase=purchase, number=1, status=Wave.Status.PRODUCTION,
        )
        product = Product.objects.create(
            wave=wave, name="Epic Quest core", kind=Product.Kind.GAME,
            game=self.game,
        )
        grid = {"show_unavailable": "1"}
        self.assertContains(
            self.client.get("/", grid), 'title="Incoming preorder"')

        # Once converted (copy linked) the game is no longer incoming.
        product.copy = Copy.objects.get(edition__game=self.game)
        product.save(update_fields=["copy"])
        self.assertNotContains(
            self.client.get("/", grid), 'title="Incoming preorder"')

    def test_incoming_mark_is_owner_scoped(self):
        """Another member's unconverted product must not flag my grid."""
        other = get_user_model().objects.create_user(
            username="other2", password="pass",
        )
        purchase = Purchase.objects.create(owner=other, name="Foreign Pledge")
        wave = Wave.objects.create(purchase=purchase, number=1)
        Product.objects.create(wave=wave, name="core", game=self.game)

        self.client.login(username="kernicek", password="pass")
        # show_unavailable keeps the (lent-out) game on the grid, so the test
        # proves the foreign product doesn't flag it — not that it's just hidden.
        self.assertNotContains(
            self.client.get("/", {"show_unavailable": "1"}),
            'title="Incoming preorder"')

    # --- §6 purchase backlink -------------------------------------------

    def test_purchases_containing_the_game_are_linked(self):
        purchase = Purchase.objects.create(
            owner=self.user, name="Epic Quest KS",
            platform=Purchase.Platform.KICKSTARTER,
            status=Purchase.Status.COMMITTED,
        )
        wave = Wave.objects.create(
            purchase=purchase, number=1, status=Wave.Status.PRODUCTION,
            expected_arrival=datetime.date(2026, 9, 1),
        )
        Product.objects.create(
            wave=wave, name="Epic Quest core", kind=Product.Kind.GAME,
            game=self.game,
        )

        response = self.get()
        self.assertContains(response, 'Purchases <span')
        self.assertContains(response, f'href="/purchases/{purchase.pk}/"')
        self.assertContains(response, "Epic Quest KS")
        self.assertContains(response, "Committed")
        self.assertContains(response, "Wave 1 — production")
        self.assertContains(response, "ETA Sep 2026")

    def test_purchase_backlink_is_owner_scoped(self):
        other = get_user_model().objects.create_user(
            username="other", password="pass",
        )
        purchase = Purchase.objects.create(owner=other, name="Foreign Pledge")
        wave = Wave.objects.create(purchase=purchase, number=1)
        Product.objects.create(wave=wave, name="Epic Quest core", game=self.game)

        response = self.get()
        self.assertNotContains(response, "Foreign Pledge")
        self.assertNotContains(response, 'Purchases <span')

    def test_one_row_per_purchase_preferring_the_incoming_wave(self):
        purchase = Purchase.objects.create(
            owner=self.user, name="Epic Quest KS",
            status=Purchase.Status.COMMITTED,
        )
        arrived = Wave.objects.create(
            purchase=purchase, number=1, status=Wave.Status.ARRIVED,
            arrived_date=datetime.date(2026, 1, 5),
        )
        incoming = Wave.objects.create(
            purchase=purchase, number=2, status=Wave.Status.PENDING,
            expected_arrival=datetime.date(2026, 12, 1),
        )
        Product.objects.create(
            wave=arrived, name="Epic Quest core", game=self.game,
            copy=Copy.objects.get(edition__game=self.game),
        )
        Product.objects.create(
            wave=incoming, name="Epic Quest extras", game=self.game,
        )

        response = self.get()
        self.assertContains(response, "Epic Quest KS", count=1)
        self.assertContains(response, "Wave 2 — pending")
        self.assertContains(response, "ETA Dec 2026")
        self.assertContains(response, "in collection")

    def test_no_purchases_card_when_nothing_references_the_game(self):
        response = self.get()
        self.assertNotContains(response, 'Purchases <span')

    def test_admin_edit_link_hidden_from_regular_user(self):
        # Issue #52: the admin change link is superuser-only.
        admin_url = reverse("admin:gamekeeper_game_change", args=[self.game.pk])
        response = self.get()
        self.assertNotContains(response, admin_url)

    def test_admin_edit_link_shown_to_superuser(self):
        # Issue #52: a superuser gets a one-click link to the admin change page.
        get_user_model().objects.create_superuser(
            username="root", password="pass",
        )
        self.client.login(username="root", password="pass")
        admin_url = reverse("admin:gamekeeper_game_change", args=[self.game.pk])
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertContains(response, admin_url)

    def test_plays_card_lists_recent_plays_with_winner(self):
        """Issue #65: the read-only plays log card on the detail page."""
        play = Play.objects.create(
            game=self.game, external_id="500",
            play_date=datetime.date(2024, 6, 1), location="Home",
            synced_at=timezone.now(),
        )
        PlayPlayer.objects.create(play=play, name="Vojta", score="42", won=True)

        response = self.get()
        self.assertEqual(len(response.context["recent_plays"]), 1)
        self.assertEqual(response.context["plays_count"], 1)
        self.assertContains(response, "Vojta")
        self.assertContains(response, "Home")


class PlaysFeedViewTests(TestCase):
    """The read-only /plays/ feed (issue #65, DESIGN §8)."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.dungeon = Game.objects.create(name="5-Minute Dungeon")
        cls.voidfall = Game.objects.create(name="Voidfall")
        now = timezone.now()
        cls.dungeon_play = Play.objects.create(
            game=cls.dungeon, external_id="1",
            play_date=datetime.date(2024, 6, 1), synced_at=now,
        )
        cls.voidfall_play = Play.objects.create(
            game=cls.voidfall, external_id="2",
            play_date=datetime.date(2024, 5, 1), synced_at=now,
        )

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/plays/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_feed_lists_all_games_plays(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/plays/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["plays"]), 2)
        self.assertContains(response, "5-Minute Dungeon")
        self.assertContains(response, "Voidfall")

    def test_game_filter_scopes_the_feed(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/plays/?game={self.dungeon.pk}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["game"], self.dungeon)
        plays = response.context["plays"]
        self.assertEqual([p.pk for p in plays], [self.dungeon_play.pk])

    def test_unknown_game_filter_is_404(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get("/plays/?game=99999")
        self.assertEqual(response.status_code, 404)


# ===========================================================================
# §3  Signup auto-group & the anonymous share link
# ===========================================================================

class AutoGroupSignalTests(TestCase):
    def test_new_user_gets_a_group_of_one_as_owner(self):
        user = get_user_model().objects.create_user(username="nova", password="pass")
        membership = user.membership
        self.assertEqual(membership.role, Membership.Role.OWNER)
        self.assertEqual(membership.group.name, "nova")
        self.assertEqual(membership.group.slug, "nova")

    def test_colliding_slug_is_suffixed(self):
        # E.g. a deleted user's orphaned group still holds the slug.
        Group.objects.create(name="old nova household", slug="nova")
        user = get_user_model().objects.create_user(username="nova", password="pass")
        self.assertEqual(user.membership.group.slug, "nova-2")

    def test_resaving_a_user_does_not_create_a_second_group(self):
        user = get_user_model().objects.create_user(username="nova", password="pass")
        user.first_name = "Nova"
        user.save()
        self.assertEqual(Group.objects.count(), 1)
        self.assertEqual(Membership.objects.count(), 1)

    def test_share_token_minting_is_idempotent(self):
        user = get_user_model().objects.create_user(username="nova", password="pass")
        group = user.membership.group
        token = group.enable_share_link()
        self.assertTrue(token)
        self.assertEqual(group.enable_share_link(), token)

    def test_location_share_token_minting_is_idempotent(self):
        # Issue #123: per-location share token mirrors the group's.
        user = get_user_model().objects.create_user(username="nova", password="pass")
        location = Location.objects.create(group=user.membership.group, name="Shelf")
        token = location.enable_share_link()
        self.assertTrue(token)
        self.assertEqual(location.enable_share_link(), token)


class ShareLinkViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.group = cls.user.membership.group  # auto-created on signup (§3)
        cls.token = cls.group.enable_share_link()
        shelf = Location.objects.create(group=cls.group, name="Secret Shelf")

        cls.owned = Game.objects.create(
            name="Epic Quest", year_published=2020,
            image_url="https://cf.geekdo-images.com/large/eq.jpg",
            min_players=2, max_players=4, min_playtime=60, max_playtime=120,
            weight=Decimal("3.5"), bgg_rating=Decimal("7.8"),
        )
        BggLink.objects.create(game=cls.owned, bgg_id=12345, is_primary=True)
        GameTag.objects.create(
            game=cls.owned,
            tag=Tag.objects.create(kind=Tag.Kind.MECHANIC, name="Deck Building"),
        )
        GameTag.objects.create(
            game=cls.owned,
            tag=Tag.objects.create(kind=Tag.Kind.THEME, name="Fantasy"),
        )
        cls._copy(cls.owned, cls.user, location=shelf,
                  excitement="8.5", keep_status=Copy.KeepStatus.KEEP)

        # An owned expansion is in the shared set (detail reachable) but the
        # grid stays base-games-only.
        cls.expansion = Game.objects.create(
            name="Epic Quest: More Heroes", type=Game.Type.EXPANSION,
        )
        cls.expansion.expands.add(cls.owned)
        cls._copy(cls.expansion, cls.user)

        # Issue #121: a known-but-unowned expansion — listed as "not owned",
        # never linked (its detail page is outside the shared set, 404s).
        cls.unowned_expansion = Game.objects.create(
            name="Epic Quest: Forgotten Realm", type=Game.Type.EXPANSION,
        )
        cls.unowned_expansion.expands.add(cls.owned)

        # Issue #121: Kickstarter is in the shared safe set; Dropbox is
        # personal file storage and must stay hidden.
        ExternalLink.objects.create(
            game=cls.owned, link_type=ExternalLink.LinkType.KICKSTARTER,
            url="https://www.kickstarter.com/projects/epic-quest",
        )
        ExternalLink.objects.create(
            game=cls.owned, link_type=ExternalLink.LinkType.DROPBOX,
            url="https://dropbox.com/personal-scans",
        )

        # Issue #121: a shareable document (rulebook link).
        cls.document = Document.objects.create(
            content_object=cls.owned, doc_type=Document.Type.RULEBOOK,
            label="Rulebook", external_url="https://example.com/rules.pdf",
        )

        # Outside the projection: no active copy (preorder / archived-only)
        # or another group's collection.
        cls.preorder = Game.objects.create(name="Preorder Mystery")
        cls.culled = Game.objects.create(name="Culled Relic")
        cls._copy(cls.culled, cls.user,
                  archive_status=Copy.ArchiveStatus.ARCHIVED)
        cls.stranger = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.foreign = Game.objects.create(name="Pavel Private Game")
        cls._copy(cls.foreign, cls.stranger)

    @classmethod
    def _copy(cls, game, owner, **fields):
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(owner=owner, edition=edition, **fields)

    def test_grid_is_anonymous_and_shows_only_active_base_games(self):
        response = self.client.get(f"/share/{self.token}/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "kernicek — collection")
        self.assertContains(response, "Epic Quest")
        self.assertContains(response, "https://cf.geekdo-images.com/large/eq.jpg")
        self.assertContains(
            response, f'href="/share/{self.token}/games/{self.owned.pk}/"',
        )
        self.assertContains(response, "1 game")
        # Not in the projection: preorders (no Copy), archived-only games,
        # other groups' games, and expansions as their own tiles.
        self.assertNotContains(response, "Preorder Mystery")
        self.assertNotContains(response, "Culled Relic")
        self.assertNotContains(response, "Pavel Private Game")
        self.assertNotContains(response, "More Heroes")

    def test_shared_grid_renders_the_chooser_without_owner_only_axes(self):
        # Issue #120: viewers get the GameChooser, minus the axes §3 hides.
        response = self.client.get(f"/share/{self.token}/")
        self.assertContains(response, 'id="chooser"')
        self.assertContains(response, 'name="players"')
        self.assertContains(response, 'name="theme"')
        # Owner-only: no location select, no availability toggle.
        self.assertNotContains(response, 'name="location"')
        self.assertNotContains(response, 'name="show_unavailable"')
        self.assertNotContains(response, "Show unavailable")

    def test_shared_grid_filters_by_players(self):
        # Epic Quest seats 2–4: 3 keeps it, 9 drops it.
        self.assertContains(
            self.client.get(f"/share/{self.token}/?players=3"), "Epic Quest")
        self.assertNotContains(
            self.client.get(f"/share/{self.token}/?players=9"), "Epic Quest")

    def test_shared_grid_ignores_availability(self):
        # Availability is owner-only (§3): a game whose only copy is lent out
        # still shows — there is no availability axis on the shared grid.
        game = Game.objects.create(name="Loaned Legends")
        copy = self._copy(game, self.user)
        Loan.objects.create(
            copy=copy, direction=Loan.Direction.LENT_OUT, counterparty_name="Mira",
        )
        self.assertContains(
            self.client.get(f"/share/{self.token}/"), "Loaned Legends")

    def test_shared_grid_collapses_series_by_default(self):
        # Issue #107 collapse default carries over to the shared grid (#120),
        # though a shared tile has no series page — it only expands.
        members = []
        for name in ("Quest Saga: Alpha", "Quest Saga: Beta"):
            member = Game.objects.create(name=name)
            self._copy(member, self.user)
            members.append(member)
        series = Series.objects.create(name="Quest Saga", primary_game=members[0])
        Game.objects.filter(pk__in=[m.pk for m in members]).update(series=series)

        collapsed = self.client.get(f"/share/{self.token}/")
        self.assertEqual(
            [t["series"].name for t in collapsed.context["tiles"] if t.get("series")],
            ["Quest Saga"],
        )
        opened = self.client.get(f"/share/{self.token}/?show_all_editions=1")
        self.assertFalse(any(t.get("series") for t in opened.context["tiles"]))
        self.assertContains(opened, "Quest Saga: Alpha")

    def test_shared_grid_dropdowns_list_only_the_shared_vocabulary(self):
        # Issue #120: filter options never leak the owner's private vocabulary —
        # only themes actually used by shared games appear.
        Tag.objects.create(kind=Tag.Kind.THEME, name="Cthulhu Secrets")
        response = self.client.get(f"/share/{self.token}/")
        self.assertContains(response, "Fantasy")  # the shared Epic Quest uses it
        self.assertNotContains(response, "Cthulhu Secrets")  # unused → private

    def test_shared_htmx_returns_just_the_grid(self):
        response = self.client.get(
            f"/share/{self.token}/", HTTP_HX_REQUEST="true",
        )
        self.assertContains(response, "of 1 game")
        self.assertNotContains(response, "<form")

    def test_shared_list_view_shows_public_stats(self):
        # Issue #92: the list view's public columns match the safe-field-set
        # already exercised by the detail page (DESIGN §3).
        response = self.client.get(f"/share/{self.token}/", {"view": "list"})
        self.assertContains(response, "2020")
        self.assertContains(response, "2–4")
        self.assertContains(response, "60–120 min")
        self.assertContains(response, "3.50 / 5")
        self.assertNotContains(response, "More Heroes")

    def test_shared_list_view_hides_location_and_keep_status(self):
        # DESIGN §3 explicitly hides storage location and keep-status from
        # the anonymous share projection; copy count follows the same rule.
        response = self.client.get(f"/share/{self.token}/", {"view": "list"})
        self.assertNotContains(response, "Secret Shelf")
        self.assertNotContains(response, "Location")
        self.assertNotContains(response, "Keep/leave")
        self.assertNotContains(response, "Copies")

    def test_shared_htmx_list_returns_just_the_results(self):
        response = self.client.get(
            f"/share/{self.token}/", {"view": "list"}, HTTP_HX_REQUEST="true",
        )
        self.assertContains(response, "<table")
        self.assertContains(response, "of 1 game")
        self.assertNotContains(response, "<form")

    def test_wrong_token_is_404(self):
        response = self.client.get("/share/not-the-token/")
        self.assertEqual(response.status_code, 404)
        response = self.client.get(
            f"/share/not-the-token/games/{self.owned.pk}/",
        )
        self.assertEqual(response.status_code, 404)

    def test_detail_shows_the_safe_field_set_only(self):
        response = self.client.get(
            f"/share/{self.token}/games/{self.owned.pk}/",
        )
        self.assertContains(response, "Epic Quest")
        self.assertContains(response, "2–4")
        self.assertContains(response, "60–120 min")
        self.assertContains(response, "3.50 / 5")
        self.assertContains(response, "7.800")
        self.assertContains(response, "https://boardgamegeek.com/boardgame/12345")
        self.assertContains(response, "Deck Building")  # mechanics: safe set
        # Hidden: personal signals and the owner's own theme vocabulary.
        self.assertNotContains(response, "Fantasy")
        self.assertNotContains(response, "8.5")
        self.assertNotContains(response, "Secret Shelf")
        # bare "Keep" collides with the "GameKeeper" brand text in every
        # page's navbar, so check the actual keep-status badge markup instead.
        self.assertNotContains(response, 'badge text-bg-secondary">Keep<')

    def test_shared_detail_bgg_rank_links_to_the_browse_page(self):
        # Issue #39: the public share page links the rank to BGG's browse page.
        self.owned.bgg_rank = 517
        self.owned.save()
        response = self.client.get(
            f"/share/{self.token}/games/{self.owned.pk}/",
        )
        self.assertContains(
            response,
            "https://boardgamegeek.com/browse/boardgame?sort=rank"
            "&amp;rankobjecttype=subtype&amp;rankobjectid=1"
            "&amp;rank=517#517",
        )

    def test_detail_404s_outside_the_shared_set(self):
        for game in (self.preorder, self.culled, self.foreign):
            response = self.client.get(
                f"/share/{self.token}/games/{game.pk}/",
            )
            self.assertEqual(response.status_code, 404, game.name)

    def test_detail_marks_expansions_owned_and_unowned(self):
        # Issue #121: the owned expansion links to its own detail page; the
        # known-but-unowned one is labelled "not owned" and never linked
        # (its detail page is outside the shared set and would 404).
        response = self.client.get(
            f"/share/{self.token}/games/{self.owned.pk}/",
        )
        self.assertContains(response, "More Heroes")
        self.assertContains(
            response, f'href="/share/{self.token}/games/{self.expansion.pk}/"',
        )
        self.assertContains(response, "Forgotten Realm")
        self.assertContains(response, "not owned")
        self.assertNotContains(
            response,
            f'href="/share/{self.token}/games/{self.unowned_expansion.pk}/"',
        )

    def test_detail_shows_shareable_external_links_only(self):
        # Issue #121: Kickstarter is public game info; Dropbox is personal
        # file storage and must stay hidden, per DESIGN §3.
        response = self.client.get(
            f"/share/{self.token}/games/{self.owned.pk}/",
        )
        self.assertContains(
            response, "https://www.kickstarter.com/projects/epic-quest",
        )
        self.assertNotContains(response, "dropbox.com")

    def test_detail_shows_documents(self):
        # Issue #121: the game's §7 documents.
        response = self.client.get(
            f"/share/{self.token}/games/{self.owned.pk}/",
        )
        self.assertContains(response, "Rulebook")
        self.assertContains(response, "https://example.com/rules.pdf")

    def test_expansion_detail_cross_links_stay_inside_the_share(self):
        response = self.client.get(
            f"/share/{self.token}/games/{self.owned.pk}/",
        )
        self.assertContains(
            response, f"/share/{self.token}/games/{self.expansion.pk}/",
        )
        response = self.client.get(
            f"/share/{self.token}/games/{self.expansion.pk}/",
        )
        self.assertContains(response, "Expansion")
        self.assertContains(
            response, f"/share/{self.token}/games/{self.owned.pk}/",
        )

    def test_group_without_token_has_no_share_page(self):
        # Pavel's auto-group never enabled a link; his slug is not a token.
        response = self.client.get("/share/pavel/")
        self.assertEqual(response.status_code, 404)


class LocationShareLinkViewTests(TestCase):
    """Issue #123: a share link pinned to one Location — same curated
    projection as the group-level tier-4 link, further narrowed to games
    with an active Copy at that one Location. Never exposes the group's
    other locations."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.group = cls.user.membership.group  # auto-created on signup (§3)
        cls.shelf = Location.objects.create(group=cls.group, name="Secret Shelf")
        cls.other_shelf = Location.objects.create(group=cls.group, name="Attic Box")
        cls.token = cls.shelf.enable_share_link()

        cls.on_shelf = Game.objects.create(name="Epic Quest")
        cls._copy(cls.on_shelf, cls.user, location=cls.shelf)

        cls.elsewhere = Game.objects.create(name="Attic Relic")
        cls._copy(cls.elsewhere, cls.user, location=cls.other_shelf)

        cls.unplaced = Game.objects.create(name="Unplaced Wanderer")
        cls._copy(cls.unplaced, cls.user)

    @classmethod
    def _copy(cls, game, owner, **fields):
        edition = Edition.objects.create(game=game, is_default=True)
        return Copy.objects.create(owner=owner, edition=edition, **fields)

    def test_grid_shows_only_games_at_the_pinned_location(self):
        response = self.client.get(f"/share/location/{self.token}/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "kernicek — Secret Shelf")
        self.assertContains(response, "Epic Quest")
        self.assertNotContains(response, "Attic Relic")
        self.assertNotContains(response, "Unplaced Wanderer")

    def test_grid_hides_the_location_chooser_and_full_list(self):
        response = self.client.get(f"/share/location/{self.token}/")
        self.assertNotContains(response, 'name="location"')
        self.assertNotContains(response, "Attic Box")

    def test_detail_404s_for_a_game_at_a_different_location(self):
        response = self.client.get(
            f"/share/location/{self.token}/games/{self.elsewhere.pk}/",
        )
        self.assertEqual(response.status_code, 404)

    def test_detail_shows_the_pinned_game(self):
        response = self.client.get(
            f"/share/location/{self.token}/games/{self.on_shelf.pk}/",
        )
        self.assertContains(response, "Epic Quest")

    def test_unknown_token_is_404(self):
        response = self.client.get("/share/location/not-a-token/")
        self.assertEqual(response.status_code, 404)
        response = self.client.get(
            f"/share/location/not-a-token/games/{self.on_shelf.pk}/",
        )
        self.assertEqual(response.status_code, 404)

    def test_works_for_a_logged_in_visitor_too(self):
        stranger = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        self.client.login(username="pavel", password="pass")
        response = self.client.get(f"/share/location/{self.token}/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Epic Quest")

    def test_grid_still_filters_by_players(self):
        self.on_shelf.min_players, self.on_shelf.max_players = 2, 4
        self.on_shelf.save()
        self.assertContains(
            self.client.get(f"/share/location/{self.token}/?players=3"),
            "Epic Quest",
        )
        self.assertNotContains(
            self.client.get(f"/share/location/{self.token}/?players=9"),
            "Epic Quest",
        )


class GroupCollectionViewTests(TestCase):
    """DESIGN §3 tiers 2+3: the logged-in /g/<slug>/ viewer surface. Access
    is modal on Group.visibility — private admits members only, shared admits
    ShareGrant targets, server_public admits any authenticated user. No
    access renders as 404 (existence not probeable), anonymous hits the
    login wall."""

    @classmethod
    def setUpTestData(cls):
        make_user = get_user_model().objects.create_user
        cls.owner = make_user(username="kernicek", password="pass")
        cls.group = cls.owner.membership.group  # auto-created on signup (§3)
        cls.viewer = make_user(username="mira", password="pass")
        cls.outsider = make_user(username="pavel", password="pass")

        cls.game = Game.objects.create(name="Epic Quest")
        edition = Edition.objects.create(game=cls.game, is_default=True)
        Copy.objects.create(
            owner=cls.owner, edition=edition,
            excitement="8.5", keep_status=Copy.KeepStatus.KEEP,
            location=Location.objects.create(group=cls.group, name="Secret Shelf"),
        )

    def grant_to(self, **grantee):
        return ShareGrant.objects.create(group=self.group, **grantee)

    def test_anonymous_is_sent_to_login(self):
        response = self.client.get(f"/g/{self.group.slug}/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_member_browses_their_own_group_at_any_tier(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/g/{self.group.slug}/")
        self.assertContains(response, "Epic Quest")
        self.assertContains(
            response, f'href="/g/{self.group.slug}/games/{self.game.pk}/"',
        )

    def test_list_view_works_and_hides_owner_only_columns(self):
        # Issue #92: ?view=list works on the /g/<slug>/ surface too, and
        # still hides the owner-only columns (owner_view is False here too).
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/g/{self.group.slug}/", {"view": "list"})
        self.assertContains(response, "<table")
        self.assertContains(response, "Epic Quest")
        self.assertNotContains(response, "Secret Shelf")
        self.assertNotContains(response, "Location")

    def test_private_group_is_404_for_other_users(self):
        self.client.login(username="mira", password="pass")
        response = self.client.get(f"/g/{self.group.slug}/")
        self.assertEqual(response.status_code, 404)

    def test_grants_are_inert_while_the_group_is_private(self):
        self.grant_to(grantee_user=self.viewer)
        self.client.login(username="mira", password="pass")
        response = self.client.get(f"/g/{self.group.slug}/")
        self.assertEqual(response.status_code, 404)

    def test_shared_admits_the_granted_user_only(self):
        self.group.visibility = Group.Visibility.SHARED
        self.group.save()
        self.grant_to(grantee_user=self.viewer)

        self.client.login(username="mira", password="pass")
        self.assertContains(self.client.get(f"/g/{self.group.slug}/"), "Epic Quest")

        self.client.login(username="pavel", password="pass")
        response = self.client.get(f"/g/{self.group.slug}/")
        self.assertEqual(response.status_code, 404)

    def test_shared_admits_members_of_a_granted_group(self):
        self.group.visibility = Group.Visibility.SHARED
        self.group.save()
        self.grant_to(grantee_group=self.viewer.membership.group)
        self.client.login(username="mira", password="pass")
        self.assertContains(self.client.get(f"/g/{self.group.slug}/"), "Epic Quest")

    def test_server_public_admits_any_logged_in_user(self):
        self.group.visibility = Group.Visibility.SERVER_PUBLIC
        self.group.save()
        self.client.login(username="pavel", password="pass")
        self.assertContains(self.client.get(f"/g/{self.group.slug}/"), "Epic Quest")

    def test_viewer_detail_keeps_the_curated_projection(self):
        # Tiers 2+3 reuse the tier-4 safe field set: a viewer is not in the
        # group, so personal signals (§3) stay hidden even though they are
        # logged in.
        self.group.visibility = Group.Visibility.SERVER_PUBLIC
        self.group.save()
        self.client.login(username="mira", password="pass")
        response = self.client.get(f"/g/{self.group.slug}/games/{self.game.pk}/")
        self.assertContains(response, "Epic Quest")
        self.assertNotContains(response, "8.5")
        self.assertNotContains(response, "Secret Shelf")
        # bare "Keep" collides with the "GameKeeper" brand text in every
        # page's navbar, so check the actual keep-status badge markup instead.
        self.assertNotContains(response, 'badge text-bg-secondary">Keep<')

    def test_viewer_detail_404s_outside_the_shared_set(self):
        self.group.visibility = Group.Visibility.SERVER_PUBLIC
        self.group.save()
        foreign = Game.objects.create(name="Pavel Private Game")
        Copy.objects.create(
            owner=self.outsider,
            edition=Edition.objects.create(game=foreign, is_default=True),
        )
        self.client.login(username="mira", password="pass")
        response = self.client.get(f"/g/{self.group.slug}/games/{foreign.pk}/")
        self.assertEqual(response.status_code, 404)

    def test_grant_needs_exactly_one_grantee(self):
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                ShareGrant.objects.create(group=self.group)
        with self.assertRaises(IntegrityError):
            ShareGrant.objects.create(
                group=self.group, grantee_user=self.viewer,
                grantee_group=self.viewer.membership.group,
            )

    def test_duplicate_grants_are_rejected(self):
        self.grant_to(grantee_user=self.viewer)
        with self.assertRaises(IntegrityError):
            self.grant_to(grantee_user=self.viewer)


class GroupSettingsViewTests(TestCase):
    """DESIGN §3 owner-facing sharing settings at /g/<slug>/settings/:
    visibility tier, tier-2 grants and the tier-4 anonymous link, managed
    without the admin. Owner-only; everyone else gets the viewer gate's 404
    stance."""

    @classmethod
    def setUpTestData(cls):
        make_user = get_user_model().objects.create_user
        cls.owner = make_user(username="kernicek", password="pass")
        cls.group = cls.owner.membership.group
        cls.outsider = make_user(username="pavel", password="pass")
        # A plain member of the owner's group (household member, not owner).
        cls.member = make_user(username="mira", password="pass")
        cls.member.membership.group = cls.group
        cls.member.membership.role = Membership.Role.MEMBER
        cls.member.membership.save()

    def url(self, tail=""):
        return f"/g/{self.group.slug}/settings/{tail}"

    def login_owner(self):
        self.client.login(username="kernicek", password="pass")

    def test_anonymous_is_sent_to_login(self):
        response = self.client.get(self.url())
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_only_the_owner_gets_in(self):
        for username in ("pavel", "mira"):
            self.client.login(username=username, password="pass")
            self.assertEqual(self.client.get(self.url()).status_code, 404)

    def test_settings_page_shows_current_state(self):
        self.login_owner()
        response = self.client.get(self.url())
        self.assertContains(response, "Who can see the collection")
        # Exactly one radio is checked: the group's current (private) tier.
        self.assertContains(response, "checked", count=1)
        self.assertContains(response, 'id="vis-private"')
        self.assertContains(response, "No grants yet.")
        self.assertContains(response, "Create share link")
        self.assertContains(response, f"/g/{self.group.slug}/")

    def test_owner_sees_settings_shortcut_on_the_viewer_page(self):
        self.login_owner()
        response = self.client.get(f"/g/{self.group.slug}/")
        self.assertContains(response, self.url())

        # A tier-3 visitor does not.
        self.group.visibility = Group.Visibility.SERVER_PUBLIC
        self.group.save()
        self.client.login(username="pavel", password="pass")
        response = self.client.get(f"/g/{self.group.slug}/")
        self.assertNotContains(response, self.url())

    def test_visibility_updates_and_rejects_unknown_tiers(self):
        self.login_owner()
        response = self.client.post(
            self.url("visibility/"), {"visibility": "server_public"},
        )
        self.group.refresh_from_db()
        self.assertEqual(self.group.visibility, Group.Visibility.SERVER_PUBLIC)
        self.assertContains(response, "Who can see the collection")

        response = self.client.post(
            self.url("visibility/"), {"visibility": "everyone"},
        )
        self.assertEqual(response.status_code, 400)
        self.group.refresh_from_db()
        self.assertEqual(self.group.visibility, Group.Visibility.SERVER_PUBLIC)

    def test_share_link_enable_and_revoke(self):
        self.login_owner()
        response = self.client.post(self.url("share-link/"), {"action": "enable"})
        self.group.refresh_from_db()
        self.assertTrue(self.group.share_token)
        self.assertContains(response, f"/share/{self.group.share_token}/")
        self.assertContains(
            response,
            f'data-copy-value="http://testserver/share/{self.group.share_token}/"',
        )

        response = self.client.post(self.url("share-link/"), {"action": "revoke"})
        self.group.refresh_from_db()
        self.assertIsNone(self.group.share_token)
        self.assertContains(response, "Create share link")

        response = self.client.post(self.url("share-link/"), {"action": "explode"})
        self.assertEqual(response.status_code, 400)

    def test_location_share_link_enable_and_revoke(self):
        # Issue #123: per-location share link, same enable/revoke shape.
        location = Location.objects.create(group=self.group, name="Shelf")
        self.login_owner()
        tail = f"locations/{location.pk}/share-link/"

        response = self.client.post(self.url(tail), {"action": "enable"})
        location.refresh_from_db()
        self.assertTrue(location.share_token)
        self.assertContains(response, f"/share/location/{location.share_token}/")
        self.assertContains(
            response,
            f'data-copy-value="http://testserver/share/location/{location.share_token}/"',
        )

        response = self.client.post(self.url(tail), {"action": "revoke"})
        location.refresh_from_db()
        self.assertIsNone(location.share_token)
        self.assertContains(response, "Create share link")

        response = self.client.post(self.url(tail), {"action": "explode"})
        self.assertEqual(response.status_code, 400)

    def test_location_share_link_404s_for_another_groups_location(self):
        foreign = Location.objects.create(
            group=self.outsider.membership.group, name="Their Shelf",
        )
        self.login_owner()
        response = self.client.post(
            self.url(f"locations/{foreign.pk}/share-link/"), {"action": "enable"},
        )
        self.assertEqual(response.status_code, 404)
        foreign.refresh_from_db()
        self.assertIsNone(foreign.share_token)

    def test_grant_to_a_user(self):
        self.login_owner()
        response = self.client.post(
            self.url("grants/"), {"grantee_type": "user", "grantee": "pavel"},
        )
        grant = ShareGrant.objects.get()
        self.assertEqual(grant.grantee_user, self.outsider)
        self.assertIsNone(grant.grantee_group)
        self.assertContains(response, "pavel")

    def test_grant_to_a_group(self):
        self.login_owner()
        target = self.outsider.membership.group
        response = self.client.post(
            self.url("grants/"), {"grantee_type": "group", "grantee": target.slug},
        )
        grant = ShareGrant.objects.get()
        self.assertEqual(grant.grantee_group, target)
        self.assertContains(response, f'group "{target.name}"')

    def test_grant_validation_messages(self):
        self.login_owner()
        cases = (
            ({"grantee_type": "user", "grantee": "nobody"}, "No user named"),
            ({"grantee_type": "user", "grantee": "mira"}, "already sees"),
            ({"grantee_type": "group", "grantee": "no-such"}, "No group with slug"),
            ({"grantee_type": "group", "grantee": self.group.slug}, "this group"),
            ({"grantee_type": "user", "grantee": " "}, "Enter a username"),
        )
        for data, message in cases:
            response = self.client.post(self.url("grants/"), data)
            self.assertContains(response, message)
        self.assertEqual(ShareGrant.objects.count(), 0)

        response = self.client.post(
            self.url("grants/"), {"grantee_type": "robot", "grantee": "pavel"},
        )
        self.assertEqual(response.status_code, 400)

    def test_duplicate_grant_is_a_message_not_a_crash(self):
        ShareGrant.objects.create(group=self.group, grantee_user=self.outsider)
        self.login_owner()
        response = self.client.post(
            self.url("grants/"), {"grantee_type": "user", "grantee": "pavel"},
        )
        self.assertContains(response, "already has a grant")
        self.assertEqual(ShareGrant.objects.count(), 1)

    def test_grant_revoke_deletes_the_row(self):
        grant = ShareGrant.objects.create(group=self.group, grantee_user=self.outsider)
        self.login_owner()
        response = self.client.post(self.url(f"grants/{grant.pk}/delete/"))
        self.assertEqual(ShareGrant.objects.count(), 0)
        self.assertContains(response, "No grants yet.")

        # Another group's grant is out of reach (404, not deleted).
        foreign = ShareGrant.objects.create(
            group=self.outsider.membership.group, grantee_user=self.owner,
        )
        response = self.client.post(self.url(f"grants/{foreign.pk}/delete/"))
        self.assertEqual(response.status_code, 404)
        self.assertEqual(ShareGrant.objects.count(), 1)

    def test_inert_grants_warning_outside_the_shared_tier(self):
        ShareGrant.objects.create(group=self.group, grantee_user=self.outsider)
        self.login_owner()
        response = self.client.get(self.url())
        self.assertContains(response, "currently inert")

        self.group.visibility = Group.Visibility.SHARED
        self.group.save()
        response = self.client.get(self.url())
        self.assertNotContains(response, "currently inert")


class InviteViewTests(TestCase):
    """Issue #61: the group owner invites an existing user (by username) to
    join as a Member; the invitee accepts/declines from /settings/."""

    @classmethod
    def setUpTestData(cls):
        make_user = get_user_model().objects.create_user
        cls.owner = make_user(username="kernicek", password="pass")
        cls.group = cls.owner.membership.group
        cls.outsider = make_user(username="pavel", password="pass")
        cls.member = make_user(username="mira", password="pass")
        cls.member.membership.group = cls.group
        cls.member.membership.role = Membership.Role.MEMBER
        cls.member.membership.save()

    def url(self, tail=""):
        return f"/g/{self.group.slug}/settings/{tail}"

    def login_owner(self):
        self.client.login(username="kernicek", password="pass")

    def test_only_the_owner_gets_in(self):
        for username in ("pavel", "mira"):
            self.client.login(username=username, password="pass")
            response = self.client.post(
                self.url("invites/"), {"invitee": "pavel"},
            )
            self.assertEqual(response.status_code, 404)

    def test_invite_unknown_username(self):
        self.login_owner()
        response = self.client.post(self.url("invites/"), {"invitee": "ghost"})
        self.assertContains(response, "No user named")
        self.assertEqual(Invite.objects.count(), 0)

    def test_invite_existing_member(self):
        self.login_owner()
        response = self.client.post(self.url("invites/"), {"invitee": "mira"})
        self.assertContains(response, "already a member")
        self.assertEqual(Invite.objects.count(), 0)

    def test_invite_duplicate(self):
        Invite.objects.create(group=self.group, invited_user=self.outsider)
        self.login_owner()
        response = self.client.post(self.url("invites/"), {"invitee": "pavel"})
        self.assertContains(response, "already has a pending invite")
        self.assertEqual(Invite.objects.count(), 1)

    def test_invite_created_and_listed(self):
        self.login_owner()
        response = self.client.post(self.url("invites/"), {"invitee": "pavel"})
        self.assertEqual(Invite.objects.count(), 1)
        invite = Invite.objects.get()
        self.assertEqual(invite.invited_user, self.outsider)
        self.assertEqual(invite.group, self.group)
        self.assertContains(response, "pavel")

    def test_owner_revokes_pending_invite(self):
        invite = Invite.objects.create(group=self.group, invited_user=self.outsider)
        self.login_owner()
        response = self.client.post(self.url(f"invites/{invite.pk}/delete/"))
        self.assertEqual(Invite.objects.count(), 0)
        self.assertContains(response, "No pending invites.")

        # Another group's invite is out of reach (404, not deleted).
        foreign_group = self.outsider.membership.group
        foreign = Invite.objects.create(group=foreign_group, invited_user=self.member)
        response = self.client.post(self.url(f"invites/{foreign.pk}/delete/"))
        self.assertEqual(response.status_code, 404)
        self.assertEqual(Invite.objects.count(), 1)

    def test_accept_moves_membership_and_deletes_old_group(self):
        invite = Invite.objects.create(group=self.group, invited_user=self.outsider)
        old_group = self.outsider.membership.group
        location = Location.objects.create(group=old_group, name="Shelf")
        copy = Copy.objects.create(
            edition=Edition.objects.create(game=Game.objects.create(name="Test Game")),
            owner=self.outsider, location=location,
        )

        self.client.login(username="pavel", password="pass")
        response = self.client.post(f"/invites/{invite.pk}/accept/")
        self.assertRedirects(response, "/settings/")

        self.outsider.membership.refresh_from_db()
        self.assertEqual(self.outsider.membership.group, self.group)
        self.assertEqual(self.outsider.membership.role, Membership.Role.MEMBER)
        self.assertEqual(Invite.objects.count(), 0)
        self.assertFalse(Group.objects.filter(pk=old_group.pk).exists())
        self.assertFalse(Location.objects.filter(pk=location.pk).exists())
        copy.refresh_from_db()
        self.assertIsNone(copy.location)

    def test_decline_deletes_invite_and_leaves_membership(self):
        invite = Invite.objects.create(group=self.group, invited_user=self.outsider)
        old_group = self.outsider.membership.group

        self.client.login(username="pavel", password="pass")
        response = self.client.post(f"/invites/{invite.pk}/decline/")
        self.assertRedirects(response, "/settings/")

        self.assertEqual(Invite.objects.count(), 0)
        self.outsider.membership.refresh_from_db()
        self.assertEqual(self.outsider.membership.group, old_group)

    def test_cannot_accept_or_decline_someone_elses_invite(self):
        invite = Invite.objects.create(group=self.group, invited_user=self.outsider)
        self.client.login(username="mira", password="pass")
        self.assertEqual(
            self.client.post(f"/invites/{invite.pk}/accept/").status_code, 404,
        )
        self.assertEqual(
            self.client.post(f"/invites/{invite.pk}/decline/").status_code, 404,
        )
        self.assertEqual(Invite.objects.count(), 1)

    def test_settings_page_shows_pending_invites(self):
        Invite.objects.create(group=self.group, invited_user=self.outsider)
        self.client.login(username="pavel", password="pass")
        response = self.client.get("/settings/")
        self.assertContains(response, "Household invites")
        self.assertContains(response, self.group.name)


class NtfySendTests(TestCase):
    """send_ntfy (issue #162): fail-soft POST to the configured ntfy server.
    No-ops (no network call) when unconfigured; swallows request failures."""

    @override_settings(NTFY_SERVER_URL="")
    def test_noop_when_server_url_blank(self):
        with mock.patch("gamekeeper.ntfy.requests.post") as post:
            with self.assertLogs("gamekeeper.ntfy", level="INFO"):
                result = send_ntfy("mytopic", "Title", "Body")
        post.assert_not_called()
        self.assertFalse(result)

    @override_settings(NTFY_SERVER_URL="http://192.168.1.17:8234")
    def test_noop_when_topic_blank(self):
        with mock.patch("gamekeeper.ntfy.requests.post") as post:
            with self.assertLogs("gamekeeper.ntfy", level="INFO"):
                result = send_ntfy("", "Title", "Body")
        post.assert_not_called()
        self.assertFalse(result)

    @override_settings(NTFY_SERVER_URL="http://192.168.1.17:8234", NTFY_AUTH_TOKEN="")
    def test_successful_post(self):
        response = mock.Mock(status_code=200)
        with mock.patch("gamekeeper.ntfy.requests.post", return_value=response) as post:
            result = send_ntfy("mytopic", "Reminder", "1 deadline this week")
        self.assertTrue(result)
        url = post.call_args[0][0]
        kwargs = post.call_args[1]
        self.assertEqual(url, "http://192.168.1.17:8234/mytopic")
        self.assertEqual(kwargs["data"], b"1 deadline this week")
        self.assertEqual(kwargs["headers"]["Title"], "Reminder")
        self.assertNotIn("Authorization", kwargs["headers"])
        response.raise_for_status.assert_called_once()

    @override_settings(
        NTFY_SERVER_URL="http://192.168.1.17:8234", NTFY_AUTH_TOKEN="tk_secret",
    )
    def test_auth_token_sets_bearer_header(self):
        response = mock.Mock(status_code=200)
        with mock.patch("gamekeeper.ntfy.requests.post", return_value=response) as post:
            send_ntfy("mytopic", "Reminder", "body")
        self.assertEqual(
            post.call_args[1]["headers"]["Authorization"], "Bearer tk_secret",
        )

    @override_settings(NTFY_SERVER_URL="http://192.168.1.17:8234/")
    def test_trailing_slash_on_server_url_is_stripped(self):
        response = mock.Mock(status_code=200)
        with mock.patch("gamekeeper.ntfy.requests.post", return_value=response) as post:
            send_ntfy("mytopic", "Reminder", "body")
        self.assertEqual(post.call_args[0][0], "http://192.168.1.17:8234/mytopic")

    @override_settings(NTFY_SERVER_URL="http://192.168.1.17:8234")
    def test_request_exception_is_swallowed(self):
        with mock.patch(
            "gamekeeper.ntfy.requests.post",
            side_effect=requests.ConnectionError("refused"),
        ):
            result = send_ntfy("mytopic", "Title", "Body")
        self.assertFalse(result)


class ReminderEmailTests(TestCase):
    """DESIGN §11 email reminders: pledge managers closing soon + watched
    campaigns ending soon, one digest per owner, idempotent via ReminderLog.
    The test runner swaps in the locmem backend, so mail.outbox sees sends.

    Also covers the ntfy push (issue #162) that complements the email."""

    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="kernicek", email="vojta@example.com",
        )
        self.today = timezone.localdate()

    def purchase(self, name, **kwargs):
        return Purchase.objects.create(owner=self.user, name=name, **kwargs)

    def in_days(self, days):
        return self.today + datetime.timedelta(days=days)

    def test_pledge_manager_closing_soon_emails_once(self):
        self.purchase(
            "Arydia KS",
            status=Purchase.Status.COMMITTED,
            pledge_manager=PledgeManager.objects.get(name="Gamefound"),
            pledge_manager_url="https://gamefound.com/arydia/pm",
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
            pledge_manager_close_date=self.in_days(3),
        )

        summary = send_reminder_emails()
        self.assertEqual(summary, "Sent 1 email(s) covering 1 reminder(s).")
        (message,) = mail.outbox
        self.assertEqual(message.to, ["vojta@example.com"])
        self.assertIn("1 deadline in the next 7 days", message.subject)
        self.assertIn("Arydia KS", message.body)
        self.assertIn("Gamefound", message.body)
        self.assertIn("https://gamefound.com/arydia/pm", message.body)
        self.assertEqual(ReminderLog.objects.count(), 1)

        # Beat fires daily; the same deadline must not email twice.
        send_reminder_emails()
        self.assertEqual(len(mail.outbox), 1)

    def test_outside_window_filled_out_or_dead_purchases_stay_quiet(self):
        base = dict(
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.NOT_YET,
        )
        self.purchase("Too Far KS", pledge_manager_close_date=self.in_days(8), **base)
        self.purchase("Missed KS", pledge_manager_close_date=self.in_days(-1), **base)
        self.purchase(
            "Done KS",
            status=Purchase.Status.COMMITTED,
            pledge_manager_status=Purchase.PledgeManagerStatus.FILLED_OUT,
            pledge_manager_close_date=self.in_days(3),
        )
        self.purchase(
            "Refunded KS",
            status=Purchase.Status.REFUNDED,
            pledge_manager_status=Purchase.PledgeManagerStatus.NOT_YET,
            pledge_manager_close_date=self.in_days(3),
        )

        send_reminder_emails()
        self.assertEqual(mail.outbox, [])
        self.assertEqual(ReminderLog.objects.count(), 0)

    def test_campaign_ending_soon_covers_watched_only(self):
        self.purchase(
            "Shiny New GF",
            status=Purchase.Status.WATCHING,
            campaign_url="https://gamefound.com/shiny",
            campaign_end_date=self.in_days(5),
        )
        # Already backed — the ending-soon nudge is for watched-but-unbacked.
        self.purchase(
            "Backed KS",
            status=Purchase.Status.COMMITTED,
            campaign_end_date=self.in_days(5),
        )

        send_reminder_emails()
        (message,) = mail.outbox
        self.assertIn("Shiny New GF", message.body)
        self.assertIn("https://gamefound.com/shiny", message.body)
        self.assertNotIn("Backed KS", message.body)

    def test_postponed_deadline_rearms_the_reminder(self):
        purchase = self.purchase(
            "Slippery KS",
            status=Purchase.Status.WATCHING,
            campaign_end_date=self.in_days(2),
        )
        send_reminder_emails()
        self.assertEqual(len(mail.outbox), 1)

        purchase.campaign_end_date = self.in_days(6)
        purchase.save()
        send_reminder_emails()
        self.assertEqual(len(mail.outbox), 2)
        self.assertEqual(ReminderLog.objects.count(), 2)

    def test_one_digest_bundles_both_kinds_per_owner(self):
        self.purchase(
            "PM Due KS",
            status=Purchase.Status.PLACEHOLDER,
            pledge_manager_status=Purchase.PledgeManagerStatus.SENT_OUT,
            pledge_manager_close_date=self.in_days(1),
        )
        self.purchase(
            "Ending GF",
            status=Purchase.Status.WATCHING,
            campaign_end_date=self.in_days(4),
        )

        summary = send_reminder_emails()
        self.assertEqual(summary, "Sent 1 email(s) covering 2 reminder(s).")
        (message,) = mail.outbox
        self.assertIn("2 deadlines", message.subject)
        self.assertIn("PM Due KS", message.body)
        self.assertIn("Ending GF", message.body)

    def test_owner_without_email_is_skipped_and_not_logged(self):
        nomail = get_user_model().objects.create_user(username="pavel")
        Purchase.objects.create(
            owner=nomail, name="Unreachable KS",
            status=Purchase.Status.WATCHING,
            campaign_end_date=self.in_days(3),
        )

        summary = send_reminder_emails()
        self.assertEqual(summary, "Sent 0 email(s) covering 0 reminder(s).")
        self.assertEqual(mail.outbox, [])
        # No log row — the reminder re-arms if an email address appears
        # while the deadline is still inside the window.
        self.assertEqual(ReminderLog.objects.count(), 0)

    def test_reminder_log_unique_per_purchase_kind_deadline(self):
        purchase = self.purchase("Twice KS")
        ReminderLog.objects.create(
            purchase=purchase, kind=ReminderLog.Kind.CAMPAIGN_END,
            deadline=self.today,
        )
        with self.assertRaises(IntegrityError):
            ReminderLog.objects.create(
                purchase=purchase, kind=ReminderLog.Kind.CAMPAIGN_END,
                deadline=self.today,
            )

    def test_owner_with_ntfy_topic_gets_a_push(self):
        self.user.membership.ntfy_topic = "kernicek-reminders"
        self.user.membership.save()
        self.purchase(
            "Arydia KS",
            status=Purchase.Status.WATCHING,
            campaign_end_date=self.in_days(3),
        )

        with mock.patch("gamekeeper.tasks.ntfy.send_ntfy") as send_ntfy_mock:
            send_reminder_emails()

        send_ntfy_mock.assert_called_once()
        topic, title, body = send_ntfy_mock.call_args[0]
        self.assertEqual(topic, "kernicek-reminders")
        self.assertIn("1 deadline", title)
        self.assertIn("Arydia KS", body)
        # The email itself is unaffected by the push.
        self.assertEqual(len(mail.outbox), 1)

    def test_owner_without_ntfy_topic_gets_no_push(self):
        self.purchase(
            "Arydia KS",
            status=Purchase.Status.WATCHING,
            campaign_end_date=self.in_days(3),
        )

        with mock.patch("gamekeeper.tasks.ntfy.send_ntfy") as send_ntfy_mock:
            send_reminder_emails()

        send_ntfy_mock.assert_not_called()
        self.assertEqual(len(mail.outbox), 1)

    def test_ntfy_failure_does_not_block_email_or_reminder_log(self):
        self.user.membership.ntfy_topic = "kernicek-reminders"
        self.user.membership.save()
        self.purchase(
            "Arydia KS",
            status=Purchase.Status.WATCHING,
            campaign_end_date=self.in_days(3),
        )

        with mock.patch(
            "gamekeeper.tasks.ntfy.send_ntfy", return_value=False,
        ):
            summary = send_reminder_emails()

        self.assertEqual(summary, "Sent 1 email(s) covering 1 reminder(s).")
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(ReminderLog.objects.count(), 1)



# ===========================================================================
# §13  Cover editing (replace + focal point)
# ===========================================================================

def image_bytes(format="PNG"):
    """A tiny real image — the cover views verify uploads with Pillow."""
    buffer = BytesIO()
    Image.new("RGB", (4, 4), "red").save(buffer, format=format)
    return buffer.getvalue()


def sized_image_bytes(width, height, color="green", format="PNG"):
    """A real solid-colour image of a chosen shape — the preview renderer
    (issue #104) crops/letterboxes non-square art, so its tests need
    non-square inputs, not the 4x4 image_bytes square."""
    buffer = BytesIO()
    Image.new("RGB", (width, height), color).save(buffer, format=format)
    return buffer.getvalue()


def patterned_image_bytes(width=8, height=16, format="PNG"):
    """A spatially-varying portrait image: top half red, bottom half blue. A
    solid colour renders the same at any focus/zoom, so content-addressed
    previews (#116) only get a new filename when the crop lands on different
    pixels — which needs an image whose regions actually differ."""
    img = Image.new("RGB", (width, height), "red")
    for y in range(height // 2, height):
        for x in range(width):
            img.putpixel((x, y), (0, 0, 255))
    buffer = BytesIO()
    img.save(buffer, format=format)
    return buffer.getvalue()


class CoverScaleTests(TestCase):
    """Issue #1: fit-mode zoom must not jump at the 100% boundary — with
    known art dimensions the scale interpolates aspect-ratio-aware."""

    def scale(self, zoom, width=None, height=None):
        return Game(name="X", cover_zoom=zoom, cover_width=width,
                    cover_height=height).cover_scale

    def test_crop_zoom_stays_a_plain_percent(self):
        self.assertEqual(self.scale(150), "1.5")
        self.assertEqual(self.scale(100, 800, 400), "1")  # dims ignored ≥ 100

    def test_fit_without_dimensions_keeps_the_old_fallback(self):
        self.assertEqual(self.scale(60), "0.6")

    def test_fit_with_dimensions_interpolates_the_aspect_ratio(self):
        # 2:1 art: just under 100 ≈ the cover-crop size (no cliff), 50 =
        # the exact contain fit, halfway lands halfway.
        self.assertEqual(self.scale(99, 800, 400), "1.98")
        self.assertEqual(self.scale(75, 800, 400), "1.5")
        self.assertEqual(self.scale(50, 800, 400), "1")
        # Orientation is irrelevant — the long/short ratio drives it.
        self.assertEqual(self.scale(75, 400, 800), "1.5")

    def test_square_art_has_no_letterbox_to_fill(self):
        self.assertEqual(self.scale(60, 500, 500), "1")


class CoverPreviewRenderTests(TestCase):
    """Issue #104: render_square_preview bakes the grid-tile crop (focus,
    zoom, fit/letterbox colour) into a small square PNG server-side."""

    def render(self, width, height, zoom=100, fit_color="",
               focus_x=50, focus_y=50, color="green"):
        data = sized_image_bytes(width, height, color)
        preview = render_square_preview(data, focus_x, focus_y, zoom, fit_color)
        return Image.open(BytesIO(preview)).convert("RGBA")

    def test_output_is_a_square_png_at_the_preview_size(self):
        out = self.render(100, 40)
        self.assertEqual(out.size, (PREVIEW_SIZE, PREVIEW_SIZE))

    def test_cover_mode_fills_the_whole_square_opaquely(self):
        # zoom >= 100 crops square (object-fit: cover) — no letterbox, so
        # every corner is opaque art.
        out = self.render(100, 40, zoom=100)
        for xy in [(0, 0), (PREVIEW_SIZE - 1, 0),
                   (0, PREVIEW_SIZE - 1), (PREVIEW_SIZE - 1, PREVIEW_SIZE - 1)]:
            self.assertEqual(out.getpixel(xy)[3], 255)

    def test_fit_mode_paints_the_letterbox_colour_in_the_margin(self):
        # A wide cover contained at zoom 50 letterboxes top/bottom with the
        # chosen colour; the art still shows through the centre.
        out = self.render(100, 40, zoom=50, fit_color="#0000ff")
        self.assertEqual(out.getpixel((0, 0)), (0, 0, 255, 255))
        centre = out.getpixel((PREVIEW_SIZE // 2, PREVIEW_SIZE // 2))
        self.assertEqual(centre[:3], (0, 128, 0))  # PIL "green"

    def test_fit_mode_without_a_colour_is_transparent(self):
        # No colour => transparent letterbox so the tile background shows.
        out = self.render(100, 40, zoom=50, fit_color="")
        self.assertEqual(out.getpixel((0, 0)), (0, 0, 0, 0))


class CoverPreviewModelTests(TestCase):
    """Issue #104: the regenerate/clear helpers on CoverArtModel that bake
    and drop the square preview file."""

    def setUp(self):
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)

    def _game_with_cover(self):
        game = Game.objects.create(name="Cover Game")
        game.cover_image.save("art.png", BytesIO(image_bytes()))
        return game

    def test_regenerate_bakes_a_square_preview_file(self):
        game = self._game_with_cover()
        game.regenerate_cover_preview()
        game.refresh_from_db()
        self.assertTrue(game.cover_preview)
        self.assertTrue(game.cover_preview.name.startswith("covers/previews/"))
        with game.cover_preview.open() as handle:
            out = Image.open(BytesIO(handle.read()))
        self.assertEqual(out.size, (PREVIEW_SIZE, PREVIEW_SIZE))

    def test_preview_name_carries_a_content_hash_token(self):
        # Issue #116: the filename is content-addressed — stem + "-" + a hex
        # hash of the rendered pixels + ".png" — so nginx can serve it immutable.
        game = self._game_with_cover()
        game.regenerate_cover_preview()
        name = PurePosixPath(game.cover_preview.name).name
        self.assertTrue(re.fullmatch(r".+-[0-9a-f]{16}\.png", name), name)

    def test_regenerate_keeps_the_name_for_identical_content(self):
        # Issue #116: re-baking the same art keeps the same URL, so a browser
        # can cache it immutable and reuse it across re-bakes.
        game = self._game_with_cover()
        game.regenerate_cover_preview()
        first = game.cover_preview.name
        game.regenerate_cover_preview()
        self.assertEqual(game.cover_preview.name, first)

    def test_regenerate_changes_the_name_when_the_image_changes(self):
        # Issue #116: different pixels -> different hash -> new URL, and the old
        # preview file is removed. This is what busts a stale cached thumbnail.
        game = self._game_with_cover()
        game.regenerate_cover_preview()
        first = game.cover_preview.name
        storage = game.cover_preview.storage
        game.cover_image.save("art2.png", BytesIO(sized_image_bytes(4, 4, "blue")))
        game.regenerate_cover_preview()
        self.assertNotEqual(game.cover_preview.name, first)
        self.assertFalse(storage.exists(first))

    def test_clear_removes_the_preview_file(self):
        game = self._game_with_cover()
        game.regenerate_cover_preview()
        name = game.cover_preview.name
        storage = game.cover_preview.storage
        game.clear_cover_preview()
        self.assertFalse(game.cover_preview)
        self.assertFalse(storage.exists(name))

    def test_regenerate_without_a_cover_clears_a_stale_preview(self):
        game = self._game_with_cover()
        game.regenerate_cover_preview()
        old = game.cover_preview.name
        storage = game.cover_preview.storage
        game.cover_image.delete(save=True)  # no local art left
        game.regenerate_cover_preview()
        self.assertFalse(game.cover_preview)
        self.assertFalse(storage.exists(old))


class GenerateCoverPreviewsTests(TestCase):
    """Issue #104: the backfill command bakes previews for existing covers
    across Game / Series / Family, and is idempotent."""

    def setUp(self):
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)

    def _cover(self, obj):
        obj.cover_image.save("art.png", BytesIO(image_bytes()))
        return obj

    def run_command(self, **extra):
        out = StringIO()
        call_command("generate_cover_previews", stdout=out, **extra)
        return out.getvalue()

    def test_backfills_previews_for_every_cover_bearing_model(self):
        game = self._cover(Game.objects.create(name="G"))
        series = self._cover(
            Series.objects.create(name="S", primary_game=game))
        family = self._cover(Family.objects.create(name="F"))
        # A cover-less object is left alone (no preview to bake).
        bare = Game.objects.create(name="Bare")

        output = self.run_command()

        for obj in (game, series, family):
            obj.refresh_from_db()
            self.assertTrue(obj.cover_preview)
        bare.refresh_from_db()
        self.assertFalse(bare.cover_preview)
        self.assertIn("preview baked", output)

    def test_is_idempotent_without_force(self):
        self._cover(Game.objects.create(name="G"))
        self.run_command()
        output = self.run_command()
        self.assertIn("already baked — skipped: 1", output)

    def test_force_rebakes_but_keeps_the_content_addressed_name(self):
        # --force re-bakes the preview, but with content-addressed filenames
        # (#116) unchanged art hashes to the same URL, so the name is stable.
        game = self._cover(Game.objects.create(name="G"))
        self.run_command()
        game.refresh_from_db()
        first = game.cover_preview.name
        self.run_command(force=True)
        game.refresh_from_db()
        self.assertEqual(game.cover_preview.name, first)

    def test_dry_run_writes_nothing(self):
        game = self._cover(Game.objects.create(name="G"))
        output = self.run_command(dry_run=True)
        self.assertIn("would bake preview", output)
        game.refresh_from_db()
        self.assertFalse(game.cover_preview)


class CoverEditMixin:
    def setUp(self):
        # Fresh MEDIA_ROOT per test, same reasoning as DownloadCoversTests:
        # files are not transactional and dedupe suffixes would leak across
        # tests.
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)
        self.client.login(username="kernicek", password="pass")

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(
            name="Epic Quest",
            image_url="https://cf.geekdo-images.com/original/img/eq.jpg",
        )
        BggLink.objects.create(game=cls.game, bgg_id=12345, is_primary=True)


class GameCoverEditTests(CoverEditMixin, TestCase):
    def post(self, data=None):
        return self.client.post(f"/games/{self.game.pk}/cover/", data or {})

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.post({"url": "https://example.com/a.png"})
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_405(self):
        response = self.client.get(f"/games/{self.game.pk}/cover/")
        self.assertEqual(response.status_code, 405)

    def test_unknown_game_is_404(self):
        response = self.client.post("/games/99999/cover/", {})
        self.assertEqual(response.status_code, 404)

    def test_upload_replaces_the_cover(self):
        self.game.cover_focus_x = 20
        self.game.cover_focus_y = 80
        self.game.cover_zoom = 180
        self.game.save()

        response = self.post({
            "file": SimpleUploadedFile("art.png", image_bytes()),
        })

        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_image.name, "covers/12345.png")
        # Art dimensions recorded for the fit-mode zoom scale (issue #1).
        self.assertEqual(self.game.cover_width, 4)
        self.assertEqual(self.game.cover_height, 4)
        with self.game.cover_image.open() as stored:
            self.assertEqual(stored.read(), image_bytes())
        # The focal point and zoom described the old art — reset.
        self.assertEqual(self.game.cover_focus_x, 50)
        self.assertEqual(self.game.cover_focus_y, 50)
        self.assertEqual(self.game.cover_zoom, 100)

    def test_url_is_fetched_server_side(self):
        with mock.patch("gamekeeper.views.requests.get") as get:
            get.return_value = mock.Mock(content=image_bytes("JPEG"))
            response = self.post({"url": "https://example.com/art?dl=1"})

        get.assert_called_once_with("https://example.com/art?dl=1", timeout=30)
        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        # Extension comes from what Pillow says the bytes are, not the URL.
        self.assertEqual(self.game.cover_image.name, "covers/12345.jpg")

    def test_replacement_changes_the_file_name_and_deletes_the_old_file(self):
        # Same-name saves would keep the URL identical and browsers would
        # show the stale cached cover.
        self.post({"file": SimpleUploadedFile("a.png", image_bytes())})
        self.game.refresh_from_db()
        first_name = self.game.cover_image.name

        self.post({"file": SimpleUploadedFile("b.png", image_bytes())})
        self.game.refresh_from_db()

        self.assertNotEqual(self.game.cover_image.name, first_name)
        self.assertFalse(self.game.cover_image.storage.exists(first_name))

    def test_game_without_bgg_link_is_named_by_pk(self):
        unlinked = Game.objects.create(name="Homebrew")
        self.client.post(f"/games/{unlinked.pk}/cover/", {
            "file": SimpleUploadedFile("art.png", image_bytes()),
        })
        unlinked.refresh_from_db()
        self.assertEqual(unlinked.cover_image.name, f"covers/game-{unlinked.pk}.png")

    def test_neither_file_nor_url_is_an_inline_error(self):
        response = self.post()
        self.assertContains(response, "Choose a file or paste an image URL.")
        self.game.refresh_from_db()
        self.assertFalse(self.game.cover_image)

    def test_both_file_and_url_is_an_inline_error(self):
        response = self.post({
            "file": SimpleUploadedFile("art.png", image_bytes()),
            "url": "https://example.com/art.png",
        })
        self.assertContains(response, "not both")

    def test_non_http_url_is_rejected(self):
        response = self.post({"url": "ftp://example.com/art.png"})
        self.assertContains(response, "must start with http")

    def test_download_failure_is_an_inline_error(self):
        with mock.patch("gamekeeper.views.requests.get") as get:
            get.side_effect = requests.ConnectionError("boom")
            response = self.post({"url": "https://example.com/art.png"})
        self.assertContains(response, "Download failed")

    def test_non_image_bytes_are_rejected(self):
        response = self.post({
            "file": SimpleUploadedFile("art.png", b"not an image at all"),
        })
        self.assertContains(response, "does not look like an image")
        self.game.refresh_from_db()
        self.assertFalse(self.game.cover_image)

    def test_non_web_formats_are_rejected(self):
        response = self.post({
            "file": SimpleUploadedFile("art.bmp", image_bytes("BMP")),
        })
        self.assertContains(response, "BMP images will not render")

    def test_upload_bakes_the_grid_preview(self):
        # Issue #104: the square grid thumbnail is baked from the new art.
        self.post({"file": SimpleUploadedFile("art.png", image_bytes())})
        self.game.refresh_from_db()
        self.assertTrue(self.game.cover_preview)
        self.assertTrue(self.game.cover_preview.name.startswith("covers/previews/"))


class GameCoverFocusTests(CoverEditMixin, TestCase):
    def post(self, data):
        return self.client.post(f"/games/{self.game.pk}/cover/focus/", data)

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.post({"x": "10", "y": "90"})
        self.assertEqual(response.status_code, 302)

    def test_get_is_405(self):
        response = self.client.get(f"/games/{self.game.pk}/cover/focus/")
        self.assertEqual(response.status_code, 405)

    def test_sets_the_focal_point(self):
        response = self.post({"x": "10", "y": "90"})
        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_focus_x, 10)
        self.assertEqual(self.game.cover_focus_y, 90)
        # The partial echoes it into the crop preview.
        self.assertContains(response, "object-position: 10% 90%")

    def test_non_numeric_focal_point_is_400(self):
        response = self.post({"x": "abc", "y": "10"})
        self.assertEqual(response.status_code, 400)

    def test_single_coordinate_updates_one_axis(self):
        # Issue #12: the X and Y number inputs post independently, so one axis
        # can move while the other stays put.
        self.game.cover_focus_y = 90
        self.game.save(update_fields=["cover_focus_y"])
        response = self.post({"x": "10"})
        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_focus_x, 10)
        self.assertEqual(self.game.cover_focus_y, 90)

    def test_out_of_range_focal_point_is_clamped(self):
        # Hand-typed values clamp to 0-100 rather than 400 (issue #12).
        response = self.post({"x": "101", "y": "-1"})
        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_focus_x, 100)
        self.assertEqual(self.game.cover_focus_y, 0)

    def test_editor_renders_number_inputs(self):
        # Issue #12: the picker exposes editable X / Y / zoom number boxes.
        response = self.post({"x": "10", "y": "90"})
        self.assertContains(response, 'name="x"')
        self.assertContains(response, 'name="y"')
        self.assertContains(response, 'id="cover-zoom-num"')
        self.assertContains(response, 'type="number"')

    def test_unknown_game_is_404(self):
        response = self.client.post("/games/99999/cover/focus/", {"x": "1", "y": "1"})
        self.assertEqual(response.status_code, 404)


    def test_zoom_alone_is_saved(self):
        response = self.post({"zoom": "175"})
        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_zoom, 175)
        # The preview scales toward the focal point.
        self.assertContains(response, "transform: scale(1.75)")

    def test_reset_posts_all_three_back_to_defaults(self):
        self.game.cover_focus_x = 10
        self.game.cover_focus_y = 90
        self.game.cover_zoom = 200
        self.game.save()

        response = self.post({"x": "50", "y": "50", "zoom": "100"})

        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_focus_x, 50)
        self.assertEqual(self.game.cover_focus_y, 50)
        self.assertEqual(self.game.cover_zoom, 100)

    def test_zoom_below_100_renders_the_fit_preview(self):
        # §13 zoom-out: the preview flips to contain-fit over the game's
        # letterbox colour (once one is picked).
        self.game.cover_fit_color = "#aabbcc"
        self.game.save(update_fields=["cover_fit_color"])
        response = self.post({"zoom": "60"})
        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_zoom, 60)
        self.assertTrue(self.game.cover_fit)
        self.assertContains(response, "cover-art-fit")
        self.assertContains(response, "transform: scale(0.6)")
        self.assertContains(
            response, 'cover-art-backdrop" style="background-color: #aabbcc"')

    def test_zoom_100_and_up_has_no_fit_treatment(self):
        self.game.cover_fit_color = "#aabbcc"
        self.game.save(update_fields=["cover_fit_color"])
        response = self.post({"zoom": "100"})
        self.assertNotContains(response, "cover-art-backdrop")
        self.assertNotContains(response, "cover-art-fit")

    def test_fit_color_is_set_validated_and_cleared(self):
        response = self.post({"fit_color": "#A1b2C3"})
        self.assertEqual(response.status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_fit_color, "#A1b2C3")

        # Bare hex (PowerToys pastes without the #) gets the # added.
        self.post({"fit_color": "1a2b3c"})
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_fit_color, "#1a2b3c")

        self.assertEqual(self.post({"fit_color": "red"}).status_code, 400)
        self.assertEqual(self.post({"fit_color": "#12345"}).status_code, 400)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_fit_color, "#1a2b3c")

        # Reset sends fit_color="" alongside the focus/zoom defaults.
        self.post({"x": "50", "y": "50", "zoom": "100", "fit_color": ""})
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_fit_color, "")

    def test_zoom_out_of_range_clamps_non_numeric_is_400(self):
        # Non-numeric zoom is still rejected; out-of-range typed zoom clamps
        # to 50-300 (issue #12).
        self.assertEqual(self.post({"zoom": "abc"}).status_code, 400)
        self.assertEqual(self.post({"zoom": "49"}).status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_zoom, 50)
        self.assertEqual(self.post({"zoom": "301"}).status_code, 200)
        self.game.refresh_from_db()
        self.assertEqual(self.game.cover_zoom, 300)

    def test_focus_edit_rebakes_the_grid_preview(self):
        # Issue #104: the preview bakes in the crop, so a focus change must
        # regenerate it. With a red/blue portrait cover the new crop lands on
        # different pixels, so the content-addressed name changes too (#116).
        self.game.cover_image.save("g.png", BytesIO(patterned_image_bytes()))
        self.game.regenerate_cover_preview()
        first = self.game.cover_preview.name

        self.post({"x": "10", "y": "90"})

        self.game.refresh_from_db()
        self.assertTrue(self.game.cover_preview)
        self.assertNotEqual(self.game.cover_preview.name, first)


class GameEditViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(
            name="Epic Quest", bgg_name="Epic Quest (EN)",
            image_url="https://cf.geekdo-images.com/original/img/eq.jpg",
            is_campaign=True, soundtrack_ambience=True,
            language_dependency=Game.LanguageDependency.EASY,
            player_conflict=2,
        )
        cls.expansion = Game.objects.create(
            name="Epic Quest: More Heroes", type=Game.Type.EXPANSION,
        )
        # Issue #78: series/family fixtures. self.game itself stays
        # series-less (form_data's default "series" is blank, matching that)
        # so the many unrelated tests above aren't affected; a dedicated
        # primary game exercises the orphan guard on its own.
        cls.series_primary = Game.objects.create(name="Series Primary")
        cls.series = Series.objects.create(
            name="Epic Quest Saga", primary_game=cls.series_primary,
        )
        Game.objects.filter(pk=cls.series_primary.pk).update(series=cls.series)
        cls.series_member = Game.objects.create(name="Series Member")
        Game.objects.filter(pk=cls.series_member.pk).update(series=cls.series)
        cls.family = Family.objects.create(name="Heist line")
        cls.other_family = Family.objects.create(name="Puzzle line")

    def setUp(self):
        self.client.login(username="kernicek", password="pass")

    def form_data(self, **overrides):
        """A minimal valid whole-form POST (checkboxes absent = off)."""
        data = {
            "name": "Epic Quest",
            "language_dependency": "", "language_dependency_note": "",
            "companion_app": "",
            "player_conflict": "", "player_conflict_note": "",
            "series": "",
        }
        data.update(overrides)
        return data

    def post(self, pk=None, **overrides):
        return self.client.post(
            f"/games/{pk or self.game.pk}/edit/", self.form_data(**overrides),
        )

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.client.get(f"/games/{self.game.pk}/edit/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_unknown_game_is_404(self):
        response = self.client.get("/games/99999/edit/")
        self.assertEqual(response.status_code, 404)

    def test_edit_page_hosts_details_form_and_cover_tools(self):
        response = self.client.get(f"/games/{self.game.pk}/edit/")
        self.assertContains(response, "Edit — Epic Quest")
        self.assertContains(response, 'value="Epic Quest"')
        self.assertContains(response, "BGG calls it")
        self.assertContains(response, "Replace cover")
        self.assertContains(response, "cover-focus-frame")
        # Base game: no expansion override inputs.
        self.assertNotContains(response, "players_min_override")

    def test_edit_page_loads_the_cover_drag_script(self):
        # Issue #13: the draggable focal-point marker needs its JS wired in.
        response = self.client.get(f"/games/{self.game.pk}/edit/")
        self.assertContains(response, "js/cover_focus.js")

    def test_detail_page_links_here_and_lost_the_cover_tools(self):
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertContains(response, f"/games/{self.game.pk}/edit/")
        self.assertNotContains(response, "Replace cover")
        self.assertNotContains(response, "cover-focus-frame")

    def test_save_updates_fields_and_redirects_to_detail(self):
        response = self.post(
            name="Epic Quest II",
            language_dependency=Game.LanguageDependency.MEDIUM,
            language_dependency_note="cards are symbols",
            companion_app=Game.AppUse.OPTIONAL,
            player_conflict="3", player_conflict_note="take that",
            is_legacy="on", has_app_version="on",
        )

        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.game.refresh_from_db()
        self.assertEqual(self.game.name, "Epic Quest II")
        self.assertEqual(self.game.language_dependency,
                         Game.LanguageDependency.MEDIUM)
        self.assertEqual(self.game.language_dependency_note, "cards are symbols")
        self.assertEqual(self.game.player_conflict, 3)
        self.assertEqual(self.game.player_conflict_note, "take that")
        self.assertTrue(self.game.is_legacy)
        self.assertTrue(self.game.has_app_version)
        # Whole-form checkbox semantics: absent means unchecked.
        self.assertFalse(self.game.is_campaign)
        self.assertFalse(self.game.soundtrack_ambience)

    def test_blank_player_conflict_clears_it(self):
        self.post(player_conflict="")
        self.game.refresh_from_db()
        self.assertIsNone(self.game.player_conflict)

    def test_alternate_names_are_added_deduped_and_pruned(self):
        # Issue #51: the textarea is one-name-per-line; blanks and
        # case-insensitive duplicates collapse, and omitting a name drops it.
        self.post(alternate_names="Safari Bar\n\n  Beasty Bar  \nsafari bar")
        self.assertEqual(
            list(self.game.alternate_names.values_list("name", flat=True)),
            ["Beasty Bar", "Safari Bar"],  # Meta.ordering = ["name"]
        )
        # A follow-up edit that omits one name prunes it.
        self.post(alternate_names="Beasty Bar")
        self.assertEqual(
            list(self.game.alternate_names.values_list("name", flat=True)),
            ["Beasty Bar"],
        )
        # An empty box clears them all.
        self.post(alternate_names="")
        self.assertFalse(self.game.alternate_names.exists())

    def test_edit_page_prefills_existing_alternate_names(self):
        AlternateName.objects.create(game=self.game, name="Safari Bar")
        response = self.client.get(f"/games/{self.game.pk}/edit/")
        self.assertContains(response, "Alternate names")
        self.assertContains(response, "Safari Bar")

    def test_empty_name_is_400(self):
        response = self.post(name="   ")
        self.assertEqual(response.status_code, 400)

    def test_unknown_select_value_is_400(self):
        response = self.post(language_dependency="klingon")
        self.assertEqual(response.status_code, 400)

    def test_garbage_and_out_of_range_conflict_are_400(self):
        self.assertEqual(self.post(player_conflict="abc").status_code, 400)
        self.assertEqual(self.post(player_conflict="4").status_code, 400)

    def test_expansion_overrides_are_saved_for_expansions_only(self):
        response = self.post(
            pk=self.expansion.pk, name="Epic Quest: More Heroes",
            players_min_override="5", players_max_override="6",
            playtime_delta_override="-10",
        )
        self.assertRedirects(response, f"/games/{self.expansion.pk}/")
        self.expansion.refresh_from_db()
        self.assertEqual(self.expansion.players_min_override, 5)
        self.assertEqual(self.expansion.players_max_override, 6)
        self.assertEqual(self.expansion.playtime_delta_override, -10)

        # The same fields on a base game are ignored, not saved.
        self.post(players_min_override="2")
        self.game.refresh_from_db()
        self.assertIsNone(self.game.players_min_override)

    def test_zero_players_override_is_400(self):
        response = self.post(
            pk=self.expansion.pk, name="Epic Quest: More Heroes",
            players_min_override="0",
        )
        self.assertEqual(response.status_code, 400)

    def test_series_and_family_selectors_render_for_base_games_only(self):
        response = self.client.get(f"/games/{self.series_primary.pk}/edit/")
        self.assertContains(response, "Series")
        self.assertContains(response, '<option value="">—</option>')
        self.assertContains(
            response,
            f'<option value="{self.series.pk}" selected>{self.series.name}</option>',
            html=True,
        )
        self.assertContains(response, "Families")
        self.assertContains(response, self.family.name)
        self.assertContains(response, self.other_family.name)

        response = self.client.get(f"/games/{self.expansion.pk}/edit/")
        self.assertNotContains(response, "edit-series")
        self.assertNotContains(response, "edit-family-")

    def test_families_checkboxes_prefill_existing_membership(self):
        self.game.families.set([self.family.pk])
        response = self.client.get(f"/games/{self.game.pk}/edit/")
        self.assertContains(
            response,
            f'<input class="form-check-input" type="checkbox" '
            f'id="edit-family-{self.family.pk}" name="families" '
            f'value="{self.family.pk}" checked>',
            html=True,
        )
        self.assertContains(
            response,
            f'<input class="form-check-input" type="checkbox" '
            f'id="edit-family-{self.other_family.pk}" name="families" '
            f'value="{self.other_family.pk}">',
            html=True,
        )

    def test_posting_a_series_moves_the_game_into_it(self):
        self.post(series=str(self.series.pk))
        self.game.refresh_from_db()
        self.assertEqual(self.game.series_id, self.series.pk)

    def test_posting_blank_series_clears_a_non_primary_member(self):
        self.post(pk=self.series_member.pk, name="Series Member", series="")
        self.series_member.refresh_from_db()
        self.assertIsNone(self.series_member.series_id)

    def test_changing_the_primary_games_series_is_400(self):
        response = self.post(
            pk=self.series_primary.pk, name="Series Primary", series="",
        )
        self.assertEqual(response.status_code, 400)
        self.series_primary.refresh_from_db()
        self.assertEqual(self.series_primary.series_id, self.series.pk)

    def test_unknown_series_value_is_400(self):
        self.assertEqual(self.post(series="99999").status_code, 400)

    def test_families_are_set_to_exactly_the_posted_pks(self):
        self.game.families.set([self.family.pk])
        self.post(families=[str(self.other_family.pk)])
        self.game.refresh_from_db()
        self.assertEqual(
            set(self.game.families.values_list("pk", flat=True)),
            {self.other_family.pk},
        )
        self.post(families=[])
        self.game.refresh_from_db()
        self.assertFalse(self.game.families.exists())

    def test_unknown_family_value_is_400(self):
        self.assertEqual(self.post(families=["99999"]).status_code, 400)

    def test_series_and_families_are_ignored_for_expansions(self):
        response = self.post(
            pk=self.expansion.pk, name="Epic Quest: More Heroes",
            series=str(self.series.pk), families=[str(self.family.pk)],
        )
        self.assertRedirects(response, f"/games/{self.expansion.pk}/")
        self.expansion.refresh_from_db()
        self.assertIsNone(self.expansion.series_id)
        self.assertFalse(self.expansion.families.exists())


# ===========================================================================
# §4  Copy add / edit
# ===========================================================================

class CopyAddViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(name="Epic Quest")
        cls.default = Edition.objects.create(game=cls.game, is_default=True)
        cls.collectors = Edition.objects.create(
            game=cls.game, name="Collector's Edition",
        )
        cls.bare = Game.objects.create(name="Editionless")

    def setUp(self):
        self.client.login(username="kernicek", password="pass")

    def post(self, pk=None, data=None):
        return self.client.post(f"/games/{pk or self.game.pk}/copies/add/",
                                data or {})

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.post()
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_405_and_unknown_game_404(self):
        response = self.client.get(f"/games/{self.game.pk}/copies/add/")
        self.assertEqual(response.status_code, 405)
        self.assertEqual(self.post(pk=99999).status_code, 404)

    def test_add_creates_copy_and_redirects_to_its_edit_page(self):
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            response = self.post(data={"edition": str(self.collectors.pk)})
        copy = Copy.objects.get(owner=self.user)
        self.assertRedirects(response, f"/copies/{copy.pk}/edit/")
        self.assertEqual(copy.edition, self.collectors)
        self.assertEqual(copy.archive_status, Copy.ArchiveStatus.ACTIVE)
        # self.game has no BggLink — nothing to push.
        delay.assert_not_called()

    def test_add_pushes_own_status_to_bgg_when_the_game_is_linked(self):
        """Issue #117: owning a copy now is the "own" signal."""
        BggLink.objects.create(game=self.game, bgg_id=207830, is_primary=True)
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.post(data={"edition": str(self.collectors.pk)})
        delay.assert_called_once_with(
            self.game.pk, Game.BggCollectionStatus.OWN, self.user.pk, priority=None,
        )

    def test_add_enqueue_failure_never_breaks_the_copy_creation(self):
        """No broker/worker (the dev reality): .delay() raises. The copy must
        still be created — only the BGG push is affected — and the failure
        surfaces as a PUSH_FAILED diff instead of a 500."""
        BggLink.objects.create(game=self.game, bgg_id=207830, is_primary=True)
        with mock.patch(
            "gamekeeper.views.push_bgg_status_task.delay",
            side_effect=RuntimeError("broker down"),
        ):
            response = self.post(data={"edition": str(self.collectors.pk)})
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Copy.objects.filter(owner=self.user).exists())
        diff = BggSyncDiff.objects.get(category=BggSyncDiff.Category.PUSH_FAILED)
        self.assertEqual(diff.game, self.game)

    def test_editionless_game_gets_its_default_edition_created(self):
        response = self.post(pk=self.bare.pk)
        copy = Copy.objects.get(owner=self.user)
        self.assertRedirects(response, f"/copies/{copy.pk}/edit/")
        self.assertTrue(copy.edition.is_default)
        self.assertEqual(copy.edition.game, self.bare)

    def test_missing_or_foreign_edition_is_400(self):
        # Editions exist, so one must be picked.
        self.assertEqual(self.post().status_code, 400)
        other_games = Edition.objects.create(game=self.bare, is_default=True)
        response = self.post(data={"edition": str(other_games.pk)})
        self.assertEqual(response.status_code, 400)

    def test_already_owned_edition_requires_confirmation(self):
        # Issue #50: dropped the unique constraint — an already-owned
        # edition (active or archived) now warns instead of outright
        # blocking. Unconfirmed, it still 400s and creates no second Copy.
        Copy.objects.create(
            owner=self.user, edition=self.default,
            archive_status=Copy.ArchiveStatus.ARCHIVED,
            archive_reason=Copy.ArchiveReason.SOLD,
        )
        response = self.post(data={"edition": str(self.default.pk)})
        self.assertEqual(response.status_code, 400)
        self.assertEqual(Copy.objects.filter(edition=self.default).count(), 1)

    def test_confirmed_duplicate_creates_a_second_copy(self):
        Copy.objects.create(owner=self.user, edition=self.default)
        response = self.post(data={
            "edition": str(self.default.pk), "confirm_duplicate_copy": "1",
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            Copy.objects.filter(owner=self.user, edition=self.default).count(), 2,
        )

    def test_detail_page_offers_every_edition_including_owned(self):
        # Issue #50: the Add-copy select now offers every edition, not just
        # unowned ones — owned_edition_pks flags which ones already have a
        # copy, for the template's duplicate-copy warning.
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertContains(response, "Add copy")
        self.assertEqual(
            set(response.context["editions"]), {self.default, self.collectors},
        )
        self.assertEqual(response.context["owned_edition_pks"], set())

        Copy.objects.create(owner=self.user, edition=self.default)
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertEqual(
            set(response.context["editions"]), {self.default, self.collectors},
        )
        self.assertEqual(
            response.context["owned_edition_pks"], {self.default.pk},
        )

        Copy.objects.create(owner=self.user, edition=self.collectors)
        response = self.client.get(f"/games/{self.game.pk}/")
        # Every edition owned: the form stays enabled — a duplicate is now
        # a confirm-past warning, not a block.
        self.assertContains(
            response, f'action="/games/{self.game.pk}/copies/add/"',
        )
        self.assertContains(response, "Add copy")
        self.assertNotContains(response, "type=\"button\" disabled")

    def test_own_active_copies_get_an_edit_link_on_the_detail_page(self):
        mine = Copy.objects.create(owner=self.user, edition=self.default)
        other = get_user_model().objects.create_user(
            username="other", password="pass",
        )
        theirs = Copy.objects.create(owner=other, edition=self.collectors)

        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertContains(response, f'href="/copies/{mine.pk}/edit/"')
        self.assertNotContains(response, f'href="/copies/{theirs.pk}/edit/"')


class LoanModelTests(TestCase):
    """Issue #43: the Loan constraints (mirrors ShareGrant's exactly-one-
    grantee XOR, models.py:164-209) and the Copy/Game helpers built on it."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="Wingspan")
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)
        cls.copy = Copy.objects.create(owner=cls.user, edition=cls.edition)

    def test_exactly_one_counterparty_required(self):
        lender = get_user_model().objects.create_user(username="pavel")
        with self.assertRaises(IntegrityError):
            Loan.objects.create(
                copy=self.copy, direction=Loan.Direction.BORROWED_IN,
                counterparty_user=lender, counterparty_name="Pavel",
            )

    def test_at_least_one_counterparty_required(self):
        with self.assertRaises(IntegrityError):
            Loan.objects.create(copy=self.copy, direction=Loan.Direction.LENT_OUT)

    def test_only_one_active_loan_per_copy(self):
        Loan.objects.create(
            copy=self.copy, direction=Loan.Direction.LENT_OUT,
            counterparty_name="Pavel",
        )
        with self.assertRaises(IntegrityError):
            Loan.objects.create(
                copy=self.copy, direction=Loan.Direction.LENT_OUT,
                counterparty_name="Mira",
            )

    def test_a_returned_loan_frees_the_copy_for_a_new_one(self):
        first = Loan.objects.create(
            copy=self.copy, direction=Loan.Direction.LENT_OUT,
            counterparty_name="Pavel",
        )
        first.returned_at = timezone.localdate()
        first.save(update_fields=["returned_at"])
        second = Loan.objects.create(
            copy=self.copy, direction=Loan.Direction.LENT_OUT,
            counterparty_name="Mira",
        )
        self.assertEqual(self.copy.active_loan, second)

    def test_active_loan_ignores_returned_loans(self):
        self.assertIsNone(self.copy.active_loan)
        loan = Loan.objects.create(
            copy=self.copy, direction=Loan.Direction.BORROWED_IN,
            counterparty_name="Pavel",
        )
        self.assertEqual(self.copy.active_loan, loan)
        loan.returned_at = timezone.localdate()
        loan.save(update_fields=["returned_at"])
        self.assertIsNone(Copy.objects.get(pk=self.copy.pk).active_loan)

    def test_is_owned_excludes_a_borrowed_in_copy(self):
        game = Game.objects.create(name="Borrowed Only")
        edition = Edition.objects.create(game=game, is_default=True)
        copy = Copy.objects.create(
            owner=self.user, edition=edition, is_borrowed_in=True,
        )
        Loan.objects.create(
            copy=copy, direction=Loan.Direction.BORROWED_IN,
            counterparty_name="Pavel",
        )
        self.assertFalse(Game.objects.get(pk=game.pk).is_owned)

    def test_owning_and_borrowing_the_same_edition_is_allowed(self):
        # Issue #43: the unique constraint only applies among non-borrowed
        # copies, so a borrowed-in duplicate of an owned edition is fine.
        borrowed = Copy.objects.create(
            owner=self.user, edition=self.edition, is_borrowed_in=True,
        )
        self.assertEqual(
            Copy.objects.filter(owner=self.user, edition=self.edition).count(), 2,
        )
        self.assertNotEqual(borrowed.pk, self.copy.pk)


class CopyAddBorrowedViewTests(TestCase):
    """Issue #43: "I'm borrowing this" — copy_add's reverse-direction
    sibling."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.lender = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.game = Game.objects.create(name="Epic Quest")
        cls.default = Edition.objects.create(game=cls.game, is_default=True)

    def setUp(self):
        self.client.login(username="kernicek", password="pass")

    def post(self, pk=None, data=None):
        return self.client.post(
            f"/games/{pk or self.game.pk}/copies/add-borrowed/", data or {})

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.post()
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_creates_a_borrowed_copy_and_redirects_to_its_edit_page(self):
        response = self.post(data={
            "edition": str(self.default.pk), "counterparty": "Mira",
        })
        copy = Copy.objects.get(owner=self.user)
        self.assertRedirects(response, f"/copies/{copy.pk}/edit/")
        self.assertTrue(copy.is_borrowed_in)
        loan = copy.active_loan
        self.assertEqual(loan.direction, Loan.Direction.BORROWED_IN)
        self.assertEqual(loan.counterparty_name, "Mira")
        self.assertIsNone(loan.counterparty_user)

    def test_counterparty_resolves_to_an_existing_username(self):
        self.post(data={"edition": str(self.default.pk), "counterparty": "pavel"})
        loan = Copy.objects.get(owner=self.user).active_loan
        self.assertEqual(loan.counterparty_user, self.lender)
        self.assertEqual(loan.counterparty_name, "")

    def test_missing_counterparty_is_400(self):
        response = self.post(data={"edition": str(self.default.pk)})
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Copy.objects.exists())

    def test_never_pushes_to_bgg(self):
        # Issue #117/#43: a borrowed copy is present but not owned.
        BggLink.objects.create(game=self.game, bgg_id=207830, is_primary=True)
        with mock.patch("gamekeeper.views.push_bgg_status_task.delay") as delay:
            self.post(data={"edition": str(self.default.pk), "counterparty": "Mira"})
        delay.assert_not_called()

    def test_can_borrow_an_edition_the_user_already_owns(self):
        Copy.objects.create(owner=self.user, edition=self.default)
        response = self.post(data={
            "edition": str(self.default.pk), "counterparty": "Mira",
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            Copy.objects.filter(owner=self.user, edition=self.default).count(), 2,
        )

    def test_can_borrow_the_same_edition_from_two_different_lenders(self):
        self.post(data={"edition": str(self.default.pk), "counterparty": "Mira"})
        response = self.post(data={
            "edition": str(self.default.pk), "counterparty": "Petr",
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            Copy.objects.filter(owner=self.user, edition=self.default).count(), 2,
        )


class CopyLoanViewTests(TestCase):
    """Issue #43: lending an owned copy out and returning either direction."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(name="Wingspan")
        cls.edition = Edition.objects.create(game=cls.game, is_default=True)

    def setUp(self):
        self.client.login(username="kernicek", password="pass")
        self.copy = Copy.objects.create(owner=self.user, edition=self.edition)

    def loan_out(self, pk=None, data=None):
        return self.client.post(
            f"/copies/{pk or self.copy.pk}/loan-out/", data or {})

    def loan_return(self, pk=None):
        return self.client.post(f"/copies/{pk or self.copy.pk}/loan-return/")

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.loan_out()
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_lend_out_creates_a_loan_and_redirects_to_edit(self):
        response = self.loan_out(data={
            "counterparty": "Mira", "since": "2026-07-01",
            "expected_return_date": "2026-08-01",
        })
        self.assertRedirects(response, f"/copies/{self.copy.pk}/edit/")
        loan = self.copy.active_loan
        self.assertEqual(loan.direction, Loan.Direction.LENT_OUT)
        self.assertEqual(loan.counterparty_name, "Mira")
        self.assertEqual(str(loan.since), "2026-07-01")
        self.assertEqual(str(loan.expected_return_date), "2026-08-01")

    def test_missing_counterparty_is_400(self):
        self.assertEqual(self.loan_out().status_code, 400)

    def test_already_on_loan_is_400(self):
        self.loan_out(data={"counterparty": "Mira"})
        response = self.loan_out(data={"counterparty": "Petr"})
        self.assertEqual(response.status_code, 400)

    def test_a_borrowed_in_copy_cannot_be_lent_out(self):
        borrowed = Copy.objects.create(
            owner=self.user, edition=Edition.objects.create(
                game=self.game, name="Borrowed Ed"),
            is_borrowed_in=True,
        )
        response = self.loan_out(pk=borrowed.pk, data={"counterparty": "Mira"})
        self.assertEqual(response.status_code, 404)

    def test_returning_a_lent_out_copy_just_clears_the_loan(self):
        self.loan_out(data={"counterparty": "Mira"})
        response = self.loan_return()
        self.assertRedirects(response, f"/copies/{self.copy.pk}/edit/")
        self.copy.refresh_from_db()
        self.assertIsNone(self.copy.active_loan)
        self.assertEqual(self.copy.archive_status, Copy.ArchiveStatus.ACTIVE)

    def test_returning_a_borrowed_in_copy_archives_it(self):
        borrowed = Copy.objects.create(
            owner=self.user, edition=Edition.objects.create(
                game=self.game, name="Borrowed Ed"),
            is_borrowed_in=True,
        )
        Loan.objects.create(
            copy=borrowed, direction=Loan.Direction.BORROWED_IN,
            counterparty_name="Mira",
        )
        response = self.loan_return(pk=borrowed.pk)
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        borrowed.refresh_from_db()
        self.assertEqual(borrowed.archive_status, Copy.ArchiveStatus.ARCHIVED)
        self.assertEqual(borrowed.archive_reason, Copy.ArchiveReason.RETURNED)
        # active_loan is None post-return by design; the Loan row itself
        # survives (returned_at set) as the loan history.
        self.assertIsNone(borrowed.active_loan)
        self.assertIsNotNone(Loan.objects.get(copy=borrowed).returned_at)

    def test_nothing_to_return_is_400(self):
        self.assertEqual(self.loan_return().status_code, 400)

    def test_not_the_owner_is_404(self):
        other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        theirs = Copy.objects.create(owner=other, edition=Edition.objects.create(
            game=self.game, name="Someone Else's Edition"))
        response = self.loan_out(pk=theirs.pk, data={"counterparty": "Mira"})
        self.assertEqual(response.status_code, 404)


class CopyEditViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="other", password="pass",
        )
        cls.game = Game.objects.create(name="Epic Quest")
        cls.default = Edition.objects.create(game=cls.game, is_default=True)
        cls.collectors = Edition.objects.create(
            game=cls.game, name="Collector's Edition",
        )
        cls.shelf = Location.objects.create(
            group=cls.user.membership.group, name="Kallax A",
        )
        cls.foreign_shelf = Location.objects.create(
            group=cls.other.membership.group, name="Elsewhere",
        )
        cls.copy = Copy.objects.create(owner=cls.user, edition=cls.default)
        cls.foreign_copy = Copy.objects.create(
            owner=cls.other, edition=cls.collectors,
        )
        # Issue #77: an expansion copy, to exercise the non-blocking
        # location warning.
        cls.expansion_game = Game.objects.create(
            name="Epic Quest: Add-on", type=Game.Type.EXPANSION,
        )
        cls.expansion_edition = Edition.objects.create(
            game=cls.expansion_game, is_default=True,
        )
        cls.expansion_copy = Copy.objects.create(
            owner=cls.user, edition=cls.expansion_edition,
        )

    def setUp(self):
        self.client.login(username="kernicek", password="pass")

    def form_data(self, **overrides):
        data = {
            "edition": str(self.copy.edition_id),
            "acquired_date": "", "location": "", "location_note": "",
            "insert_3d": "none", "card_dividers": "none",
            "accessories_3d": "none", "other_accessories": "none",
            "upgrades_note": "", "notes": "",
            # Curation (issues #31, #56) — merged onto this one save (#136).
            # immune is a checkbox: absent means unchecked, so it's not
            # defaulted here (pass immune="on" to check it).
            "excitement": "", "keep_status": "", "why_might_leave": "",
        }
        data.update(overrides)
        return data

    def post(self, pk=None, **overrides):
        return self.client.post(
            f"/copies/{pk or self.copy.pk}/edit/", self.form_data(**overrides),
        )

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.client.get(f"/copies/{self.copy.pk}/edit/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_foreign_archived_and_unknown_copies_are_404(self):
        self.assertEqual(
            self.client.get(f"/copies/{self.foreign_copy.pk}/edit/").status_code,
            404,
        )
        self.copy.archive_status = Copy.ArchiveStatus.ARCHIVED
        self.copy.save(update_fields=["archive_status"])
        self.assertEqual(
            self.client.get(f"/copies/{self.copy.pk}/edit/").status_code, 404,
        )
        self.assertEqual(self.client.get("/copies/99999/edit/").status_code, 404)

    def test_edit_page_renders_the_form(self):
        response = self.client.get(f"/copies/{self.copy.pk}/edit/")
        self.assertContains(response, "My copy — Epic Quest")
        self.assertContains(response, "default edition")
        self.assertContains(response, "Collector&#x27;s Edition")
        self.assertContains(response, "Kallax A")
        self.assertNotContains(response, "Elsewhere")  # other group's location

    def test_curation_fields_prefill_and_render(self):
        # Issues #31, #56: the §11 cull signals live on this same page.
        self.copy.excitement = Decimal("6.0")
        self.copy.immune = True
        self.copy.save()
        response = self.client.get(f"/copies/{self.copy.pk}/edit/")
        self.assertContains(response, "Curation")
        self.assertContains(response, 'value="6.0"')
        self.assertContains(response, "Immune from culling")

    def test_game_edit_page_does_not_show_curation(self):
        # Issues #56/#136: curation lives solely on the copy edit page.
        response = self.client.get(f"/games/{self.game.pk}/edit/")
        self.assertNotContains(response, "Save curation")

    def test_save_updates_fields_and_redirects_to_the_game(self):
        # Issue #136: campaign/details fields and curation fields save
        # together in one POST — the actual point of the unification.
        response = self.post(
            edition=str(self.collectors.pk), acquired_date="2024-05-01",
            location=str(self.shelf.pk), location_note="bottom row",
            insert_3d="done", card_dividers="todo", accessories_3d="maybe",
            other_accessories="not_necessary",
            upgrades_note="metal coins", notes="signed by the designer",
            excitement="7.5", keep_status=Copy.KeepStatus.MIGHT_CYCLE,
            why_might_leave="shelf hog", immune="on",
        )
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.copy.refresh_from_db()
        self.assertEqual(self.copy.edition, self.collectors)
        self.assertEqual(self.copy.acquired_date, datetime.date(2024, 5, 1))
        self.assertEqual(self.copy.location, self.shelf)
        self.assertEqual(self.copy.location_note, "bottom row")
        self.assertEqual(self.copy.insert_3d, Copy.UpgradeStatus.DONE)
        self.assertEqual(self.copy.card_dividers, Copy.UpgradeStatus.TODO)
        self.assertEqual(self.copy.accessories_3d, Copy.UpgradeStatus.MAYBE)
        self.assertEqual(self.copy.other_accessories,
                         Copy.UpgradeStatus.NOT_NECESSARY)
        self.assertEqual(self.copy.upgrades_note, "metal coins")
        self.assertEqual(self.copy.notes, "signed by the designer")
        self.assertEqual(self.copy.excitement, Decimal("7.5"))
        self.assertEqual(self.copy.keep_status, Copy.KeepStatus.MIGHT_CYCLE)
        self.assertEqual(self.copy.why_might_leave, "shelf hog")
        self.assertTrue(self.copy.immune)

    def test_marking_will_leave_pushes_fortrade(self):
        # Issue #82: same marker/push as the curation table, reached from
        # the copy edit page's merged save (#136).
        BggLink.objects.create(game=self.game, bgg_id=300200, is_primary=True)
        with mock.patch("gamekeeper.views.push_bgg_fortrade_task.delay") as delay:
            self.post(keep_status=Copy.KeepStatus.WILL_LEAVE)
        delay.assert_called_once_with(self.game.pk, True, self.user.pk)

    def test_saving_without_a_bgglink_does_not_push(self):
        with mock.patch("gamekeeper.views.push_bgg_fortrade_task.delay") as delay:
            self.post(keep_status=Copy.KeepStatus.WILL_LEAVE)
        delay.assert_not_called()

    def test_a_copy_reached_from_a_purchase_returns_to_it_on_save(self):
        # Issue #45: converting a purchase item lands here with
        # ?from_purchase=<pk>; the origin rides a hidden field so save returns
        # to the purchase (to handle its remaining items), not the game.
        purchase = Purchase.objects.create(
            owner=self.user, name="Origin KS", status=Purchase.Status.COMMITTED,
        )
        get = self.client.get(
            f"/copies/{self.copy.pk}/edit/?from_purchase={purchase.pk}")
        self.assertContains(
            get, f'name="return_purchase" value="{purchase.pk}"')
        self.assertContains(get, "Save &amp; back to purchase")

        response = self.post(return_purchase=str(purchase.pk))
        self.assertRedirects(response, f"/purchases/{purchase.pk}/")

    def test_a_foreign_or_bogus_from_purchase_falls_back_to_the_game(self):
        # A from_purchase that isn't one of my purchases is ignored — no hidden
        # field on GET, and a forged one on POST still lands on the game.
        foreign = Purchase.objects.create(
            owner=self.other, name="Pavel's KS", status=Purchase.Status.COMMITTED,
        )
        get = self.client.get(
            f"/copies/{self.copy.pk}/edit/?from_purchase={foreign.pk}")
        self.assertNotContains(get, 'name="return_purchase"')

        response = self.post(return_purchase=str(foreign.pk))
        self.assertRedirects(response, f"/games/{self.game.pk}/")

    def test_blank_fields_clear_them(self):
        self.copy.acquired_date = datetime.date(2020, 1, 1)
        self.copy.location = self.shelf
        self.copy.excitement = Decimal("6.0")
        self.copy.keep_status = Copy.KeepStatus.KEEP
        self.copy.immune = True
        self.copy.save()

        self.post()
        self.copy.refresh_from_db()
        self.assertIsNone(self.copy.acquired_date)
        self.assertIsNone(self.copy.location)
        self.assertIsNone(self.copy.excitement)
        self.assertEqual(self.copy.keep_status, "")
        # Whole-form checkbox semantics: absent means unchecked.
        self.assertFalse(self.copy.immune)

    def test_bad_values_are_rejected_inline(self):
        # CopyForm (issue #136): invalid input re-renders at 200 with
        # inline errors instead of a blunt 400.
        for overrides in (
            {"acquired_date": "soonish"},
            {"edition": "99999"},
            {"location": str(self.foreign_shelf.pk)},
            {"insert_3d": "golden"},
            {"excitement": "hot"},
            {"excitement": "11"},
            {"keep_status": "forever"},
        ):
            response = self.post(**overrides)
            self.assertEqual(response.status_code, 200, overrides)

    def test_switching_to_an_edition_i_already_own_is_rejected_inline(self):
        Copy.objects.create(owner=self.user, edition=self.collectors)
        # The queryset excludes already-owned editions (below), so a
        # tampered POST for one fails Django's own "not a valid choice"
        # field validation rather than reaching CopyForm.clean_edition's
        # custom message (that one only fires on the underlying race).
        response = self.post(edition=str(self.collectors.pk))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "not one of the available choices")

        # And the select does not offer it in the first place.
        response = self.client.get(f"/copies/{self.copy.pk}/edit/")
        self.assertNotContains(response, f'value="{self.collectors.pk}"')

    def test_expansion_copy_page_carries_the_location_warning(self):
        # Issue #77: the JS confirm modal only renders for expansion copies.
        response = self.client.get(f"/copies/{self.expansion_copy.pk}/edit/")
        self.assertContains(response, "expansion-location-modal")
        self.assertContains(
            response, "Expansions usually travel with their base game")

    def test_base_game_copy_page_has_no_location_warning(self):
        response = self.client.get(f"/copies/{self.copy.pk}/edit/")
        self.assertNotContains(response, "expansion-location-modal")
        self.assertNotContains(
            response, "Expansions usually travel with their base game")

    def test_setting_a_location_on_an_expansion_copy_still_saves(self):
        # The warning is non-blocking — the Django test client never runs
        # the JS modal, so this proves the save is never server-side gated.
        response = self.post(
            pk=self.expansion_copy.pk,
            edition=str(self.expansion_edition.pk),
            location=str(self.shelf.pk),
        )
        self.assertRedirects(
            response, f"/games/{self.expansion_game.pk}/")
        self.expansion_copy.refresh_from_db()
        self.assertEqual(self.expansion_copy.location, self.shelf)


class EditionAddViewTests(TestCase):
    """The edition create page (issue #53): /games/<pk>/editions/add/."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(name="Epic Quest")
        cls.default = Edition.objects.create(game=cls.game, is_default=True)

    def setUp(self):
        self.client.login(username="kernicek", password="pass")

    def form_data(self, **overrides):
        data = {
            "name": "Kickstarter Edition", "bgg_version_id": "",
            "components_language": "", "size_category": "", "num_boxes": "",
            "box_length_mm": "", "box_width_mm": "", "box_height_mm": "",
        }
        data.update(overrides)
        return data

    def post(self, pk=None, **overrides):
        return self.client.post(
            f"/games/{pk or self.game.pk}/editions/add/",
            self.form_data(**overrides),
        )

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.client.get(f"/games/{self.game.pk}/editions/add/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_unknown_game_is_404(self):
        self.assertEqual(
            self.client.get("/games/99999/editions/add/").status_code, 404,
        )

    def test_add_page_renders_the_form(self):
        response = self.client.get(f"/games/{self.game.pk}/editions/add/")
        self.assertContains(response, "New edition — Epic Quest")
        self.assertContains(
            response, f'action="/games/{self.game.pk}/editions/add/"',
        )

    def test_add_creates_the_edition_and_redirects_to_the_game(self):
        response = self.post(
            bgg_version_id="123456", components_language="en",
            size_category="huge", num_boxes="3", box_length_mm="400",
            box_width_mm="300", box_height_mm="200",
        )
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        edition = self.game.editions.get(name="Kickstarter Edition")
        self.assertFalse(edition.is_default)
        self.assertEqual(edition.bgg_version_id, 123456)
        self.assertEqual(edition.components_language,
                         Edition.ComponentsLanguage.ENGLISH)
        self.assertEqual(edition.size_category, Edition.SizeCategory.HUGE)
        self.assertEqual(edition.num_boxes, 3)
        self.assertEqual(
            (edition.box_length_mm, edition.box_width_mm,
             edition.box_height_mm),
            (400, 300, 200),
        )

    def test_blank_name_is_allowed(self):
        response = self.post(name="")
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.assertEqual(self.game.editions.count(), 2)

    def test_bad_numbers_and_choices_are_400(self):
        self.assertEqual(self.post(num_boxes="many").status_code, 400)
        self.assertEqual(self.post(box_length_mm="-5").status_code, 400)
        self.assertEqual(
            self.post(components_language="klingon").status_code, 400,
        )
        self.assertEqual(self.post(size_category="colossal").status_code, 400)
        self.assertEqual(self.game.editions.count(), 1)

    def test_creating_as_default_needs_confirmation_when_one_exists(self):
        # The modal's ride-along is missing (JS bypassed): 400, no write.
        self.assertEqual(self.post(is_default="on").status_code, 400)
        self.assertEqual(self.game.editions.count(), 1)

        response = self.post(is_default="on", confirm_default_switch="1",
                             old_default_name="Retail Edition")
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.default.refresh_from_db()
        self.assertFalse(self.default.is_default)
        self.assertEqual(self.default.name, "Retail Edition")
        self.assertTrue(
            self.game.editions.get(name="Kickstarter Edition").is_default,
        )

    def test_creating_a_default_without_clash_needs_no_confirmation(self):
        bare = Game.objects.create(name="Editionless")
        response = self.post(pk=bare.pk, is_default="on")
        self.assertRedirects(response, f"/games/{bare.pk}/")
        self.assertTrue(bare.editions.get().is_default)

    def test_detail_page_lists_editions_with_edit_links_and_add_button(self):
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertContains(response, "Editions")
        self.assertContains(
            response, f'href="/games/{self.game.pk}/editions/add/"',
        )
        self.assertContains(
            response, f'href="/editions/{self.default.pk}/edit/"',
        )


class EditionEditViewTests(TestCase):
    """The edition edit page (issue #53): /editions/<pk>/edit/."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.game = Game.objects.create(name="Epic Quest")
        cls.default = Edition.objects.create(game=cls.game, is_default=True)
        cls.collectors = Edition.objects.create(
            game=cls.game, name="Collector's Edition",
            components_language=Edition.ComponentsLanguage.ENGLISH,
            size_category=Edition.SizeCategory.LARGE,
            num_boxes=2, box_length_mm=300,
        )

    def setUp(self):
        self.client.login(username="kernicek", password="pass")

    def form_data(self, **overrides):
        data = {
            "name": "Collector's Edition", "bgg_version_id": "",
            "components_language": "en", "size_category": "large",
            "num_boxes": "2", "box_length_mm": "300",
            "box_width_mm": "", "box_height_mm": "",
        }
        data.update(overrides)
        return data

    def post(self, pk=None, **overrides):
        return self.client.post(
            f"/editions/{pk or self.collectors.pk}/edit/",
            self.form_data(**overrides),
        )

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.client.get(f"/editions/{self.collectors.pk}/edit/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_unknown_edition_is_404(self):
        self.assertEqual(
            self.client.get("/editions/99999/edit/").status_code, 404,
        )

    def test_edit_page_renders_current_values(self):
        response = self.client.get(f"/editions/{self.collectors.pk}/edit/")
        self.assertContains(response, "Edit edition — Epic Quest")
        self.assertContains(response, "Collector&#x27;s Edition")
        self.assertContains(response, 'value="300"')
        # A different default exists, so the switch-confirm modal renders.
        self.assertContains(response, "default-switch-modal")

        # Editing the default itself: no possible clash, no modal.
        response = self.client.get(f"/editions/{self.default.pk}/edit/")
        self.assertNotContains(response, "default-switch-modal")

    def test_save_updates_fields_and_redirects_to_the_game(self):
        response = self.post(
            name="Deluxe Edition", bgg_version_id="654321",
            components_language="cs_en", size_category="medium",
            num_boxes="1", box_length_mm="310", box_width_mm="220",
            box_height_mm="70",
        )
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.collectors.refresh_from_db()
        self.assertEqual(self.collectors.name, "Deluxe Edition")
        self.assertEqual(self.collectors.bgg_version_id, 654321)
        self.assertEqual(self.collectors.components_language,
                         Edition.ComponentsLanguage.CZECH_ENGLISH)
        self.assertEqual(self.collectors.size_category,
                         Edition.SizeCategory.MEDIUM)
        self.assertEqual(self.collectors.num_boxes, 1)
        self.assertEqual(
            (self.collectors.box_length_mm, self.collectors.box_width_mm,
             self.collectors.box_height_mm),
            (310, 220, 70),
        )

    def test_blank_inputs_clear_the_fields(self):
        self.post(name="", components_language="", size_category="",
                  num_boxes="", box_length_mm="")
        self.collectors.refresh_from_db()
        self.assertEqual(self.collectors.name, "")
        self.assertEqual(self.collectors.components_language, "")
        self.assertEqual(self.collectors.size_category, "")
        self.assertIsNone(self.collectors.num_boxes)
        self.assertIsNone(self.collectors.box_length_mm)

    def test_bad_numbers_and_choices_are_400(self):
        self.assertEqual(self.post(num_boxes="many").status_code, 400)
        self.assertEqual(self.post(bgg_version_id="-1").status_code, 400)
        self.assertEqual(
            self.post(components_language="klingon").status_code, 400,
        )
        self.assertEqual(self.post(size_category="colossal").status_code, 400)
        self.collectors.refresh_from_db()
        self.assertEqual(self.collectors.num_boxes, 2)

    def test_default_switch_needs_confirmation(self):
        self.assertEqual(self.post(is_default="on").status_code, 400)
        self.default.refresh_from_db()
        self.collectors.refresh_from_db()
        self.assertTrue(self.default.is_default)
        self.assertFalse(self.collectors.is_default)

    def test_confirmed_default_switch_demotes_and_renames_the_old_default(self):
        response = self.post(is_default="on", confirm_default_switch="1",
                             old_default_name="First printing")
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.default.refresh_from_db()
        self.collectors.refresh_from_db()
        self.assertFalse(self.default.is_default)
        self.assertEqual(self.default.name, "First printing")
        self.assertTrue(self.collectors.is_default)

    def test_saving_the_default_as_default_needs_no_confirmation(self):
        response = self.post(pk=self.default.pk, name="", is_default="on",
                             components_language="", size_category="",
                             num_boxes="", box_length_mm="")
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.default.refresh_from_db()
        self.assertTrue(self.default.is_default)

    def test_unchecking_default_just_saves(self):
        response = self.post(pk=self.default.pk, name="", num_boxes="",
                             components_language="", size_category="",
                             box_length_mm="")
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.default.refresh_from_db()
        self.assertFalse(self.default.is_default)

    def test_edit_page_renders_the_pnp_checkbox(self):
        # PnP is an edition-level flag (#138), edited here rather than on the game.
        self.collectors.is_pnp = True
        self.collectors.save(update_fields=["is_pnp"])
        response = self.client.get(f"/editions/{self.collectors.pk}/edit/")
        self.assertContains(response, 'id="edition-pnp"')
        self.assertContains(response, 'name="is_pnp" checked')

    def test_save_sets_and_clears_pnp(self):
        # Absent checkbox on the whole-form POST reads as unchecked (safe here).
        self.post()
        self.collectors.refresh_from_db()
        self.assertFalse(self.collectors.is_pnp)
        # A present checkbox sets it.
        self.post(is_pnp="on")
        self.collectors.refresh_from_db()
        self.assertTrue(self.collectors.is_pnp)

    def test_has_pnp_edition_reflects_per_edition_flags(self):
        # A game with both a PnP edition and a store edition is PnP at the game
        # level only via "has any PnP edition" (#138) — never one flag per title.
        self.assertFalse(self.game.has_pnp_edition)
        self.collectors.is_pnp = True
        self.collectors.save(update_fields=["is_pnp"])
        self.assertTrue(self.game.has_pnp_edition)
        self.assertFalse(self.default.is_pnp)


class SeriesModelTests(TestCase):
    """DESIGN §4 Series (issue #21): the cover_source fallback and the
    duck-typed cover_art.html contract for custom series art."""

    @classmethod
    def setUpTestData(cls):
        cls.crime_city = Game.objects.create(
            name="MicroMacro: Crime City",
            image_url="https://cf.geekdo-images.com/original/mm1.jpg",
            cover_focus_x=30, cover_focus_y=70,
        )
        cls.full_house = Game.objects.create(name="MicroMacro: Full House")
        cls.series = Series.objects.create(
            name="MicroMacro", primary_game=cls.crime_city,
        )
        Game.objects.filter(
            pk__in=[cls.crime_city.pk, cls.full_house.pk],
        ).update(series=cls.series)

    def test_cover_source_is_the_primary_member_without_custom_art(self):
        # The primary's full focus/zoom machinery represents the series.
        self.assertEqual(self.series.cover_source, self.crime_city)

    def test_custom_cover_wins_and_duck_types_the_tile_contract(self):
        self.series.cover_image = "covers/micromacro.jpg"
        self.assertIs(self.series.cover_source, self.series)
        self.assertIn("covers/micromacro.jpg", self.series.cover_url)
        # cover_art.html's contract: plain centered crop, no fit backdrop.
        self.assertFalse(self.series.cover_fit)
        self.assertEqual(self.series.cover_focus_x, 50)
        self.assertEqual(self.series.cover_focus_y, 50)
        self.assertEqual(self.series.cover_scale, "1")

    def test_clean_rejects_a_primary_outside_the_members(self):
        self.series.primary_game = Game.objects.create(name="Timeline")
        with self.assertRaises(ValidationError):
            self.series.full_clean()

    def test_series_covers_carry_the_full_fit_machinery(self):
        # Issue #54: CoverArtModel gives series the same focus/zoom/fit
        # treatment as games — including the aspect-aware fit scale.
        self.series.cover_zoom = 60
        self.series.cover_width = 800
        self.series.cover_height = 400
        self.assertTrue(self.series.cover_fit)
        self.assertEqual(self.series.cover_scale, "1.2")


class SeriesCollapseTests(TestCase):
    """Issue #21: the collection grid collapses a series to one tile whose
    overlay lists every member; a series matches the chooser filters when
    ANY member does."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.inventions = Game.objects.create(
            name="Timeline: Inventions", min_players=2, max_players=8,
            image_url="https://cf.geekdo-images.com/original/tl1.jpg",
            cover_focus_x=30, cover_focus_y=60,
        )
        cls.events = Game.objects.create(
            name="Timeline: Events", min_players=2, max_players=4,
        )
        cls.series = Series.objects.create(
            name="Timeline", primary_game=cls.inventions,
        )
        Game.objects.filter(
            pk__in=[cls.inventions.pk, cls.events.pk],
        ).update(series=cls.series)
        cls.standalone = Game.objects.create(
            name="Duel", min_players=2, max_players=2,
        )

    def get(self, params=None, **extra):
        self.client.login(username="kernicek", password="pass")
        # These fixtures carry no copies; show_unavailable opts past the
        # available-only default (issue #107) so collapse is what's under test.
        query = {"show_unavailable": "1"}
        query.update(params or {})
        return self.client.get("/", query, **extra)

    def test_series_collapses_to_one_tile(self):
        # The tile class, not data-series-tile — the page's delegated
        # script mentions the attribute too.
        response = self.get()
        self.assertContains(response, 'class="cover cover-series"', count=1)
        self.assertContains(response, "cover-flag-count")
        # Members render only inside the overlay, never as top-level tiles:
        # the sole top-level <a class="cover"> is the standalone game.
        self.assertContains(response, '<a class="cover" ', count=1)
        # The header keeps counting GAMES, not tiles.
        self.assertContains(response, "3 of 3 games")

    def test_member_count_renders_as_a_corner_badge(self):
        # Issue #15: the count moved from the caption to a bottom-left
        # badge, so it stays readable while names are hover-only.
        response = self.get()
        self.assertContains(response, ">2</span>")
        self.assertContains(response, "2 games in series")
        self.assertNotContains(response, "(2)")

    def test_tile_carries_series_url_for_middle_click(self):
        # Issue #57: the tile is a <div>, so the delegated auxclick/ctrl+click
        # handler needs the target URL on the tile itself.
        response = self.get()
        self.assertContains(
            response, f'data-series-url="/series/{self.series.pk}/"', count=1)

    def test_overlay_links_the_series_page_and_each_member(self):
        response = self.get()
        self.assertContains(response, f"/series/{self.series.pk}/")
        self.assertContains(response, f"/games/{self.inventions.pk}/")
        self.assertContains(response, f"/games/{self.events.pk}/")

    def test_collapsed_tile_uses_the_primary_members_cover(self):
        response = self.get()
        self.assertContains(
            response, "https://cf.geekdo-images.com/original/tl1.jpg")
        self.assertContains(response, "object-position: 30% 60%")

    def test_custom_series_cover_beats_the_primary(self):
        self.series.cover_image = "covers/timeline-custom.jpg"
        self.series.save()
        self.assertContains(self.get(), "covers/timeline-custom.jpg")

    def test_series_matches_when_any_member_matches(self):
        # 8 players: only Inventions (2-8) qualifies — the tile still shows.
        response = self.get({"players": "8"})
        self.assertContains(response, 'class="cover cover-series"', count=1)
        self.assertNotContains(response, "Duel")

    def test_series_hides_when_no_member_matches(self):
        response = self.get({"players": "20"})
        self.assertNotContains(response, "cover-series")

    def test_name_search_by_one_member_surfaces_the_series(self):
        response = self.get({"q": "Events"})
        self.assertContains(response, 'class="cover cover-series"', count=1)

    def test_search_matches_the_series_name_itself(self):
        # Issue #86: searching a series name surfaces its collapsed tile even
        # when no member name contains the query (the MicroMacro case — the
        # existing "Timeline" fixture can't prove this, as its name is a
        # substring of both member names).
        crime = Game.objects.create(name="Crime City")
        house = Game.objects.create(name="Full House")
        micromacro = Series.objects.create(name="MicroMacro", primary_game=crime)
        Game.objects.filter(pk__in=[crime.pk, house.pk]).update(series=micromacro)

        response = self.get({"q": "MicroMacro"})
        self.assertContains(
            response, f'data-series-url="/series/{micromacro.pk}/"', count=1)
        # q-filtering is genuinely active: the standalone game and the
        # unrelated Timeline series (name/members lack "MicroMacro") drop out.
        self.assertNotContains(response, "Duel")
        self.assertNotContains(response, f'data-series-url="/series/{self.series.pk}/"')

    def test_incoming_member_flags_the_collapsed_tile(self):
        purchase = Purchase.objects.create(
            owner=self.user, name="Timeline restock",
            status=Purchase.Status.COMMITTED,
        )
        wave = Wave.objects.create(purchase=purchase, number=1)
        Product.objects.create(wave=wave, name="Timeline: Events",
                               game=self.events)
        self.assertContains(self.get(), "cover-flag-incoming")

    def test_shared_prev_owned_status_flags_the_collapsed_tile(self):
        Game.objects.filter(pk__in=[self.inventions.pk, self.events.pk]).update(
            bgg_collection_status=Game.BggCollectionStatus.PREV_OWNED)
        self.assertContains(self.get(), "cover-flag-prev-owned")

    def test_shared_wishlist_status_flags_the_collapsed_tile(self):
        Game.objects.filter(pk__in=[self.inventions.pk, self.events.pk]).update(
            bgg_collection_status=Game.BggCollectionStatus.WISHLIST)
        self.assertContains(self.get(), "cover-flag-wishlist")

    def test_shared_own_status_flags_the_collapsed_tile(self):
        Game.objects.filter(pk__in=[self.inventions.pk, self.events.pk]).update(
            bgg_collection_status=Game.BggCollectionStatus.OWN)
        self.assertContains(self.get(), "cover-flag-own")

    def test_shared_preordered_status_flags_the_collapsed_tile(self):
        Game.objects.filter(pk__in=[self.inventions.pk, self.events.pk]).update(
            bgg_collection_status=Game.BggCollectionStatus.PREORDERED)
        self.assertContains(self.get(), "cover-flag-preordered")

    def test_mismatched_statuses_show_no_shared_badge(self):
        # Issue #105: no "mixed" treatment — a shared badge only renders
        # when EVERY member matches. Members individually keep showing their
        # own status in the overlay's mini-tiles (issue #8, unaffected here)
        # so this checks only the collapsed tile's own badge area, not the
        # whole page.
        Game.objects.filter(pk=self.inventions.pk).update(
            bgg_collection_status=Game.BggCollectionStatus.PREV_OWNED)
        Game.objects.filter(pk=self.events.pk).update(
            bgg_collection_status=Game.BggCollectionStatus.WISHLIST)
        response = self.get()
        tile_html = response.content.decode().split('<div class="series-overlay">')[0]
        for status_class in ("cover-flag-own", "cover-flag-preordered",
                              "cover-flag-prev-owned", "cover-flag-wishlist"):
            self.assertNotIn(status_class, tile_html)

    def test_blank_statuses_show_no_shared_badge(self):
        # Both members carry their default blank status — still "all equal"
        # but there's nothing to badge.
        response = self.get()
        for status_class in ("cover-flag-own", "cover-flag-preordered",
                              "cover-flag-prev-owned", "cover-flag-wishlist"):
            self.assertNotContains(response, status_class)

    def test_htmx_partial_carries_the_overlay(self):
        response = self.get(HTTP_HX_REQUEST="true")
        self.assertContains(response, "data-series-tile")
        self.assertContains(response, "series-overlay")

    def test_list_view_collapses_series_to_one_row_with_member_count(self):
        # Issue #92: the list view's series row reuses the grid's own
        # wording/count so this stays consistent between the two views.
        response = self.get({"view": "list"})
        self.assertContains(response, "2 games in series", count=1)
        self.assertContains(response, f'href="/series/{self.series.pk}/"')

    def test_list_row_uses_primary_members_stats(self):
        # Issue #92: a series row can't show two members' differing stats at
        # once, so it commits to the primary member — the same
        # representative-game concept the grid uses for the cover.
        # update(), not .save() on the cached instance: setUpTestData set
        # series via a bypassing update() too, so the in-memory instances
        # never picked up series_id — a plain .save() here would clobber it.
        Game.objects.filter(pk=self.inventions.pk).update(
            weight=Decimal("2.5"), year_published=2015)
        Game.objects.filter(pk=self.events.pk).update(
            weight=Decimal("4.0"), year_published=2021)

        response = self.get({"view": "list"})
        self.assertContains(response, "2015")
        self.assertContains(response, "2.50 / 5")
        self.assertNotContains(response, "2021")
        self.assertNotContains(response, "4.00 / 5")


class SeriesDetailViewTests(TestCase):
    """Issue #21: the series page — member grid, primary-member stats,
    Copies/Purchases unions over members, plays summed with breakdown."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.dale1 = Game.objects.create(
            name="Dale of Merchants", min_players=2, max_players=4,
            min_playtime=30, max_playtime=45, bgg_numplays=12,
        )
        cls.dale2 = Game.objects.create(
            name="Dale of Merchants 2", bgg_numplays=5,
        )
        cls.dale3 = Game.objects.create(name="Dale of Merchants 3")
        cls.series = Series.objects.create(
            name="Dale of Merchants", primary_game=cls.dale1,
        )
        Game.objects.filter(
            pk__in=[cls.dale1.pk, cls.dale2.pk, cls.dale3.pk],
        ).update(series=cls.series)

        edition1 = Edition.objects.create(game=cls.dale1, is_default=True)
        Copy.objects.create(owner=cls.user, edition=edition1)
        edition2 = Edition.objects.create(
            game=cls.dale2, is_default=True, name="Second printing",
        )
        Copy.objects.create(owner=cls.user, edition=edition2)

        # A purchase of MINE containing member 2 — must pool in; another
        # user's purchase of the primary must not (owner-scoped like §6).
        mine = Purchase.objects.create(
            owner=cls.user, name="Snowdale order",
            status=Purchase.Status.COMMITTED,
        )
        Product.objects.create(
            wave=Wave.objects.create(purchase=mine, number=1),
            name="Dale 2", game=cls.dale2,
        )
        theirs = Purchase.objects.create(
            owner=cls.other, name="Pavel's Dale order",
            status=Purchase.Status.COMMITTED,
        )
        Product.objects.create(
            wave=Wave.objects.create(purchase=theirs, number=1),
            name="Dale 1", game=cls.dale1,
        )

        # Issue #127: expansions of the members, pooled onto the series page.
        # One owned (an active Copy exists) on member 1, one unowned on member 2.
        cls.exp_owned = Game.objects.create(
            name="Dale of Merchants: Bonus Deck", type=Game.Type.EXPANSION,
        )
        cls.exp_owned.expands.add(cls.dale1)
        exp_edition = Edition.objects.create(game=cls.exp_owned, is_default=True)
        Copy.objects.create(owner=cls.user, edition=exp_edition)
        cls.exp_unowned = Game.objects.create(
            name="Dale of Merchants 2: Extra Guilds", type=Game.Type.EXPANSION,
        )
        cls.exp_unowned.expands.add(cls.dale2)

        # Issue #133: actual Play rows on two different members, pooled into the
        # recent-plays feed (distinct from the summed bgg_numplays count above).
        now = timezone.now()
        play1 = Play.objects.create(
            game=cls.dale1, external_id="900",
            play_date=datetime.date(2024, 6, 1), synced_at=now,
        )
        PlayPlayer.objects.create(play=play1, name="Alice", won=True)
        play2 = Play.objects.create(
            game=cls.dale2, external_id="901",
            play_date=datetime.date(2024, 5, 1), synced_at=now,
        )
        PlayPlayer.objects.create(play=play2, name="Bob")

    def get(self, pk=None):
        self.client.login(username="kernicek", password="pass")
        return self.client.get(f"/series/{pk or self.series.pk}/")

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get(f"/series/{self.series.pk}/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_unknown_series_is_404(self):
        self.assertEqual(self.get(99999).status_code, 404)

    def test_member_grid_links_every_member(self):
        response = self.get()
        self.assertContains(response, "Games in this series")
        for member in (self.dale1, self.dale2, self.dale3):
            self.assertContains(response, f"/games/{member.pk}/")

    def test_hero_stats_come_from_the_primary_member(self):
        response = self.get()
        self.assertContains(response, "2–4")
        self.assertContains(response, "30–45 min")
        self.assertContains(response, f"/series/{self.series.pk}/edit/")

    def test_primary_bgg_rank_links_to_the_browse_page(self):
        # Issue #39: the primary member's rank links to BGG's rank-browse page.
        self.dale1.bgg_rating = "7.100"
        self.dale1.bgg_rank = 842
        self.dale1.save()
        self.assertContains(
            self.get(),
            "https://boardgamegeek.com/browse/boardgame?sort=rank"
            "&amp;rankobjecttype=subtype&amp;rankobjectid=1"
            "&amp;rank=842#842",
        )

    def test_copies_union_names_each_members_game(self):
        response = self.get()
        # show_game adds the Game column so pooled rows stay attributable.
        self.assertContains(response, "<th>Game</th>")
        self.assertContains(response, "Second printing")
        self.assertContains(response, "Copies <span")

    def test_purchases_union_is_owner_scoped(self):
        response = self.get()
        self.assertContains(response, "Snowdale order")
        self.assertNotContains(response, "Pavel&#x27;s Dale order")

    def test_plays_sum_with_per_member_breakdown(self):
        response = self.get()
        self.assertContains(response, "logged on BGG")
        self.assertContains(response, "Dale of Merchants: 12")
        self.assertContains(response, "Dale of Merchants 2: 5")
        # No plays — omitted from the breakdown.
        self.assertNotContains(response, "Dale of Merchants 3: ")

    def test_game_detail_links_back_via_a_series_badge(self):
        self.client.login(username="kernicek", password="pass")
        response = self.client.get(f"/games/{self.dale2.pk}/")
        self.assertContains(response, f"/series/{self.series.pk}/")
        self.assertContains(response, "Series: Dale of Merchants")

    def test_expansions_pool_across_members_naming_the_base(self):
        # Issue #127: both members' expansions surface, each row linking to its
        # own game page and naming the base member it expands. Badge counts owned.
        response = self.get()
        self.assertContains(response, "Expansions <span")
        self.assertContains(response, f"/games/{self.exp_owned.pk}/")
        self.assertContains(response, f"/games/{self.exp_unowned.pk}/")
        self.assertContains(response, "(Dale of Merchants)")
        self.assertContains(response, "(Dale of Merchants 2)")
        self.assertEqual(response.context["owned_expansion_count"], 1)

    def test_unowned_member_expansion_shown_muted(self):
        # All expansions show (consistent with the game page), unowned muted.
        response = self.get()
        self.assertContains(response, "not owned")
        self.assertContains(response, "kernicek")  # owner chip on the owned one

    def test_recent_plays_feed_pools_actual_play_rows_across_members(self):
        # Issue #133: real Play rows pooled across members, most-recent-first,
        # naming each member game — distinct from the summed bgg_numplays count.
        response = self.get()
        self.assertContains(response, "Recent plays")
        self.assertEqual(len(response.context["recent_plays"]), 2)
        self.assertEqual(response.context["plays_count"], 2)
        self.assertContains(response, "Alice")
        self.assertContains(response, "Bob")


class SeriesListViewTests(TestCase):
    """Issue #80: the series overview grid — cover, name, member count, each
    tile linking to its detail page."""

    @classmethod
    def setUpTestData(cls):
        get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.dale1 = Game.objects.create(name="Dale of Merchants")
        cls.dale2 = Game.objects.create(name="Dale of Merchants 2")
        cls.series = Series.objects.create(
            name="Dale of Merchants", primary_game=cls.dale1,
        )
        Game.objects.filter(
            pk__in=[cls.dale1.pk, cls.dale2.pk],
        ).update(series=cls.series)

    def get(self):
        self.client.login(username="kernicek", password="pass")
        return self.client.get("/series/")

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/series/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_lists_series_with_member_count_and_link(self):
        response = self.get()
        self.assertContains(response, "Dale of Merchants")
        self.assertContains(response, f"/series/{self.series.pk}/")
        self.assertEqual(
            response.context["series_rows"][0].member_count, 2,
        )

    def test_empty_state_when_no_series_exist(self):
        self.series.delete()
        response = self.get()
        self.assertContains(response, "No series yet.")


class SeriesBulkLocationTests(TestCase):
    """Issue #58: bulk-move the current user's copies of a series' members to
    one location. Hard-scoped to owner=request.user and active copies — a
    co-owner's copies of the same members are never touched."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.other = get_user_model().objects.create_user(
            username="pavel", password="pass",
        )
        cls.game1 = Game.objects.create(name="Dale of Merchants")
        cls.game2 = Game.objects.create(name="Dale of Merchants 2")
        cls.series = Series.objects.create(
            name="Dale of Merchants", primary_game=cls.game1,
        )
        Game.objects.filter(
            pk__in=[cls.game1.pk, cls.game2.pk],
        ).update(series=cls.series)

        cls.ed1 = Edition.objects.create(game=cls.game1, is_default=True)
        cls.ed2 = Edition.objects.create(game=cls.game2, is_default=True)
        # A non-member game the user also owns — must never be swept in.
        cls.outsider = Game.objects.create(name="Unrelated Game")
        cls.ed_out = Edition.objects.create(game=cls.outsider, is_default=True)

        cls.shelf = Location.objects.create(
            group=cls.user.membership.group, name="Kallax A",
        )
        cls.other_shelf = Location.objects.create(
            group=cls.user.membership.group, name="Kallax B",
        )
        cls.foreign_shelf = Location.objects.create(
            group=cls.other.membership.group, name="Elsewhere",
        )

        # The user's active member copies (the bulk targets).
        cls.copy1 = Copy.objects.create(owner=cls.user, edition=cls.ed1)
        cls.copy2 = Copy.objects.create(owner=cls.user, edition=cls.ed2)
        # An archived member copy — untouched.
        cls.archived = Copy.objects.create(
            owner=cls.user, edition=Edition.objects.create(game=cls.game1),
            archive_status=Copy.ArchiveStatus.ARCHIVED,
        )
        # The user's copy of a non-member game — untouched.
        cls.outside_copy = Copy.objects.create(
            owner=cls.user, edition=cls.ed_out,
        )
        # Another owner's copy of a member — untouched (the core #58 rule).
        cls.their_copy = Copy.objects.create(
            owner=cls.other, edition=cls.ed1, location=cls.foreign_shelf,
        )

        # Issue #77: an expansion that is itself a series member — its
        # copies are skipped by the bulk move.
        cls.expansion_member = Game.objects.create(
            name="Dale of Merchants: Team Play", type=Game.Type.EXPANSION,
            series=cls.series,
        )
        cls.ed_expansion = Edition.objects.create(
            game=cls.expansion_member, is_default=True,
        )
        cls.expansion_copy = Copy.objects.create(
            owner=cls.user, edition=cls.ed_expansion,
        )

    def setUp(self):
        self.client.login(username="kernicek", password="pass")

    def url(self, pk=None):
        return f"/series/{pk or self.series.pk}/location/"

    def test_bulk_move_sets_all_own_active_member_copies(self):
        response = self.client.post(self.url(), {"location": str(self.shelf.pk)})
        self.assertRedirects(
            response, f"/series/{self.series.pk}/?moved=2",
            fetch_redirect_response=False,
        )
        self.copy1.refresh_from_db()
        self.copy2.refresh_from_db()
        self.assertEqual(self.copy1.location, self.shelf)
        self.assertEqual(self.copy2.location, self.shelf)

    def test_other_owners_copies_are_never_touched(self):
        self.client.post(self.url(), {"location": str(self.shelf.pk)})
        self.their_copy.refresh_from_db()
        self.assertEqual(self.their_copy.location, self.foreign_shelf)

    def test_archived_and_non_member_copies_are_untouched(self):
        self.client.post(self.url(), {"location": str(self.shelf.pk)})
        self.archived.refresh_from_db()
        self.outside_copy.refresh_from_db()
        self.assertIsNone(self.archived.location)
        self.assertIsNone(self.outside_copy.location)

    def test_foreign_location_is_a_400(self):
        response = self.client.post(
            self.url(), {"location": str(self.foreign_shelf.pk)},
        )
        self.assertEqual(response.status_code, 400)
        self.copy1.refresh_from_db()
        self.assertIsNone(self.copy1.location)

    def test_empty_or_missing_location_is_a_400(self):
        self.assertEqual(self.client.post(self.url(), {"location": ""}).status_code, 400)
        self.assertEqual(self.client.post(self.url(), {}).status_code, 400)
        self.copy1.refresh_from_db()
        self.assertIsNone(self.copy1.location)

    def test_moved_count_ignores_copies_already_at_target(self):
        self.copy1.location = self.shelf
        self.copy1.save()
        response = self.client.post(self.url(), {"location": str(self.shelf.pk)})
        # Only copy2 actually moves.
        self.assertRedirects(
            response, f"/series/{self.series.pk}/?moved=1",
            fetch_redirect_response=False,
        )

    def test_each_move_is_logged_by_simple_history(self):
        before = self.copy1.history.count()
        self.client.post(self.url(), {"location": str(self.shelf.pk)})
        self.assertEqual(self.copy1.history.count(), before + 1)

    def test_bulk_move_skips_expansion_member_copies(self):
        # Issue #77: expansions usually travel with their base game, so the
        # bulk move never relocates them, and they're not counted as moved.
        response = self.client.post(self.url(), {"location": str(self.shelf.pk)})
        self.assertRedirects(
            response, f"/series/{self.series.pk}/?moved=2",
            fetch_redirect_response=False,
        )
        self.expansion_copy.refresh_from_db()
        self.assertIsNone(self.expansion_copy.location)

    def test_my_copy_count_excludes_expansion_members(self):
        response = self.client.get(f"/series/{self.series.pk}/")
        self.assertEqual(response.context["my_copy_count"], 2)

    def test_detail_renders_form_when_user_has_copies_and_locations(self):
        response = self.client.get(f"/series/{self.series.pk}/")
        self.assertEqual(response.context["my_copy_count"], 2)
        self.assertContains(response, "Move all 2 of your")
        self.assertContains(response, f'action="/series/{self.series.pk}/location/"')

    def test_detail_hides_form_when_user_has_no_member_copies(self):
        self.client.logout()
        self.client.login(username="pavel", password="pass")
        # pavel owns a member copy, so create a locationless scenario: remove it.
        self.their_copy.delete()
        response = self.client.get(f"/series/{self.series.pk}/")
        self.assertEqual(response.context["my_copy_count"], 0)
        self.assertNotContains(response, f'action="/series/{self.series.pk}/location/"')

    def test_moved_banner_renders_from_query_param(self):
        response = self.client.get(f"/series/{self.series.pk}/?moved=4")
        self.assertContains(response, "Moved 4 of your copies")

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.client.post(self.url(), {"location": str(self.shelf.pk)})
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])


class SeriesEditViewTests(TestCase):
    """Issue #21: the in-app series editor — candidate scoping, membership
    reconciliation (the form edits the reverse side of Game.series) and the
    optional custom cover."""

    def setUp(self):
        # Fresh MEDIA_ROOT per test (the CoverEditMixin reasoning): files are
        # not transactional and dedupe suffixes would leak across tests.
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)
        self.client.login(username="kernicek", password="pass")

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.beasty = Game.objects.create(name="Beasty Bar")
        cls.beasty2 = Game.objects.create(name="Beasty Bar: New Beasts")
        cls.claimed = Game.objects.create(name="MicroMacro: Crime City")
        cls.other_series = Series.objects.create(
            name="MicroMacro", primary_game=cls.claimed,
        )
        Game.objects.filter(pk=cls.claimed.pk).update(series=cls.other_series)
        cls.expansion = Game.objects.create(
            name="Beasty Bar expansion pack", type=Game.Type.EXPANSION,
        )

    def add(self, **data):
        return self.client.post("/series/add/", data)

    def test_candidates_exclude_claimed_games_and_expansions(self):
        response = self.client.get("/series/add/")
        self.assertContains(response, f'id="member-{self.beasty.pk}"')
        self.assertContains(response, f'id="member-{self.beasty2.pk}"')
        self.assertNotContains(response, f'id="member-{self.claimed.pk}"')
        self.assertNotContains(response, f'id="member-{self.expansion.pk}"')

    def test_create_claims_members_and_redirects_to_the_detail_page(self):
        response = self.add(
            name="Beasty Bar", primary_game=str(self.beasty.pk),
            members=[str(self.beasty.pk), str(self.beasty2.pk)],
        )
        series = Series.objects.get(name="Beasty Bar")
        self.assertRedirects(response, f"/series/{series.pk}/")
        self.assertEqual(series.primary_game, self.beasty)
        self.beasty.refresh_from_db()
        self.beasty2.refresh_from_db()
        self.assertEqual(self.beasty.series, series)
        self.assertEqual(self.beasty2.series, series)

    def test_blank_name_is_400(self):
        response = self.add(
            name="  ", primary_game=str(self.beasty.pk),
            members=[str(self.beasty.pk)],
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Series.objects.filter(primary_game=self.beasty).exists())

    def test_primary_outside_the_members_is_400(self):
        response = self.add(
            name="Beasty Bar", primary_game=str(self.beasty.pk),
            members=[str(self.beasty2.pk)],
        )
        self.assertEqual(response.status_code, 400)

    def test_member_claimed_by_another_series_is_400(self):
        response = self.add(
            name="Beasty Bar", primary_game=str(self.beasty.pk),
            members=[str(self.beasty.pk), str(self.claimed.pk)],
        )
        self.assertEqual(response.status_code, 400)
        self.claimed.refresh_from_db()
        self.assertEqual(self.claimed.series, self.other_series)

    def test_edit_reconciles_membership_both_ways(self):
        series = Series.objects.create(name="Beasty Bar",
                                       primary_game=self.beasty)
        Game.objects.filter(pk=self.beasty.pk).update(series=series)
        newcomer = Game.objects.create(name="Beasty Bar Deluxe")

        response = self.client.post(f"/series/{series.pk}/edit/", {
            "name": "Beasty Bar", "primary_game": str(newcomer.pk),
            "members": [str(newcomer.pk)],
        })
        self.assertRedirects(response, f"/series/{series.pk}/")
        self.beasty.refresh_from_db()
        newcomer.refresh_from_db()
        self.assertIsNone(self.beasty.series)  # unchecked → released
        self.assertEqual(newcomer.series, series)

    def test_cover_upload_sets_the_custom_cover(self):
        response = self.add(
            name="Beasty Bar", primary_game=str(self.beasty.pk),
            members=[str(self.beasty.pk)],
            cover=SimpleUploadedFile("art.png", image_bytes()),
        )
        series = Series.objects.get(name="Beasty Bar")
        self.assertRedirects(response, f"/series/{series.pk}/")
        self.assertEqual(series.cover_image.name,
                         f"covers/series-{series.pk}.png")

    def test_edit_page_shows_the_cover_editor_column(self):
        # Issue #54: an existing series gets the shared htmx cover editor
        # (upload/URL/focus/zoom), not the old plain file input.
        series = Series.objects.create(name="Beasty Bar",
                                       primary_game=self.beasty)
        response = self.client.get(f"/series/{series.pk}/edit/")
        self.assertContains(response, 'id="cover-column"')
        # No cover yet, so only the replace form shows — the focus/zoom
        # tools appear after the first upload swaps the column.
        self.assertContains(response, f'hx-post="/series/{series.pk}/cover/"')
        self.assertNotContains(response, 'name="cover"')

    def test_add_page_keeps_the_plain_cover_input(self):
        # No pk yet, so no htmx endpoints — the create form still takes an
        # optional upload directly.
        response = self.client.get("/series/add/")
        self.assertContains(response, 'name="cover"')
        self.assertNotContains(response, 'id="cover-column"')

    def test_member_filter_markup_is_present(self):
        # Issue #54: client-side filter — rows carry the searchable name.
        response = self.client.get("/series/add/")
        self.assertContains(response, 'id="member-filter"')
        self.assertContains(response, 'data-name="beasty bar ')

    def test_non_image_upload_is_400(self):
        response = self.add(
            name="Beasty Bar", primary_game=str(self.beasty.pk),
            members=[str(self.beasty.pk)],
            cover=SimpleUploadedFile("art.png", b"not an image at all"),
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Series.objects.filter(name="Beasty Bar").exists())


class SeriesCoverMixin(CoverEditMixin):
    """CoverEditMixin plus a series claiming the fixture game — the series
    flavour of the shared cover editor (issue #54)."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.series = Series.objects.create(
            name="Epic Saga", primary_game=cls.game,
        )
        Game.objects.filter(pk=cls.game.pk).update(series=cls.series)

    def give_cover(self):
        # Series cover_url has no remote fallback (unlike the game fixture's
        # image_url), so partial-rendering assertions need a real file —
        # saved here, after setUp's MEDIA_ROOT override is active.
        self.series.cover_image.save(
            f"series-{self.series.pk}.png", BytesIO(image_bytes()))


class SeriesCoverEditTests(SeriesCoverMixin, TestCase):
    """Issue #54: /series/<pk>/cover/ — upload, URL fetch and clear. The
    shared _handle_cover_edit plumbing (both-inputs, bad bytes, download
    failures…) is exhaustively covered by GameCoverEditTests; here the
    series-specific wiring is what's under test."""

    def post(self, data=None):
        return self.client.post(f"/series/{self.series.pk}/cover/", data or {})

    def test_anonymous_users_are_redirected_to_login(self):
        self.client.logout()
        response = self.post({"url": "https://example.com/a.png"})
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_get_is_405(self):
        response = self.client.get(f"/series/{self.series.pk}/cover/")
        self.assertEqual(response.status_code, 405)

    def test_unknown_series_is_404(self):
        response = self.client.post("/series/99999/cover/", {})
        self.assertEqual(response.status_code, 404)

    def test_upload_sets_the_custom_cover(self):
        self.series.cover_focus_x = 20
        self.series.cover_zoom = 180
        self.series.save()

        response = self.post({
            "file": SimpleUploadedFile("art.png", image_bytes()),
        })

        self.assertEqual(response.status_code, 200)
        self.series.refresh_from_db()
        self.assertEqual(self.series.cover_image.name,
                         f"covers/series-{self.series.pk}.png")
        self.assertEqual(self.series.cover_width, 4)
        self.assertEqual(self.series.cover_height, 4)
        # The focal point and zoom described the old art — reset.
        self.assertEqual(self.series.cover_focus_x, 50)
        self.assertEqual(self.series.cover_zoom, 100)

    def test_url_is_fetched_server_side(self):
        with mock.patch("gamekeeper.views.requests.get") as get:
            get.return_value = mock.Mock(content=image_bytes("JPEG"))
            response = self.post({"url": "https://example.com/art?dl=1"})

        get.assert_called_once_with("https://example.com/art?dl=1", timeout=30)
        self.assertEqual(response.status_code, 200)
        self.series.refresh_from_db()
        # Extension comes from what Pillow says the bytes are, not the URL.
        self.assertEqual(self.series.cover_image.name,
                         f"covers/series-{self.series.pk}.jpg")

    def test_neither_file_nor_url_is_an_inline_error(self):
        response = self.post()
        self.assertContains(response, "Choose a file or paste an image URL.")
        self.series.refresh_from_db()
        self.assertFalse(self.series.cover_image)

    def test_non_image_bytes_are_rejected(self):
        response = self.post({
            "file": SimpleUploadedFile("art.png", b"not an image at all"),
        })
        self.assertContains(response, "does not look like an image")
        self.series.refresh_from_db()
        self.assertFalse(self.series.cover_image)

    def test_clear_removes_the_cover_and_falls_back_to_the_primary(self):
        self.give_cover()
        stored = self.series.cover_image.name

        response = self.post({"clear": "1"})

        self.assertEqual(response.status_code, 200)
        self.series.refresh_from_db()
        self.assertFalse(self.series.cover_image)
        self.assertEqual(self.series.cover_source, self.game)
        self.assertFalse(self.series.cover_image.storage.exists(stored))

    def test_clear_also_removes_the_baked_preview(self):
        # Issue #104: clearing the art drops its grid thumbnail too.
        self.give_cover()
        self.series.regenerate_cover_preview()
        preview = self.series.cover_preview.name
        storage = self.series.cover_preview.storage

        self.post({"clear": "1"})

        self.series.refresh_from_db()
        self.assertFalse(self.series.cover_preview)
        self.assertFalse(storage.exists(preview))

    def test_clear_without_a_cover_is_a_no_op(self):
        response = self.post({"clear": "1"})
        self.assertEqual(response.status_code, 200)

    def test_partial_wires_the_series_endpoints_and_the_remove_button(self):
        self.give_cover()
        response = self.post({"url": "", "file": ""})  # inline-error re-render
        self.assertContains(response, f"/series/{self.series.pk}/cover/")
        self.assertContains(response, f"/series/{self.series.pk}/cover/focus/")
        self.assertContains(response, "Remove custom cover")


class SeriesCoverFocusTests(SeriesCoverMixin, TestCase):
    """Issue #54: /series/<pk>/cover/focus/ — the same focus/zoom/fit
    treatment games get, on the series' own fields."""

    def post(self, data):
        return self.client.post(
            f"/series/{self.series.pk}/cover/focus/", data)

    def test_get_is_405(self):
        response = self.client.get(f"/series/{self.series.pk}/cover/focus/")
        self.assertEqual(response.status_code, 405)

    def test_unknown_series_is_404(self):
        response = self.client.post(
            "/series/99999/cover/focus/", {"x": "1", "y": "1"})
        self.assertEqual(response.status_code, 404)

    def test_sets_the_focal_point(self):
        self.give_cover()
        response = self.post({"x": "10", "y": "90"})
        self.assertEqual(response.status_code, 200)
        self.series.refresh_from_db()
        self.assertEqual(self.series.cover_focus_x, 10)
        self.assertEqual(self.series.cover_focus_y, 90)
        # The partial echoes it into the crop preview.
        self.assertContains(response, "object-position: 10% 90%")

    def test_zoom_alone_is_saved(self):
        self.give_cover()
        response = self.post({"zoom": "175"})
        self.assertEqual(response.status_code, 200)
        self.series.refresh_from_db()
        self.assertEqual(self.series.cover_zoom, 175)
        self.assertContains(response, "transform: scale(1.75)")

    def test_out_of_range_input_is_clamped(self):
        # Same clamp-on-typed-entry behaviour games get (issues #12/#13).
        self.assertEqual(self.post({"x": "101", "y": "0"}).status_code, 200)
        self.series.refresh_from_db()
        self.assertEqual(self.series.cover_focus_x, 100)
        self.assertEqual(self.post({"zoom": "49"}).status_code, 200)
        self.series.refresh_from_db()
        self.assertEqual(self.series.cover_zoom, 50)
        self.assertEqual(self.post({"zoom": "301"}).status_code, 200)
        self.series.refresh_from_db()
        self.assertEqual(self.series.cover_zoom, 300)

    def test_zoom_below_100_renders_the_fit_preview(self):
        # §13 zoom-out, now on a series: contain-fit over the letterbox
        # colour — the full game treatment, not the old centered crop.
        self.give_cover()
        self.series.cover_fit_color = "#aabbcc"
        self.series.save(update_fields=["cover_fit_color"])
        response = self.post({"zoom": "60"})
        self.assertEqual(response.status_code, 200)
        self.series.refresh_from_db()
        self.assertEqual(self.series.cover_zoom, 60)
        self.assertTrue(self.series.cover_fit)
        self.assertContains(response, "cover-art-fit")
        self.assertContains(
            response, 'cover-art-backdrop" style="background-color: #aabbcc"')


class FamilyModelTests(TestCase):
    """Issue #42: the cover_source chain — custom art, else the first
    member alphabetically (article-blind sort_name order), else None."""

    @classmethod
    def setUpTestData(cls):
        cls.family = Family.objects.create(name="Fowers heists")
        # "The Climbers" files under C (issue #6) — alphabetically it beats
        # "Fowl Play", so the article-blind order is what's under test.
        cls.climbers = Game.objects.create(name="The Climbers")
        cls.fowl = Game.objects.create(name="Fowl Play")
        cls.climbers.families.add(cls.family)
        cls.fowl.families.add(cls.family)

    def test_cover_source_is_the_first_member_alphabetically(self):
        self.assertEqual(self.family.cover_source, self.climbers)

    def test_custom_cover_wins(self):
        self.family.cover_image = "covers/family-custom.jpg"
        self.assertEqual(self.family.cover_source, self.family)

    def test_cover_source_is_none_without_members(self):
        empty = Family.objects.create(name="Empty line")
        self.assertIsNone(empty.cover_source)


class FamilyFacetTests(TestCase):
    """Issue #42: the family GameChooser facet — members filter in, and a
    family NEVER collapses the grid (that is Series' job)."""

    @classmethod
    def setUpTestData(cls):
        get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.burgle = Game.objects.create(name="Burgle Bros")
        cls.burgle2 = Game.objects.create(name="Burgle Bros 2")
        cls.paperback = Game.objects.create(name="Paperback")
        cls.standalone = Game.objects.create(name="Duel")
        cls.heists = Family.objects.create(name="Fowers heists")
        cls.fowers = Family.objects.create(name="Fowers Games")
        cls.burgle.families.add(cls.heists, cls.fowers)
        cls.burgle2.families.add(cls.heists)
        cls.paperback.families.add(cls.fowers)

    def get(self, params=None, **extra):
        self.client.login(username="kernicek", password="pass")
        # Copy-less fixtures: opt past the available-only default (issue #107)
        # so the family facet, not availability, is what's under test.
        query = {"show_unavailable": "1"}
        query.update(params or {})
        return self.client.get("/", query, **extra)

    def test_facet_select_lists_the_families(self):
        response = self.get()
        self.assertContains(response, 'id="f-family"')
        self.assertContains(response, "Fowers heists")
        self.assertContains(response, "Fowers Games")

    def test_family_filter_keeps_members_only(self):
        response = self.get({"family": str(self.heists.pk)})
        self.assertContains(response, "Burgle Bros")
        self.assertContains(response, "Burgle Bros 2")
        self.assertNotContains(response, "Paperback")
        self.assertNotContains(response, "Duel")

    def test_a_game_in_several_families_matches_via_either(self):
        for family in (self.heists, self.fowers):
            response = self.get({"family": str(family.pk)})
            self.assertContains(response, "Burgle Bros")

    def test_selected_family_is_echoed_back(self):
        response = self.get({"family": str(self.fowers.pk)})
        self.assertEqual(response.context["filters"]["families"],
                         {self.fowers.pk})

    def test_family_membership_never_collapses_the_grid(self):
        # Both heist games keep their own top-level tiles (plus Paperback
        # and Duel = 4) — no series-style collapse, no overlay.
        response = self.get()
        self.assertContains(response, '<a class="cover" ', count=4)
        self.assertNotContains(response, "cover-series")


class GameDetailFamilyTests(TestCase):
    """Issue #106: family memberships surface as header badges on the game
    detail page, each linking to the family's own page."""

    @classmethod
    def setUpTestData(cls):
        get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.burgle = Game.objects.create(name="Burgle Bros")
        cls.burgle2 = Game.objects.create(name="Burgle Bros 2")
        cls.paperback = Game.objects.create(name="Paperback")
        cls.heists = Family.objects.create(name="Fowers heists")
        cls.burgle.families.add(cls.heists)
        cls.burgle2.families.add(cls.heists)
        cls.solo = Family.objects.create(name="Solo line")
        cls.burgle.families.add(cls.solo)

    def get(self, game):
        self.client.login(username="kernicek", password="pass")
        return self.client.get(f"/games/{game.pk}/")

    def test_badge_links_to_the_family(self):
        response = self.get(self.burgle2)
        self.assertContains(response, "Family: Fowers heists")
        self.assertContains(response, f"/families/{self.heists.pk}/")

    def test_multiple_families_each_get_a_badge(self):
        response = self.get(self.burgle)
        self.assertContains(response, "Family: Fowers heists")
        self.assertContains(response, "Family: Solo line")

    def test_no_badge_without_families(self):
        response = self.get(self.paperback)
        self.assertNotContains(response, "Family:")


class FamilyDetailViewTests(TestCase):
    """Issue #42: the family page — member grid, note, BGG family link. No
    pooled stats/Copies/Purchases: members are distinct games."""

    @classmethod
    def setUpTestData(cls):
        get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.burgle = Game.objects.create(name="Burgle Bros")
        cls.burgle2 = Game.objects.create(name="Burgle Bros 2")
        cls.family = Family.objects.create(
            name="Fowers heists", bgg_family_id=27373,
            note="Same world, different gameplay.",
        )
        cls.burgle.families.add(cls.family)
        cls.burgle2.families.add(cls.family)

    def get(self, pk=None):
        self.client.login(username="kernicek", password="pass")
        return self.client.get(f"/families/{pk or self.family.pk}/")

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get(f"/families/{self.family.pk}/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_unknown_family_is_404(self):
        self.assertEqual(self.get(99999).status_code, 404)

    def test_member_grid_links_every_member(self):
        response = self.get()
        self.assertContains(response, "Games in this family")
        for member in (self.burgle, self.burgle2):
            self.assertContains(response, f"/games/{member.pk}/")
        self.assertContains(response, f"/families/{self.family.pk}/edit/")

    def test_note_and_bgg_family_link_render(self):
        response = self.get()
        self.assertContains(response, "Same world, different gameplay.")
        self.assertContains(
            response, "https://boardgamegeek.com/boardgamefamily/27373")

    def test_hero_falls_back_to_the_first_members_art(self):
        # No custom cover: the alphabetically-first member (Burgle Bros)
        # represents the family, not the placeholder.
        Game.objects.filter(pk=self.burgle.pk).update(
            image_url="https://cf.geekdo-images.com/original/bb1.jpg")
        Game.objects.filter(pk=self.burgle2.pk).update(
            image_url="https://cf.geekdo-images.com/original/bb2.jpg")
        response = self.get()
        self.assertContains(
            response, 'class="hero-cover" '
            'src="https://cf.geekdo-images.com/original/bb1.jpg"')

    def test_custom_cover_beats_the_members(self):
        Game.objects.filter(pk=self.burgle.pk).update(
            image_url="https://cf.geekdo-images.com/original/bb1.jpg")
        Family.objects.filter(pk=self.family.pk).update(
            cover_image="covers/heists-custom.jpg")
        response = self.get()
        # The member's art still renders its grid tile below — the HERO is
        # what must carry the custom cover.
        self.assertContains(
            response, 'class="hero-cover" src="/media/covers/heists-custom.jpg"')

    def test_hero_placeholder_when_no_member_has_art(self):
        response = self.get()
        self.assertContains(response, "hero-placeholder")
        self.assertNotContains(response, "hero-cover")


class FamilyListViewTests(TestCase):
    """Issue #80: the family overview grid — cover, name, member count, each
    tile linking to its detail page."""

    @classmethod
    def setUpTestData(cls):
        get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.burgle = Game.objects.create(name="Burgle Bros")
        cls.burgle2 = Game.objects.create(name="Burgle Bros 2")
        cls.family = Family.objects.create(name="Fowers heists")
        cls.burgle.families.add(cls.family)
        cls.burgle2.families.add(cls.family)

    def get(self):
        self.client.login(username="kernicek", password="pass")
        return self.client.get("/families/")

    def test_anonymous_users_are_redirected_to_login(self):
        response = self.client.get("/families/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])

    def test_lists_families_with_member_count_and_link(self):
        response = self.get()
        self.assertContains(response, "Fowers heists")
        self.assertContains(response, f"/families/{self.family.pk}/")
        self.assertEqual(
            response.context["family_rows"][0].member_count, 2,
        )

    def test_empty_state_when_no_families_exist(self):
        self.family.delete()
        response = self.get()
        self.assertContains(response, "No families yet.")


class FamilyEditViewTests(TestCase):
    """Issue #42: the in-app family editor — every base game is a candidate
    (loose M2M, no claimed-elsewhere restriction) and membership is
    reconciled with one set()."""

    def setUp(self):
        # Fresh MEDIA_ROOT per test (the CoverEditMixin reasoning): files are
        # not transactional and dedupe suffixes would leak across tests.
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)
        self.client.login(username="kernicek", password="pass")

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        cls.burgle = Game.objects.create(name="Burgle Bros")
        cls.burgle2 = Game.objects.create(name="Burgle Bros 2")
        cls.elsewhere = Game.objects.create(name="Paperback")
        cls.other_family = Family.objects.create(name="Fowers Games")
        cls.elsewhere.families.add(cls.other_family)
        cls.expansion = Game.objects.create(
            name="Burgle Bros: The Fence", type=Game.Type.EXPANSION,
        )

    def add(self, **data):
        return self.client.post("/families/add/", data)

    def test_candidates_are_all_base_games_even_claimed_ones(self):
        # Loose M2M: membership elsewhere is no obstacle — only expansions
        # stay out (they hang off their base).
        response = self.client.get("/families/add/")
        self.assertContains(response, f'id="member-{self.burgle.pk}"')
        self.assertContains(response, f'id="member-{self.elsewhere.pk}"')
        self.assertNotContains(response, f'id="member-{self.expansion.pk}"')

    def test_create_sets_membership_and_redirects_to_the_detail_page(self):
        response = self.add(
            name="Fowers heists", note="Heist line", bgg_family_id="27373",
            members=[str(self.burgle.pk), str(self.burgle2.pk)],
        )
        family = Family.objects.get(name="Fowers heists")
        self.assertRedirects(response, f"/families/{family.pk}/")
        self.assertEqual(family.note, "Heist line")
        self.assertEqual(family.bgg_family_id, 27373)
        self.assertEqual(
            set(family.members.all()), {self.burgle, self.burgle2})
        self.assertIn(family, self.burgle.families.all())

    def test_joining_a_second_family_keeps_the_first(self):
        response = self.add(
            name="Fowers heists", members=[str(self.elsewhere.pk)],
        )
        family = Family.objects.get(name="Fowers heists")
        self.assertRedirects(response, f"/families/{family.pk}/")
        self.assertEqual(
            set(self.elsewhere.families.all()), {family, self.other_family})

    def test_blank_name_is_400(self):
        response = self.add(name="  ", members=[str(self.burgle.pk)])
        self.assertEqual(response.status_code, 400)
        self.assertEqual(Family.objects.count(), 1)  # only the fixture

    def test_non_integer_member_ids_are_400(self):
        response = self.add(name="Fowers heists", members=["abc"])
        self.assertEqual(response.status_code, 400)

    def test_non_numeric_bgg_family_id_is_400(self):
        response = self.add(name="Fowers heists", bgg_family_id="soon",
                            members=[str(self.burgle.pk)])
        self.assertEqual(response.status_code, 400)

    def test_an_expansion_member_is_400(self):
        response = self.add(
            name="Fowers heists",
            members=[str(self.burgle.pk), str(self.expansion.pk)],
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Family.objects.filter(name="Fowers heists").exists())

    def test_edit_reconciles_membership_both_ways(self):
        family = Family.objects.create(name="Fowers heists")
        self.burgle.families.add(family)

        response = self.client.post(f"/families/{family.pk}/edit/", {
            "name": "Fowers heists", "members": [str(self.burgle2.pk)],
        })
        self.assertRedirects(response, f"/families/{family.pk}/")
        self.assertEqual(set(family.members.all()), {self.burgle2})
        self.assertNotIn(family, self.burgle.families.all())

    def test_cover_upload_sets_the_cover(self):
        response = self.add(
            name="Fowers heists", members=[str(self.burgle.pk)],
            cover=SimpleUploadedFile("art.png", image_bytes()),
        )
        family = Family.objects.get(name="Fowers heists")
        self.assertRedirects(response, f"/families/{family.pk}/")
        self.assertEqual(family.cover_image.name,
                         f"covers/family-{family.pk}.png")

    def test_non_image_upload_is_400(self):
        response = self.add(
            name="Fowers heists", members=[str(self.burgle.pk)],
            cover=SimpleUploadedFile("art.png", b"not an image at all"),
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(Family.objects.filter(name="Fowers heists").exists())

    def test_edit_page_shows_the_cover_editor_column(self):
        # The #54 shared htmx cover editor, family flavour.
        family = Family.objects.create(name="Fowers heists")
        response = self.client.get(f"/families/{family.pk}/edit/")
        self.assertContains(response, 'id="cover-column"')
        self.assertContains(response, f'hx-post="/families/{family.pk}/cover/"')
        self.assertNotContains(response, 'name="cover"')

    def test_add_page_keeps_the_plain_cover_input(self):
        response = self.client.get("/families/add/")
        self.assertContains(response, 'name="cover"')
        self.assertNotContains(response, 'id="cover-column"')

    def test_member_filter_markup_is_present(self):
        response = self.client.get("/families/add/")
        self.assertContains(response, 'id="member-filter"')
        self.assertContains(response, 'data-name="burgle bros ')


class FamilyCoverEditTests(CoverEditMixin, TestCase):
    """Issue #42: /families/<pk>/cover/ — the family flavour of the shared
    cover editor. The _handle_cover_edit plumbing is exhaustively covered by
    GameCoverEditTests; here the family wiring (endpoints, clear-to-
    placeholder) is what's under test."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.family = Family.objects.create(name="Epic line")
        cls.game.families.add(cls.family)

    def post(self, data=None):
        return self.client.post(
            f"/families/{self.family.pk}/cover/", data or {})

    def test_upload_sets_the_cover(self):
        response = self.post({
            "file": SimpleUploadedFile("art.png", image_bytes()),
        })
        self.assertEqual(response.status_code, 200)
        self.family.refresh_from_db()
        self.assertEqual(self.family.cover_image.name,
                         f"covers/family-{self.family.pk}.png")

    def test_clear_removes_the_cover(self):
        self.family.cover_image.save(
            f"family-{self.family.pk}.png", BytesIO(image_bytes()))
        stored = self.family.cover_image.name

        response = self.post({"clear": "1"})

        self.assertEqual(response.status_code, 200)
        self.family.refresh_from_db()
        self.assertFalse(self.family.cover_image)
        self.assertFalse(self.family.cover_image.storage.exists(stored))

    def test_clear_also_removes_the_baked_preview(self):
        # Issue #104: clearing the art drops its grid thumbnail too.
        self.family.cover_image.save(
            f"family-{self.family.pk}.png", BytesIO(image_bytes()))
        self.family.regenerate_cover_preview()
        preview = self.family.cover_preview.name
        storage = self.family.cover_preview.storage

        self.post({"clear": "1"})

        self.family.refresh_from_db()
        self.assertFalse(self.family.cover_preview)
        self.assertFalse(storage.exists(preview))

    def test_partial_wires_the_family_endpoints_and_the_remove_button(self):
        self.family.cover_image.save(
            f"family-{self.family.pk}.png", BytesIO(image_bytes()))
        response = self.post({"url": "", "file": ""})  # inline-error re-render
        self.assertContains(response, f"/families/{self.family.pk}/cover/")
        self.assertContains(
            response, f"/families/{self.family.pk}/cover/focus/")
        self.assertContains(response, "Remove custom cover")

    def test_focus_sets_the_focal_point(self):
        self.family.cover_image.save(
            f"family-{self.family.pk}.png", BytesIO(image_bytes()))
        response = self.client.post(
            f"/families/{self.family.pk}/cover/focus/", {"x": "10", "y": "90"})
        self.assertEqual(response.status_code, 200)
        self.family.refresh_from_db()
        self.assertEqual(self.family.cover_focus_x, 10)
        self.assertEqual(self.family.cover_focus_y, 90)


# ===========================================================================
# Issue #28  Mobile layout
# ===========================================================================

class MobileLayoutTests(TestCase):
    """The base template collapses its navbar on small screens and data
    tables scroll rather than overflow the viewport."""

    def test_base_navbar_collapses_behind_a_toggler(self):
        # Any page renders base.html; the login page needs no fixtures.
        response = self.client.get("/accounts/login/")
        self.assertContains(response, 'class="navbar-toggler"')
        self.assertContains(response, 'id="mainNav"')
        self.assertContains(response, "navbar-collapse")
        # Regression: base.html's {# #} comments are single-line only, so a
        # stray multi-line comment would leak into the page as literal text.
        self.assertNotContains(response, "Bootstrap bundle JS drives the toggle")

    def test_data_tables_get_a_responsive_wrapper(self):
        html = render_to_string(
            "partials/purchases_table.html", {"purchase_rows": []})
        self.assertIn("table-responsive", html)

    def test_game_list_partial_gets_a_responsive_wrapper(self):
        # Issue #92: the collection list view follows the same table
        # convention as every other data table in the app.
        game = Game.objects.create(name="Wrapper Test")
        tile = {
            "game": game, "detail_url": f"/games/{game.pk}/",
            "stat_game": game, "players_low": None, "players_high": None,
            "copy_count": 0, "location_label": "—", "keep_label": "—",
        }
        html = render_to_string(
            "partials/game_list.html", {"tiles": [tile], "owner_view": True})
        self.assertIn("table-responsive", html)


# ===========================================================================
# Issue #80  Navbar "Browse" dropdown
# ===========================================================================

class NavbarBrowseDropdownTests(TestCase):
    """Series/Families/Sleeves group under one "Browse" dropdown instead of
    growing the flat nav; the flat Collection link is dropped since the
    navbar-brand link already covers it."""

    def setUp(self):
        get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )
        self.client.login(username="kernicek", password="pass")

    def test_browse_dropdown_links_series_families_and_sleeves(self):
        response = self.client.get("/dashboard/")
        self.assertContains(response, 'id="browseMenu"')
        self.assertContains(response, '<a class="dropdown-item" href="/series/">Series</a>')
        self.assertContains(
            response, '<a class="dropdown-item" href="/families/">Families</a>')
        self.assertContains(
            response, '<a class="dropdown-item" href="/sleeves/">Sleeves</a>')

    def test_flat_collection_link_is_dropped(self):
        response = self.client.get("/dashboard/")
        self.assertNotContains(
            response, '<a class="nav-link" href="/">Collection</a>')
        # The brand link still reaches the collection page.
        self.assertContains(response, 'class="navbar-brand" href="/"')


# ===========================================================================
# §7 documents (issue #60)
# ===========================================================================


class DocumentModelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.game = Game.objects.create(name="Wingspan")

    def test_document_requires_a_url_or_a_file(self):
        # §7 invariant: a document is a link and/or a file, never neither.
        doc = Document(content_object=self.game, doc_type=Document.Type.OTHER)
        with self.assertRaises(ValidationError):
            doc.clean()

    def test_a_url_only_document_is_valid(self):
        doc = Document(content_object=self.game, external_url="https://x/y.pdf")
        doc.clean()  # does not raise

    def test_ordering_surfaces_pinned_documents_first(self):
        plain = Document(content_object=self.game, doc_type=Document.Type.RULEBOOK,
                         label="A", external_url="https://a")
        plain.save()
        pinned = Document(content_object=self.game, doc_type=Document.Type.RULEBOOK,
                          label="B", is_primary=True, external_url="https://b")
        pinned.save()
        self.assertEqual(list(self.game.documents.all()), [pinned, plain])


class DocumentPathTests(TestCase):
    """Issue #99: an expansion's documents nest under its base game's folder,
    and its editions one level deeper still, while base games and their editions
    keep the pre-#99 top-level layout."""

    @classmethod
    def setUpTestData(cls):
        cls.base = Game.objects.create(name="Wingspan")
        cls.expansion = Game.objects.create(
            name="Wingspan Oceania", type=Game.Type.EXPANSION)
        cls.expansion.expands.add(cls.base)

    def _path(self, host, filename="upload.pdf", label="Rules"):
        from gamekeeper.models import document_upload_path
        return document_upload_path(
            Document(content_object=host, label=label), filename)

    def test_base_game_document_stays_top_level(self):
        self.assertEqual(
            self._path(self.base),
            f"documents/games/Wingspan [{self.base.pk}]/Rules.pdf")

    def test_base_edition_document_keeps_pre_99_layout(self):
        edition = Edition.objects.create(game=self.base, name="KS")
        self.assertEqual(
            self._path(edition),
            f"documents/games/Wingspan [{self.base.pk}]/"
            f"KS [{edition.pk}]/Rules.pdf")

    def test_expansion_document_nests_under_the_base(self):
        self.assertEqual(
            self._path(self.expansion),
            f"documents/games/Wingspan [{self.base.pk}]/"
            f"Wingspan Oceania [{self.expansion.pk}]/Rules.pdf")

    def test_expansion_edition_document_nests_one_level_deeper(self):
        edition = Edition.objects.create(game=self.expansion, name="Deluxe")
        self.assertEqual(
            self._path(edition),
            f"documents/games/Wingspan [{self.base.pk}]/"
            f"Wingspan Oceania [{self.expansion.pk}]/"
            f"Deluxe [{edition.pk}]/Rules.pdf")

    def test_multi_base_expansion_picks_the_first_linked_base(self):
        # Second base linked later and alphabetically earlier — the folder still
        # roots at the first-linked (lowest-pk) base, not the alphabetical one.
        other = Game.objects.create(name="Aviary")
        self.expansion.expands.add(other)
        self.assertEqual(self.expansion.primary_base, self.base)
        self.assertTrue(
            self._path(self.expansion).startswith(
                f"documents/games/Wingspan [{self.base.pk}]/"))

    def test_unlinked_expansion_stays_top_level(self):
        orphan = Game.objects.create(
            name="Homebrew Expansion", type=Game.Type.EXPANSION)
        self.assertIsNone(orphan.primary_base)
        self.assertEqual(
            self._path(orphan),
            f"documents/games/Homebrew Expansion [{orphan.pk}]/Rules.pdf")


class DocumentViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass")
        cls.game = Game.objects.create(name="Wingspan")

    def setUp(self):
        # Fresh MEDIA_ROOT per test — file writes aren't transactional and the
        # dedupe suffix would otherwise leak across tests (CoverEditMixin note).
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)
        self.client.login(username="kernicek", password="pass")

    @property
    def add_url(self):
        return f"/games/{self.game.pk}/documents/add/"

    def test_add_with_an_external_url_only(self):
        response = self.client.post(self.add_url, {
            "doc_type": "rulebook",
            "external_url": "https://example.com/rules.pdf",
        })
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        document = self.game.documents.get()
        self.assertEqual(document.external_url, "https://example.com/rules.pdf")
        self.assertFalse(document.file)

    def test_add_with_an_uploaded_file_uses_the_human_readable_path(self):
        response = self.client.post(self.add_url, {
            "doc_type": "rulebook",
            "label": "Oceania Rulebook",
            "file": SimpleUploadedFile("upload.pdf", b"%PDF-1.4 fake"),
        })
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        document = self.game.documents.get()
        self.assertEqual(
            document.file.name,
            f"documents/games/Wingspan [{self.game.pk}]/Oceania Rulebook.pdf")

    def test_add_with_both_a_url_and_a_file(self):
        response = self.client.post(self.add_url, {
            "doc_type": "other",
            "external_url": "https://example.com/official",
            "file": SimpleUploadedFile("copy.pdf", b"data"),
        })
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        document = self.game.documents.get()
        self.assertTrue(document.external_url)
        self.assertTrue(document.file)

    def test_add_can_pin_as_primary(self):
        self.client.post(self.add_url, {
            "doc_type": "rulebook",
            "external_url": "https://example.com/rules",
            "is_primary": "on",
        })
        self.assertTrue(self.game.documents.get().is_primary)

    def test_add_with_neither_url_nor_file_is_rejected(self):
        response = self.client.post(self.add_url, {"doc_type": "other"})
        self.assertEqual(response.status_code, 400)
        self.assertEqual(self.game.documents.count(), 0)

    @override_settings(DOCUMENT_MAX_UPLOAD_SIZE=10)
    def test_oversized_upload_is_rejected(self):
        response = self.client.post(self.add_url, {
            "doc_type": "other",
            "file": SimpleUploadedFile("big.pdf", b"x" * 50),
        })
        self.assertEqual(response.status_code, 400)
        self.assertEqual(self.game.documents.count(), 0)

    def test_disallowed_extension_is_rejected(self):
        response = self.client.post(self.add_url, {
            "doc_type": "other",
            "file": SimpleUploadedFile("virus.exe", b"data"),
        })
        self.assertEqual(response.status_code, 400)
        self.assertEqual(self.game.documents.count(), 0)

    def test_edit_page_prefills_existing_values(self):
        document = Document(content_object=self.game, doc_type="rulebook",
                            label="Rules", is_primary=True,
                            external_url="https://example.com/r")
        document.save()
        response = self.client.get(f"/documents/{document.pk}/edit/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="Rules"')
        self.assertContains(response, 'value="https://example.com/r"')
        self.assertContains(response, "checked")
        # Delete lives on the edit page (issue #97).
        self.assertContains(response, f"/documents/{document.pk}/delete/")

    def test_edit_updates_every_field(self):
        document = Document(content_object=self.game, doc_type="other",
                            label="Old", external_url="https://example.com/old")
        document.save()
        response = self.client.post(f"/documents/{document.pk}/edit/", {
            "doc_type": "rulebook",
            "label": "New",
            "external_url": "https://example.com/new",
            "is_primary": "on",
        })
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        document.refresh_from_db()
        self.assertEqual(document.doc_type, "rulebook")
        self.assertEqual(document.label, "New")
        self.assertEqual(document.external_url, "https://example.com/new")
        self.assertTrue(document.is_primary)

    def test_edit_keeps_existing_file_when_url_cleared(self):
        # An already-stored file satisfies the URL/file invariant on edit.
        document = Document(content_object=self.game, doc_type="rulebook",
                            external_url="https://example.com/r")
        document.file = SimpleUploadedFile("rules.pdf", b"data")
        document.save()
        response = self.client.post(f"/documents/{document.pk}/edit/", {
            "doc_type": "rulebook", "external_url": "",
        })
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        document.refresh_from_db()
        self.assertEqual(document.external_url, "")
        self.assertTrue(document.file)

    def test_delete_removes_the_document_and_its_file(self):
        self.client.post(self.add_url, {
            "doc_type": "rulebook",
            "file": SimpleUploadedFile("rules.pdf", b"data"),
        })
        document = self.game.documents.get()
        path = document.file.path
        self.assertTrue(Path(path).exists())
        response = self.client.post(f"/documents/{document.pk}/delete/")
        self.assertRedirects(response, f"/games/{self.game.pk}/")
        self.assertEqual(self.game.documents.count(), 0)
        self.assertFalse(Path(path).exists())

    def test_game_detail_lists_documents_with_the_pin_marker(self):
        Document(content_object=self.game, doc_type="rulebook", label="Rules",
                 is_primary=True, external_url="https://example.com/r").save()
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertContains(response, "Rules")
        self.assertContains(response, "★")

    def test_game_detail_document_row_links_to_the_edit_page(self):
        # Issue #97: own-doc rows offer Edit only — no inline delete form.
        document = Document(content_object=self.game, doc_type="rulebook",
                            label="Rules", external_url="https://example.com/r")
        document.save()
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertContains(response, f"/documents/{document.pk}/edit/")
        self.assertNotContains(response, f"/documents/{document.pk}/delete/")

    def _owned_expansion(self, name="Wingspan Oceania"):
        """A base-game expansion owned via an active copy (issue #97)."""
        expansion = Game.objects.create(name=name)
        expansion.expands.add(self.game)
        edition = Edition.objects.create(game=expansion, is_default=True)
        Copy.objects.create(owner=self.user, edition=edition)
        return expansion

    def test_owned_expansion_document_surfaces_on_the_base_page(self):
        expansion = self._owned_expansion()
        Document(content_object=expansion, doc_type="rulebook",
                 label="Oceania Rules",
                 external_url="https://example.com/oceania").save()
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertContains(response, "Oceania Rules")
        # Badged with the source expansion, linked to its page.
        self.assertContains(response, expansion.name)
        self.assertContains(response, f"/games/{expansion.pk}/")

    def test_unowned_expansion_document_does_not_surface(self):
        expansion = Game.objects.create(name="Wingspan: Europe")
        expansion.expands.add(self.game)  # linked but no active copy → unowned
        Document(content_object=expansion, doc_type="rulebook",
                 label="Europe Rules",
                 external_url="https://example.com/europe").save()
        response = self.client.get(f"/games/{self.game.pk}/")
        self.assertNotContains(response, "Europe Rules")

    def test_surfaced_expansion_document_is_read_only(self):
        expansion = self._owned_expansion()
        document = Document(content_object=expansion, doc_type="rulebook",
                            label="Oceania Rules",
                            external_url="https://example.com/oceania")
        document.save()
        # Base page: no edit/delete controls for the surfaced doc.
        base = self.client.get(f"/games/{self.game.pk}/")
        self.assertNotContains(base, f"/documents/{document.pk}/edit/")
        self.assertNotContains(base, f"/documents/{document.pk}/delete/")
        # Expansion's own page: full controls (Edit link) remain.
        own = self.client.get(f"/games/{expansion.pk}/")
        self.assertContains(own, f"/documents/{document.pk}/edit/")

    def test_add_requires_login(self):
        self.client.logout()
        response = self.client.post(self.add_url, {
            "doc_type": "other", "external_url": "https://example.com/x",
        })
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response["Location"])


class DiscoverDocumentsCommandTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.game = Game.objects.create(name="Wingspan")

    def setUp(self):
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)

    def _write(self, relpath, content=b"data"):
        full = Path(settings.MEDIA_ROOT) / relpath
        full.parent.mkdir(parents=True, exist_ok=True)
        full.write_bytes(content)
        return full

    def test_adopts_a_file_under_the_game_folder_in_place(self):
        rel = f"documents/games/Wingspan [{self.game.pk}]/Rulebook.pdf"
        self._write(rel)
        call_command("discover_documents", stdout=StringIO())
        document = Document.objects.get()
        self.assertEqual(document.file.name, rel)  # points at the file in place
        self.assertEqual(document.content_object, self.game)
        self.assertEqual(document.label, "Rulebook")

    def test_adopts_a_file_under_an_edition_subfolder(self):
        edition = Edition.objects.create(game=self.game, name="KS")
        rel = (f"documents/games/Wingspan [{self.game.pk}]/"
               f"KS [{edition.pk}]/Insert.pdf")
        self._write(rel)
        call_command("discover_documents", stdout=StringIO())
        self.assertEqual(Document.objects.get().content_object, edition)

    def test_rerun_is_idempotent(self):
        self._write(f"documents/games/Wingspan [{self.game.pk}]/Rulebook.pdf")
        call_command("discover_documents", stdout=StringIO())
        call_command("discover_documents", stdout=StringIO())
        self.assertEqual(Document.objects.count(), 1)

    def test_unparseable_paths_are_skipped(self):
        self._write("documents/random/loose.pdf")  # unknown top folder
        self._write("documents/games/NoTrailingId/file.pdf")  # no [pk]
        self._write("documents/games/Ghost [999999]/file.pdf")  # pk has no row
        call_command("discover_documents", stdout=StringIO())
        self.assertEqual(Document.objects.count(), 0)

    def test_dry_run_reports_but_writes_nothing(self):
        self._write(f"documents/games/Wingspan [{self.game.pk}]/Rulebook.pdf")
        out = StringIO()
        call_command("discover_documents", "--dry-run", stdout=out)
        self.assertEqual(Document.objects.count(), 0)
        self.assertIn("would adopt", out.getvalue())

    def _expansion(self, name="Wingspan Oceania"):
        expansion = Game.objects.create(name=name, type=Game.Type.EXPANSION)
        expansion.expands.add(self.game)
        return expansion

    def test_adopts_an_expansion_file_nested_under_its_base(self):
        # Issue #99: games/<base>/<expansion>/<file> -> the expansion.
        expansion = self._expansion()
        rel = (f"documents/games/Wingspan [{self.game.pk}]/"
               f"Wingspan Oceania [{expansion.pk}]/Rules.pdf")
        self._write(rel)
        call_command("discover_documents", stdout=StringIO())
        self.assertEqual(Document.objects.get().content_object, expansion)

    def test_adopts_an_expansion_edition_file(self):
        # games/<base>/<expansion>/<edition>/<file> -> the expansion's edition.
        expansion = self._expansion()
        edition = Edition.objects.create(game=expansion, name="Deluxe")
        rel = (f"documents/games/Wingspan [{self.game.pk}]/"
               f"Wingspan Oceania [{expansion.pk}]/"
               f"Deluxe [{edition.pk}]/Insert.pdf")
        self._write(rel)
        call_command("discover_documents", stdout=StringIO())
        self.assertEqual(Document.objects.get().content_object, edition)

    def test_base_edition_still_resolves_at_three_segments(self):
        # #99 disambiguation: a len-3 path is an edition of the base first,
        # only an expansion if no matching base-edition exists.
        edition = Edition.objects.create(game=self.game, name="KS")
        rel = (f"documents/games/Wingspan [{self.game.pk}]/"
               f"KS [{edition.pk}]/Insert.pdf")
        self._write(rel)
        call_command("discover_documents", stdout=StringIO())
        self.assertEqual(Document.objects.get().content_object, edition)

    def test_legacy_top_level_expansion_folder_still_adopts(self):
        # Backward compat: a pre-#99 folder that put the expansion at the top
        # level still adopts to the expansion (games/<expansion> is just a game).
        expansion = self._expansion()
        rel = f"documents/games/Wingspan Oceania [{expansion.pk}]/Rules.pdf"
        self._write(rel)
        call_command("discover_documents", stdout=StringIO())
        self.assertEqual(Document.objects.get().content_object, expansion)


class RestowDocumentsCommandTests(TestCase):
    """Issue #99: relocate document files whose canonical §7 folder changed —
    chiefly expansion docs that predate the base-nesting rule."""

    @classmethod
    def setUpTestData(cls):
        cls.base = Game.objects.create(name="Wingspan")
        cls.expansion = Game.objects.create(
            name="Wingspan Oceania", type=Game.Type.EXPANSION)
        cls.expansion.expands.add(cls.base)

    def setUp(self):
        media_override = override_settings(MEDIA_ROOT=tempfile.mkdtemp())
        media_override.enable()
        self.addCleanup(media_override.disable)

    def _doc_at(self, host, relpath, content=b"data"):
        """A Document whose file is planted at an explicit (possibly stale) path,
        bypassing upload_to so we can simulate a pre-#99 location."""
        full = Path(settings.MEDIA_ROOT) / relpath
        full.parent.mkdir(parents=True, exist_ok=True)
        full.write_bytes(content)
        document = Document(content_object=host, label="Rules")
        document.file.name = relpath
        document.save()
        return document

    def test_moves_an_expansion_doc_to_its_base_nested_path(self):
        old = f"documents/games/Wingspan Oceania [{self.expansion.pk}]/Rules.pdf"
        document = self._doc_at(self.expansion, old)
        new = (f"documents/games/Wingspan [{self.base.pk}]/"
               f"Wingspan Oceania [{self.expansion.pk}]/Rules.pdf")
        call_command("restow_documents", stdout=StringIO())
        document.refresh_from_db()
        self.assertEqual(document.file.name, new)
        self.assertTrue((Path(settings.MEDIA_ROOT) / new).exists())
        self.assertFalse((Path(settings.MEDIA_ROOT) / old).exists())

    def test_leaves_base_game_docs_untouched(self):
        canonical = f"documents/games/Wingspan [{self.base.pk}]/Rules.pdf"
        document = self._doc_at(self.base, canonical)
        call_command("restow_documents", stdout=StringIO())
        document.refresh_from_db()
        self.assertEqual(document.file.name, canonical)
        self.assertTrue((Path(settings.MEDIA_ROOT) / canonical).exists())

    def test_rerun_is_idempotent(self):
        old = f"documents/games/Wingspan Oceania [{self.expansion.pk}]/Rules.pdf"
        self._doc_at(self.expansion, old)
        call_command("restow_documents", stdout=StringIO())
        call_command("restow_documents", stdout=StringIO())
        self.assertEqual(Document.objects.count(), 1)
        new = (f"documents/games/Wingspan [{self.base.pk}]/"
               f"Wingspan Oceania [{self.expansion.pk}]/Rules.pdf")
        self.assertEqual(Document.objects.get().file.name, new)

    def test_dry_run_moves_nothing(self):
        old = f"documents/games/Wingspan Oceania [{self.expansion.pk}]/Rules.pdf"
        document = self._doc_at(self.expansion, old)
        out = StringIO()
        call_command("restow_documents", "--dry-run", stdout=out)
        document.refresh_from_db()
        self.assertEqual(document.file.name, old)
        self.assertTrue((Path(settings.MEDIA_ROOT) / old).exists())
        self.assertIn("would move", out.getvalue())


class ToolsViewTests(TestCase):
    """The superuser Tools page (issue #90): access gate, enqueueing a run,
    the overlap guard, and the task's success/failure recording."""

    @classmethod
    def setUpTestData(cls):
        cls.superuser = get_user_model().objects.create_superuser(
            username="admin", password="pass",
        )
        cls.plain = get_user_model().objects.create_user(
            username="plain", password="pass",
        )

    def test_anonymous_and_non_superuser_are_redirected(self):
        for url in ("/tools/", "/tools/status/"):
            response = self.client.get(url)
            self.assertEqual(response.status_code, 302)
            self.assertIn("/accounts/login/", response.url)
        self.client.login(username="plain", password="pass")
        response = self.client.get("/tools/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_superuser_sees_both_tool_buttons(self):
        self.client.login(username="admin", password="pass")
        response = self.client.get("/tools/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Run BGG sync")
        self.assertContains(response, "Run Cover download")

    def test_post_enqueues_run_and_creates_running_row(self):
        self.client.login(username="admin", password="pass")
        with mock.patch("gamekeeper.views.run_tool_command.delay") as delay:
            response = self.client.post(f"/tools/run/{ToolRun.Kind.BGG_SYNC}/")
        self.assertEqual(response.status_code, 200)
        run = ToolRun.objects.get()
        self.assertEqual(run.kind, ToolRun.Kind.BGG_SYNC)
        self.assertEqual(run.status, ToolRun.Status.RUNNING)
        self.assertEqual(run.triggered_by, self.superuser)
        delay.assert_called_once_with(run.pk)

    def test_enqueue_failure_marks_run_failed_and_returns_partial(self):
        # No broker/worker (the dev reality): .delay() raises. The run must
        # not be left stuck `running`, and the user must get feedback.
        self.client.login(username="admin", password="pass")
        with mock.patch("gamekeeper.views.run_tool_command.delay",
                        side_effect=RuntimeError("broker down")):
            response = self.client.post(
                f"/tools/run/{ToolRun.Kind.DOWNLOAD_COVERS}/")
        self.assertEqual(response.status_code, 200)
        run = ToolRun.objects.get()
        self.assertEqual(run.status, ToolRun.Status.FAILED)
        self.assertIsNotNone(run.finished_at)
        self.assertContains(response, "Could not queue the job")
        # The guard is released — a retry is possible once the worker is back.
        self.assertFalse(ToolRun.is_running(ToolRun.Kind.DOWNLOAD_COVERS))

    def test_unknown_kind_is_rejected(self):
        self.client.login(username="admin", password="pass")
        with mock.patch("gamekeeper.views.run_tool_command.delay") as delay:
            response = self.client.post("/tools/run/bogus/")
        self.assertEqual(response.status_code, 400)
        self.assertEqual(ToolRun.objects.count(), 0)
        delay.assert_not_called()

    def test_overlap_guard_blocks_second_run_of_same_kind(self):
        ToolRun.objects.create(
            kind=ToolRun.Kind.BGG_SYNC, triggered_by=self.superuser,
        )
        self.client.login(username="admin", password="pass")
        with mock.patch("gamekeeper.views.run_tool_command.delay") as delay:
            response = self.client.post(f"/tools/run/{ToolRun.Kind.BGG_SYNC}/")
        self.assertEqual(response.status_code, 200)
        # No second run created, and nothing re-enqueued.
        self.assertEqual(
            ToolRun.objects.filter(kind=ToolRun.Kind.BGG_SYNC).count(), 1,
        )
        delay.assert_not_called()

    def test_task_records_success_and_captures_output(self):
        run = ToolRun.objects.create(
            kind=ToolRun.Kind.DOWNLOAD_COVERS, triggered_by=self.superuser,
        )

        def fake_command(name, *args, **options):
            options["stdout"].write("Downloaded 3 covers.")

        with mock.patch("gamekeeper.tasks.call_command", side_effect=fake_command):
            run_tool_command(run.pk)

        run.refresh_from_db()
        self.assertEqual(run.status, ToolRun.Status.SUCCESS)
        self.assertIn("Downloaded 3 covers.", run.summary)
        self.assertIsNotNone(run.finished_at)

    def test_task_records_failure_with_traceback(self):
        run = ToolRun.objects.create(
            kind=ToolRun.Kind.BGG_SYNC, triggered_by=self.superuser,
        )
        with mock.patch("gamekeeper.tasks.call_command",
                        side_effect=CommandError("Set BGG_USERNAME")):
            run_tool_command(run.pk)

        run.refresh_from_db()
        self.assertEqual(run.status, ToolRun.Status.FAILED)
        self.assertIn("Set BGG_USERNAME", run.summary)
        self.assertIsNotNone(run.finished_at)

    def test_bgg_sync_task_scopes_to_triggering_user(self):
        run = ToolRun.objects.create(
            kind=ToolRun.Kind.BGG_SYNC, triggered_by=self.superuser,
        )
        with mock.patch("gamekeeper.tasks.call_command") as command:
            run_tool_command(run.pk)
        _, kwargs = command.call_args
        self.assertEqual(command.call_args.args[0], "sync_bgg")
        self.assertEqual(kwargs["user"], "admin")

    def test_superuser_sees_cover_previews_button(self):
        # Issue #112: the previews rebuild is a third Tools action.
        self.client.login(username="admin", password="pass")
        response = self.client.get("/tools/")
        self.assertContains(response, "Run Cover previews")

    def test_post_enqueues_cover_previews_run(self):
        self.client.login(username="admin", password="pass")
        with mock.patch("gamekeeper.views.run_tool_command.delay") as delay:
            response = self.client.post(
                f"/tools/run/{ToolRun.Kind.GENERATE_PREVIEWS}/")
        self.assertEqual(response.status_code, 200)
        run = ToolRun.objects.get()
        self.assertEqual(run.kind, ToolRun.Kind.GENERATE_PREVIEWS)
        self.assertEqual(run.status, ToolRun.Status.RUNNING)
        delay.assert_called_once_with(run.pk)

    def test_previews_task_forces_full_regeneration(self):
        # The default backfill only fills missing previews; this action must
        # rebuild every one, so it invokes the command with force=True (#112).
        run = ToolRun.objects.create(
            kind=ToolRun.Kind.GENERATE_PREVIEWS, triggered_by=self.superuser,
        )
        with mock.patch("gamekeeper.tasks.call_command") as command:
            run_tool_command(run.pk)
        run.refresh_from_db()
        self.assertEqual(run.status, ToolRun.Status.SUCCESS)
        self.assertEqual(command.call_args.args[0], "generate_cover_previews")
        self.assertTrue(command.call_args.kwargs["force"])


class PushBggStatusTaskTests(TestCase):
    """push_bgg_status_task (issue #117): the thin Celery wrapper views
    enqueue via .delay() — just loads game/user by pk and delegates."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)

    def test_task_loads_game_and_user_and_delegates(self):
        with mock.patch("gamekeeper.tasks.bgg_sync.push_bgg_status") as push:
            push_bgg_status_task(
                self.game.pk, Game.BggCollectionStatus.WISHLIST, self.user.pk,
                priority=3,
            )
        push.assert_called_once_with(
            self.game, Game.BggCollectionStatus.WISHLIST, priority=3, user=self.user,
        )


class PushBggFortradeTaskTests(TestCase):
    """push_bgg_fortrade_task (issue #82) — the fortrade counterpart of
    PushBggStatusTaskTests, same thin-wrapper contract."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username="kernicek")
        cls.game = Game.objects.create(name="5 Minute Dungeon")
        BggLink.objects.create(game=cls.game, bgg_id=207830, is_primary=True)

    def test_task_loads_game_and_user_and_delegates(self):
        with mock.patch("gamekeeper.tasks.bgg_sync.push_bgg_fortrade") as push:
            push_bgg_fortrade_task(self.game.pk, True, self.user.pk)
        push.assert_called_once_with(self.game, True, user=self.user)


class ImpersonationTests(TestCase):
    """Superuser impersonation (issue #108, via django-impersonate): the
    superuser-only gate, no privilege escalation, the always-visible banner,
    and the Tools-page picker."""

    @classmethod
    def setUpTestData(cls):
        User = get_user_model()
        cls.superuser = User.objects.create_superuser(
            username="admin", password="pass",
        )
        cls.other_super = User.objects.create_superuser(
            username="admin2", password="pass",
        )
        cls.member = User.objects.create_user(
            username="member", password="pass",
        )

    def test_superuser_can_impersonate_member(self):
        self.client.login(username="admin", password="pass")
        response = self.client.get(f"/impersonate/{self.member.pk}/")
        self.assertEqual(response.status_code, 302)
        # Subsequent requests render as the member.
        response = self.client.get("/")
        self.assertTrue(response.context["user"].is_impersonate)
        self.assertEqual(response.context["user"].username, "member")

    def test_banner_shows_while_impersonating(self):
        self.client.login(username="admin", password="pass")
        self.client.get(f"/impersonate/{self.member.pk}/")
        response = self.client.get("/")
        self.assertContains(response, "Impersonating")
        self.assertContains(response, "Exit impersonation")

    def test_stop_restores_superuser_identity(self):
        self.client.login(username="admin", password="pass")
        self.client.get(f"/impersonate/{self.member.pk}/")
        self.client.get("/impersonate/stop/")
        response = self.client.get("/")
        self.assertFalse(response.context["user"].is_impersonate)
        self.assertEqual(response.context["user"].username, "admin")

    def test_non_superuser_cannot_impersonate(self):
        # REQUIRE_SUPERUSER=True: a plain member's start attempt establishes no
        # impersonation.
        self.client.login(username="member", password="pass")
        self.client.get(f"/impersonate/{self.superuser.pk}/")
        response = self.client.get("/")
        self.assertFalse(response.context["user"].is_impersonate)

    def test_cannot_impersonate_another_superuser(self):
        # ALLOW_SUPERUSER=False: no privilege escalation — a superuser target is
        # never applied, even to a superuser impersonator.
        self.client.login(username="admin", password="pass")
        self.client.get(f"/impersonate/{self.other_super.pk}/")
        response = self.client.get("/")
        self.assertFalse(response.context["user"].is_impersonate)

    def test_tools_picker_offers_only_non_superusers(self):
        self.client.login(username="admin", password="pass")
        response = self.client.get("/tools/")
        self.assertContains(response, "Impersonate a user")
        self.assertContains(response, f"/impersonate/{self.member.pk}/")
        self.assertNotContains(response, f"/impersonate/{self.other_super.pk}/")

    def test_picker_hidden_from_non_superusers(self):
        self.client.login(username="member", password="pass")
        response = self.client.get("/")
        self.assertEqual(list(response.context["impersonatable_users"]), [])


class EnvBannerTests(TestCase):
    """The non-production banner (issue #109) is gated on ENVIRONMENT."""

    def test_processor_shows_banner_off_production(self):
        with override_settings(ENVIRONMENT="development"):
            ctx = environment(None)
        self.assertTrue(ctx["SHOW_ENV_BANNER"])
        self.assertEqual(ctx["ENVIRONMENT"], "development")

    def test_processor_hides_banner_on_production(self):
        with override_settings(ENVIRONMENT="production"):
            ctx = environment(None)
        self.assertFalse(ctx["SHOW_ENV_BANNER"])

    def test_banner_rendered_off_production(self):
        with override_settings(ENVIRONMENT="staging"):
            response = self.client.get("/accounts/login/")
        self.assertContains(response, "not production")
        self.assertContains(response, "staging")

    def test_banner_absent_on_production(self):
        with override_settings(ENVIRONMENT="production"):
            response = self.client.get("/accounts/login/")
        self.assertNotContains(response, "not production")


class HeaderActionsDropdownTests(TestCase):
    """Issue #128: the multi-action page headers fold their right-side links
    into a single 'Actions' dropdown (mobile-friendly). Each page still exposes
    the same action links, now as dropdown items."""

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="kernicek", password="pass",
        )

    def setUp(self):
        self.client.login(username="kernicek", password="pass")

    def assert_dropdown_actions(self, url, links):
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        # The header 'Actions' dropdown replaces the sprawling inline link row.
        # (Match the button specifically — the navbar username menu, issue #137,
        # is also a data-bs-toggle="dropdown" and must not be what we detect.)
        self.assertContains(response, '>Actions</button>')
        # Every action survives the fold-in, as a dropdown item.
        for href, label in links:
            self.assertContains(response, f'href="{href}"')
            self.assertContains(response, label)

    def test_collection_header_actions_in_dropdown(self):
        self.assert_dropdown_actions("/", [
            ("/games/add/", "Add game"),
            ("/games/import/", "Import from BGG"),
            ("/series/add/", "New series"),
            ("/families/add/", "New family"),
        ])

    def test_curation_keeps_nav_button_and_moves_reset_into_filters(self):
        # Like purchases: the nav action ('Archived copies') stays as a header
        # button, and 'Reset filters' moves into the filter card.
        response = self.client.get("/curation/")
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, '>Actions</button>')
        self.assertContains(response, 'href="/curation/archived/"')
        self.assertContains(response, "Archived copies")
        self.assertContains(response, 'href="/curation/"')
        self.assertContains(response, "Reset filters")

    def test_archive_keeps_nav_button_and_moves_reset_into_filters(self):
        response = self.client.get("/curation/archived/")
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, '>Actions</button>')
        self.assertContains(response, 'href="/curation/"')
        self.assertContains(response, "Back to curation")
        self.assertContains(response, 'href="/curation/archived/"')
        self.assertContains(response, "Reset filters")

    def test_purchases_keeps_add_button_and_moves_reset_into_filters(self):
        # Purchases is the exception: a plain 'Add purchase' button stays in
        # the header (no dropdown), and 'Reset filters' lives in the filter card.
        response = self.client.get("/purchases/")
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, '>Actions</button>')
        self.assertContains(response, 'href="/purchases/add/"')
        self.assertContains(response, "Add purchase")
        self.assertContains(response, 'href="/purchases/"')
        self.assertContains(response, "Reset filters")


class CoverWritePermissionsTests(SimpleTestCase):
    """#113: cover/preview writes must produce world-traversable directories so a
    cover command run from a root shell (Portainer console) can't bake a 700
    covers/previews/ dir that nginx (a different uid) 403s. Verified at the storage
    layer rather than via POSIX stat bits, which os.chmod can't set on the Windows
    dev suite."""

    def test_settings_pin_world_readable_media_permissions(self):
        self.assertEqual(settings.FILE_UPLOAD_PERMISSIONS, 0o644)
        self.assertEqual(settings.FILE_UPLOAD_DIRECTORY_PERMISSIONS, 0o755)

    def test_cover_storage_will_chmod_dirs_and_files(self):
        from django.core.files.storage import default_storage
        from gamekeeper.models import DocumentStorage
        # Every cover/preview write funnels through the default FileSystemStorage;
        # 755 dirs + 644 files are applied on save regardless of the writer umask.
        self.assertEqual(default_storage.directory_permissions_mode, 0o755)
        self.assertEqual(default_storage.file_permissions_mode, 0o644)
        # The §7 document storage subclass inherits the same global modes, so the
        # storage layer (not one command) is what is healed.
        self.assertEqual(DocumentStorage().directory_permissions_mode, 0o755)


def _fake_ghcr_session(head_map, tags=None):
    """Build a Mock requests.Session for the GHCR update check (issue #95).

    head_map maps a tag name -> (status_code, digest); a missing tag answers 404.
    GET serves the anonymous token endpoint and the tags/list endpoint by URL.
    """
    session = mock.Mock()

    def head(url, **kwargs):
        tag = url.rsplit("/", 1)[-1]
        status, digest = head_map.get(tag, (404, None))
        resp = mock.Mock(status_code=status)
        resp.headers = {"Docker-Content-Digest": digest} if digest else {}
        resp.raise_for_status = mock.Mock()
        return resp

    def get(url, **kwargs):
        resp = mock.Mock(status_code=200)
        resp.raise_for_status = mock.Mock()
        if url.endswith("/tags/list"):
            resp.json = mock.Mock(return_value={"tags": tags or []})
        else:  # the anonymous token-exchange endpoint
            resp.json = mock.Mock(return_value={"token": "anon-token"})
        return resp

    session.head.side_effect = head
    session.get.side_effect = get
    session.close = mock.Mock()
    return session


@override_settings(
    APP_VERSION="v5",
    GHCR_IMAGE="ghcr.io/kernicek/gamekeeper",
    GHCR_TOKEN="tok",
)
class UpdateCheckTests(TestCase):
    """Issue #95: superuser navbar icon when a newer GHCR image is published."""

    def setUp(self):
        # get_update_status caches in the default LocMemCache, which persists
        # across tests in-process — clear it so each test sees a fresh check.
        cache.clear()

    def test_update_available_reports_running_and_named_latest(self):
        session = _fake_ghcr_session(
            head_map={
                "latest": (200, "sha256:NEW"),
                "v5": (200, "sha256:OLD"),
                "v6": (200, "sha256:NEW"),
            },
            tags=["latest", "v5", "v6"],
        )
        with mock.patch("gamekeeper.updates.requests.Session", return_value=session):
            from gamekeeper.updates import check_for_update
            result = check_for_update()
        self.assertEqual(
            result,
            {"update_available": True, "running": "v5", "latest": "v6"},
        )

    def test_up_to_date_reports_no_update(self):
        session = _fake_ghcr_session(
            head_map={"latest": (200, "sha256:SAME"), "v5": (200, "sha256:SAME")},
        )
        with mock.patch("gamekeeper.updates.requests.Session", return_value=session):
            from gamekeeper.updates import check_for_update
            result = check_for_update()
        self.assertEqual(result["update_available"], False)
        # No tags/list scan when the running image already matches :latest.
        session.get.assert_not_called()

    def test_check_returns_none_on_request_error(self):
        session = _fake_ghcr_session(head_map={})
        session.head.side_effect = requests.RequestException("boom")
        with mock.patch("gamekeeper.updates.requests.Session", return_value=session):
            from gamekeeper.updates import check_for_update
            result = check_for_update()
        self.assertIsNone(result)

    @override_settings(GHCR_TOKEN="")
    def test_anonymous_token_used_when_no_ghcr_token(self):
        session = _fake_ghcr_session(
            head_map={
                "latest": (200, "sha256:NEW"),
                "v5": (200, "sha256:OLD"),
                "v6": (200, "sha256:NEW"),
            },
            tags=["latest", "v5", "v6"],
        )
        with mock.patch("gamekeeper.updates.requests.Session", return_value=session):
            from gamekeeper.updates import check_for_update
            result = check_for_update()
        self.assertTrue(result["update_available"])
        # Without a configured token the anonymous token endpoint is hit.
        token_calls = [c for c in session.get.call_args_list if "/token" in c.args[0]]
        self.assertTrue(token_calls)

    def test_get_update_status_skips_check_without_app_version(self):
        with override_settings(APP_VERSION=""):
            with mock.patch("gamekeeper.updates.check_for_update") as checker:
                from gamekeeper.updates import get_update_status
                self.assertIsNone(get_update_status())
                checker.assert_not_called()

    def test_get_update_status_caches_result(self):
        payload = {"update_available": True, "running": "v5", "latest": "v6"}
        with mock.patch(
            "gamekeeper.updates.check_for_update", return_value=payload,
        ) as checker:
            from gamekeeper.updates import get_update_status
            first = get_update_status()
            second = get_update_status()
        self.assertEqual(first, payload)
        self.assertEqual(second, payload)
        # The expensive check runs once; the second read comes from cache.
        checker.assert_called_once()


class UpdateNoticeNavbarTests(TestCase):
    """Issue #95: the navbar update icon is superuser-only and update-gated."""

    @classmethod
    def setUpTestData(cls):
        cls.superuser = get_user_model().objects.create_superuser(
            username="admin", password="pass",
        )
        cls.plain = get_user_model().objects.create_user(
            username="plain", password="pass",
        )

    def test_superuser_sees_update_icon_when_available(self):
        status = {"update_available": True, "running": "v5", "latest": "v6"}
        self.client.login(username="admin", password="pass")
        with mock.patch(
            "gamekeeper.context_processors.get_update_status", return_value=status,
        ):
            body = self.client.get("/settings/").content.decode()
        self.assertIn("bi-arrow-up-circle-fill", body)
        self.assertIn("Update available: v5", body)
        self.assertIn("v6", body)

    def test_non_superuser_never_sees_update_icon(self):
        status = {"update_available": True, "running": "v5", "latest": "v6"}
        self.client.login(username="plain", password="pass")
        with mock.patch(
            "gamekeeper.context_processors.get_update_status", return_value=status,
        ):
            body = self.client.get("/settings/").content.decode()
        self.assertNotIn("bi-arrow-up-circle-fill", body)

    def test_no_icon_when_up_to_date(self):
        status = {"update_available": False, "running": "v5", "latest": "v5"}
        self.client.login(username="admin", password="pass")
        with mock.patch(
            "gamekeeper.context_processors.get_update_status", return_value=status,
        ):
            body = self.client.get("/settings/").content.decode()
        self.assertNotIn("bi-arrow-up-circle-fill", body)


# ===========================================================================
# Issue #25  Site footer
# ===========================================================================

class SiteFooterTests(TestCase):
    """base.html renders a persistent footer with app name/version, a link to
    the source repo, and a copyright line, on every page."""

    def test_footer_shows_app_name_and_repo_link_and_copyright(self):
        # Any page renders base.html; the login page needs no fixtures.
        response = self.client.get("/accounts/login/")
        self.assertContains(response, "<footer")
        self.assertContains(response, "GameKeeper")
        self.assertContains(
            response, 'href="https://github.com/kernicek/GameKeeper"')
        self.assertContains(response, "&copy; 2026 Vojta Karen")

    def test_footer_adds_v_prefix_to_calver_tag(self):
        with override_settings(APP_VERSION="2026.07.17.1"):
            response = self.client.get("/accounts/login/")
        self.assertContains(response, "GameKeeper v2026.07.17.1")

    def test_footer_does_not_double_prefix_a_v_already_present(self):
        with override_settings(APP_VERSION="v5"):
            response = self.client.get("/accounts/login/")
        self.assertContains(response, "GameKeeper v5")
        self.assertNotContains(response, "vv5")

    def test_footer_omits_bare_version_when_app_version_is_unset(self):
        # Dev has no APP_VERSION baked in; the footer must not render a bare
        # "GameKeeper v" or "GameKeeper " with nothing after it.
        with override_settings(APP_VERSION=""):
            response = self.client.get("/accounts/login/")
        self.assertContains(response, "GameKeeper")
        self.assertNotContains(response, "GameKeeper v")

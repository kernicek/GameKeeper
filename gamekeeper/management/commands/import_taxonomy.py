"""Re-runnable import of the Mastersheet's taxonomy columns (DESIGN §10 + §12).

Two sources in one workbook:
  * "Overview" sheet — the curated per-game taxonomy the first import pass
    deferred: language columns 17-19 (Lang(components), Difficulty for
    non-speakers, Lang(how much)), game-type flag columns 32-38 (values
    'y' / 'opt' / 'app'; columns 39-42 are derived Pure/All aggregates —
    skipped), player-conflict 43, app/soundtrack 45-48, campaign-structure
    flags 76-79, theme flags 169-203 plus the Adapts-from flags 206-209
    (folded in as "Adapts: ..." themes per §10). Theme cells hold 'y' or
    'f' (favourite). Data starts at row 4.
  * "APPs" sheet — one row per game with Android/Steam 'y' flags, imported
    as DigitalImplementations. Data starts at row 2.

Rows are matched to existing Games via the BGG hyperlink in column 3 — the
identity import_mastersheet established — never by title. Game-level fields
land on the Game; Lang(components) lands on the Edition the --user's Copy
points at (a Czech edition is a different edition). Mechanics columns 80-125
are NOT imported: DESIGN §10 makes mechanics BGG-synced tags, populated by
sync_bgg's /thing pass instead (§15) — Tag(kind=mechanic) lands there.

Idempotent: Game fields are assigned from the row wholesale (the sheet is
the source of truth for taxonomy), tags/game-types/implementations are
upserted by natural keys, so re-running does not duplicate.
"""

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from gamekeeper.models import (
    BggLink,
    Copy,
    DigitalImplementation,
    Edition,
    Game,
    GameType,
    GameTag,
    Tag,
)
from gamekeeper.bgg import extract_bgg_id
from gamekeeper.management.commands.import_mastersheet import (
    cell_link,
    cell_text,
)

DATA_START_ROW = 4

# 1-based column numbers in the Overview sheet.
COL_TITLE = 2
COL_BGG = 3
COL_LANG_COMPONENTS = 17
COL_NON_SPEAKERS = 18
COL_LANG_HOW_MUCH = 19
COL_PLAYER_CONFLICT = 43
COL_APP = 45
COL_APP_VERSION = 46
COL_SOUNDTRACK_AMBIENCE = 47
COL_SOUNDTRACK_TIMER = 48

GAME_TYPE_COLS = [
    (32, GameType.Type.ONE_VS_ALL),
    (33, GameType.Type.COMPETITIVE),
    (34, GameType.Type.COOPERATIVE),
    (35, GameType.Type.SEMI_COOP),
    (36, GameType.Type.SOLO),
    (37, GameType.Type.TEAM),
    (38, GameType.Type.TRAITOR),
]

CAMPAIGN_COLS = [
    (76, "is_campaign"),
    (77, "is_legacy"),
    (78, "has_scenarios"),
    (79, "is_one_off"),
]

# Theme flag columns; the header (row 1) carries the theme name. The two
# ranges bracket the derived "All" aggregate columns, which are skipped.
THEME_COLS = list(range(169, 204))
ADAPTS_COLS = list(range(206, 210))  # headers already read "Adapts: Book" etc.

LANG_COMPONENTS_MAP = {
    "EN": Edition.ComponentsLanguage.ENGLISH,
    "CZ": Edition.ComponentsLanguage.CZECH,
    "C+E": Edition.ComponentsLanguage.CZECH_ENGLISH,
    "N/A": Edition.ComponentsLanguage.NONE,
}

NON_SPEAKERS_MAP = {
    "(no text)": Game.LanguageDependency.NO_TEXT,
    "trivial": Game.LanguageDependency.TRIVIAL,
    "easy": Game.LanguageDependency.EASY,
    "medium": Game.LanguageDependency.MEDIUM,
    "difficult": Game.LanguageDependency.DIFFICULT,
}

# Leading keyword of a "Lang (how much)" cell -> dependency level, on the
# merged scale (issue #2); the rest of the cell is detail kept in
# language_dependency_note. "doable" has no direct counterpart on the
# merged 5-value scale, so it folds into MEDIUM — this level is only used
# when the "Difficulty for non-speakers" column doesn't resolve the game.
LANG_DEPENDENCY_MAP = {
    "trivial": Game.LanguageDependency.TRIVIAL,
    "easy": Game.LanguageDependency.EASY,
    "doable": Game.LanguageDependency.MEDIUM,
    "hard": Game.LanguageDependency.DIFFICULT,
    "difficult": Game.LanguageDependency.DIFFICULT,
    "diffucult": Game.LanguageDependency.DIFFICULT,  # the sheet's recurring typo
}

GAME_TYPE_QUALIFIER_MAP = {
    "y": "",
    "opt": GameType.Qualifier.OPTIONAL,
    "app": GameType.Qualifier.APP,
}

APP_MAP = {
    "y": Game.AppUse.REQUIRED,
    "opt": Game.AppUse.OPTIONAL,
}

# APPs sheet layout.
APPS_DATA_START_ROW = 2
APPS_COL_TITLE = 2
APPS_COL_BGG = 3
APPS_PLATFORM_COLS = [
    (6, DigitalImplementation.Platform.ANDROID),
    (7, DigitalImplementation.Platform.STEAM),
]

DECISION_NOTES = [
    "Game-type qualifiers: 'opt' -> optional mode, 'app' -> only with the "
    "companion app; the derived Pure Comp/Coop/Solo and All columns are not "
    "imported.",
    "Campaign structure is four independent booleans, not one choice — the "
    "sheet freely combines them (Scenarios + One-off is common).",
    "Language dependency (issue #2): 'Difficulty for non-speakers' and 'Lang "
    "(how much)' are merged into one language_dependency field. The "
    "non-speakers column wins when both resolve; otherwise the lang column's "
    "leading keyword is used ('doable' folds into 'medium', 'hard' into "
    "'difficult'). The full lang cell survives in the note when it says more.",
    "Player conflict: clean 0-3 values fill player_conflict; the user's own "
    "uncertainty markers ('0-1?', '1/0') go to player_conflict_note verbatim, "
    "value left empty.",
    "A '?' in any taxonomy cell means unknown -> field left empty, not "
    "reported.",
    "Theme cell 'f' -> tagged as favourite (GameTag.is_favourite).",
]

DEFERRED_NOTES = [
    "Mechanics columns 80-125: DESIGN §10 makes mechanics BGG-synced "
    "Tag(kind=mechanic) rows, populated by sync_bgg's /thing pass (§15) — "
    "NOT imported from the sheet.",
    "Player-elimination column 44 (3 marks): already covered by the BGG "
    "'Player elimination' mechanic tag.",
    "APPs sheet 'Owned physically?' column: ownership is already modelled by "
    "Copies.",
    "GameChooser / Plays / Selling / Gamebooks+ sheets: derived views or out "
    "of scope per DESIGN §12.",
]


class Command(BaseCommand):
    help = "Import the Mastersheet's taxonomy columns and APPs sheet (DESIGN §10)."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the Mastersheet .xlsx file.")
        parser.add_argument(
            "--user", required=True,
            help="Username whose Copy's Edition receives the components language.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Parse and report, but write nothing to the database.",
        )

    def handle(self, *args, **options):
        try:
            workbook = openpyxl.load_workbook(options["path"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {options['path']}")
        for sheet_name in ("Overview", "APPs"):
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f'The workbook has no "{sheet_name}" sheet.')

        User = get_user_model()
        try:
            user = User.objects.get(username=options["user"])
        except User.DoesNotExist:
            raise CommandError(f"User {options['user']!r} does not exist.")

        self.counts = {}
        self.skipped = []    # (where, title, reason) — nothing imported
        self.ambiguous = []  # (where, title, note) — imported with a caveat
        self._tags = {}      # (kind, name) -> Tag

        overview = workbook["Overview"]
        self._theme_names = {}
        for col in THEME_COLS + ADAPTS_COLS:
            header = overview.cell(row=1, column=col).value
            if header:
                self._theme_names[col] = " ".join(str(header).split())

        with transaction.atomic():
            for row in range(DATA_START_ROW, overview.max_row + 1):
                self._import_overview_row(overview, row, user)
            for row in range(APPS_DATA_START_ROW, workbook["APPs"].max_row + 1):
                self._import_apps_row(workbook["APPs"], row)
            if options["dry_run"]:
                transaction.set_rollback(True)

        self._print_report(dry_run=options["dry_run"])

    def _bump(self, key, created=None):
        if created is not None:
            key = f"{key} {'created' if created else 'updated'}"
        self.counts[key] = self.counts.get(key, 0) + 1

    def _get_tag(self, kind, name):
        if (kind, name) not in self._tags:
            tag, created = Tag.objects.get_or_create(kind=kind, name=name)
            if created:
                self._bump("tags created")
            self._tags[(kind, name)] = tag
        return self._tags[(kind, name)]

    # --- Overview sheet -------------------------------------------------------

    def _import_overview_row(self, sheet, row, user):
        def cell(col):
            return sheet.cell(row=row, column=col)

        title = cell_text(cell(COL_TITLE))
        bgg_url = cell_link(cell(COL_BGG))
        if not title and not bgg_url:
            return  # entirely blank row
        bgg_id = extract_bgg_id(bgg_url)
        if not bgg_id:
            self.skipped.append((f"row {row}", title, "no BGG link — cannot establish identity"))
            return
        link = BggLink.objects.filter(bgg_id=bgg_id, is_primary=True).select_related("game").first()
        if not link:
            self.skipped.append(
                (f"row {row}", title, f"BGG id {bgg_id} not in database — run import_mastersheet first"),
            )
            return
        game = link.game

        # --- Game-level curated fields (assigned wholesale: the sheet is the
        # source of truth, so an emptied cell empties the field on re-run) ---
        game.language_dependency, game.language_dependency_note = (
            self._map_language_and_difficulty(
                row, title, cell(COL_NON_SPEAKERS), cell(COL_LANG_HOW_MUCH),
            )
        )
        game.player_conflict, game.player_conflict_note = (
            self._map_player_conflict(cell(COL_PLAYER_CONFLICT))
        )
        game.companion_app = self._map_app(row, title, cell(COL_APP))
        game.has_app_version = bool(cell_text(cell(COL_APP_VERSION)))
        game.soundtrack_ambience = bool(cell_text(cell(COL_SOUNDTRACK_AMBIENCE)))
        game.soundtrack_timer = bool(cell_text(cell(COL_SOUNDTRACK_TIMER)))
        for col, field in CAMPAIGN_COLS:
            setattr(game, field, bool(cell_text(cell(col))))
        game.save(update_fields=[
            "language_dependency", "language_dependency_note",
            "player_conflict", "player_conflict_note",
            "companion_app", "has_app_version", "soundtrack_ambience",
            "soundtrack_timer", "is_campaign", "is_legacy", "has_scenarios",
            "is_one_off", "updated_at",
        ])
        self._bump("games updated")

        # --- Game types ---
        for col, game_type in GAME_TYPE_COLS:
            raw = cell_text(cell(col))
            if not raw:
                continue
            qualifier = GAME_TYPE_QUALIFIER_MAP.get(raw.lower())
            if qualifier is None:
                self.ambiguous.append(
                    (f"row {row}", title, f"unrecognised game-type mark {raw!r} for {game_type} — skipped"),
                )
                continue
            _, created = GameType.objects.update_or_create(
                game=game, game_type=game_type, defaults={"qualifier": qualifier},
            )
            self._bump("game types", created=created)

        # --- Themes (incl. Adapts-from) ---
        for col, name in self._theme_names.items():
            raw = cell_text(cell(col))
            if not raw:
                continue
            lowered = raw.lower()
            if lowered not in ("y", "f"):
                self.ambiguous.append(
                    (f"row {row}", title, f"unrecognised theme mark {raw!r} for {name!r} — tagged anyway"),
                )
            tag = self._get_tag(Tag.Kind.THEME, name)
            _, created = GameTag.objects.update_or_create(
                game=game, tag=tag, defaults={"is_favourite": lowered == "f"},
            )
            self._bump("theme tags", created=created)

        # --- Components language (per-Edition, via the user's Copy) ---
        self._import_components_language(row, title, cell(COL_LANG_COMPONENTS), game, user)

    def _map_non_speakers(self, row, title, cell):
        raw = cell_text(cell)
        if not raw or raw == "?":
            return ""
        mapped = NON_SPEAKERS_MAP.get(raw.lower())
        if mapped is None:
            self.ambiguous.append(
                (f"row {row}", title, f"unrecognised non-speaker difficulty {raw!r} — left empty"),
            )
            return ""
        return mapped

    def _map_language_note(self, cell):
        """Parse the "Lang (how much)" cell only — no side effects. The
        caller decides whether the "kept in note only" report applies once
        merged with the non-speakers column (issue #2)."""
        raw = cell_text(cell)
        if not raw or raw == "?":
            return "", ""
        first_word = raw.split()[0].strip(",;:-(").lower()
        level = LANG_DEPENDENCY_MAP.get(first_word, "")
        note = raw if (not level or raw.lower() != first_word) else ""
        return level, note[:300]

    def _map_language_and_difficulty(self, row, title, non_speaker_cell, lang_cell):
        """Merge the sheet's two language-difficulty columns (issue #2):
        "Difficulty for non-speakers" wins when it resolves; otherwise fall
        back to the "Lang (how much)" cell's leading keyword. The ambiguous
        "kept in note only" report only fires when the final merged value is
        still empty — the non-speakers column can rescue an unparseable lang
        cell."""
        non_speaker_level = self._map_non_speakers(row, title, non_speaker_cell)
        lang_level, note = self._map_language_note(lang_cell)
        merged = non_speaker_level or lang_level
        if not merged and note:
            self.ambiguous.append(
                (f"row {row}", title,
                 f"language-dependency {note!r} has no recognised level — kept in note only"),
            )
        return merged, note

    def _map_player_conflict(self, cell):
        value = cell.value
        if value is None:
            return None, ""
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if int(value) == value and 0 <= value <= 3:
                return int(value), ""
        raw = cell_text(cell)
        if raw == "?":
            return None, ""
        # The user's own uncertainty markers ("0-1?", "1/0") — note, no value.
        return None, raw[:300]

    def _map_app(self, row, title, cell):
        raw = cell_text(cell)
        if not raw:
            return ""
        mapped = APP_MAP.get(raw.lower())
        if mapped is None:
            self.ambiguous.append(
                (f"row {row}", title, f"unrecognised app mark {raw!r} — left empty"),
            )
            return ""
        return mapped

    def _import_components_language(self, row, title, cell, game, user):
        raw = cell_text(cell)
        if not raw or raw == "?":
            return
        language = LANG_COMPONENTS_MAP.get(raw.upper())
        if language is None:
            self.ambiguous.append(
                (f"row {row}", title, f"unrecognised components language {raw!r} — skipped"),
            )
            return
        copies = list(
            Copy.objects.filter(owner=user, edition__game=game).select_related("edition"),
        )
        if not copies:
            self.skipped.append(
                (f"row {row}", title,
                 f"{user.username} owns no Copy of this game — components language not placed"),
            )
            return
        edition = copies[0].edition
        if len(copies) > 1:
            self.ambiguous.append(
                (f"row {row}", title,
                 f"{len(copies)} copies owned — components language put on {edition}"),
            )
        if edition.components_language != language:
            edition.components_language = language
            edition.save(update_fields=["components_language"])
            self._bump("edition languages set")

    # --- APPs sheet -----------------------------------------------------------

    def _import_apps_row(self, sheet, row):
        def cell(col):
            return sheet.cell(row=row, column=col)

        title = cell_text(cell(APPS_COL_TITLE))
        bgg_url = cell_link(cell(APPS_COL_BGG))
        if not title and not bgg_url:
            return  # entirely blank row
        bgg_id = extract_bgg_id(bgg_url)
        if not bgg_id:
            self.skipped.append(
                (f"APPs row {row}", title, "no BGG link — cannot establish identity"),
            )
            return
        link = BggLink.objects.filter(bgg_id=bgg_id, is_primary=True).select_related("game").first()
        if not link:
            self.skipped.append(
                (f"APPs row {row}", title,
                 f"BGG id {bgg_id} not in database (likely not owned physically) — skipped"),
            )
            return

        for col, platform in APPS_PLATFORM_COLS:
            raw = cell_text(cell(col))
            if not raw:
                continue
            if raw.lower() != "y":
                self.ambiguous.append(
                    (f"APPs row {row}", title,
                     f"unrecognised {platform} mark {raw!r} — imported anyway"),
                )
            _, created = DigitalImplementation.objects.get_or_create(
                game=link.game, platform=platform,
            )
            if created:
                self._bump("digital implementations created")

    # --- report ---------------------------------------------------------------

    def _print_report(self, dry_run):
        write = self.stdout.write
        if dry_run:
            write(self.style.WARNING("DRY RUN — nothing was written to the database.\n"))

        write(self.style.MIGRATE_HEADING("Summary"))
        for key in sorted(self.counts):
            write(f"  {key}: {self.counts[key]}")

        write(self.style.MIGRATE_HEADING("Skipped rows"))
        if self.skipped:
            for where, title, reason in self.skipped:
                write(f"  {where} ({title!r}): {reason}")
        else:
            write("  none")

        write(self.style.MIGRATE_HEADING("Ambiguous (imported with caveats)"))
        if self.ambiguous:
            for where, title, note in self.ambiguous:
                write(f"  {where} ({title!r}): {note}")
        else:
            write("  none")

        write(self.style.MIGRATE_HEADING("Interpretation decisions"))
        for note in DECISION_NOTES:
            write(f"  - {note}")

        write(self.style.MIGRATE_HEADING("Deferred (not imported)"))
        for note in DEFERRED_NOTES:
            write(f"  - {note}")

        write(self.style.SUCCESS("Done." if not dry_run else "Dry run complete."))

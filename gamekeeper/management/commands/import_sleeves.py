"""Re-runnable import of the Mastersheet's sleeve data (DESIGN §5 + §12).

Two sources in one workbook:
  * "Sleeves" sheet — the CardSize catalog (row 1 display names, row 2 WxH
    dimensions, sizes across columns C..AA), the Tlama product catalog
    (row 40 product names, row 41 shop URLs) and the pack inventory
    (row 20 "full packs owned"). Column B is a TOTAL column — skipped.
    Rows 4-19/21-38 are derived aggregates / prices — cross-checks only.
  * "Overview" sheet — per-game sleeve requirements and sleeved status as
    column pairs from 130 to 158 (even column = card count, odd column =
    brand/status code), plus two free-text "Other" pairs: 160 (num) +
    162 (WxH text) and 163 (num) + 165 (WxH text). Data starts at row 4.

Rows are matched to existing Games via the BGG hyperlink in column 3 — the
same identity import_mastersheet established — never by title. Requirements
land on the Edition the user's Copy points at; statuses land on the Copy.

CardSizes are keyed by (width, height): the sheet names the same size
differently in different places (Teal vs Azur 45x68), so other spellings
become aliases. Idempotent throughout: everything is upserted by natural
keys, so re-running does not duplicate.

Brand/status code interpretation (documented decisions, cf. the Sleeves
sheet's own per-code aggregate rows 9-15):
  * tl / at / up / upq -> sleeved with that brand's product for the size.
  * D<letter> (Tlama Diamond colour shorthand, e.g. DR) -> sleeved, Tlama.
  * 'c'  -> sleeved, product unknown (the sheet counts 'c' under "sleeved").
  * '?'  -> to-sleeve (wants sleeving, not done).
  * a count with NO code -> not-sleeved (the sheet records no status).
  * anything else (stray numbers/sizes) -> data noise: treated as no code,
    reported as ambiguous.
"""

import re
from decimal import Decimal, InvalidOperation

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from gamekeeper.models import (
    BggLink,
    CardSize,
    Copy,
    CopySleeveStatus,
    SleeveInventory,
    SleeveProduct,
    SleeveRequirement,
)
from gamekeeper.bgg import extract_bgg_id
from gamekeeper.management.commands.import_mastersheet import (
    cell_link,
    cell_text,
)

DATA_START_ROW = 4

# 1-based column numbers in the Overview sheet.
COL_TITLE = 2
COL_BGG = 3

# Overview requirement columns: (count column, canonical name, W, H).
# The brand/status code sits in count column + 1.
OVERVIEW_SIZES = [
    (130, "Mini", "41", "63"),
    (132, "Small*", "44", "63"),
    (134, "Small", "43", "66"),
    (136, "Teal", "45", "68"),
    (138, "White", "50", "75"),
    (140, "Purple", "56", "87"),
    (142, "Medium", "57.5", "89"),
    (144, "Large", "59", "91"),
    (146, "Standard", "63", "88"),
    (148, "Extra large", "65", "100"),
    (150, "Square", "70", "70"),
    (152, "Tarot*", "70", "110"),
    (154, "Tarot", "70", "120"),
    (156, "Square", "80", "80"),
    (158, "Oversized", "80", "120"),
]
# Free-text size pairs: (count column, WxH-text column).
OVERVIEW_OTHER_PAIRS = [(160, 162), (163, 165)]

# Sleeves sheet layout (pivot: sizes across columns, metrics down rows).
SLEEVES_NAME_ROW = 1
SLEEVES_DIM_ROW = 2
SLEEVES_PACKS_ROW = 20      # "full packs owned" -> the inventory to import
SLEEVES_PRODUCT_ROW = 40    # Tlama product names ("Diamond Yellow", ...)
SLEEVES_URL_ROW = 41        # tlamagames.com product URLs
SLEEVES_FIRST_SIZE_COL = 3  # column B is a TOTAL column — skip it

# "41x63", "57.5x89", "70x120-clear" (suffix tolerated), "76x101,5".
DIM_RE = re.compile(r"^(\d+(?:[.,]\d+)?)\s*[x×]\s*(\d+(?:[.,]\d+)?)(?:-\w+)?$", re.IGNORECASE)

BRAND_CODE_MAP = {"tl": "Tlama", "at": "AT", "up": "UP", "upq": "UPQ"}
# Tlama Diamond colour shorthand seen in stray cells (DR = Diamond Red, ...).
DIAMOND_CODE_RE = re.compile(r"^d[a-z]$", re.IGNORECASE)

DEFERRED_NOTES = [
    "BGG card-size scrape pre-fill (§5 'fetch from BGG' button): blocked with "
    "the rest of the §8 BGG work (API token pending); requirements are "
    "manual/import-only.",
    "(Pre)orders sheet and the 'TOTAL incl. preorders' rows: imported by "
    "import_preorders (§6); sleeve_shortfall(include_preorders=True) is the "
    "toggle.",
    "Sleeves sheet rows 4-19 and 21-38 (per-brand sleeved totals, to-sleeve "
    "counts, extra-sleeve/price math): derived aggregates of the Overview "
    "columns — cross-checks, not source data.",
    "Loose-sleeve leftovers: the sheet tracks whole packs only, so "
    "SleeveInventory.loose stays 0.",
]

DECISION_NOTES = [
    "Codes tl/at/up/upq -> status 'sleeved' with that brand's product; "
    "D<letter> Diamond colour codes -> sleeved with the Tlama product.",
    "Code 'c' -> 'sleeved' with product unknown (the Sleeves sheet's own "
    "aggregates count 'c' under 'sleeved').",
    "Code '?' -> 'to-sleeve'.",
    "A card count with no code -> 'not-sleeved' (no status recorded in the "
    "sheet).",
]


def parse_dimensions(raw):
    """Parse a 'WxH' text into (Decimal, Decimal), or None. Tolerates decimal
    commas and a variant suffix ('70x120-clear' -> 70x120)."""
    match = DIM_RE.match(raw.strip())
    if not match:
        return None
    try:
        return tuple(
            Decimal(part.replace(",", ".")).quantize(Decimal("0.1"))
            for part in match.groups()
        )
    except InvalidOperation:
        return None


def as_count(value):
    """Coerce a count cell to a positive int, or None (floats like 365.0 are
    ints; strings and non-positives are noise)."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        count = int(value)
        return count if count > 0 and count == value else None
    return None


class Command(BaseCommand):
    help = "Import sleeve requirements, inventory and products from the Mastersheet (DESIGN §5)."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the Mastersheet .xlsx file.")
        parser.add_argument(
            "--user", required=True,
            help="Username whose Copies get statuses and who owns the inventory.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Parse and report, but write nothing to the database.",
        )

    def handle(self, *args, **options):
        try:
            workbook = openpyxl.load_workbook(options["path"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {options['path']}")
        for sheet_name in ("Sleeves", "Overview"):
            if sheet_name not in workbook.sheetnames:
                raise CommandError(f'The workbook has no "{sheet_name}" sheet.')

        User = get_user_model()
        try:
            user = User.objects.get(username=options["user"])
        except User.DoesNotExist:
            raise CommandError(f"User {options['user']!r} does not exist.")

        self.counts = {}
        self.skipped = []    # (row, title, reason) — nothing imported
        self.ambiguous = []  # (where, title, note) — imported with a caveat
        self._sizes = {}     # (width, height) Decimal pair -> CardSize
        self._products = {}  # (brand, card_size pk) -> SleeveProduct

        with transaction.atomic():
            # Seed the canonical Overview size names first so the Sleeves
            # sheet's Tlama colour labels (Azur, Black, Pink...) become
            # aliases, not primary names.
            for _, name, width, height in OVERVIEW_SIZES:
                self._get_or_create_size(self._quantize(width, height), name)
            self._import_size_catalog(workbook["Sleeves"], user)
            overview = workbook["Overview"]
            for row in range(DATA_START_ROW, overview.max_row + 1):
                self._import_overview_row(overview, row, user)
            if options["dry_run"]:
                transaction.set_rollback(True)

        self._print_report(dry_run=options["dry_run"])

    def _bump(self, key, created=None):
        if created is not None:
            key = f"{key} {'created' if created else 'updated'}"
        self.counts[key] = self.counts.get(key, 0) + 1

    @staticmethod
    def _quantize(width, height):
        return (Decimal(width).quantize(Decimal("0.1")),
                Decimal(height).quantize(Decimal("0.1")))

    # --- Sleeves sheet: size catalog, Tlama products, pack inventory ---------

    def _import_size_catalog(self, sheet, user):
        for col in range(SLEEVES_FIRST_SIZE_COL, sheet.max_column + 1):
            dim_raw = cell_text(sheet.cell(row=SLEEVES_DIM_ROW, column=col))
            if not dim_raw or dim_raw == "num":  # blank spacer / the Other column
                continue
            dims = parse_dimensions(dim_raw)
            if not dims:
                self.ambiguous.append(
                    ("Sleeves sheet", f"column {col}", f"unparseable size {dim_raw!r}"),
                )
                continue

            name = cell_text(sheet.cell(row=SLEEVES_NAME_ROW, column=col))
            if name in ("#REF!", "Other"):  # broken/aggregate headers carry no name
                name = ""
            size = self._get_or_create_size(dims, name)

            product = None
            product_name = cell_text(sheet.cell(row=SLEEVES_PRODUCT_ROW, column=col))
            if product_name:
                url_cell = sheet.cell(row=SLEEVES_URL_ROW, column=col)
                product, created = SleeveProduct.objects.update_or_create(
                    brand="Tlama", name=product_name, card_size=size,
                    defaults={"url": cell_link(url_cell) or ""},
                )
                self._products[("Tlama", size.pk)] = product
                self._bump("sleeve products", created=created)

            packs = as_count(sheet.cell(row=SLEEVES_PACKS_ROW, column=col).value)
            if packs:
                if not product:
                    self.ambiguous.append((
                        "Sleeves sheet", f"column {col}",
                        f"{packs} packs owned for {size} but no product listed — inventory skipped",
                    ))
                    continue
                _, created = SleeveInventory.objects.update_or_create(
                    owner=user, product=product, defaults={"packs": packs},
                )
                self._bump("inventory rows", created=created)

    def _get_or_create_size(self, dims, name):
        """Upsert a CardSize by its (width, height) natural key. A different
        display name for known dimensions becomes an alias (DESIGN §5:
        Teal vs Azur, Extra large vs Bronze)."""
        if dims not in self._sizes:
            size, created = CardSize.objects.get_or_create(
                width_mm=dims[0], height_mm=dims[1], defaults={"name": name},
            )
            if created:
                self._bump("card sizes created")
            self._sizes[dims] = size
        size = self._sizes[dims]
        if name and not size.name:
            size.name = name
            size.save(update_fields=["name"])
        elif name and name != size.name and name not in size.alias_list:
            size.aliases = ", ".join(size.alias_list + [name])
            size.save(update_fields=["aliases"])
            self._bump("card size aliases added")
        return size

    # --- Overview sheet: requirements + per-copy sleeved status --------------

    def _import_overview_row(self, sheet, row, user):
        def cell(col):
            return sheet.cell(row=row, column=col)

        title = cell_text(cell(COL_TITLE))
        bgg_url = cell_link(cell(COL_BGG))
        if not title and not bgg_url:
            return  # entirely blank row
        bgg_id = extract_bgg_id(bgg_url)

        # Does the row carry any sleeve data at all? (count OR code OR other-size)
        has_data = any(
            cell(col).value is not None or cell_text(cell(col + 1))
            for col, *_ in OVERVIEW_SIZES
        ) or any(
            cell(num_col).value is not None or cell_text(cell(size_col))
            for num_col, size_col in OVERVIEW_OTHER_PAIRS
        )
        if not has_data:
            return

        if not bgg_id:
            self.skipped.append((row, title, "no BGG link — cannot establish identity"))
            return
        link = BggLink.objects.filter(bgg_id=bgg_id, is_primary=True).select_related("game").first()
        if not link:
            self.skipped.append(
                (row, title, f"BGG id {bgg_id} not in database — run import_mastersheet first"),
            )
            return
        game = link.game

        copies = list(Copy.objects.filter(owner=user, edition__game=game).select_related("edition"))
        if not copies:
            self.skipped.append((row, title, f"{user.username} owns no Copy of this game"))
            return
        copy = copies[0]
        if len(copies) > 1:
            self.ambiguous.append(
                (row, title, f"{len(copies)} copies owned — requirements/status put on {copy.edition}"),
            )

        # A size can appear twice in one row (a fixed column AND an "Other"
        # pair, e.g. 60×63x88 'tl' + 1280×"63x88" on Arydia) — those are
        # separate batches of the same size, so collect (count, code) batches
        # per size first and write each (edition, size) exactly once.
        batches = {}  # CardSize -> [(count, code), ...]

        for count_col, name, width, height in OVERVIEW_SIZES:
            count = as_count(cell(count_col).value)
            code = cell_text(cell(count_col + 1))
            if count is None:
                if code:
                    self.ambiguous.append(
                        (row, title, f"code {code!r} for {name} without a card count — pair skipped"),
                    )
                elif cell(count_col).value is not None:
                    self.ambiguous.append(
                        (row, title, f"non-numeric count {cell(count_col).value!r} for {name} — pair skipped"),
                    )
                continue
            size = self._get_or_create_size(self._quantize(width, height), name)
            batches.setdefault(size, []).append((count, code))

        for num_col, size_col in OVERVIEW_OTHER_PAIRS:
            raw_count = cell(num_col).value
            size_raw = cell_text(cell(size_col))
            if raw_count is None and not size_raw:
                continue
            dims = parse_dimensions(size_raw) if size_raw else None
            count = as_count(raw_count)
            if not dims or count is None:
                self.ambiguous.append(
                    (row, title, f"Other pair not importable (num={raw_count!r}, size={size_raw!r})"),
                )
                continue
            size = self._get_or_create_size(dims, "")
            # The Other columns carry no brand/status code.
            batches.setdefault(size, []).append((count, ""))

        for size, size_batches in batches.items():
            self._import_requirement(copy, size, size_batches, row, title)

    def _import_requirement(self, copy, size, batches, row, title):
        total = sum(count for count, _ in batches)
        requirement, created = SleeveRequirement.objects.update_or_create(
            edition=copy.edition, card_size=size, defaults={"count": total},
        )
        self._bump("requirements", created=created)

        # One status per (copy, size): take it from the largest batch and
        # flag rows whose batches disagree (e.g. partly sleeved, partly not).
        count, code = max(batches, key=lambda batch: batch[0])
        status, product, note = self._map_code(code, size)
        if note:
            self.ambiguous.append((row, title, note))
        if len(batches) > 1:
            detail = " + ".join(
                f"{c}× {code_ or 'no code'}" for c, code_ in batches
            )
            self.ambiguous.append((
                row, title,
                f"{size} appears in several batches ({detail}) — counts summed, "
                f"status '{status}' taken from the largest batch",
            ))
        _, created = CopySleeveStatus.objects.update_or_create(
            copy=copy, requirement=requirement,
            defaults={"status": status, "product": product},
        )
        self._bump("sleeve statuses", created=created)

    def _map_code(self, code, size):
        """Interpret a brand/status code cell -> (status, product, caveat).
        The decisions are documented in the module docstring / report."""
        Status = CopySleeveStatus.Status
        lowered = code.lower()
        if not code:
            return Status.NOT_SLEEVED, None, None
        if lowered == "?":
            return Status.TO_SLEEVE, None, None
        if lowered == "c":
            return Status.SLEEVED, None, None
        if lowered in BRAND_CODE_MAP:
            return Status.SLEEVED, self._product_for_brand(BRAND_CODE_MAP[lowered], size), None
        if DIAMOND_CODE_RE.match(code):
            product = self._product_for_brand("Tlama", size)
            return Status.SLEEVED, product, f"Diamond colour code {code!r} -> sleeved with {product}"
        return Status.NOT_SLEEVED, None, f"unrecognised sleeve code {code!r} for {size} — treated as no code"

    def _product_for_brand(self, brand, size):
        """A brand's product for a size — the catalog one if the Sleeves sheet
        seeded it (Tlama), else a bare get_or_create placeholder."""
        key = (brand, size.pk)
        if key not in self._products:
            product = SleeveProduct.objects.filter(brand=brand, card_size=size).first()
            if not product:
                product, _ = SleeveProduct.objects.get_or_create(
                    brand=brand, name="", card_size=size,
                )
                self._bump("sleeve products created")
            self._products[key] = product
        return self._products[key]

    # --- report ---------------------------------------------------------------

    def _print_report(self, dry_run):
        write = self.stdout.write
        if dry_run:
            write(self.style.WARNING("DRY RUN — nothing was written to the database.\n"))

        write(self.style.MIGRATE_HEADING("Summary"))
        for key in sorted(self.counts):
            write(f"  {key}: {self.counts[key]}")

        write(self.style.MIGRATE_HEADING("Skipped rows"))
        if self.skipped:
            for row, title, reason in self.skipped:
                write(f"  row {row} ({title!r}): {reason}")
        else:
            write("  none")

        write(self.style.MIGRATE_HEADING("Ambiguous (imported with caveats)"))
        if self.ambiguous:
            for where, title, note in self.ambiguous:
                write(f"  {where} ({title!r}): {note}")
        else:
            write("  none")

        write(self.style.MIGRATE_HEADING("Interpretation decisions"))
        for note in DECISION_NOTES:
            write(f"  - {note}")

        write(self.style.MIGRATE_HEADING("Deferred (not imported)"))
        for note in DEFERRED_NOTES:
            write(f"  - {note}")

        write(self.style.SUCCESS("Done." if not dry_run else "Dry run complete."))

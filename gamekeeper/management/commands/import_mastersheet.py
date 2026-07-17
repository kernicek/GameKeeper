"""One-time, re-runnable import of the Boardgame Mastersheet (DESIGN §12).

Parses the "Overview" sheet and creates Games / Editions / Copies plus the
personal & curation data the sheet carries. Idempotent: Games are upserted by
their primary BGG id, Editions/Copies by natural keys, so re-running does not
duplicate. BGG-synced stats/images are intentionally left blank — the §8 sync
engine backfills them from the primary BggLink.

Sheet layout (validated against the real file):
  row 1 = header labels, row 2 = aggregate stats, row 3 = column letters,
  data starts at row 4. The BGG/CF/Drive columns keep their payload in the
  cell hyperlink, not the cell value.
"""

import re

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

import openpyxl

from gamekeeper.bgg import extract_bgg_id
from gamekeeper.models import (
    BggLink,
    Copy,
    Edition,
    ExternalLink,
    Game,
    Group,
    Location,
    Membership,
)

DATA_START_ROW = 4

# 1-based column numbers in the Overview sheet.
COL_TITLE = 2
COL_BGG = 3
COL_CF = 4
COL_DRIVE = 5
COL_EDITION = 14
COL_PNP = 15
COL_CHYNICE = 16
COL_SIZE = 20
COL_NUM_BOXES = 21
COL_KEEP_STATUS = 216
# Col 217 ("do I want to play it?") is skipped: excitement carries the same
# signal, so the field was dropped from Copy (decision 2026-07).
COL_EXCITEMENT = 218  # the 0–10 column; col 219 is a derived 1–5 duplicate
COL_IMMUNE = 221
COL_FAVOURITE = 222
COL_BRINGS_EXTRA = 223
COL_WHY_LEAVE = 224
# col 225 (1-in-1-out) dropped with issue #29 — the field no longer exists.
COL_PLAY_UNTIL = 231
COL_INSERT_3D = 233
COL_CARD_DIVIDERS = 235
COL_ACC_3D = 237
COL_OTHER_ACC = 239

PNP_SUFFIX_RE = re.compile(r"\s*\(PnP\)\s*$", re.IGNORECASE)

# "M (Medium)" etc. — the leading letter carries the category.
SIZE_MAP = {
    "T": Edition.SizeCategory.TINY,
    "S": Edition.SizeCategory.SMALL,
    "M": Edition.SizeCategory.MEDIUM,
    "N": Edition.SizeCategory.NORMAL,
    "L": Edition.SizeCategory.LARGE,
    "H": Edition.SizeCategory.HUGE,
}

# Keep status cells look like "2.2 - Keep (PnP)"; the integer part of the
# leading number selects the bucket.
KEEP_MAP = {
    1: Copy.KeepStatus.ALWAYS_KEEP,
    2: Copy.KeepStatus.KEEP,
    3: Copy.KeepStatus.UNDECIDED,
    4: Copy.KeepStatus.MIGHT_CYCLE,
    5: Copy.KeepStatus.WILL_LEAVE,
}

# Leading keyword of an upgrade cell -> status. Anything after the keyword
# (or a wholly unmatched cell) is free-text detail kept in upgrades_note.
UPGRADE_PREFIX_MAP = [
    ("not necessary", Copy.UpgradeStatus.NOT_NECESSARY),
    ("included", Copy.UpgradeStatus.INCLUDED),
    ("done", Copy.UpgradeStatus.DONE),
    ("gametrayz", Copy.UpgradeStatus.DONE),  # manufacturer insert == solved
    ("to-do", Copy.UpgradeStatus.TODO),
    ("todo", Copy.UpgradeStatus.TODO),
    ("partly", Copy.UpgradeStatus.TODO),
    ("maybe", Copy.UpgradeStatus.MAYBE),
    ("possib", Copy.UpgradeStatus.MAYBE),  # "Possibly" / "Possible, but ..."
    ("probably", Copy.UpgradeStatus.MAYBE),
    ("???", Copy.UpgradeStatus.MAYBE),
]

# Cells that are exactly a bare status carry no detail worth keeping.
UPGRADE_BARE_VALUES = {"not necessary", "included", "done", "to-do", "todo", "maybe"}

EXTERNAL_LINK_DOMAIN_MAP = [
    ("kickstarter.com", ExternalLink.LinkType.KICKSTARTER),
    ("gamefound.com", ExternalLink.LinkType.GAMEFOUND),
    ("drive.google.com", ExternalLink.LinkType.GOOGLE_DRIVE),
    ("dropbox.com", ExternalLink.LinkType.DROPBOX),
    ("zatrolene-hry.cz", ExternalLink.LinkType.ZATROLENE),
]

# Columns the sheet has but the schema doesn't (yet) — reported, not imported.
DEFERRED_NOTES = [
    "Taxonomy & language columns (themes/game-type, Lang(components), "
    "difficulty-for-non-speakers, Lang(how-much)): imported by import_taxonomy "
    "(§10).",
    "Sleeves sheet and (Pre)orders sheet: §5/§6 models not built yet.",
    "Plays, Minis, Stats/Utils sheets: out of scope per DESIGN §12.",
    "'Rulebook downloaded' column: documents (§7) not built yet.",
    "Expansions column is a count, not identities — expansion Games and "
    "expands-relations come from the §8 BGG sync, not this import.",
    "Players / optimal players / length columns: BGG-synced Game stats are left "
    "blank and backfilled by the §8 sync engine.",
]


def cell_link(cell):
    """Return the URL carried by a cell: hyperlink target, else an http value."""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    if isinstance(cell.value, str) and cell.value.startswith("http"):
        return cell.value
    return None


def cell_text(cell):
    return str(cell.value).strip() if cell.value is not None else ""


def link_type_for_url(url):
    for domain, link_type in EXTERNAL_LINK_DOMAIN_MAP:
        if domain in url:
            return link_type
    return ExternalLink.LinkType.OTHER


def map_upgrade(raw):
    """Map an upgrade cell to (status, detail). Unmatched -> (None, raw)."""
    if not raw:
        return Copy.UpgradeStatus.NONE, ""
    lowered = raw.lower()
    for prefix, status in UPGRADE_PREFIX_MAP:
        if lowered.startswith(prefix):
            detail = "" if lowered in UPGRADE_BARE_VALUES else raw
            return status, detail
    return None, raw


class Command(BaseCommand):
    help = 'Import the Boardgame Mastersheet "Overview" sheet (DESIGN §12).'

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the Mastersheet .xlsx file.")
        parser.add_argument(
            "--user", required=True,
            help="Username that will own all imported Copies.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Parse and report, but write nothing to the database.",
        )

    def handle(self, *args, **options):
        try:
            workbook = openpyxl.load_workbook(options["path"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {options['path']}")
        if "Overview" not in workbook.sheetnames:
            raise CommandError('The workbook has no "Overview" sheet.')
        sheet = workbook["Overview"]

        User = get_user_model()
        try:
            user = User.objects.get(username=options["user"])
        except User.DoesNotExist:
            raise CommandError(f"User {options['user']!r} does not exist.")

        self.counts = {}
        self.skipped = []    # (row, title, reason) — nothing imported
        self.ambiguous = []  # (row, title, note) — imported with a caveat
        self._locations = {}

        with transaction.atomic():
            group = self._get_or_create_group(user)
            for row in range(DATA_START_ROW, sheet.max_row + 1):
                self._import_row(sheet, row, user, group)
            if options["dry_run"]:
                transaction.set_rollback(True)

        self._print_report(dry_run=options["dry_run"])

    # --- setup -------------------------------------------------------------

    def _get_or_create_group(self, user):
        membership = Membership.objects.filter(user=user).first()
        if membership:
            return membership.group
        # The signup auto-group signal (DESIGN §3, signals.py) normally makes
        # this on user creation; keep the fallback for users predating it.
        group, _ = Group.objects.get_or_create(
            slug=slugify(user.username), defaults={"name": user.username},
        )
        Membership.objects.create(user=user, group=group, role=Membership.Role.OWNER)
        self._bump("groups created")
        return group

    def _get_or_create_location(self, group, name):
        if name not in self._locations:
            location, created = Location.objects.get_or_create(group=group, name=name)
            self._locations[name] = location
            if created:
                self._bump("locations created")
        return self._locations[name]

    def _bump(self, key, created=None):
        if created is not None:
            key = f"{key} {'created' if created else 'updated'}"
        self.counts[key] = self.counts.get(key, 0) + 1

    # --- row import ----------------------------------------------------------

    def _import_row(self, sheet, row, user, group):
        def cell(col):
            return sheet.cell(row=row, column=col)

        title = cell_text(cell(COL_TITLE))
        bgg_url = cell_link(cell(COL_BGG))
        if not title and not bgg_url:
            return  # entirely blank row
        bgg_id = extract_bgg_id(bgg_url)
        if not bgg_id:
            self.skipped.append((row, title, "no BGG link — cannot establish identity"))
            return

        # --- Game (upsert by primary BGG id) ---
        name = PNP_SUFFIX_RE.sub("", title).strip()
        is_pnp = bool(cell_text(cell(COL_PNP)))
        if not is_pnp and PNP_SUFFIX_RE.search(title):
            # One real row has "(PnP)" in the title but a clobbered PnP cell.
            is_pnp = True
            self.ambiguous.append((row, name, "PnP column empty; inferred from title"))

        link = BggLink.objects.filter(bgg_id=bgg_id, is_primary=True).select_related("game").first()
        if link:
            game = link.game
            game.name = name
            game.save(update_fields=["name", "updated_at"])
            self._bump("games", created=False)
        else:
            game = Game.objects.create(name=name)
            BggLink.objects.create(game=game, bgg_id=bgg_id, is_primary=True)
            self._bump("games", created=True)

        # --- External links (CF / Drive / a URL pasted into the Edition cell) ---
        cf_url = cell_link(cell(COL_CF))
        if cf_url:
            self._add_external_link(game, cf_url)
        drive_url = cell_link(cell(COL_DRIVE))
        if drive_url:
            self._add_external_link(game, drive_url)

        # --- Edition ---
        edition_url = cell_link(cell(COL_EDITION))
        edition_name = "" if edition_url else cell_text(cell(COL_EDITION))
        if edition_url:
            # Overloaded Edition cell: a pasted URL is a link, not an edition name.
            self._add_external_link(game, edition_url, label="from Edition column")
            self.ambiguous.append((row, name, "Edition cell held a URL — routed to an ExternalLink, edition left default"))
        if edition_name == "PnP":
            edition_name = ""  # duplicate of the PnP flag, not an edition name

        location_notes = []
        size_raw = cell_text(cell(COL_SIZE))
        size_category = ""
        if size_raw:
            size_category = SIZE_MAP.get(size_raw[0].upper(), "")
            if not size_category:
                if "in other box" in size_raw:
                    location_notes.append("in other box")
                else:
                    self.ambiguous.append((row, name, f"unrecognised size {size_raw!r}"))

        num_boxes = cell(COL_NUM_BOXES).value
        edition, created = Edition.objects.update_or_create(
            game=game, name=edition_name,
            defaults={
                "is_default": edition_name == "",
                "is_pnp": is_pnp,
                "size_category": size_category,
                "num_boxes": int(num_boxes) if num_boxes else None,
            },
        )
        self._bump("editions", created=created)

        # --- Copy fields ---
        keep_raw = cell_text(cell(COL_KEEP_STATUS))
        keep_status = ""
        if keep_raw:
            match = re.match(r"(\d+)", keep_raw)
            keep_status = KEEP_MAP.get(int(match.group(1))) if match else None
            if not keep_status:
                keep_status = ""
                self.ambiguous.append((row, name, f"unrecognised keep status {keep_raw!r}"))

        upgrade_notes = []
        upgrade_statuses = {}
        for field, col, label in (
            ("insert_3d", COL_INSERT_3D, "3D insert"),
            ("card_dividers", COL_CARD_DIVIDERS, "card dividers"),
            ("accessories_3d", COL_ACC_3D, "3D accessories"),
            ("other_accessories", COL_OTHER_ACC, "other accessories"),
        ):
            raw = cell_text(cell(col))
            status, detail = map_upgrade(raw)
            if status is None:
                status = Copy.UpgradeStatus.NONE
                self.ambiguous.append((row, name, f"unrecognised {label} value {raw!r}"))
            upgrade_statuses[field] = status
            if detail:
                upgrade_notes.append(f"{label}: {detail}")

        location = None
        location_raw = cell_text(cell(COL_CHYNICE))
        if location_raw == "y":
            location = self._get_or_create_location(group, "Chynice")
        elif location_raw == "kancl":
            location = self._get_or_create_location(group, "Office")
        elif location_raw and location_raw != "n":
            if location_raw.startswith("y"):
                location = self._get_or_create_location(group, "Chynice")
            location_notes.append(location_raw)
            self.ambiguous.append((row, name, f"unmapped location value {location_raw!r} kept in location_note"))

        excitement = cell(COL_EXCITEMENT).value

        _, created = Copy.objects.update_or_create(
            owner=user, edition=edition,
            defaults={
                "excitement": excitement,
                "keep_status": keep_status,
                "immune": bool(cell_text(cell(COL_IMMUNE))),
                "play_until_or_leaves": cell_text(cell(COL_PLAY_UNTIL))[:300],
                "favourite_thing": cell_text(cell(COL_FAVOURITE)),
                "brings_extra": cell_text(cell(COL_BRINGS_EXTRA)),
                "why_might_leave": cell_text(cell(COL_WHY_LEAVE)),
                "upgrades_note": "\n".join(upgrade_notes),
                "location": location,
                "location_note": "; ".join(location_notes)[:300],
                **upgrade_statuses,
            },
        )
        self._bump("copies", created=created)

    def _add_external_link(self, game, url, label=""):
        _, created = ExternalLink.objects.get_or_create(
            game=game, link_type=link_type_for_url(url), url=url,
            defaults={"label": label},
        )
        if created:
            self._bump("external links created")

    # --- report ------------------------------------------------------------

    def _print_report(self, dry_run):
        write = self.stdout.write
        if dry_run:
            write(self.style.WARNING("DRY RUN — nothing was written to the database.\n"))

        write(self.style.MIGRATE_HEADING("Summary"))
        for key in sorted(self.counts):
            write(f"  {key}: {self.counts[key]}")

        write(self.style.MIGRATE_HEADING("Skipped rows"))
        if self.skipped:
            for row, title, reason in self.skipped:
                write(f"  row {row} ({title!r}): {reason}")
        else:
            write("  none")

        write(self.style.MIGRATE_HEADING("Ambiguous rows (imported with caveats)"))
        if self.ambiguous:
            for row, title, note in self.ambiguous:
                write(f"  row {row} ({title!r}): {note}")
        else:
            write("  none")

        write(self.style.MIGRATE_HEADING("Deferred (not imported — schema not built yet)"))
        for note in DEFERRED_NOTES:
            write(f"  - {note}")

        write(self.style.SUCCESS("Done." if not dry_run else "Dry run complete."))

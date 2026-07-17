"""Re-runnable import of the Mastersheet's "(Pre)orders" sheet (DESIGN §6 + §12).

Creates the Purchase -> Wave -> Product hierarchy plus per-Product sleeve
needs (the §5 include-preorders shortfall toggle feeds off those).

Sheet layout (validated against the real file):
  row 1 = header labels, row 2 = aggregate stats, row 3 = column letters,
  data starts at row 4. One row = one product; the Purchase column (A) groups
  rows into campaigns, with an optional "(Wave N)" suffix splitting a
  campaign into shipment waves — no suffix means the auto "Wave 1"
  (DESIGN §6). BGG/campaign/PM cells keep their payload in the hyperlink.

Purchase-level fields (platform, links, PM, dates, excitement, comments)
repeat on every row of a group; the first non-empty value wins and
disagreements are reported as ambiguous.

Game matching mirrors the other importers: via the BGG hyperlink (column C),
never by title. Game-kind rows whose BGG id is unknown CREATE a Game +
primary BggLink — but never a Copy: the row only proves the game was
ordered, ownership comes from the Overview import / §8 BGG sync. Arrived
game products are linked to the owner's Copy of the matched Game when there
is exactly one (the §6 product->Copy conversion, backfilled).

Sleeve-need columns AF..AQ are mapped to CardSizes by the Sleeves sheet's
own meta rows ("(Pre)orders-end", row 37), NOT by their header labels — two
headers are mislabeled (AH says 44x68 but aggregates into 43x66 "Small",
AI says 56x89 but aggregates into 57.5x89 "Medium"). AO (70x70) is absent
from the sheet's formula map and is taken at header value. AR/AS and AT/AU
are free-text "Other" pairs like the Overview sheet's.
"""

import re
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from gamekeeper.models import (
    BggLink,
    CardSize,
    Copy,
    Edition,
    Game,
    PledgeManager,
    Product,
    ProductSleeveRequirement,
    Purchase,
    Wave,
)
from gamekeeper.bgg import extract_bgg_id
from gamekeeper.management.commands.import_mastersheet import (
    cell_link,
    cell_text,
)
from gamekeeper.management.commands.import_sleeves import (
    as_count,
    parse_dimensions,
)

SHEET_NAME = "(Pre)orders"
DATA_START_ROW = 4

# 1-based column numbers in the (Pre)orders sheet.
COL_PURCHASE = 1
COL_NAME = 2
COL_BGG = 3
COL_CAMPAIGN = 4
COL_DRIVE = 5
# 6 "Game name" is a display/grouping label — matching by title is forbidden.
COL_PLATFORM = 7
COL_KIND = 8
COL_STATUS = 9
COL_ORDERED = 10
COL_ORIG_ETA = 11
COL_ARRIVAL = 12
# 13-19 (duration/delay/arrives-in) are derived in-sheet; delay is computed
# from the dates (Wave.delay_days), not imported.
COL_PM = 20
COL_PM_STATUS = 21
COL_PM_CLOSE = 22
COL_ADDRESS = 24
COL_COMMENTS = 25
COL_EXCITEMENT = 26
COL_MINIATURES = 27
COL_INSERT_3D = 28
COL_CONTAINS_CARDS = 29
COL_NEEDS_SLEEVES = 30
COL_FITS_SLEEVED = 31

# Sleeve-need count columns -> canonical (name, W, H), per the Sleeves
# sheet's meta map (module docstring). Names match the existing CardSize
# catalog so no new sizes are created for the fixed columns.
PREORDER_SIZES = [
    (32, "Mini", "41", "63"),
    (33, "Small*", "44", "63"),
    (34, "Small", "43", "66"),          # header mislabeled "44x68"
    (35, "Medium", "57.5", "89"),       # header mislabeled "56x89"
    (36, "Large", "59", "91"),
    (37, "Standard", "63", "88"),
    (38, "Extra large", "65", "100"),
    (39, "Tarot*", "70", "110"),
    (40, "Tarot", "70", "120"),
    (41, "Black", "70", "70"),
    (42, "Square", "80", "80"),
    (43, "Oversized", "80", "120"),
]
# Free-text size pairs: (count column, WxH-text column).
OTHER_PAIRS = [(44, 45), (46, 47)]

WAVE_SUFFIX_RE = re.compile(r"\s*\(Wave\s*(\d+)\)\s*$", re.IGNORECASE)
# Issue #32: the "KS"/"GF"/"BK" marker is redundant with the platform column
# (G), so it is stripped from the stored Purchase name — display derives the
# platform from Purchase.platform, not the title.
PLATFORM_SUFFIX_RE = re.compile(r"\s+(KS|GF|BK)$")
PNP_NAME_RE = re.compile(r"\s*[-–]?\s*\(?PnP\)?\s*$", re.IGNORECASE)
MULTIPLIER_RE = re.compile(r"\s*\(x\d+\)\s*$", re.IGNORECASE)
DIGITAL_NAME_RE = re.compile(r"\(digital\)\s*$", re.IGNORECASE)

PLATFORM_MAP = {
    "KS": Purchase.Platform.KICKSTARTER,
    "GF": Purchase.Platform.GAMEFOUND,
    "BK": Purchase.Platform.BACKERKIT,
    "Other": Purchase.Platform.OTHER,
}

# Column H, with the sheet's "-show"/"-hide" display suffixes stripped.
KIND_MAP = {
    "Board Game(s)": Product.Kind.GAME,
    "Game+Expansion(s)": Product.Kind.GAME_AND_EXPANSIONS,
    "Expansion(s)": Product.Kind.EXPANSION,
    "Game(s)-PnP": Product.Kind.PNP_GAME,
    "Gamebook": Product.Kind.GAMEBOOK,
    "Accessories": Product.Kind.ACCESSORY,
    "Promo(s)": Product.Kind.PROMO,
    "Book(s)": Product.Kind.BOOK,
    "Puzzle(s)": Product.Kind.PUZZLE,
    "$1 Pledge": Product.Kind.PLACEHOLDER_PLEDGE,
    "$1 Pledge-just support": Product.Kind.PLACEHOLDER_PLEDGE,
    "$1 Pledge-to upgrade": Product.Kind.PLACEHOLDER_PLEDGE,
}

STATUS_MAP = {
    "Arrived": Wave.Status.ARRIVED,
    "Pre-production": Wave.Status.PRE_PRODUCTION,
    "Production": Wave.Status.PRODUCTION,
    "Failed": Wave.Status.NEVER_ARRIVED,
    # "Just support" is a purchase-lifecycle signal (placeholder), not a
    # shipment state; the wave stays pending.
    "Just support": Wave.Status.PENDING,
}
# Least-progressed first: a wave with mixed row statuses is only as far
# along as its slowest product.
STATUS_PRIORITY = [
    Wave.Status.PENDING,
    Wave.Status.PRE_PRODUCTION,
    Wave.Status.PRODUCTION,
    Wave.Status.FULFILMENT,
    Wave.Status.NEVER_ARRIVED,
    Wave.Status.CANCELLED,
    Wave.Status.ARRIVED,
]

PM_STATUS_MAP = {
    "Filled out": Purchase.PledgeManagerStatus.FILLED_OUT,
    "Not necessary": Purchase.PledgeManagerStatus.NOT_NECESSARY,
    "Won't fill out": Purchase.PledgeManagerStatus.WONT_FILL_OUT,
    "Not yet": Purchase.PledgeManagerStatus.NOT_YET,
    "Sent out": Purchase.PledgeManagerStatus.SENT_OUT,
}

# Issue #159/#181: pledge_manager is a fixed choice (now a PledgeManager FK),
# not free text. "JetBacker" has no shared default link and folds into Other.
PM_MAP = {
    "Gamefound": "Gamefound",
    "CrowdOx": "CrowdOx",
    "BackerKit": "BackerKit",
    "PledgeManager": "PledgeManager",
    "Pledg.it": "Pledg.it",
    "PledgeBox": "PledgeBox",
    "Kickstarter": "Kickstarter",
    "other": "Other",
    "N/A": "",
}

TRISTATE_MAP = {
    "Yes": Product.TriState.YES,
    "No": Product.TriState.NO,
    "?": Product.TriState.UNKNOWN,
}

DECISION_NOTES = [
    "Waves come from the '(Wave N)' suffix in the Purchase column; without a "
    "suffix everything is the auto 'Wave 1' (DESIGN §6).",
    "The manual 'KS'/'GF'/'BK' name suffix is stripped (issue #32) — the "
    "platform lives in column G (Purchase.platform) and is shown separately.",
    "Purchase status: placeholder when every row of the campaign is a $1 "
    "pledge / 'Just support'; committed otherwise. 'Failed' rows -> wave "
    "never-arrived.",
    "A wave is digital when ALL its products are PnP games or '(Digital)' "
    "items; mixed waves stay physical (splitting out §6 digital waves is a "
    "manual cleanup).",
    "Game-kind rows (games/expansions/PnP) with an unknown BGG id create "
    "Game+BggLink but never a Copy — ownership comes from the Overview "
    "import / §8 sync. Created Games keep the sheet's product name (minus "
    "PnP/'(xN)' suffixes) until BGG sync corrects it.",
    "Gamebooks, accessories, promos, books and puzzles are NOT Games even "
    "when they carry a BGG link; the link is kept on Product.bgg_url.",
    "Arrived game products are linked to the owner's Copy of the matched "
    "Game when exactly one exists (the product->Copy conversion, "
    "backfilled). Rows that reuse a base game's BGG id (upgrade packs) "
    "therefore link to the base game's Copy.",
    "Sleeve columns are mapped per the Sleeves sheet's formula map, not "
    "their headers: AH -> 43x66 'Small' (header says 44x68), AI -> "
    "57.5x89 'Medium' (header says 56x89). AO (70x70) is not in the "
    "formula map and is taken at header value.",
    "Pledge manager 'N/A' -> blank (none needed).",
    "Numeric excitement goes to Purchase.excitement, prose to "
    "excitement_note (the sheet mixes both in one column).",
]

DEFERRED_NOTES = [
    "campaign_end_date / 'watching' lifecycle: the sheet has no watched-but-"
    "unbacked rows and no campaign end dates — starts empty (§11 reminders "
    "need it filled in-app).",
    "Money / per-product cost: deferred by DESIGN §6.",
    "Drive links are kept raw on Product.drive_url; Document records are §7.",
    "Tracking links: the sheet carries none; Wave.tracking_url starts blank.",
    "Duration/delay/arrives-in columns (M-S): derived from the dates "
    "(Wave.delay_days), not imported.",
    "BGG stats/names for created Games: backfilled by the §8 sync engine.",
    "Editions for game products: unknown to the sheet; Product.edition "
    "starts empty — except PnP products, which get a default PnP-flagged "
    "Edition (#138) the product links to.",
]


def as_date(value):
    """Coerce an openpyxl cell value to a date, or None (the date columns are
    real datetimes in the sheet; anything else is noise)."""
    return value.date() if hasattr(value, "date") else None


class Command(BaseCommand):
    help = 'Import the Mastersheet "(Pre)orders" sheet into Purchases/Waves/Products (DESIGN §6).'

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the Mastersheet .xlsx file.")
        parser.add_argument(
            "--user", required=True,
            help="Username that will own all imported Purchases.",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Parse and report, but write nothing to the database.",
        )

    def handle(self, *args, **options):
        try:
            workbook = openpyxl.load_workbook(options["path"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {options['path']}")
        if SHEET_NAME not in workbook.sheetnames:
            raise CommandError(f'The workbook has no "{SHEET_NAME}" sheet.')
        sheet = workbook[SHEET_NAME]

        User = get_user_model()
        try:
            user = User.objects.get(username=options["user"])
        except User.DoesNotExist:
            raise CommandError(f"User {options['user']!r} does not exist.")

        self.counts = {}
        self.skipped = []    # (row, title, reason) — nothing imported
        self.ambiguous = []  # (row, title, note) — imported with a caveat
        self._sizes = {}     # (width, height) Decimal pair -> CardSize
        self._pledge_managers = {}  # name -> PledgeManager

        rows = self._parse_rows(sheet)
        with transaction.atomic():
            self._import(rows, user)
            if options["dry_run"]:
                transaction.set_rollback(True)

        self._print_report(dry_run=options["dry_run"])

    def _bump(self, key, created=None):
        if created is not None:
            key = f"{key} {'created' if created else 'updated'}"
        self.counts[key] = self.counts.get(key, 0) + 1

    # --- parse: sheet rows -> plain dicts, grouped later --------------------

    def _parse_rows(self, sheet):
        rows = []
        for row in range(DATA_START_ROW, sheet.max_row + 1):
            def cell(col):
                return sheet.cell(row=row, column=col)

            purchase_raw = cell_text(cell(COL_PURCHASE))
            name = cell_text(cell(COL_NAME))
            if not purchase_raw and not name:
                continue  # entirely blank row
            if not purchase_raw:
                self.skipped.append((row, name, "no Purchase (column A) value — cannot group"))
                continue

            match = WAVE_SUFFIX_RE.search(purchase_raw)
            purchase_name = WAVE_SUFFIX_RE.sub("", purchase_raw).strip()
            purchase_name = PLATFORM_SUFFIX_RE.sub("", purchase_name).strip()
            rows.append({
                "row": row,
                "purchase": purchase_name,
                "wave": int(match.group(1)) if match else 1,
                "name": name,
                "bgg_url": cell_link(cell(COL_BGG)),
                "campaign_url": cell_link(cell(COL_CAMPAIGN)),
                "drive_url": cell_link(cell(COL_DRIVE)),
                "platform": cell_text(cell(COL_PLATFORM)),
                "kind": cell_text(cell(COL_KIND)),
                "status": cell_text(cell(COL_STATUS)),
                "ordered": as_date(cell(COL_ORDERED).value),
                "orig_eta": as_date(cell(COL_ORIG_ETA).value),
                "arrival": as_date(cell(COL_ARRIVAL).value),
                "pm": cell_text(cell(COL_PM)),
                "pm_url": cell_link(cell(COL_PM)),
                "pm_status": cell_text(cell(COL_PM_STATUS)),
                "pm_close": as_date(cell(COL_PM_CLOSE).value),
                "address": cell_text(cell(COL_ADDRESS)),
                "comments": cell_text(cell(COL_COMMENTS)),
                "excitement": cell(COL_EXCITEMENT).value,
                "miniatures": cell(COL_MINIATURES).value,
                "insert_3d": cell_text(cell(COL_INSERT_3D)),
                "contains_cards": cell_text(cell(COL_CONTAINS_CARDS)),
                "needs_sleeves": cell_text(cell(COL_NEEDS_SLEEVES)),
                "fits_sleeved": cell_text(cell(COL_FITS_SLEEVED)),
                "sizes": [
                    (col, name_, width, height, cell(col).value)
                    for col, name_, width, height in PREORDER_SIZES
                    if cell(col).value is not None
                ],
                "others": [
                    (num_col, cell(num_col).value, cell_text(cell(size_col)))
                    for num_col, size_col in OTHER_PAIRS
                    if cell(num_col).value is not None or cell_text(cell(size_col))
                ],
            })
        return rows

    # --- import --------------------------------------------------------------

    def _import(self, rows, user):
        purchases = {}  # purchase name -> [row dict, ...]
        for row in rows:
            purchases.setdefault(row["purchase"], []).append(row)

        for name, group in purchases.items():
            purchase = self._import_purchase(name, group, user)
            waves = {}
            for row in group:
                waves.setdefault(row["wave"], []).append(row)
            for number, wave_rows in sorted(waves.items()):
                wave = self._import_wave(purchase, number, wave_rows)
                for row in wave_rows:
                    self._import_product(wave, row, user)

    def _first(self, group, key, label):
        """First non-empty value of a purchase-level column; disagreements are
        reported once and the first value wins."""
        values = [row[key] for row in group if row[key] not in (None, "")]
        distinct = {str(v) for v in values}
        if len(distinct) > 1:
            first = group[0]
            self.ambiguous.append((
                first["row"], first["purchase"],
                f"rows disagree on {label} ({', '.join(sorted(distinct))}) — first value wins",
            ))
        return values[0] if values else None

    def _import_purchase(self, name, group, user):
        platform_raw = self._first(group, "platform", "platform")
        platform = PLATFORM_MAP.get(platform_raw)
        if platform is None:
            platform = Purchase.Platform.OTHER
            if platform_raw:
                self.ambiguous.append((
                    group[0]["row"], name, f"unrecognised platform {platform_raw!r} -> other",
                ))

        # Placeholder when every row is a $1 pledge / "Just support"; the
        # sheet has no watched-but-unbacked rows, so everything else was
        # committed to (DESIGN §6 lifecycle).
        placeholder = all(
            KIND_MAP.get(self._base_kind(row["kind"])) == Product.Kind.PLACEHOLDER_PLEDGE
            or row["status"] == "Just support"
            for row in group
        )

        pm_raw = self._first(group, "pm", "pledge manager") or ""
        pm_name = PM_MAP.get(pm_raw, "")
        if pm_raw and pm_raw not in PM_MAP:
            pm_name = "Other"
            self.ambiguous.append((
                group[0]["row"], name, f"unrecognised pledge manager {pm_raw!r} -> other",
            ))
        pm = self._get_pledge_manager(pm_name) if pm_name else None
        pm_status_raw = self._first(group, "pm_status", "PM status")
        pm_status = PM_STATUS_MAP.get(pm_status_raw, "")
        if pm_status_raw and not pm_status:
            self.ambiguous.append((
                group[0]["row"], name, f"unrecognised PM status {pm_status_raw!r} — left blank",
            ))

        excitement, excitement_note = None, ""
        raw_excitement = self._first(group, "excitement", "excitement")
        if isinstance(raw_excitement, (int, float)):
            excitement = Decimal(str(raw_excitement))
        elif raw_excitement:
            excitement_note = str(raw_excitement).strip()[:300]

        ordered_dates = [row["ordered"] for row in group if row["ordered"]]
        purchase, created = Purchase.objects.update_or_create(
            owner=user, name=name,
            defaults={
                "platform": platform,
                "status": (
                    Purchase.Status.PLACEHOLDER if placeholder else Purchase.Status.COMMITTED
                ),
                "campaign_url": self._first(group, "campaign_url", "campaign link") or "",
                "ordered_date": min(ordered_dates) if ordered_dates else None,
                "pledge_manager": pm,
                "pledge_manager_url": self._first(group, "pm_url", "PM link") or "",
                "pledge_manager_status": pm_status,
                "pledge_manager_close_date": self._first(group, "pm_close", "PM close date"),
                "excitement": excitement,
                "excitement_note": excitement_note,
            },
        )
        self._bump("purchases", created=created)
        return purchase

    @staticmethod
    def _base_kind(raw):
        return re.sub(r"-(show|hide)$", "", raw)

    def _import_wave(self, purchase, number, wave_rows):
        statuses = []
        for row in wave_rows:
            status = STATUS_MAP.get(row["status"])
            if status is None:
                status = Wave.Status.PENDING
                if row["status"]:
                    self.ambiguous.append((
                        row["row"], row["name"],
                        f"unrecognised status {row['status']!r} -> pending",
                    ))
            statuses.append(status)
        if len(set(statuses)) > 1:
            self.ambiguous.append((
                wave_rows[0]["row"], purchase.name,
                f"wave {number} rows disagree on status ({sorted(set(statuses))}) "
                "— least-progressed wins",
            ))
        status = min(statuses, key=STATUS_PRIORITY.index)

        orig_etas = [row["orig_eta"] for row in wave_rows if row["orig_eta"]]
        arrivals = [row["arrival"] for row in wave_rows if row["arrival"]]
        arrival = max(arrivals) if arrivals else None  # a wave lands with its last product
        if len(set(arrivals)) > 1:
            self.ambiguous.append((
                wave_rows[0]["row"], purchase.name,
                f"wave {number} rows disagree on arrival date — latest wins",
            ))

        digital = all(self._is_digital_row(row) for row in wave_rows)
        wave, created = Wave.objects.update_or_create(
            purchase=purchase, number=number,
            defaults={
                "delivery_type": (
                    Wave.DeliveryType.DIGITAL if digital else Wave.DeliveryType.PHYSICAL
                ),
                "status": status,
                "original_eta": min(orig_etas) if orig_etas else None,
                "expected_arrival": None if status == Wave.Status.ARRIVED else arrival,
                "arrived_date": arrival if status == Wave.Status.ARRIVED else None,
                "address": self._first(wave_rows, "address", "address") or "",
            },
        )
        self._bump("waves", created=created)
        return wave

    def _is_digital_row(self, row):
        kind = KIND_MAP.get(self._base_kind(row["kind"]))
        return kind == Product.Kind.PNP_GAME or bool(DIGITAL_NAME_RE.search(row["name"]))

    def _import_product(self, wave, row, user):
        kind_raw = self._base_kind(row["kind"])
        kind = KIND_MAP.get(kind_raw)
        if kind is None:
            kind = Product.Kind.OTHER
            if kind_raw:
                self.ambiguous.append((
                    row["row"], row["name"], f"unrecognised type {kind_raw!r} -> other",
                ))

        game, copy, edition, bgg_url = None, None, None, ""
        bgg_id = extract_bgg_id(row["bgg_url"])
        if kind in Product.GAME_KINDS:
            game, copy = self._match_game(row, kind, wave, user, bgg_id)
            if game is not None and kind == Product.Kind.PNP_GAME:
                # PnP is an edition-level property (#138). A fully-PnP campaign's
                # only wave is digital and may never physically "arrive", so
                # record the PnP edition now and link the product to it, rather
                # than leaving a bare is_pnp on the title.
                edition, created = Edition.objects.get_or_create(
                    game=game, is_default=True, defaults={"name": ""},
                )
                if created:
                    self._bump("PnP editions created")
                if not edition.is_pnp:
                    edition.is_pnp = True
                    edition.save(update_fields=["is_pnp"])
        elif row["bgg_url"]:
            # On BGG but not modelled as a Game (gamebooks, accessories...).
            bgg_url = row["bgg_url"]

        miniatures, notes = row["miniatures"], row["comments"].strip()
        miniatures_count = as_count(miniatures)
        if miniatures is not None and miniatures_count is None:
            if str(miniatures).strip().isdigit():
                miniatures_count = int(str(miniatures).strip())
            else:
                notes = f"{notes}\nminiatures: {miniatures}".strip()
                self.ambiguous.append((
                    row["row"], row["name"],
                    f"non-numeric miniatures value {miniatures!r} kept in notes",
                ))

        for field, label in (("contains_cards", "contains cards"),
                             ("needs_sleeves", "needs sleeves")):
            if row[field] and row[field] not in TRISTATE_MAP:
                self.ambiguous.append((
                    row["row"], row["name"],
                    f"unrecognised {label} value {row[field]!r} — left blank",
                ))

        product, created = Product.objects.update_or_create(
            wave=wave, name=row["name"],
            defaults={
                "kind": kind,
                "game": game,
                "edition": edition,
                "copy": copy,
                "bgg_url": bgg_url or "",
                "drive_url": row["drive_url"] or "",
                "contains_cards": TRISTATE_MAP.get(row["contains_cards"], ""),
                "needs_sleeves": TRISTATE_MAP.get(row["needs_sleeves"], ""),
                "fits_sleeved_note": row["fits_sleeved"][:100],
                "miniatures_count": miniatures_count,
                "insert_3d_note": row["insert_3d"][:300],
                "notes": notes,
            },
        )
        self._bump("products", created=created)
        self._import_sleeve_needs(product, row)

    def _match_game(self, row, kind, wave, user, bgg_id):
        """Resolve a game-kind row to a Game via its BGG id (never by title),
        creating Game+BggLink for unknown ids; and to the owner's Copy for
        arrived products. Returns (game, copy)."""
        if not bgg_id:
            self.ambiguous.append((
                row["row"], row["name"],
                "game-kind row without a usable BGG link — product not linked to a Game",
            ))
            return None, None

        link = BggLink.objects.filter(bgg_id=bgg_id, is_primary=True).select_related("game").first()
        if link:
            game = link.game
        else:
            name = PNP_NAME_RE.sub("", MULTIPLIER_RE.sub("", row["name"])).strip()
            game = Game.objects.create(
                name=name,
                type=(
                    Game.Type.EXPANSION
                    if kind == Product.Kind.EXPANSION
                    or (row["bgg_url"] and "boardgameexpansion" in row["bgg_url"])
                    else Game.Type.BASE
                ),
            )
            BggLink.objects.create(game=game, bgg_id=bgg_id, is_primary=True)
            self._bump("games created")

        copy = None
        if wave.status == Wave.Status.ARRIVED:
            copies = list(Copy.objects.filter(owner=user, edition__game=game))
            if len(copies) == 1:
                copy = copies[0]
            elif len(copies) > 1:
                self.ambiguous.append((
                    row["row"], row["name"],
                    f"{len(copies)} copies of {game.name!r} owned — product not linked to a Copy",
                ))
        return game, copy

    def _import_sleeve_needs(self, product, row):
        batches = {}  # CardSize -> count
        for col, name, width, height, value in row["sizes"]:
            count = as_count(value)
            if count is None:
                self.ambiguous.append((
                    row["row"], row["name"],
                    f"non-numeric count {value!r} for {name} — skipped",
                ))
                continue
            size = self._get_or_create_size(
                (Decimal(width).quantize(Decimal("0.1")),
                 Decimal(height).quantize(Decimal("0.1"))),
                name,
            )
            batches[size] = batches.get(size, 0) + count

        for num_col, raw_count, size_raw in row["others"]:
            dims = parse_dimensions(size_raw) if size_raw else None
            count = as_count(raw_count)
            if not dims or count is None:
                self.ambiguous.append((
                    row["row"], row["name"],
                    f"Other pair not importable (num={raw_count!r}, size={size_raw!r})",
                ))
                continue
            size = self._get_or_create_size(dims, "")
            batches[size] = batches.get(size, 0) + count

        for size, count in batches.items():
            _, created = ProductSleeveRequirement.objects.update_or_create(
                product=product, card_size=size, defaults={"count": count},
            )
            self._bump("sleeve needs", created=created)

    def _get_or_create_size(self, dims, name):
        """CardSizes are keyed by dimensions (DESIGN §5); after the sleeves
        import all fixed-column sizes exist, so this mostly resolves. New dims
        (free-text Other sizes) are created nameless — naming them is manual."""
        if dims not in self._sizes:
            size, created = CardSize.objects.get_or_create(
                width_mm=dims[0], height_mm=dims[1], defaults={"name": name},
            )
            if created:
                self._bump("card sizes created")
            self._sizes[dims] = size
        return self._sizes[dims]

    def _get_pledge_manager(self, name):
        if name not in self._pledge_managers:
            self._pledge_managers[name] = PledgeManager.objects.get(name=name)
        return self._pledge_managers[name]

    # --- report ---------------------------------------------------------------

    def _print_report(self, dry_run):
        write = self.stdout.write
        if dry_run:
            write(self.style.WARNING("DRY RUN — nothing was written to the database.\n"))

        write(self.style.MIGRATE_HEADING("Summary"))
        for key in sorted(self.counts):
            write(f"  {key}: {self.counts[key]}")

        write(self.style.MIGRATE_HEADING("Skipped rows"))
        if self.skipped:
            for row, title, reason in self.skipped:
                write(f"  row {row} ({title!r}): {reason}")
        else:
            write("  none")

        write(self.style.MIGRATE_HEADING("Ambiguous (imported with caveats)"))
        if self.ambiguous:
            for row, title, note in self.ambiguous:
                write(f"  row {row} ({title!r}): {note}")
        else:
            write("  none")

        write(self.style.MIGRATE_HEADING("Interpretation decisions"))
        for note in DECISION_NOTES:
            write(f"  - {note}")

        write(self.style.MIGRATE_HEADING("Deferred (not imported)"))
        for note in DEFERRED_NOTES:
            write(f"  - {note}")

        write(self.style.SUCCESS("Done." if not dry_run else "Dry run complete."))
